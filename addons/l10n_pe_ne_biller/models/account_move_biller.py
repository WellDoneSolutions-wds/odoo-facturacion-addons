import base64
import io
import json
import logging
import re
import zipfile
from datetime import timedelta

import pytz
import requests

try:  # SQS para el modo asíncrono (l10n_pe_ne_biller.async_enabled); si falta
    import boto3  # boto3, el modo síncrono sigue funcionando igual.
except ImportError:  # pragma: no cover
    boto3 = None

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools import float_round

from ..tools.amount_to_words import leyenda_monto

_logger = logging.getLogger(__name__)

# Cache de clientes boto3 por (service, region) — ver _l10n_pe_boto_client.
# Guarda (módulo_boto3, cliente) para invalidarse solo si boto3 fue parcheado.
_BOTO_CLIENTS = {}

# Descuento global que NO afecta la base imponible del IGV (AllowanceChargeReasonCode, cat. 53).
# CONFIRMADO contra el validador SUNAT (ValidaExprRegFactura-2.0.1.xsl) y beta (spike 2026-07-21):
# el código "03" es el que el validador cuenta en `descuentosGlobalesNOAfectaBI` (línea 335) y NO
# resta de la base del IGV; el "02" cae en `MontoDescuentoAfectoBI` (SÍ afecta la base) → daba
# error 3291 (esperaba el IGV sobre la base ya descontada). Con "03" el IGV queda sobre el precio
# lleno y baja solo el MtoImpVenta.
DESC_GLOBAL_NO_AFECTA_COD = "03"

# Catálogo SUNAT de descripciones de motivo para Nota de Débito (08).
ND_MOTIVO_DESC = {
    "01": "Intereses por mora",
    "02": "Aumento en el valor",
    "03": "Penalidades/otros conceptos",
    "11": "Ajustes de operaciones de exportación",
    "12": "Ajustes afectos al IVAP",
}

# Mapeo del código de tributo de la tax de Odoo (`account.tax.l10n_pe_edi_tax_code`, cat. 05 de
# SUNAT, provisto por la localización community l10n_pe) a la tupla que exige el contrato SFS por
# línea/tributo:  (tipAfeIGV cat.07, codTriIGV cat.05, nomTributo, codTipTributo UN/ECE 5153,
# codCatTributo UN/ECE 5305). Valores tomados de los mocks ya aceptados E2E por el microservicio.
TAX_CODE_MAP = {
    "1000": ("10", "1000", "IGV", "VAT", "S"),  # Gravado - operación onerosa
    "1016": ("17", "1016", "IVAP", "VAT", "S"),  # IVAP (arroz pilado)
    "9997": ("20", "9997", "EXO", "VAT", "E"),  # Exonerado
    "9998": ("30", "9998", "INA", "FRE", "O"),  # Inafecto
    "9995": ("40", "9995", "EXP", "FRE", "G"),  # Exportación
    "9996": (
        "11",
        "9996",
        "GRA",
        "FRE",
        "Z",
    ),  # Gratuita (retiro/transferencia gratuita)
}
DEFAULT_TAX_CODE = "1000"  # Sin tax reconocida -> gravado IGV (caso más común).

# Código de unidad de medida de SUNAT (cat. 03 / UN-ECE Rec. 20) por XMLID de la unidad estándar
# de Odoo. Mapeo replicado de l10n_pe_edi (enterprise). Se resuelve en runtime porque las UoM base
# son `noupdate` y un data file de otro módulo no las actualiza; para unidades personalizadas, el
# usuario fija el override en `uom.uom.l10n_pe_ne_unit_code`.
UOM_CODE_BY_XMLID = {
    "uom.product_uom_unit": "NIU",
    "uom.product_uom_dozen": "DPC",
    "uom.product_uom_kgm": "KGM",
    "uom.product_uom_gram": "GRM",
    "uom.product_uom_day": "DAY",
    "uom.product_uom_hour": "HUR",
    "uom.product_uom_ton": "TNE",
    "uom.product_uom_meter": "MTR",
    "uom.product_uom_km": "KTM",
    "uom.product_uom_cm": "CMT",
    "uom.product_uom_litre": "LTR",
    "uom.product_uom_lb": "LBR",
    "uom.product_uom_oz": "ONZ",
    "uom.product_uom_inch": "INH",
    "uom.product_uom_foot": "FOT",
    "uom.product_uom_mile": "M52",
    "uom.product_uom_floz": "OZ",
    "uom.product_uom_qt": "QTI",
    "uom.product_uom_gal": "GLL",
}
DEFAULT_UNIT_CODE = "NIU"

# Importación de productos por Excel: mapeo tolerante de TEXTO en español (o el propio código
# cat.03) → código SUNAT cat.03. Clave = texto normalizado (minúsculas, sin tildes). Espejo del
# catálogo del front (lib/unidades.ts) más sinónimos comunes de ferretería/bodega.
UNIDAD_IMPORT = {
    "unidad": "NIU", "unidades": "NIU", "und": "NIU", "unid": "NIU", "niu": "NIU", "u": "NIU",
    "servicio": "ZZ", "servicios": "ZZ", "serv": "ZZ", "zz": "ZZ",
    "kilogramo": "KGM", "kilogramos": "KGM", "kilo": "KGM", "kilos": "KGM", "kg": "KGM", "kgm": "KGM",
    "gramo": "GRM", "gramos": "GRM", "gr": "GRM", "grm": "GRM",
    "libra": "LBR", "libras": "LBR", "lb": "LBR", "lbr": "LBR",
    "tonelada": "TNE", "toneladas": "TNE", "tonelada metrica": "TNE", "ton": "TNE", "tne": "TNE",
    "litro": "LTR", "litros": "LTR", "lt": "LTR", "ltr": "LTR",
    "galon": "GLL", "galones": "GLL", "gln": "GLL", "gll": "GLL",
    "barril": "BLL", "barriles": "BLL", "bll": "BLL",
    "lata": "CA", "latas": "CA", "ca": "CA",
    "caja": "BX", "cajas": "BX", "bx": "BX",
    "millar": "MLL", "millares": "MLL", "mll": "MLL",
    "metro": "MTR", "metros": "MTR", "mt": "MTR", "mtr": "MTR", "m": "MTR",
    "centimetro": "CMT", "centimetros": "CMT", "cm": "CMT", "cmt": "CMT",
    "metro cuadrado": "MTK", "m2": "MTK", "mtk": "MTK",
    "metro cubico": "MTQ", "m3": "MTQ", "mtq": "MTQ",
    "dia": "DAY", "dias": "DAY", "day": "DAY",
    "hora": "HUR", "horas": "HUR", "hr": "HUR", "hur": "HUR",
    "juego": "SET", "juegos": "SET", "set": "SET",
    "docena": "DPC", "docenas": "DPC", "dpc": "DPC",
    "onza": "ONZ", "onzas": "ONZ", "onz": "ONZ",
}
# Afectación IGV: texto (cat.07 humano) → código cat.07 que espera el producto.
AFECT_IMPORT = {
    "gravado": "1000", "gravada": "1000",
    "exonerado": "9997", "exonerada": "9997",
    "inafecto": "9998", "inafecta": "9998",
    "exportacion": "9995",
    "gratuito": "9996", "gratuita": "9996",
}
# Códigos cat.03 válidos (para aceptar el código directo en el Excel, ej. "KGM").
_UNIDAD_CODES = set(UNIDAD_IMPORT.values())


def _percep_float(v):
    """float() de percepTasa que no revienta con un 500 críptico: tolera coma decimal (igual
    que el import masivo) y, si no es numérico, da un UserError legible en vez de un
    ValueError sin traducir. Vacío/None/False → 0.0 (limpia el campo, mismo criterio de
    siempre)."""
    if v in (None, "", False):
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        raise UserError(_("La percepción sugerida debe ser un número (ej. 2 o 1.5)."))


class AccountMoveLine(models.Model):
    _inherit = "account.move.line"

    # Cantidad con 3 decimales (SUNAT admite hasta 10 en ctdUnidadItem). Por defecto la precisión
    # de UoM de Odoo es 2 y truncaba la venta al peso de balanza (18.375 kg -> 18.38). Ver QA-020.
    quantity = fields.Float(digits=(16, 3))

    # Overrides SUNAT por línea para el flujo rápido (sin depender de la UoM/producto de Odoo,
    # evitando el problema de categorías de unidad de medida).
    l10n_pe_ne_unit_code = fields.Char(
        string="Unidad SUNAT (cat.03)",
        copy=False,
        help="Código de unidad de medida SUNAT de la línea (ej. NIU, KGM, ZZ). "
        "Si está vacío se deriva de la unidad de medida del producto.",
    )
    l10n_pe_ne_cod_producto_sunat = fields.Char(
        string="Cód. producto SUNAT (cat.25)",
        copy=False,
        help="Código de producto SUNAT (UNSPSC, catálogo 25) de la línea, si aplica.",
    )
    # Lote/serie de una línea de COMPRA. El lote entra con la mercadería, así que se captura
    # en la compra y viaja con la línea hasta que _l10n_pe_ne_mover_stock_compra crea el
    # movimiento. En la VENTA no se pide: Odoo reserva y asigna el lote solo, por su
    # estrategia de salida (lo que vence antes sale primero). Verificado.
    l10n_pe_ne_lote = fields.Char(
        string="Lote / serie",
        copy=False,
        help="Número de lote o serie de la mercadería que ingresa por esta línea.",
    )
    l10n_pe_ne_vence = fields.Date(
        string="Vencimiento del lote",
        copy=False,
        help="Fecha de vencimiento del lote que ingresa por esta línea.",
    )
    # Sub-tipo de operación gratuita (cat. 07 SUNAT). Solo aplica a líneas gratuitas (9996):
    # afina el genérico "11" al motivo real (retiro, bonificación, donación…). Vacío = 11.
    l10n_pe_ne_afectacion_gratuita = fields.Selection(
        [
            ("11", "Retiro por premio"),
            ("12", "Retiro por donación"),
            ("13", "Retiro de bienes"),
            ("14", "Retiro por publicidad"),
            ("15", "Bonificación"),
            ("16", "Retiro por entrega a trabajadores"),
        ],
        string="Tipo de operación gratuita",
        copy=False,
        help="Solo para líneas gratuitas: precisa el motivo (catálogo 07 de SUNAT). "
        "Si se deja vacío se usa 'Retiro por premio' (11).",
    )


class AccountMove(models.Model):
    _inherit = "account.move"

    l10n_pe_biller_state = fields.Selection(
        selection=[
            ("por_enviar", "Por enviar"),
            ("en_proceso", "En proceso"),
            ("enviado", "Enviado"),
            ("anulado", "Anulado"),
            ("rechazado", "Rechazado"),
            ("error", "Error"),
        ],
        string="Estado Facturador",
        default="por_enviar",
        copy=False,
        tracking=True,  # cambios visibles en el chatter (que sí refresca en vivo)
    )
    l10n_pe_ne_baja_motivo = fields.Char(
        string="Motivo de baja",
        copy=False,
        help="Motivo de la comunicación de baja (RA) del comprobante ante SUNAT.",
    )
    l10n_pe_ne_baja_correlativo = fields.Char(
        string="Correlativo RA", copy=False, readonly=True
    )
    l10n_pe_ne_baja_fecha = fields.Date(string="Fecha RA", copy=False, readonly=True)
    l10n_pe_ne_baja_doc = fields.Char(
        string="Comunicación de baja",
        compute="_compute_l10n_pe_ne_baja_doc",
        store=True,
        help="Identificador de la comunicación de baja: RA-AAAAMMDD-correlativo.",
    )
    l10n_pe_ne_baja_cdr = fields.Many2one(
        "ir.attachment", string="CDR de baja", copy=False
    )
    # Identidad realmente emitida, congelada al enviar: la baja debe referenciar EXACTAMENTE lo emitido,
    # no recomputar del partner/nombre (que podrían cambiar y anular el comprobante equivocado).
    l10n_pe_ne_tipo_doc = fields.Char(
        string="Tipo doc. emitido", copy=False, readonly=True
    )
    l10n_pe_ne_serie_emit = fields.Char(
        string="Serie emitida", copy=False, readonly=True
    )
    l10n_pe_ne_corr_emit = fields.Char(
        string="Correlativo emitido", copy=False, readonly=True
    )
    l10n_pe_serie = fields.Char(
        string="Serie",
        compute="_compute_l10n_pe_serie",
        store=True,
        readonly=False,
        copy=False,
        help="Serie del comprobante. Por defecto, la del diario (l10n_pe_ne_serie) con la "
        "letra ajustada a la familia del comprobante (F factura / B boleta).",
    )
    l10n_pe_correlativo = fields.Char(
        string="Correlativo",
        copy=False,
        help="Correlativo del comprobante. Si se deja vacío, se auto-incrementa del número del "
        "asiento (folio gestionado por Odoo por diario).",
    )

    l10n_pe_ne_detraccion = fields.Boolean(string="Sujeto a detracción", copy=False)
    l10n_pe_ne_detraccion_code = fields.Char(
        string="Código detracción",
        copy=False,
        help="Código del bien/servicio sujeto a detracción (catálogo 54 de SUNAT, ej. 037).",
    )
    l10n_pe_ne_detraccion_rate = fields.Float(
        string="% Detracción", digits=(5, 2), copy=False
    )
    l10n_pe_ne_detraccion_medio_pago = fields.Char(
        string="Medio de pago detracción",
        default="001",
        copy=False,
        help="Catálogo 59 (001 = depósito en cuenta del Banco de la Nación).",
    )
    l10n_pe_ne_detraccion_cuenta = fields.Char(
        string="Cuenta de detracción (Banco de la Nación)",
        copy=False,
        help="Cuenta del Banco de la Nación de ESTE comprobante. Si se deja vacía, "
        "se usa la cuenta de detracciones configurada en la empresa.",
    )
    l10n_pe_ne_percepcion = fields.Boolean(string="Aplica percepción", copy=False)
    l10n_pe_ne_percepcion_rate = fields.Float(
        string="% Percepción", digits=(5, 2), default=2.0, copy=False
    )
    l10n_pe_ne_anticipos = fields.Json(
        string="Anticipos regularizados",
        copy=False,
        help="Lista de anticipos (doc. A) que ESTA venta final regulariza/deduce. Cada elemento: "
        "{doc: serie-correlativo, monto: importe con IGV, tipo: '02'/'03' (cat. 12), "
        "origenId: id del anticipo local enlazado o null si se emitió por fuera}. SUNAT permite "
        "varios (pagos escalonados): se emiten como N documentos relacionados con numIdeAnticipo 1..N.",
    )
    l10n_pe_ne_es_anticipo = fields.Boolean(
        string="Es pago anticipado",
        copy=False,
        help="Marca que ESTE comprobante se emite por un pago anticipado (doc. A del ciclo de "
        "anticipos, equivale al 'Sí' de SEE-SOL). Es una venta interna normal (0101) por el monto "
        "anticipado; su descripción lleva 'PAGO ANTICIPADO' y queda disponible para regularizarse "
        "luego en la factura de venta final. No confundir con el anticipo aplicado, que descuenta "
        "un anticipo ya emitido (doc. B).",
    )
    l10n_pe_ne_anticipo_aplicado = fields.Monetary(
        string="Anticipo ya aplicado",
        compute="_compute_l10n_pe_ne_anticipo_saldo",
        help="Suma de las regularizaciones vivas que apuntan a este anticipo (doc. A).",
    )
    l10n_pe_ne_anticipo_saldo = fields.Monetary(
        string="Saldo del anticipo",
        compute="_compute_l10n_pe_ne_anticipo_saldo",
        help="Importe del anticipo aún disponible para regularizar = total − aplicado. Solo "
        "aplica a comprobantes marcados como pago anticipado (doc. A).",
    )
    l10n_pe_ne_desc_no_afecta = fields.Monetary(
        string="Descuento que no afecta el IGV",
        copy=False,
        help="Descuento (S/, CON IGV incluido en el sentido de que baja el total a pagar) que NO "
        "afecta la base imponible del IGV (cat. 53 'no afecta'): la gravada y el IGV se calculan "
        "sobre el precio lleno; este importe solo reduce el total (MtoImpVenta). Agrega el "
        "descuento por ítem 'no afecta' y el descuento global 'no afecta' del comprobante.",
    )

    @api.depends("l10n_pe_ne_es_anticipo", "amount_total")
    def _compute_l10n_pe_ne_anticipo_saldo(self):
        """Saldo por anticipo (doc. A) = total − regularizaciones vivas que lo aplican. Mismo
        criterio de 'vivo' que las NC (posteadas y no rechazadas/anuladas/con error; las en cola
        cuentan, para que dos regularizaciones simultáneas no consuman más que el total).
        El dato vive en una lista JSON (`l10n_pe_ne_anticipos`): no se puede agrupar por
        contenido JSON con `_read_group`, así que se busca las regularizaciones vivas y se
        agrega en Python sumando el `monto` de cada anticipo cuyo `origenId` matchee."""
        anticipos = self.filtered("l10n_pe_ne_es_anticipo")
        aplicado = {a.id: 0.0 for a in anticipos}
        if anticipos:
            # Nota de escala: search + loop Python sobre TODAS las regularizaciones vivas del
            # sistema; para volúmenes altos se podría filtrar por partner/JSONB, pero YAGNI.
            regs = self.env["account.move"].search([
                ("l10n_pe_ne_anticipos", "!=", False),
                ("state", "=", "posted"),
                ("l10n_pe_biller_state", "not in", ("rechazado", "error", "anulado")),
            ])
            for reg in regs:
                for a in reg._l10n_pe_ne_anticipos_list():
                    oid = a["origenId"]
                    if oid in aplicado:
                        aplicado[oid] += a["monto"]
        for move in self:
            ap = round(aplicado.get(move.id, 0.0), 2) if move.l10n_pe_ne_es_anticipo else 0.0
            move.l10n_pe_ne_anticipo_aplicado = ap
            move.l10n_pe_ne_anticipo_saldo = (
                round(move.amount_total - ap, 2) if move.l10n_pe_ne_es_anticipo else 0.0
            )

    def l10n_pe_ne_anticipos_pendientes(self, ruc=None, partner_id=None, moneda=None):
        """Anticipos (doc. A) ACEPTADOS por SUNAT y con saldo pendiente, para autocompletar la
        regularización en la venta final. Filtra por cliente (id o RUC/DNI) y, opcionalmente,
        moneda. Devuelve serie-correlativo, tipo (cat. 12: 02 factura / 03 boleta) y saldo."""
        domain = [
            ("l10n_pe_ne_es_anticipo", "=", True),
            ("l10n_pe_biller_state", "=", "enviado"),
            ("move_type", "=", "out_invoice"),
        ]
        if partner_id:
            domain.append(("partner_id", "=", int(partner_id)))
        elif ruc:
            domain.append(("partner_id.vat", "=", (ruc or "").strip()))
        if moneda:
            domain.append(("currency_id.name", "=", moneda))
        out = []
        for m in self.search(domain, order="id desc", limit=100):
            if round(m.l10n_pe_ne_anticipo_saldo, 2) <= 0:
                continue
            out.append(
                {
                    "id": m.id,
                    "doc": "%s-%s" % (m.l10n_pe_ne_serie_emit or "", m.l10n_pe_ne_corr_emit or ""),
                    "tipo": "02" if (m.l10n_pe_ne_tipo_doc or "01") == "01" else "03",
                    "total": m.amount_total,
                    "aplicado": m.l10n_pe_ne_anticipo_aplicado,
                    "saldo": m.l10n_pe_ne_anticipo_saldo,
                    "moneda": m.currency_id.name or "PEN",
                    "cliente": m.partner_id.name or "",
                    "fechaEmision": m.invoice_date.strftime("%Y-%m-%d") if m.invoice_date else "",
                }
            )
        return out

    def _l10n_pe_importe_cobrar(self):
        """Importe neto a cobrar = total − anticipo aplicado − descuento que no afecta el IGV −
        bienes gratuitos (lo que el cliente paga)."""
        self.ensure_one()
        ant = self._l10n_pe_anticipo()
        return round(
            self.amount_total
            - (ant[2] if ant else 0.0)
            - self._l10n_pe_desc_no_afecta()
            - self._l10n_pe_gratuito_base(),
            2,
        )

    def _l10n_pe_percepcion_monto(self):
        self.ensure_one()
        # La percepción se calcula sobre el neto a cobrar (descontado el anticipo): así la base de la
        # percepción nunca supera el PayableAmount (evita rechazo SUNAT 2797 al combinar con anticipo).
        return round(
            self._l10n_pe_importe_cobrar()
            * (self.l10n_pe_ne_percepcion_rate or 0.0)
            / 100.0,
            2,
        )

    def _l10n_pe_anticipo_gravado(self):
        """Categoría y tasa de la operación gravada del anticipo: (cod_tri, tasa, motivo_no_soportado).
        El anticipo solo es representable como descuento global código 04 sobre una operación gravada
        homogénea (un único régimen IGV '1000' o IVAP '1016'). Devuelve un motivo si no lo es."""
        self.ensure_one()
        lines = self._l10n_pe_product_lines()
        codes = {self._l10n_pe_tax_info(l)[0][1] for l in lines}
        gravado = codes & {"1000", "1016"}
        if not gravado:
            return None, 0.0, "no_gravado"
        if len(gravado) > 1 or (codes - {"1000", "1016"}):
            return None, 0.0, "mixto"
        cod_tri = next(iter(gravado))
        tasa = next(
            (
                self._l10n_pe_tax_info(l)[1]
                for l in lines
                if self._l10n_pe_tax_info(l)[0][1] == cod_tri
            ),
            18.0,
        )
        return cod_tri, tasa, None

    def _l10n_pe_ne_anticipos_list(self):
        """Lista normalizada de anticipos de esta factura: [{doc, monto, tipo, origenId}]. Vacía
        si no aplica (no out_invoice, NC/ND o sin anticipos).
        `origenId` se coerciona con seguridad (nunca explota con basura tipo "abc"): esta lista
        la recorre también `_compute_l10n_pe_ne_anticipo_saldo` para TODAS las regularizaciones
        vivas del sistema, así que una fila envenenada en OTRA factura no debe romper el
        saldo/pendientes de todas las demás — se trata como sin origen local (None)."""
        self.ensure_one()
        if self.move_type != "out_invoice" or self.debit_origin_id:
            return []
        out = []
        for a in (self.l10n_pe_ne_anticipos or []):
            monto = round(float(a.get("monto") or 0.0), 2)
            if monto <= 0:
                continue
            origen_raw = a.get("origenId")
            try:
                origen_id = int(origen_raw) if origen_raw not in (None, "", False) else None
            except (TypeError, ValueError):
                origen_id = None
            out.append({
                "doc": (a.get("doc") or "").strip(),
                "monto": monto,
                "tipo": a.get("tipo") or "02",
                "origenId": origen_id,
            })
        return out

    def _l10n_pe_anticipos_montos(self):
        """(valor, igv, total) AGREGADO de los anticipos: (0.0, 0.0, 0.0) si no aplica (no
        out_invoice, NC/ND o sin anticipos). El valor de CADA anticipo se separa con la tasa
        gravada homogénea real de la factura (no asume 18%) y el agregado es la SUMA de los
        valores/igv por anticipo (no el total dividido una sola vez), así que con montos que no
        son fracción redonda de la base el agregado no arrastra un desvío de redondeo distinto al
        que vería cada `AdditionalDocumentReference` individual — de ahí que el loop por ítem se
        mantenga aunque solo se devuelva el agregado (nadie más consume el desglose por ítem)."""
        self.ensure_one()
        lst = self._l10n_pe_ne_anticipos_list()
        if not lst:
            return (0.0, 0.0, 0.0)
        _cod, tasa, _m = self._l10n_pe_anticipo_gravado()
        vt = it = tt = 0.0
        for a in lst:
            total = round(a["monto"], 2)
            valor = round(total / (1.0 + (tasa or 0.0) / 100.0), 2)
            igv = round(total - valor, 2)
            vt += valor
            it += igv
            tt += total
        return round(vt, 2), round(it, 2), round(tt, 2)

    def _l10n_pe_anticipo(self):
        """(valor, igv, total) AGREGADO de los anticipos, o None si no hay ninguno. Wrapper de
        `_l10n_pe_anticipos_montos()` para no romper a los llamadores previos (percepción, importe
        a cobrar, cabecera, tributos, variable global 04) que solo necesitan el agregado."""
        self.ensure_one()
        v, i, t = self._l10n_pe_anticipos_montos()
        return (v, i, t) if t > 0 else None

    def _l10n_pe_desc_no_afecta(self):
        """Monto del descuento global que NO afecta la base del IGV, topeado para no dejar el total
        negativo. Es un ajuste SOLO de emisión (como el anticipo): no agrega una línea a Odoo, así
        que la gravada/IGV quedan sobre el precio lleno; solo baja el MtoImpVenta (total a pagar).
        Devuelve 0.0 si no aplica (notas, o sin descuento)."""
        self.ensure_one()
        monto = self.l10n_pe_ne_desc_no_afecta or 0.0
        if monto <= 0 or self.move_type not in ("out_invoice",) or self.debit_origin_id:
            return 0.0
        # Tope: no puede superar el total menos lo ya deducido por anticipo (dejaría MtoImpVenta < 0).
        ant = self._l10n_pe_anticipo()
        tope = round((self.amount_total or 0.0) - (ant[2] if ant else 0.0), 2)
        return round(min(monto, max(0.0, tope)), 2)

    def _l10n_pe_check_anticipo(self):
        """Valida que los anticipos sean representables (N documentos relacionados + un descuento
        global código 04 AGREGADO) antes de emitir el XML. Rechaza con un mensaje claro los casos no
        soportados, en vez de generar un comprobante inválido. Cada anticipo de la lista se valida
        individualmente (doc, origen, partner, moneda y saldo propio); la SUMA se valida contra el
        total de la factura."""
        self.ensure_one()
        if self.l10n_pe_ne_es_anticipo and self._l10n_pe_anticipo():
            raise UserError(
                _(
                    "Un comprobante que se emite por un pago anticipado no puede a la vez regularizar "
                    "otro anticipo. Desmarque una de las dos opciones."
                )
            )
        if not self._l10n_pe_anticipo():
            return
        lst = self._l10n_pe_ne_anticipos_list()
        total_aplicado = round(sum(a["monto"] for a in lst), 2)
        if total_aplicado > self.amount_total + 0.01:
            raise UserError(
                _(
                    "El total de anticipos (%.2f) no puede exceder el total de la factura (%.2f)."
                )
                % (total_aplicado, self.amount_total)
            )
        # Otras regularizaciones vivas que ya consumen anticipos (excluye esta factura). El dato vive
        # en la lista JSON: se busca ampliamente UNA sola vez y se filtra/agrega en Python por origen
        # (mismo patrón que `_compute_l10n_pe_ne_anticipo_saldo`).
        otras = self.env["account.move"].search([
            ("id", "!=", self.id),
            ("l10n_pe_ne_anticipos", "!=", False),
            ("state", "=", "posted"),
            ("l10n_pe_biller_state", "not in", ("rechazado", "error", "anulado")),
        ])
        aplicado_en_esta = {}  # origenId -> suma de monto ya visto en ESTA factura (mismo origen 2x).
        for idx, a in enumerate(lst, start=1):
            if a["tipo"] not in ("02", "03"):
                raise UserError(
                    _(
                        "Tipo de documento de anticipo inválido (debe ser 02 factura o 03 boleta)."
                    )
                )
            if not a["doc"]:
                raise UserError(
                    _(
                        "Indique el comprobante del anticipo #%d (serie-correlativo, ej. F001-00000100)."
                    )
                    % idx
                )
            # Si la regularización enlaza un anticipo local (doc. A), valida moneda y saldo
            # disponible: el importe aplicado no puede exceder lo que le queda al anticipo (evita
            # doble consumo), sea desde otra factura o desde otra línea de esta misma lista.
            origen = (
                self.env["account.move"].browse(a["origenId"])
                if a["origenId"]
                else self.env["account.move"]
            )
            if not origen:
                continue
            if not origen.l10n_pe_ne_es_anticipo:
                raise UserError(
                    _("El documento enlazado (%s) no está marcado como pago anticipado.")
                    % origen.display_name
                )
            if origen.partner_id != self.partner_id:
                raise UserError(
                    _("El anticipo (%s) pertenece a otro cliente: solo puede regularizarlo su titular.")
                    % origen.display_name
                )
            if origen.currency_id != self.currency_id:
                raise UserError(
                    _(
                        "El anticipo (%s) y la factura deben estar en la misma moneda: "
                        "regularice el anticipo en un comprobante de su misma moneda."
                    )
                    % origen.display_name
                )
            aplicado_otras = round(
                sum(
                    oa["monto"]
                    for m in otras
                    for oa in m._l10n_pe_ne_anticipos_list()
                    if oa["origenId"] == origen.id
                ),
                2,
            )
            aplicado_en_esta[origen.id] = round(
                aplicado_en_esta.get(origen.id, 0.0) + a["monto"], 2
            )
            disponible = round(origen.amount_total - aplicado_otras, 2)
            if aplicado_en_esta[origen.id] > disponible + 0.01:
                raise UserError(
                    _(
                        "El anticipo %s ya no tiene saldo suficiente: disponible %.2f, "
                        "intentas aplicar %.2f."
                    )
                    % (origen.display_name, disponible, aplicado_en_esta[origen.id])
                )
        _cod, _tasa, motivo = self._l10n_pe_anticipo_gravado()
        if motivo:
            raise UserError(
                _(
                    "El anticipo solo se soporta sobre una operación gravada homogénea (IGV o IVAP). "
                    "No es aplicable a operaciones exoneradas/inafectas/exportación ni a facturas con "
                    "regímenes mixtos: regularice el anticipo en un comprobante separado."
                )
            )

    def _l10n_pe_check_lineas_impuesto(self):
        """Ninguna línea con importe llega al XML sin su tax cat-05: `_l10n_pe_tax_info`
        la clasificaría con el default 'gravado (1000)' a tasa 0 y SUNAT rechaza con 3111
        (TaxableAmount>0 + TaxAmount=0.00), un mensaje críptico que además llega recién
        del validador. Se corta aquí, con el dato que el usuario sí puede arreglar.
        Las líneas de importe 0 (p.ej. NC de corrección de texto) no tienen base imponible
        —no hay 3111 posible— y pasan."""
        self.ensure_one()
        for line in self._l10n_pe_product_lines():
            if not line.price_subtotal:
                continue
            if not any(t.l10n_pe_edi_tax_code in TAX_CODE_MAP for t in line.tax_ids):
                raise UserError(
                    _(
                        "La línea «%s» no tiene impuesto SUNAT asignado (IGV, exonerado, "
                        "inafecto…). Asigna la afectación IGV en el producto o en la línea "
                        "y vuelve a emitir."
                    )
                    % (line.name or line.product_id.display_name or "?")
                )

    def _l10n_pe_relacionados(self):
        """Documentos relacionados de la factura: guía de remisión (indDocRelacionado 1,
        DespatchDocumentReference) y/o comprobante de anticipo (indDocRelacionado 2)."""
        rels = []
        guia = (self.l10n_pe_ne_guia_ref or "").strip()
        if guia:
            rels.append(
                {
                    "indDocRelacionado": "1",
                    "tipDocRelacionado": self.l10n_pe_ne_guia_tipo or "09",
                    "numDocRelacionado": guia,
                    "tipDocEmisor": "6",
                    "numDocEmisor": self.company_id.vat or "",
                }
            )
        # N AdditionalDocumentReference (uno por anticipo), numIdeAnticipo correlativo 1..N en el
        # orden de la lista — así SUNAT liga cada PrepaidPayment con su propio documento relacionado.
        lst = self._l10n_pe_ne_anticipos_list()
        for idx, a in enumerate(lst, start=1):
            rels.append(
                {
                    "indDocRelacionado": "2",
                    "tipDocRelacionado": a["tipo"] or "02",
                    "numDocRelacionado": a["doc"],
                    "numIdeAnticipo": str(idx),
                    "mtoDocRelacionado": self._l10n_pe_fmt(a["monto"]),
                    "tipDocEmisor": "6",
                    "numDocEmisor": self.company_id.vat or "",
                }
            )
        return rels

    def _l10n_pe_variables_globales(self):
        """Variables globales de la factura:
        - código 51: percepción (el agente percibe un % sobre la venta; el cliente paga total + percepción).
        - código 04: descuento global por anticipo (regulariza uno o más anticipos ya facturados;
          reduce la base del IGV en el valor AGREGADO de todos los anticipos). Exigido por SUNAT
          (regla 3287) cuando hay anticipo. Con N>1 anticipos se emite UN solo 04 con la suma —no uno
          por anticipo—, en línea con los N documentos relacionados (`_l10n_pe_relacionados`) que sí
          van uno por cada `AdditionalDocumentReference`/`numIdeAnticipo`."""
        fmt = self._l10n_pe_fmt
        moneda = self.currency_id.name or "PEN"
        out = []
        if self.l10n_pe_ne_percepcion:
            out.append(
                {
                    "tipVariableGlobal": "true",
                    "codTipoVariableGlobal": "51",
                    "porVariableGlobal": "%.2f"
                    % (self.l10n_pe_ne_percepcion_rate / 100.0),
                    "monMontoVariableGlobal": moneda,
                    "mtoVariableGlobal": fmt(self._l10n_pe_percepcion_monto()),
                    "monBaseImponibleVariableGlobal": moneda,
                    # Base de la percepción = neto a cobrar (descontado el anticipo): sin anticipo es el total.
                    "mtoBaseImpVariableGlobal": fmt(self._l10n_pe_importe_cobrar()),
                }
            )
        ant = self._l10n_pe_anticipo()
        if ant:
            valor, _igv, _total = ant
            base = self.amount_untaxed or 0.0
            out.append(
                {
                    "tipVariableGlobal": "false",
                    "codTipoVariableGlobal": "04",
                    # Factor con 5 decimales: SUNAT valida mtoVariableGlobal ≈ base × por (tolerancia
                    # ~±1, error 3307). Con solo 2 decimales, un anticipo parcial cuyo valor no es
                    # fracción redonda de la base (p.ej. 254.24 sobre 1000 → 0.25) descuadra y se
                    # rechaza; 5 decimales reconstruyen el monto dentro de la tolerancia.
                    "porVariableGlobal": "%.5f" % (valor / base if base else 0.0),
                    "monMontoVariableGlobal": moneda,
                    "mtoVariableGlobal": fmt(valor),
                    "monBaseImponibleVariableGlobal": moneda,
                    "mtoBaseImpVariableGlobal": fmt(base),
                }
            )
        # Descuento global que NO afecta la base del IGV (código del facturador en
        # DESC_GLOBAL_NO_AFECTA_COD, pendiente de confirmar contra beta). La base es el precio de
        # venta con IGV (amount_total): el descuento NO reduce gravada/IGV, solo el MtoImpVenta.
        desc_na = self._l10n_pe_desc_no_afecta()
        if desc_na > 0:
            base = self.amount_total or 0.0
            out.append(
                {
                    "tipVariableGlobal": "false",
                    "codTipoVariableGlobal": DESC_GLOBAL_NO_AFECTA_COD,
                    # 5 decimales: misma tolerancia SUNAT que el anticipo (mtoVariable ≈ base × por).
                    "porVariableGlobal": "%.5f" % (desc_na / base if base else 0.0),
                    "monMontoVariableGlobal": moneda,
                    "mtoVariableGlobal": fmt(desc_na),
                    "monBaseImponibleVariableGlobal": moneda,
                    "mtoBaseImpVariableGlobal": fmt(base),
                }
            )
        return out

    @api.depends("journal_id", "partner_id", "move_type", "debit_origin_id",
                 "reversed_entry_id", "l10n_latam_document_type_id")
    def _compute_l10n_pe_serie(self):
        for move in self:
            serie = move.l10n_pe_serie or move.journal_id.l10n_pe_ne_serie or "F001"
            # La letra de la serie la manda la familia del comprobante (F factura / B boleta),
            # no el diario: con un solo diario de ventas la serie del diario es de una familia
            # y la boleta (cliente sin RUC) necesita la otra.
            if (
                move.state == "draft"
                and move.move_type in ("out_invoice", "out_refund")
                and move.partner_id
                and serie[:1].upper() in ("F", "B")
            ):
                prefix = move._l10n_pe_serie_prefix()
                if serie[:1].upper() != prefix:
                    serie = prefix + serie[1:]
            move.l10n_pe_serie = serie

    def _l10n_pe_detraccion_monto(self):
        self.ensure_one()
        # SUNAT (SPOT): el monto de la detracción se redondea al ENTERO más próximo
        # (sin decimales), medio hacia arriba. Ej.: 12% de 25 386.52 = 3046.38 -> 3046.
        return float_round(
            self.amount_total * (self.l10n_pe_ne_detraccion_rate or 0.0) / 100.0,
            precision_digits=0,
            rounding_method="HALF-UP",
        )

    def _l10n_pe_neto_pendiente(self):
        """Neto pendiente de pago = total − detracción (si aplica). Con detracción el
        cliente solo paga el neto; el monto detraído se deposita en el Banco de la Nación,
        así que el pendiente/cuotas van sobre el neto, no sobre el total."""
        self.ensure_one()
        det = self._l10n_pe_detraccion_monto() if self.l10n_pe_ne_detraccion else 0.0
        # Venta con inicial al contado: el saldo a crédito (lo que suman las cuotas) es el total
        # menos la detracción y menos la inicial ya pagada.
        inicial = self.l10n_pe_ne_inicial_contado or 0.0
        return round((self.amount_total or 0.0) - det - inicial, 2)

    def _l10n_pe_adicional_cabecera(self):
        """Bloque adicional de la cabecera: detracción y/o total a cobrar de la percepción."""
        fmt = self._l10n_pe_fmt
        block = {}
        if self.l10n_pe_ne_detraccion:
            block.update(
                {
                    "ctaBancoNacionDetraccion": self.l10n_pe_ne_detraccion_cuenta
                    or self.company_id.l10n_pe_ne_cuenta_detraccion
                    or "",
                    "codBienDetraccion": self.l10n_pe_ne_detraccion_code or "",
                    "porDetraccion": fmt(self.l10n_pe_ne_detraccion_rate),
                    "mtoDetraccion": fmt(self._l10n_pe_detraccion_monto()),
                    "codMedioPago": self.l10n_pe_ne_detraccion_medio_pago or "001",
                }
            )
        if self.l10n_pe_ne_percepcion:
            # Total a cobrar = neto a cobrar (descontado el anticipo) + la percepción.
            block["mtoTotPercepcion"] = fmt(
                self._l10n_pe_importe_cobrar() + self._l10n_pe_percepcion_monto()
            )
        # Exportación (tipOperacion 0200): el adquirente es no domiciliado. SUNAT pide el país del
        # cliente (cat. país, ISO 3166 alpha-2 = el mismo code de res.country). El biller lo mapea a
        # codPaisCliente del AdditionalHeader. Se omite si el partner no tiene país (evita "" inútil).
        if self._l10n_pe_tipo_operacion() == "0200":
            pais = (self.partner_id.country_id.code or "").strip().upper()
            if pais:
                block["codPaisCliente"] = pais
        return block or None

    def _l10n_pe_dato_pago(self):
        moneda = self.currency_id.name or "PEN"
        if self.l10n_pe_ne_forma_pago == "Credito":
            return {
                "formaPago": "Credito",
                "mtoNetoPendientePago": self._l10n_pe_fmt(
                    self._l10n_pe_credito_pendiente()
                ),
                "tipMonedaMtoNetoPendientePago": moneda,
            }
        dato = {"formaPago": "Contado"}
        if self.l10n_pe_ne_detraccion:
            # Operación al contado con detracción: el neto pendiente es total − detracción
            # (lo que el cliente paga; la detracción va al Banco de la Nación).
            dato["mtoNetoPendientePago"] = self._l10n_pe_fmt(
                self._l10n_pe_neto_pendiente()
            )
            dato["tipMonedaMtoNetoPendientePago"] = moneda
        return dato

    def _l10n_pe_cuotas_netas(self):
        """Cuotas guardadas AJUSTADAS al neto pendiente. Con detracción, las cuotas pueden
        venir sobre el TOTAL (front antiguo, emisión masiva, API); se escalan al neto para
        que sumen exactamente el pendiente — la última absorbe el redondeo. Sin detracción
        el neto == total, así que no cambian. Garantiza sum(cuotas) == mtoNetoPendientePago
        pase lo que pase (SUNAT lo exige) y que el cliente no pague la parte detraída."""
        cuotas = [
            c
            for c in (self.l10n_pe_ne_cuotas or [])
            if c.get("fecha") and float(c.get("monto") or 0) > 0
        ]
        if not cuotas:
            return []
        neto = self._l10n_pe_neto_pendiente()
        suma = sum(float(c["monto"]) for c in cuotas)
        if suma <= 0 or abs(suma - neto) < 0.01:
            return [{"fecha": c["fecha"], "monto": round(float(c["monto"]), 2)} for c in cuotas]
        factor = neto / suma
        out, acc = [], 0.0
        for i, c in enumerate(cuotas):
            if i < len(cuotas) - 1:
                monto = round(float(c["monto"]) * factor, 2)
                acc += monto
            else:  # la última cuota cuadra el total al neto exacto
                monto = round(neto - acc, 2)
            out.append({"fecha": c["fecha"], "monto": monto})
        return out

    def _l10n_pe_credito_pendiente(self):
        """Monto neto pendiente del crédito = suma de las cuotas (ya ajustadas al neto);
        si no hay cuotas, el neto (total − detracción)."""
        netas = self._l10n_pe_cuotas_netas()
        return sum(c["monto"] for c in netas) if netas else self._l10n_pe_neto_pendiente()

    def _l10n_pe_detalle_pago(self):
        """detallePago (cuotas) para crédito: cuotas ajustadas al neto, o una = neto."""
        moneda = self.currency_id.name or "PEN"
        out = [
            {
                "mtoCuotaPago": self._l10n_pe_fmt(c["monto"]),
                "fecCuotaPago": c["fecha"],
                "tipMonedaCuotaPago": moneda,
            }
            for c in self._l10n_pe_cuotas_netas()
        ]
        if not out:
            fecha = self.invoice_date_due or self.invoice_date
            out = [
                {
                    "mtoCuotaPago": self._l10n_pe_fmt(self._l10n_pe_neto_pendiente()),
                    "fecCuotaPago": fecha.strftime("%Y-%m-%d") if fecha else "",
                    "tipMonedaCuotaPago": moneda,
                }
            ]
        return out

    # Establecimiento anexo emisor (código SUNAT de 4 dígitos). Va como codLocalEmisor en el XML;
    # "0000" = domicilio fiscal. Para negocios con sucursales, cada comprobante declara su local.
    l10n_pe_ne_cod_establecimiento = fields.Char(
        string="Establecimiento emisor",
        default="0000",
        copy=False,
        help="Código de establecimiento anexo SUNAT (4 dígitos). '0000' = domicilio fiscal.",
    )
    # Guía de remisión que sustenta el traslado: va como cac:DespatchDocumentReference en el XML
    # de la factura (indDocRelacionado 1). QA-031.
    l10n_pe_ne_guia_ref = fields.Char(
        string="Guía de remisión referenciada",
        copy=False,
        help="Serie-número de la GRE que sustenta el traslado (ej. T001-00000123).",
    )
    l10n_pe_ne_guia_tipo = fields.Selection(
        [("09", "Guía de remisión remitente"), ("31", "Guía de remisión transportista")],
        string="Tipo de guía referenciada",
        default="09",
    )
    # Proyecto/contrato (facturación por avance de obra): controla que la suma de las
    # valorizaciones no supere el valor total del contrato (QA-039).
    l10n_pe_ne_proyecto_id = fields.Many2one(
        "l10n_pe_ne.proyecto", string="Proyecto / contrato", copy=False,
        help="Contrato al que pertenece esta valorización. El total facturado no puede superar "
        "el valor del contrato.",
    )
    l10n_pe_ne_forma_pago = fields.Selection(
        [("Contado", "Contado"), ("Credito", "Crédito")],
        default="Contado",
        copy=False,
        string="Forma de pago",
        help="Forma de pago SUNAT (cac:PaymentTerms). 'Crédito' emite cuotas.",
    )
    l10n_pe_ne_cuotas = fields.Json(
        string="Cuotas de crédito", copy=False
    )  # [{'fecha','monto'}]
    # Forma de pago MIXTA: parte pagada al contado (inicial) + saldo a crédito en cuotas. El neto
    # pendiente (y por ende las cuotas y el mtoNetoPendientePago SUNAT) se reduce en esta inicial.
    l10n_pe_ne_inicial_contado = fields.Monetary(
        string="Inicial al contado",
        copy=False,
        help="Parte del total pagada al contado al emitir (venta con inicial + saldo a crédito). "
        "El saldo a crédito = total − detracción − inicial y es lo que suman las cuotas.",
    )
    l10n_pe_ne_medios_pago = fields.Json(
        string="Medios de pago (POS)", copy=False
    )  # [{'medio','monto'}]
    # Redondeo de efectivo (Ley 29571 + retiro de monedas < S/ 0.10): ajuste ≤ 0 a favor del
    # consumidor sobre el total a cobrar EN EFECTIVO. NO va al XML/comprobante (amount_total sigue
    # exacto); es un dato de caja: el arqueo espera 'amount_total + redondeo' de efectivo, y el
    # ticket muestra 'A pagar efectivo'. Ver _l10n_pe_ne_ticket_adicional y l10n_pe_ne_caja.
    l10n_pe_ne_redondeo = fields.Monetary(
        string="Redondeo efectivo",
        copy=False,
        help="Ajuste (≤ 0) del importe cobrado en efectivo por redondeo al décimo. No altera el "
        "comprobante ni las bases/IGV; solo el efectivo cobrado y el arqueo de caja.",
    )

    l10n_pe_motivo_code = fields.Char(
        string="Cód. motivo NC/ND",
        default="01",
        copy=False,
        help="Código SUNAT del motivo de la nota de crédito (cat. 09) o débito (cat. 10).",
    )
    l10n_pe_motivo_desc = fields.Char(
        string="Motivo/sustento NC/ND",
        copy=False,
        help="Motivo o sustento (texto libre) de la nota. Si se omite, se usa la "
             "descripción del catálogo correspondiente al código de motivo.",
    )
    l10n_pe_biller_xml = fields.Many2one(
        "ir.attachment", string="XML UBL firmado", copy=False
    )
    l10n_pe_biller_cdr = fields.Many2one(
        "ir.attachment", string="CDR SUNAT", copy=False
    )
    # Modo instantáneo: tras FIRMAR se guarda el ZIP de ENVI + el filename/canal para que el
    # cron envíe a SUNAT en 2º plano. Se limpian al recibir el CDR (ya no hay nada pendiente).
    l10n_pe_ne_envi_zip = fields.Text(
        string="ZIP ENVI pendiente (base64)", copy=False,
        help="ZIP de ENVI firmado, aún no enviado a SUNAT. El cron lo envía y lo limpia al aceptarse.")
    l10n_pe_ne_biller_filename = fields.Char(string="Nombre de archivo del facturador", copy=False)
    l10n_pe_ne_biller_canal = fields.Char(string="Canal SUNAT (GEM/OTROS_CPE)", copy=False)
    l10n_pe_ne_envio_intentos = fields.Integer(string="Intentos de envío a SUNAT", default=0, copy=False)
    l10n_pe_ne_stock_aviso = fields.Char(
        string="Aviso de inventario",
        copy=False,
        readonly=True,
        help="Por qué no se pudo mover el inventario de este documento. El comprobante es "
        "válido igual: el stock nunca lo tumba. Vacío = el movimiento se hizo.",
    )

    # Resumen Diario de boletas (RC) idempotente: al enviar se guarda el TICKET; el poll usa el
    # ticket (no re-envía → no duplica). Correlativo/fecha del RC al que pertenece la boleta.
    l10n_pe_ne_rc_ticket = fields.Char(string="Ticket del Resumen Diario", copy=False)
    l10n_pe_ne_rc_correlativo = fields.Char(string="Correlativo del Resumen Diario", copy=False)
    l10n_pe_ne_rc_fecha = fields.Date(string="Fecha del Resumen Diario", copy=False)
    l10n_pe_biller_pdf = fields.Many2one(
        "ir.attachment", string="PDF (representación impresa)", copy=False
    )
    l10n_pe_biller_pdf_ticket = fields.Many2one(
        "ir.attachment", string="PDF ticket 80mm (representación impresa)", copy=False
    )
    l10n_pe_biller_message = fields.Text(string="Mensaje Facturador", copy=False)

    # ----------------------------------------------------------------- helpers
    def _l10n_pe_fmt(self, amount):
        return "%.2f" % (amount or 0.0)

    def _l10n_pe_fmt_cant(self, qty):
        """Cantidad para SUNAT (ctdUnidadItem): hasta 3 decimales, sin ceros de relleno más allá
        de 2. Conserva la venta al peso de balanza (18.375) sin ensuciar los conteos (2 -> 2.00).
        SUNAT admite hasta 10 decimales; `_l10n_pe_fmt` (2 dec) es solo para montos."""
        entero, _p, dec = ("%.3f" % (qty or 0.0)).partition(".")
        dec = dec.rstrip("0")
        if len(dec) < 2:
            dec = (dec + "00")[:2]
        return "%s.%s" % (entero, dec)

    def _l10n_pe_document_type(self):
        """Código SUNAT del comprobante: 01 Factura, 03 Boleta, 07 NC, 08 ND."""
        self.ensure_one()
        if self.move_type == "out_refund":
            return "07"
        if self.move_type == "out_invoice" and self.debit_origin_id:
            return "08"
        # Una exportación es siempre Factura (01), aunque el adquirente extranjero no tenga RUC
        # (si fuese Boleta 03 con serie F, el validador de factura la rechaza por tipo/serie).
        if self.move_type == "out_invoice" and self._l10n_pe_tipo_operacion() == "0200":
            return "01"
        # El tipo elegido en el comprobante manda: a un cliente con RUC se le puede emitir
        # Boleta (compra como consumidor final). El documento de identidad solo decide
        # cuando no hay tipo elegido (diario sin documentos latam, flujos por código).
        if self.move_type == "out_invoice":
            code = self.l10n_latam_document_type_id.code
            if code in ("01", "03"):
                return code
        vat_code = (
            self.partner_id.l10n_latam_identification_type_id.l10n_pe_vat_code or ""
        )
        return "01" if vat_code == "6" else "03"

    def _l10n_pe_serie_prefix(self):
        """Letra que SUNAT exige en la serie: F para Factura (01) y sus notas, B para Boleta (03)
        y las suyas. En NC/ND manda la familia del documento afectado, no el partner."""
        self.ensure_one()
        origin = self.reversed_entry_id or self.debit_origin_id
        if origin:
            tipo = origin.l10n_pe_ne_tipo_doc or origin._l10n_pe_document_type()
        else:
            tipo = self._l10n_pe_document_type()
        if tipo not in ("01", "03"):  # NC/ND sin documento afectado: decide el cliente
            vat_code = (
                self.partner_id.l10n_latam_identification_type_id.l10n_pe_vat_code or ""
            )
            tipo = "01" if vat_code == "6" else "03"
        return "B" if tipo == "03" else "F"

    def _l10n_pe_check_serie(self):
        """Serie de familia equivocada (p.ej. F001 en una boleta) es rechazo seguro de SUNAT;
        se corta aquí antes de enviar/encolar."""
        self.ensure_one()
        serie, _corr = self._l10n_pe_serie_correlativo()
        prefix = self._l10n_pe_serie_prefix()
        if (serie or "")[:1].upper() != prefix:
            docname = {
                "01": _("Factura"),
                "03": _("Boleta"),
                "07": _("Nota de Crédito"),
                "08": _("Nota de Débito"),
            }.get(self._l10n_pe_document_type(), "")
            raise UserError(
                _(
                    "La serie '%(serie)s' no corresponde al tipo de comprobante: una %(doc)s "
                    "debe usar una serie que empiece con '%(prefix)s' (p.ej. %(prefix)s001)."
                )
                % {"serie": serie, "doc": docname, "prefix": prefix}
            )
        # QA-074: la serie debe estar HABILITADA para el emisor. Una serie inventada (p.ej. F099
        # tecleada a mano) la acepta la beta de SUNAT pero en producción se rechaza; se corta aquí.
        habilitadas = self._l10n_pe_ne_series_habilitadas()
        if (serie or "").upper() not in habilitadas:
            raise UserError(
                _(
                    "La serie '%(serie)s' no está habilitada para %(ruc)s. Configúrala en un "
                    "diario de venta (campo Serie) o usa una serie registrada: %(lista)s."
                )
                % {
                    "serie": serie,
                    "ruc": self.company_id.vat or self.company_id.display_name or "",
                    "lista": ", ".join(sorted(habilitadas)),
                }
            )

    def _l10n_pe_ne_series_habilitadas(self):
        """Series válidas del emisor (QA-074): las configuradas en sus diarios de venta
        (l10n_pe_ne_serie) con su variante de familia (F↔B), más los defaults que genera el
        sistema (F001/B001 y las notas FC01/FD01/BC01/BD01). No se usa el histórico de series
        ya emitidas a propósito: una serie inventada usada por error no debe volverse 'válida'."""
        self.ensure_one()
        validas = {"F001", "B001", "FC01", "FD01", "BC01", "BD01"}
        journals = self.env["account.journal"].sudo().search(
            [
                ("company_id", "=", self.company_id.id),
                ("type", "=", "sale"),
                ("l10n_pe_ne_serie", "!=", False),
            ]
        )
        for j in journals:
            base = (j.l10n_pe_ne_serie or "").upper().strip()
            if len(base) >= 2 and base[0] in ("F", "B"):
                validas.add("F" + base[1:])
                validas.add("B" + base[1:])
        return validas

    def _l10n_pe_product_lines(self):
        return self.invoice_line_ids.filtered(
            lambda l: not l.display_type or l.display_type == "product"
        )

    def _l10n_pe_tax_info(self, line):
        """Afectación IGV de la línea según la tax de Odoo. Devuelve
        ((tipAfeIGV, codTriIGV, nomTributo, codTipTributo, codCatTributo), porcentaje_igv).
        Lee `account.tax.l10n_pe_edi_tax_code` (cat. 05) de la localización l10n_pe; si la línea no
        trae una tax reconocida, asume gravado (IGV)."""
        for tax in line.tax_ids:
            if tax.l10n_pe_edi_tax_code in TAX_CODE_MAP:
                return TAX_CODE_MAP[tax.l10n_pe_edi_tax_code], tax.amount
        return TAX_CODE_MAP[DEFAULT_TAX_CODE], 0.0

    def _l10n_pe_icbper_tax(self, line):
        """La tax ICBPER (impuesto a las bolsas, cat. 05 = 7152) de la línea, si la trae. Es una
        tax de monto fijo (amount_type='fixed') = soles por bolsa."""
        return line.tax_ids.filtered(lambda t: t.l10n_pe_edi_tax_code == "7152")[:1]

    def _l10n_pe_isc_tax(self, line):
        """La tax ISC (Impuesto Selectivo al Consumo, cat. 05 = 2000) de la línea, si la trae.
        Debe estar marcada 'Afecta la base de los impuestos posteriores' para que el IGV se compute
        sobre valor+ISC."""
        return line.tax_ids.filtered(lambda t: t.l10n_pe_edi_tax_code == "2000")[:1]

    def _l10n_pe_line_amounts(self, line):
        """Descompone los tributos de la línea: (base, igv, isc, icbper).

        price_total - price_subtotal incluye los tres. El ICBPER = nº bolsas × monto fijo. El ISC
        'al valor' (sis. 01) = base × tasa; 'monto fijo' (02) = cantidad × monto. El IGV es el resto
        (Odoo ya lo computa sobre valor+ISC si la tax ISC afecta la base)."""
        base = line.price_subtotal
        total_tax = line.price_total - line.price_subtotal
        icbper_tax = self._l10n_pe_icbper_tax(line)
        icbper = (
            round(int(round(line.quantity or 0)) * icbper_tax.amount, 2)
            if icbper_tax
            else 0.0
        )
        isc_tax = self._l10n_pe_isc_tax(line)
        if isc_tax:
            if isc_tax.amount_type == "fixed":
                isc = round((line.quantity or 0.0) * isc_tax.amount, 2)
            else:
                isc = round(base * isc_tax.amount / 100.0, 2)
        else:
            isc = 0.0
        return base, total_tax - isc - icbper, isc, icbper

    def _l10n_pe_total_icbper(self):
        return sum(
            self._l10n_pe_line_amounts(l)[3] for l in self._l10n_pe_product_lines()
        )

    def _l10n_pe_unit_code(self, line):
        """Código de unidad SUNAT (cat. 03) de la línea: override por línea, luego el guardado en el
        producto (POS/masiva no mandan unidad por línea), luego override manual en la UoM, si no el
        mapeo por XMLID de la unidad estándar de Odoo, si no 'NIU'."""
        if line.l10n_pe_ne_unit_code:
            return line.l10n_pe_ne_unit_code
        if line.product_id.l10n_pe_ne_unit_code:
            return line.product_id.l10n_pe_ne_unit_code
        uom = line.product_uom_id
        if not uom:
            return DEFAULT_UNIT_CODE
        if uom.l10n_pe_ne_unit_code:
            return uom.l10n_pe_ne_unit_code
        xmlid = uom.get_external_id().get(uom.id, "")
        return UOM_CODE_BY_XMLID.get(xmlid, DEFAULT_UNIT_CODE)

    _L10N_PE_ANTICIPO_PREFIX = "PAGO ANTICIPADO"

    def _l10n_pe_des_item(self, line):
        """Descripción del ítem para el XML. En un comprobante marcado como pago anticipado
        (doc. A) antepone 'PAGO ANTICIPADO' para que el documento identifique la operación sin
        depender de una leyenda cat. 52 (que no existe para anticipos)."""
        desc = line.name or line.product_id.display_name or ""
        if self.l10n_pe_ne_es_anticipo and not desc.startswith(self._L10N_PE_ANTICIPO_PREFIX):
            desc = ("%s - %s" % (self._L10N_PE_ANTICIPO_PREFIX, desc)).strip(" -")
        return desc

    def _l10n_pe_detalle(self):
        fmt = self._l10n_pe_fmt
        detalle = []
        for line in self._l10n_pe_product_lines():
            (tip_afe, cod_tri, nom_trib, cod_tip_trib, _cod_cat), por_igv = (
                self._l10n_pe_tax_info(line)
            )
            # Gratuita: si la línea precisa el sub-tipo (retiro 13, bonificación 15, …) se usa ese
            # código de catálogo 07 en vez del genérico 11. La estructura UBL gratuita es idéntica.
            if cod_tri == "9996" and line.l10n_pe_ne_afectacion_gratuita:
                tip_afe = line.l10n_pe_ne_afectacion_gratuita
            qty = line.quantity or 1.0
            base, igv, isc, icbper = self._l10n_pe_line_amounts(line)
            # Valor unitario BRUTO (antes del descuento): regla SUNAT 3271 exige
            # mtoValorVentaItem = mtoValorUnitario*cantidad - descuento. El descuento sale aparte
            # en adicionalDetalle; mtoValorVentaItem (LineExtensionAmount) queda neto.
            disc = (
                round(line.price_unit * line.quantity - base, 2)
                if line.discount
                else 0.0
            )
            gross = base + disc
            item = {
                "tipAfeIGV": tip_afe,
                "codProducto": line.product_id.default_code or "-",
                "codProductoSUNAT": line.l10n_pe_ne_cod_producto_sunat or "-",
                "codUnidadMedida": self._l10n_pe_unit_code(line),
                "ctdUnidadItem": self._l10n_pe_fmt_cant(qty),
                "desItem": self._l10n_pe_des_item(line),
                "mtoValorUnitario": fmt(gross / qty if qty else 0.0),
                "mtoValorVentaItem": fmt(base),
                # Precio de venta unitario = (valor venta + ISC + IGV) / cantidad; NO incluye el ICBPER.
                "mtoPrecioVentaUnitario": fmt((base + isc + igv) / qty if qty else 0.0),
                "mtoValorReferencialUnitario": "0.00",
                "porIgvItem": fmt(por_igv),
                # La base del IGV incluye el ISC (el IGV se computa sobre valor venta + ISC).
                "mtoBaseIgvItem": fmt(base + isc),
                "mtoIgvItem": fmt(igv),
                "sumTotTributosItem": fmt(igv + isc + icbper),
                "codTriIGV": cod_tri,
                "nomTributoIgvItem": nom_trib,
                "codTipTributoIgvItem": cod_tip_trib,
            }
            # Operación gratuita (cat. 05 = 9996). Estructura SUNAT (ref: enterprise invoice_free.xml):
            # Price/PriceAmount=0; valor de mercado en mtoValorReferencialUnitario (PricingReference 02);
            # LineExtensionAmount(mtoValorVentaItem)=valor de mercado; TaxSubtotal 9996 con base y el IGV
            # teórico 18% (mtoBaseIgvItem/mtoIgvItem); pero el TaxTotal/TaxAmount de la LÍNEA
            # (sumTotTributosItem) = 0 — el IGV gratuito NO se cobra (clave del fault 3272).
            if cod_tri == "9996":
                igv_grat = round(base * 0.18, 2)
                item.update(
                    {
                        "mtoValorUnitario": "0.00",
                        "mtoValorVentaItem": fmt(base),
                        "mtoPrecioVentaUnitario": "0.00",
                        "mtoValorReferencialUnitario": fmt(gross / qty if qty else 0.0),
                        "porIgvItem": "18.00",
                        "mtoBaseIgvItem": fmt(base),
                        "mtoIgvItem": fmt(igv_grat),
                        "sumTotTributosItem": "0.00",
                    }
                )
            isc_tax = self._l10n_pe_isc_tax(line)
            if isc_tax:
                por_isc = (
                    isc_tax.amount
                    if isc_tax.amount_type != "fixed"
                    else (isc / base * 100.0 if base else 0.0)
                )
                item.update(
                    {
                        "codTriISC": "2000",
                        "nomTributoIscItem": "ISC",
                        "codTipTributoIscItem": "EXC",
                        "tipSisISC": isc_tax.l10n_pe_edi_isc_type or "01",
                        "mtoBaseIscItem": fmt(base),
                        "mtoIscItem": fmt(isc),
                        "porIscItem": fmt(por_isc),
                    }
                )
            icbper_tax = self._l10n_pe_icbper_tax(line)
            if icbper_tax:
                item.update(
                    {
                        "codTriIcbper": "7152",
                        "nomTributoIcbperItem": "ICBPER",
                        "codTipTributoIcbperItem": "OTH",
                        "ctdBolsasTriIcbperItem": str(int(round(qty))),
                        "mtoTriIcbperUnidad": fmt(icbper_tax.amount),
                        "mtoTriIcbperItem": fmt(icbper),
                    }
                )
            detalle.append(item)
        return detalle

    def _l10n_pe_tributos(self):
        """Un tributo por categoría presente (IGV/EXO/INA/EXP/GRA/IVAP), con la base y el monto
        sumados de las líneas de esa categoría."""
        fmt = self._l10n_pe_fmt
        grupos = {}  # codTriIGV -> [base, monto, (nomTributo, codTipTributo, codCatTributo)]
        isc_base = isc_total = 0.0
        for line in self._l10n_pe_product_lines():
            (_tip, cod_tri, nom_trib, cod_tip_trib, cod_cat), _por = (
                self._l10n_pe_tax_info(line)
            )
            base, igv, isc, _icbper = self._l10n_pe_line_amounts(line)
            # Base del IGV de cabecera = valor venta (no incluye el ISC, a diferencia de la línea).
            g = grupos.setdefault(
                cod_tri, [0.0, 0.0, (nom_trib, cod_tip_trib, cod_cat)]
            )
            g[0] += base
            # Gratuito (9996): el IGV teórico (18% del valor de mercado) va en el tributo de cabecera
            # aunque no se cobre. En las demás categorías es el IGV real (el grupo no incluye ICBPER).
            g[1] += round(base * 0.18, 2) if cod_tri == "9996" else igv
            if isc:
                isc_base += base
                isc_total += isc
        # Anticipo: el descuento global código 04 reduce la base y el impuesto de cabecera del grupo
        # gravado (no las líneas, que declaran la operación completa). El validador computa el impuesto
        # sobre la base ya reducida. Se reduce el régimen real (IGV '1000' o IVAP '1016').
        ant = self._l10n_pe_anticipo()
        if ant:
            cod_tri, _tasa, _motivo = self._l10n_pe_anticipo_gravado()
            if cod_tri and cod_tri in grupos:
                valor, igv, _total = ant
                grupos[cod_tri][0] -= valor
                grupos[cod_tri][1] -= igv
        tributos = [
            {
                "ideTributo": cod_tri,
                "nomTributo": meta[0],
                "codTipTributo": meta[1],
                "codCatTributo": meta[2],
                "mtoBaseImponible": fmt(b),
                "mtoTributo": fmt(m),
            }
            for cod_tri, (b, m, meta) in grupos.items()
        ]
        if isc_total:
            tributos.append(
                {
                    "ideTributo": "2000",
                    "nomTributo": "ISC",
                    "codTipTributo": "EXC",
                    "codCatTributo": "S",
                    "mtoBaseImponible": fmt(isc_base),
                    "mtoTributo": fmt(isc_total),
                }
            )
        # ICBPER (7152): TaxSubtotal de cabecera SIN TaxableAmount (el FTL lo omite para 7152), solo el
        # monto. Necesario para que TaxInclusive = LineExt + TaxTotal (regla SUNAT 3279).
        icbper_total = self._l10n_pe_total_icbper()
        if icbper_total:
            tributos.append(
                {
                    "ideTributo": "7152",
                    "nomTributo": "ICBPER",
                    "codTipTributo": "OTH",
                    "codCatTributo": "S",
                    "mtoBaseImponible": "0.00",
                    "mtoTributo": fmt(icbper_total),
                }
            )
        return tributos

    def _l10n_pe_leyendas(self):
        # El monto en letras corresponde al importe a cobrar (total − anticipo aplicado).
        leyendas = [
            {
                "codLeyenda": "1000",
                "desLeyenda": leyenda_monto(self._l10n_pe_importe_cobrar()),
            }
        ]
        if self.l10n_pe_ne_detraccion:
            leyendas.append(
                {"codLeyenda": "2006", "desLeyenda": "Operacion sujeta a detraccion"}
            )
        if self._l10n_pe_gratuito_base() > 0:
            leyendas.append(
                {"codLeyenda": "1002", "desLeyenda": "TRANSFERENCIA GRATUITA"}
            )
        return leyendas

    def _l10n_pe_gratuito_base(self):
        """Suma de las bases (valor de mercado) de las líneas gratuitas (cat. 05 = 9996)."""
        self.ensure_one()
        total = 0.0
        for line in self._l10n_pe_product_lines():
            if self._l10n_pe_tax_info(line)[0][1] == "9996":
                total += self._l10n_pe_line_amounts(line)[0]
        return round(total, 2)

    def _l10n_pe_tipo_operacion(self):
        """1001 detracción, 2001 percepción, 0200 exportación; si no, 0101 (venta interna)."""
        if self.l10n_pe_ne_detraccion:
            return "1001"
        if self.l10n_pe_ne_percepcion:
            return "2001"
        lineas = self._l10n_pe_product_lines()
        afectaciones = {self._l10n_pe_tax_info(l)[0][0] for l in lineas}
        return "0200" if afectaciones == {"40"} else "0101"

    def _l10n_pe_cliente_doc(self):
        """(tipDocUsuario, numDocUsuario) del cliente. Consumidor final sin documento → ('0','00000000');
        si trae número pero no tipo, se infiere (11 dígitos→RUC '6', si no DNI '1')."""
        self.ensure_one()
        p = self.partner_id
        vat = (p.vat or "").strip()
        cod = p.l10n_latam_identification_type_id.l10n_pe_vat_code or ""
        if not vat:
            return "0", "00000000"
        if not cod:
            cod = "6" if (len(vat) == 11 and vat.isdigit()) else "1"
        return cod, vat

    @api.model
    def _l10n_pe_ne_today_lima(self):
        """Fecha de HOY en hora local de Perú (América/Lima, UTC-5).

        Evita el descuadre de zona horaria: `fields.Date.context_today` cae a UTC
        cuando el usuario no tiene tz configurada, así que de noche (después de las
        7pm Lima = medianoche UTC) devuelve el día SIGUIENTE. Eso hacía que fecEmision
        saltara un día respecto a horEmision (que sí fuerza América/Lima)."""
        return (
            pytz.utc.localize(fields.Datetime.now())
            .astimezone(pytz.timezone("America/Lima"))
            .date()
        )

    def _l10n_pe_cabecera(self):
        fmt = self._l10n_pe_fmt
        partner = self.partner_id
        # El ICBPER (cat. 05 = 7152) SÍ entra en el total de tributos (sumTotTributos), en el precio de
        # venta (TaxInclusiveAmount) y en el importe a cobrar — regla SUNAT 3279/3280 (ref. enterprise:
        # ICBPER es tributo 'OTH', no allowance-charge). Ademas se emite como su propio TaxSubtotal de
        # cabecera (ver _l10n_pe_tributos). amount_tax/amount_total de Odoo ya lo incluyen.
        # Anticipo aplicado: el IGV de cabecera se reduce por el IGV del anticipo; el importe a cobrar
        # (PayableAmount) = precio de venta completo − total del anticipo (que va como PrepaidAmount).
        ant = self._l10n_pe_anticipo()
        anticipo_total = ant[2] if ant else 0.0
        anticipo_igv = ant[1] if ant else 0.0
        # Operación gratuita: el valor de los bienes regalados NO se cobra → se excluye de valor venta,
        # precio, importe Y del total de tributos de cabecera. El IGV teórico (18%) solo vive en la
        # TaxSubtotal 9996 (línea y cabecera); el cbc:TaxAmount de cabecera (sumTotTributos) NO lo
        # incluye: la regla 4301 suma únicamente los tributos 1000/1016/7152/9999/2000 (no el 9996),
        # y la referencia SUNAT aceptada consigna sumTotTributos = IGV real, sin el 18% gratuito.
        grat_base = self._l10n_pe_gratuito_base()
        # Descuento global que NO afecta la base del IGV: baja el importe a cobrar (MtoImpVenta) y va
        # como AllowanceCharge global (sumDescTotal), SIN tocar la base gravada ni el IGV. Mismo estilo
        # de ajuste solo-de-emisión que el anticipo (no agrega línea a Odoo).
        desc_no_afecta = self._l10n_pe_desc_no_afecta()
        cabecera = {
            "tipOperacion": self._l10n_pe_tipo_operacion(),
            "fecEmision": self.invoice_date.strftime("%Y-%m-%d")
            if self.invoice_date
            else "",
            # Hora de emisión en hora local de Perú (América/Lima, UTC-5). `fields.Datetime.now()`
            # es UTC-naive: sin convertir, el comprobante salía +5h (bug de zona horaria).
            "horEmision": pytz.utc.localize(fields.Datetime.now())
            .astimezone(pytz.timezone("America/Lima"))
            .strftime("%H:%M:%S"),
            "fecVencimiento": self.invoice_date_due.strftime("%Y-%m-%d")
            if self.invoice_date_due
            else "",
            "codLocalEmisor": (self.l10n_pe_ne_cod_establecimiento or "0000"),
            "tipDocUsuario": self._l10n_pe_cliente_doc()[0],
            "numDocUsuario": self._l10n_pe_cliente_doc()[1],
            "rznSocialUsuario": partner.name or "",
            "tipMoneda": self.currency_id.name or "PEN",
            # El IGV teórico del gratuito NO se cobra: NO entra en el total de tributos de cabecera
            # (regla 4301: el TaxAmount de cabecera excluye el 9996). El 9996 va solo como TaxSubtotal.
            "sumTotTributos": fmt(self.amount_tax - anticipo_igv),
            "sumTotValVenta": fmt(self.amount_untaxed - grat_base),
            # TaxInclusiveAmount: INCLUYE el ICBPER (igual que la ref. enterprise: PayableAmount =
            # TaxInclusive − anticipo, ambos con el ICBPER). Excluirlo de aquí pero incluirlo en
            # sumImpVenta desbalancea el comprobante → SUNAT Client.3280.
            "sumPrecioVenta": fmt(self.amount_total - grat_base),
            "sumImpVenta": fmt(
                self.amount_total - anticipo_total - grat_base - desc_no_afecta
            ),
            "sumDescTotal": fmt(desc_no_afecta),
            "sumOtrosCargos": "0.00",
            "sumTotalAnticipos": fmt(anticipo_total),
            "ublVersionId": "2.1",
            "customizationId": "2.0",
        }
        if grat_base:
            cabecera["sumValVentaGratuito"] = fmt(grat_base)
        adicional = self._l10n_pe_adicional_cabecera()
        if adicional:
            cabecera["adicionalCabecera"] = adicional
        return cabecera

    def _l10n_pe_serie_correlativo(self):
        """Serie y correlativo del comprobante. Una vez emitido, la identidad fiscal es
        inmutable: se devuelve la serie/correlativo CONGELADOS (l10n_pe_ne_serie/corr_emit),
        que ahora salen de una secuencia POR SERIE (ver _l10n_pe_ne_assign_numero). Para un
        move aún no emitido (previsualización) se cae al comportamiento anterior: el manual si
        se fijó; si no, el folio (parte numérica final) del número del asiento; si no hay, '1'."""
        self.ensure_one()
        # Retrocompatible: en los comprobantes históricos corr_emit == folio, así que esto
        # devuelve el mismo valor de antes; solo las emisiones nuevas usan la secuencia por serie.
        if self.l10n_pe_ne_serie_emit and self.l10n_pe_ne_corr_emit:
            try:
                return self.l10n_pe_ne_serie_emit, str(int(self.l10n_pe_ne_corr_emit))
            except (TypeError, ValueError):
                return self.l10n_pe_ne_serie_emit, self.l10n_pe_ne_corr_emit
        name = (self.name or "").replace(" ", "")
        matches = list(re.finditer(r"\d+", name))
        folio = matches[-1].group() if matches else None
        serie = self.l10n_pe_serie or self.journal_id.l10n_pe_ne_serie or "F001"
        correlativo = self.l10n_pe_correlativo or folio or "1"
        return serie, correlativo

    def _l10n_pe_ne_next_correlativo(self, company, serie):
        """Correlativo por (compañía, serie): SUNAT exige numeración correlativa POR SERIE y por
        RUC. Con un contador global (el folio del diario) la serie F001 se saltaba números cuando
        una boleta B001 o una nota FC01 tomaban el correlativo intermedio (hueco por serie → riesgo
        de observación en el RVIE). Crea una ir.sequence 'no_gap' al primer uso, sembrada tras el
        correlativo más alto ya emitido en esa serie (migración transparente desde el folio global).
        Mismo patrón, ya probado, que las Guías de Remisión (l10n_pe_ne_guia_remision)."""
        code = "l10n_pe.ne.cpe.%s" % serie
        # Lock consultivo: serializa el primer uso de una (serie, compañía) para no crear la
        # secuencia dos veces en concurrencia; después la unicidad la garantiza 'no_gap' (que
        # bloquea la fila de ir_sequence en cada next_by_id → dos cajas no obtienen el mismo nº).
        self.env.cr.execute(
            "SELECT pg_advisory_xact_lock(hashtext(%s))",
            ("%s/%s" % (code, company.id),),
        )
        Seq = self.env["ir.sequence"].sudo()
        seq = Seq.search(
            [("code", "=", code), ("company_id", "=", company.id)], limit=1
        )
        if not seq:
            ultimo = 0
            for m in self.sudo().search(
                [
                    ("company_id", "=", company.id),
                    ("l10n_pe_ne_serie_emit", "=", serie),
                    ("l10n_pe_ne_corr_emit", "!=", False),
                ]
            ):
                try:
                    ultimo = max(ultimo, int(m.l10n_pe_ne_corr_emit or 0))
                except (TypeError, ValueError):
                    pass
            seq = Seq.create(
                {
                    "name": "CPE %s (%s)" % (serie, company.display_name),
                    "code": code,
                    "company_id": company.id,
                    "padding": 1,
                    "number_increment": 1,
                    "implementation": "no_gap",
                    "number_next": ultimo + 1,
                }
            )
        return str(seq.next_by_id())

    def _l10n_pe_ne_assign_numero(self):
        """Fija (una sola vez) la serie+correlativo FISCAL antes de construir el payload/firmar.
        Idempotente: si ya está asignado no hace nada. Respeta un correlativo manual si se fijó;
        si no, lo toma de la secuencia POR SERIE. A partir de aquí _l10n_pe_serie_correlativo()
        devuelve estos valores congelados en todo el flujo (payload, XML, QR, PDF, baja)."""
        self.ensure_one()
        if self.l10n_pe_ne_corr_emit:
            return
        serie = self.l10n_pe_serie or self.journal_id.l10n_pe_ne_serie or "F001"
        if self.l10n_pe_correlativo:
            corr = str(self.l10n_pe_correlativo).strip()
        else:
            corr = self._l10n_pe_ne_next_correlativo(self.company_id, serie)
        self.l10n_pe_ne_serie_emit = serie
        self.l10n_pe_ne_corr_emit = corr.zfill(8)

    def _l10n_pe_id_block(self, with_document_type=True):
        serie, correlativo = self._l10n_pe_serie_correlativo()
        block = {
            "ruc": self.company_id.vat or "",
            "serie": serie,
            "correlativo": correlativo.zfill(8),
        }
        if with_document_type:
            block["documentType"] = self._l10n_pe_document_type()
        return block

    # ----------------------------------------------------------- constructores
    def _l10n_pe_emisor(self):
        """Datos de empresa del emisor (desde res.company) para el request. Las credenciales y el
        certificado de firma quedan en el servidor indexados por RUC; aquí solo van datos NO secretos.
        El microservicio prefiere estos sobre su registro por RUC, campo a campo."""
        self.ensure_one()
        company = self.company_id
        partner = company.partner_id
        emisor = {
            "razonSocial": company.name or "",
            "nombreComercial": company.name or "",
        }
        # Dirección todo-o-nada: solo se envía si el distrito (ubigeo) está configurado, para no mezclar
        # datos reales con los del registro del micro campo a campo (coalesce).
        distrito = partner.l10n_pe_district
        if distrito:
            emisor["direccion"] = {
                "ubigeo": distrito.code or "",
                "direccion": partner.street or "",
                "departamento": partner.state_id.name or "",
                "provincia": (distrito.city_id.name or partner.city or ""),
                "distrito": distrito.name or "",
                "urbanizacion": partner.street2 or "",
            }
        return emisor

    def _l10n_pe_build_invoice_request(self):
        """Factura (01) / Boleta (03) — endpoint /generator/factura."""
        _logger.info("---------------------------------------- Invoice request ------------------------------------------------")
        _logger.info(
            "%s %s %s",
            self._l10n_pe_id_block(with_document_type=True),
            self._l10n_pe_emisor(),
            self._l10n_pe_cabecera(),
        )
        _logger.info("---------------------------------------- Invoice request ------------------------------------------------")
        self.ensure_one()
        self._l10n_pe_check_lineas_impuesto()
        self._l10n_pe_check_anticipo()
        # Boleta > S/700 exige el documento de identidad del cliente (SUNAT la rechaza sin él en prod).
        _logger.info("Product lines: %s", len(self._l10n_pe_product_lines()))
        if (
            self._l10n_pe_document_type() == "03"
            and (self.amount_total or 0.0) > 700
            and not (self.partner_id.vat or "").strip()
        ):
            raise UserError(
                _(
                    "Una boleta mayor a S/ 700 requiere el documento de identidad del cliente."
                )
            )
        req = {
            "id": self._l10n_pe_id_block(with_document_type=True),
            "emisor": self._l10n_pe_emisor(),
            "cabecera": self._l10n_pe_cabecera(),
            "datoPago": self._l10n_pe_dato_pago(),
            "tributos": self._l10n_pe_tributos(),
            "detalle": self._l10n_pe_detalle(),
            "adicionalDetalle": self._l10n_pe_adicional_detalle(),
            "variablesGlobales": self._l10n_pe_variables_globales(),
            "leyendas": self._l10n_pe_leyendas(),
        }
        if self.l10n_pe_ne_forma_pago == "Credito":
            req["detallePago"] = self._l10n_pe_detalle_pago()
        relacionados = self._l10n_pe_relacionados()
        if relacionados:
            req["relacionados"] = relacionados
        return req

    def _l10n_pe_adicional_detalle(self):
        """Descuentos por ítem (cat. 53 código 00, que afecta la base del IGV) — hace explícito en
        el comprobante el descuento de cada línea con `discount` > 0. La línea ya va por su valor
        neto (IGV sobre el neto); este bloque solo lo muestra, no cambia los totales."""
        fmt = self._l10n_pe_fmt
        moneda = self.currency_id.name or "PEN"
        out = []
        idx = 0
        for line in self._l10n_pe_product_lines():
            idx += 1
            if not line.discount:
                continue
            gross = round(line.price_unit * line.quantity, 2)
            disc = round(gross - line.price_subtotal, 2)
            out.append(
                {
                    "idLinea": str(idx),
                    # "-" en las propiedades para que la plantilla salte el bloque AdditionalItemProperty
                    # (la misma lista sirve para descuentos y propiedades; sin esto el render falla).
                    "nomPropiedad": "-",
                    "codBienPropiedad": "-",
                    "tipVariable": "false",
                    "codTipoVariable": "00",
                    # Factor con 5 decimales: SUNAT valida mtoVariable ≈ base × porVariable (error 3290,
                    # "cargo/descuento por ítem difiere"). Con 2 decimales, un descuento en monto fijo
                    # (p.ej. S/50 sobre 470 → 10.6383% → 0.11) descuadra y se rechaza; 5 decimales
                    # reconstruyen el monto dentro de la tolerancia.
                    "porVariable": "%.5f" % (line.discount / 100.0),
                    "monMontoVariable": moneda,
                    "mtoVariable": fmt(disc),
                    "monBaseImponibleVariable": moneda,
                    "mtoBaseImpVariable": fmt(gross),
                }
            )
        return out

    def _l10n_pe_build_note_request(self):
        """Nota de Crédito (07) / Débito (08) — referencia al documento afectado."""
        self.ensure_one()
        self._l10n_pe_check_lineas_impuesto()
        dt = self._l10n_pe_document_type()
        origin = self.reversed_entry_id if dt == "07" else self.debit_origin_id
        cabecera = self._l10n_pe_cabecera()
        if origin:
            o_serie, o_corr = origin._l10n_pe_serie_correlativo()
            cabecera["numDocAfectado"] = "%s-%s" % (o_serie, o_corr.zfill(8))
            cabecera["tipDocAfectado"] = origin._l10n_pe_document_type()
        else:
            cabecera["numDocAfectado"] = ""
            cabecera["tipDocAfectado"] = "01"
        cabecera["codMotivo"] = self.l10n_pe_motivo_code or (
            "01" if dt == "07" else "02"
        )
        if dt == "08":
            # Sustento libre si el usuario lo escribió; si no, descripción del catálogo.
            cabecera["desMotivo"] = (self.l10n_pe_motivo_desc or "").strip() or ND_MOTIVO_DESC.get(
                cabecera["codMotivo"], "Aumento en el valor"
            )
        req = {
            "id": self._l10n_pe_id_block(with_document_type=False),
            "emisor": self._l10n_pe_emisor(),
            "cabecera": cabecera,
            "tributos": self._l10n_pe_tributos(),
            "detalle": self._l10n_pe_detalle(),
            "leyendas": self._l10n_pe_leyendas(),
        }
        # Nota de Crédito (07): el CreditNoteMapper del biller exige forma de pago y
        # fuerza el <cbc:Amount> de PaymentTerms. "Contado" rebota (errorCode 2071/3246)
        # y omitirlo rebota (3245). El único patrón que valida en el SFS es "Credito"
        # con una cuota = total (campos válidos del contrato SFS, no se toca el biller).
        # La ND (08) valida sin datoPago, así que no se le agrega.
        # EXCEPCIÓN: una NC de importe 0 (motivo 03, corrección de descripción) NO puede
        # llevar el Amount de la cuota Crédito (SUNAT rechaza cac:PaymentTerms/cbc:Amount
        # "0.00"), y omitir la FormaPago rebota con errorCode 3245. El patrón válido es
        # "Contado" SIN <cbc:Amount>. El mapper del biller (GenericBillingMapper) defaultea
        # el monto a "0.00" y la moneda a "" salvo que se le mande el sentinel "-", que le
        # dice que NO setee esos campos → el FTL entonces omite el <cbc:Amount>.
        if dt == "07":
            if self.amount_total:
                total = self._l10n_pe_fmt(self.amount_total)
                fecha = self.invoice_date.strftime("%Y-%m-%d") if self.invoice_date else ""
                moneda = self.currency_id.name or "PEN"
                req["datoPago"] = {
                    "formaPago": "Credito",
                    "mtoNetoPendientePago": total,
                    "tipMonedaMtoNetoPendientePago": moneda,
                }
                req["detallePago"] = [
                    {
                        "mtoCuotaPago": total,
                        "fecCuotaPago": fecha,
                        "tipMonedaCuotaPago": moneda,
                    }
                ]
            else:
                # NC de importe 0 (motivo 03): SUNAT exige FormaPago=Credito con Amount>0
                # (Contado→3246, omitir→3245, Amount 0.00→2071). Se referencia el total del
                # comprobante afectado como monto de la cuota (el documento en sí va en 0).
                ref = self._l10n_pe_fmt((origin.amount_total if origin else 0) or 0)
                fecha = self.invoice_date.strftime("%Y-%m-%d") if self.invoice_date else ""
                moneda = self.currency_id.name or "PEN"
                req["datoPago"] = {
                    "formaPago": "Credito",
                    "mtoNetoPendientePago": ref,
                    "tipMonedaMtoNetoPendientePago": moneda,
                }
                req["detallePago"] = [
                    {"mtoCuotaPago": ref, "fecCuotaPago": fecha, "tipMonedaCuotaPago": moneda}
                ]
        return req

    def _l10n_pe_target(self):
        """(endpoint, payload) según el tipo de comprobante."""
        self._l10n_pe_check_serie()
        dt = self._l10n_pe_document_type()
        if dt == "07":
            return ("notaCredito", self._l10n_pe_build_note_request())
        if dt == "08":
            return ("notaDebito", self._l10n_pe_build_note_request())
        return ("factura", self._l10n_pe_build_invoice_request())

    def _l10n_pe_store_cdr(self, cdr_b64):
        """Guarda el CDR de SUNAT (zip en base64, del header X-Sunat-Cdr) como adjunto en
        l10n_pe_biller_cdr y devuelve (responseCode, description) del ApplicationResponse."""
        self.ensure_one()
        try:
            cdr_bytes = base64.b64decode(cdr_b64)
        except Exception:
            return "", ""
        serie, correlativo = self._l10n_pe_serie_correlativo()
        name = "R%s-%s-%s.zip" % (
            self.company_id.vat or "",
            serie,
            correlativo.zfill(8),
        )
        att = self.env["ir.attachment"].create(
            {
                "name": name,
                "res_model": "account.move",
                "res_id": self.id,
                "mimetype": "application/zip",
                "raw": cdr_bytes,
            }
        )
        self.l10n_pe_biller_cdr = att.id
        return self._l10n_pe_parse_cdr_codes(cdr_bytes)

    def _l10n_pe_parse_cdr_codes(self, cdr_bytes):
        """(responseCode, description) del ApplicationResponse dentro del zip CDR."""
        code = desc = ""
        try:
            with zipfile.ZipFile(io.BytesIO(cdr_bytes)) as zf:
                xml_name = next(
                    (n for n in zf.namelist() if n.lower().endswith(".xml")), None
                )
                content = zf.read(xml_name) if xml_name else b""
            m = re.search(rb"<cbc:ResponseCode>([^<]*)</cbc:ResponseCode>", content)
            code = m.group(1).decode() if m else ""
            m = re.search(rb"<cbc:Description>([^<]*)</cbc:Description>", content)
            desc = m.group(1).decode("utf-8", "replace") if m else ""
        except Exception:
            pass
        return code, desc

    def _l10n_pe_apply_emission_response(self, ok, body_text, cdr_b64):
        """Aplica al move el resultado de una emisión — mismo tratamiento para el
        flujo síncrono (respuesta HTTP directa) y el asíncrono (cron que recoge
        XML/CDR desde S3 vía el worker): adjunta el XML firmado, guarda el CDR,
        congela la identidad emitida y fija estado + mensaje."""
        self.ensure_one()
        signed = ok and any(
            tag in (body_text or "")
            for tag in ("<Invoice", "<CreditNote", "<DebitNote")
        )
        if not signed:
            self.l10n_pe_biller_state = "rechazado"
            self.l10n_pe_biller_message = (body_text or "")[:2000]
            return
        serie, correlativo = self._l10n_pe_serie_correlativo()
        # Si el XML ya se adjuntó estando en_proceso (firma del modo instant o
        # item "firmado" del worker async), reemplazarlo: sin esto quedaban DOS
        # adjuntos idénticos colgados del move (el viejo huérfano en el panel).
        if self.l10n_pe_biller_xml:
            # Contenido distinto = re-emisión con XML corregido: los PDFs
            # cacheados renderizan el XML viejo y quedarían servidos por
            # siempre (el cache pdfver no detecta cambios de contenido).
            if (self.l10n_pe_biller_xml.raw or b"") != body_text.encode("utf-8"):
                self._l10n_pe_invalidar_pdfs()
            self.l10n_pe_biller_xml.unlink()
        att = self.env["ir.attachment"].create(
            {
                "name": "%s-%s-%s.xml"
                % (self.company_id.vat, serie, correlativo.zfill(8)),
                "res_model": "account.move",
                "res_id": self.id,
                "mimetype": "application/xml",
                "raw": body_text.encode("utf-8"),
            }
        )
        self.l10n_pe_biller_xml = att.id
        self.l10n_pe_biller_state = "enviado"
        # Congela la identidad emitida para una eventual baja (no recomputar luego del partner/nombre).
        self.l10n_pe_ne_tipo_doc = self._l10n_pe_document_type()
        self.l10n_pe_ne_serie_emit = serie
        self.l10n_pe_ne_corr_emit = correlativo.zfill(8)
        code, desc = self._l10n_pe_store_cdr(cdr_b64) if cdr_b64 else ("", "")
        if code == "0":
            self.l10n_pe_biller_message = _(
                "Aceptado por SUNAT — CDR ResponseCode 0. %s"
            ) % (desc or "")
            # Automatización (opt-in): al aceptarse, enviar el comprobante (XML + PDF + CDR) al
            # correo del cliente. Gateado por config para no mandar correos sin querer; nunca
            # rompe la emisión (un fallo de correo se loguea y sigue).
            if self.env["ir.config_parameter"].sudo().get_param(
                "l10n_pe_ne_biller.email_on_accept", ""
            ).strip().lower() in ("1", "true"):
                try:
                    self._l10n_pe_ne_email_comprobante()
                except Exception as e:  # noqa: BLE001
                    _logger.warning("email comprobante %s: %s", self.name, e)
        elif code:
            self.l10n_pe_biller_message = _(
                "CDR de SUNAT (ResponseCode %s). %s"
            ) % (code, desc or "")
        else:
            self.l10n_pe_biller_message = _("Aceptado por el facturador (HTTP 200).")

    def _l10n_pe_ne_email_comprobante(self):
        """Envía el comprobante aceptado (XML firmado + PDF A4 + CDR) al correo del cliente.
        Automatiza la entrega manual. No-op si el cliente no tiene correo; nunca lanza (el
        llamador lo envuelve, pero igual usamos send sin excepción)."""
        self.ensure_one()
        email = (self.partner_id.email or "").strip()
        if not email:
            _logger.info("email comprobante %s: cliente sin correo, se omite", self.name)
            return False
        atts = self.env["ir.attachment"]
        if self.l10n_pe_biller_xml:
            atts |= self.l10n_pe_biller_xml
        try:
            pdf = self._l10n_pe_get_pdf_attachment(formato="A4")
            if pdf:
                atts |= pdf
        except Exception:  # noqa: BLE001 — el PDF es deseable pero no bloquea el correo
            pass
        if self.l10n_pe_biller_cdr:
            atts |= self.l10n_pe_biller_cdr
        serie, corr = self._l10n_pe_serie_correlativo()
        num = "%s-%s" % (serie, corr)
        subject = _("Comprobante electrónico %s") % num
        body = _(
            "<p>Estimado cliente,</p>"
            "<p>Adjuntamos su comprobante electrónico <b>%(num)s</b> emitido por "
            "<b>%(emisor)s</b> y aceptado por SUNAT.</p>"
            "<p>Se incluyen el XML firmado, la representación impresa (PDF) y el CDR.</p>"
        ) % {"num": num, "emisor": self.company_id.name or ""}
        mail = self.env["mail.mail"].sudo().create({
            "subject": subject,
            "body_html": body,
            "email_to": email,
            "email_from": self.company_id.email or self.env.user.email_formatted,
            "attachment_ids": [(6, 0, atts.ids)],
            "auto_delete": False,
        })
        mail.send(raise_exception=False)
        _logger.info("email comprobante %s enviado a %s (%d adjuntos)", num, email, len(atts))
        return True

    def _l10n_pe_apply_signed(self, firma):
        """Modo instantáneo: aplica el resultado de la FIRMA (sin enviar a SUNAT). Adjunta el
        XML firmado (con eso el ticket/PDF ya funcionan), congela la identidad, guarda el ZIP
        de ENVI + filename/canal para el envío en 2º plano y deja el estado en 'en_proceso'."""
        self.ensure_one()
        firma = firma or {}
        xml = firma.get("xmlFirmado") or ""
        if not any(tag in xml for tag in ("<Invoice", "<CreditNote", "<DebitNote")):
            self.l10n_pe_biller_state = "error"
            self.l10n_pe_biller_message = _("La firma no devolvió un XML válido.")
            return False
        serie, correlativo = self._l10n_pe_serie_correlativo()
        # Re-firma (re-emisión tras rechazo/error en modo instant): reemplaza el
        # XML anterior (evita el adjunto huérfano) e invalida los PDFs cacheados
        # del intento previo antes de pre-generar los nuevos.
        if self.l10n_pe_biller_xml:
            if (self.l10n_pe_biller_xml.raw or b"") != xml.encode("utf-8"):
                self._l10n_pe_invalidar_pdfs()
            self.l10n_pe_biller_xml.unlink()
        att = self.env["ir.attachment"].create(
            {
                "name": "%s-%s-%s.xml" % (self.company_id.vat, serie, correlativo.zfill(8)),
                "res_model": "account.move",
                "res_id": self.id,
                "mimetype": "application/xml",
                "raw": xml.encode("utf-8"),
            }
        )
        self.l10n_pe_biller_xml = att.id
        self.l10n_pe_ne_tipo_doc = self._l10n_pe_document_type()
        self.l10n_pe_ne_serie_emit = serie
        self.l10n_pe_ne_corr_emit = correlativo.zfill(8)
        self.l10n_pe_ne_envi_zip = firma.get("enviZip") or ""
        self.l10n_pe_ne_biller_filename = firma.get("filename") or ""
        self.l10n_pe_ne_biller_canal = firma.get("canal") or "GEM"
        self.l10n_pe_ne_envio_intentos = 0
        self.l10n_pe_biller_state = "en_proceso"
        self.l10n_pe_biller_message = _("Firmado — ticket listo. Pendiente de envío a SUNAT.")
        # Pre-generar la representación impresa YA (con el XML firmado) para que la
        # descarga sea instantánea: así el adjunto existe cuando el usuario da clic y
        # no depende de un cold-start del micro en ese momento (que llegaba a expirar y
        # dejaba la sensación de "no se puede descargar mientras procesa"). No es fatal:
        # si el micro falla aquí, queda como fallback la generación on-demand.
        try:
            self._l10n_pe_get_pdf_attachment()  # A4
            if self.l10n_pe_ne_tipo_doc in ("01", "03"):
                self._l10n_pe_get_pdf_attachment(formato="TICKET")  # 80mm
        except Exception as exc:  # noqa: BLE001
            _logger.warning(
                "No se pudo pre-generar el PDF tras firmar %s: %s",
                self.name or self.id, exc,
            )
        return True

    @api.model
    def _l10n_pe_cron_enviar_pendientes(self):
        """Modo instantáneo: envía a SUNAT (2º plano) los comprobantes ya FIRMADOS que quedaron
        en 'en_proceso' con ZIP pendiente. Al recibir el CDR pasa a aceptado/rechazado y limpia
        el ZIP. Reintentable: un fallo de red deja el move en en_proceso para la próxima corrida
        (con tope de intentos para no reintentar por siempre un rechazo)."""
        icp = self.env["ir.config_parameter"].sudo()
        if icp.get_param("l10n_pe_ne_biller.instant_enabled", "").strip().lower() not in ("1", "true"):
            return
        base = icp.get_param("l10n_pe_ne_biller.url", "http://localhost:8090").rstrip("/")
        timeout = int(icp.get_param("l10n_pe_ne_biller.timeout", "240"))
        max_intentos = int(icp.get_param("l10n_pe_ne_biller.max_intentos_envio", "30"))
        domain = [("l10n_pe_biller_state", "=", "en_proceso"), ("l10n_pe_ne_envi_zip", "!=", False)]
        # Si las boletas van por Resumen Diario, se excluyen del envío individual (las manda el RC).
        if icp.get_param("l10n_pe_ne_biller.boletas_resumen", "").strip().lower() in ("1", "true"):
            domain.append(("l10n_pe_ne_tipo_doc", "!=", "03"))
        pend = self.search(domain, limit=50)
        for move in pend:
            headers = {"X-Api-Key": move.company_id.sudo().l10n_pe_ne_api_key or ""}
            signed_xml = (move.l10n_pe_biller_xml.raw or b"").decode("utf-8") if move.l10n_pe_biller_xml else ""
            body = {
                "ruc": move.company_id.vat or "",
                "filename": move.l10n_pe_ne_biller_filename or "",
                "canal": move.l10n_pe_ne_biller_canal or "GEM",
                "enviZip": move.l10n_pe_ne_envi_zip or "",
                "signedXml": signed_xml,
            }
            ok = False
            try:
                resp = requests.post(base + "/generator/enviar", json=body, headers=headers, timeout=(5, timeout))
                if resp.status_code == 200:
                    data = resp.json() or {}
                    if data.get("rechazado"):
                        # SUNAT rechazó (regla de negocio) → estado final, NO reintentar.
                        move.l10n_pe_biller_state = "rechazado"
                        move.l10n_pe_biller_message = (_("Rechazado por SUNAT: %s") % (data.get("motivo") or ""))[:2000]
                        move.l10n_pe_ne_envi_zip = False
                    else:
                        move._l10n_pe_apply_emission_response(True, signed_xml, data.get("cdr") or "")
                        move.l10n_pe_ne_envi_zip = False  # enviado; nada pendiente
                    ok = True
                else:
                    move.l10n_pe_biller_message = ("Envío HTTP %s: %s" % (resp.status_code, resp.text))[:2000]
            except Exception as e:  # noqa: BLE001 — red/SUNAT: reintentar
                _logger.warning("enviar pendiente %s: %s (reintenta)", move.name, e)
                move.l10n_pe_biller_message = ("Reintentando envío: %s" % e)[:2000]
            if not ok:
                move.l10n_pe_ne_envio_intentos = (move.l10n_pe_ne_envio_intentos or 0) + 1
                if move.l10n_pe_ne_envio_intentos >= max_intentos:
                    move.l10n_pe_biller_state = "error"
            self.env["bus.bus"]._sendone(
                "l10n_pe_biller_updates",
                "l10n_pe_biller_update",
                {"move_id": move.id, "state": move.l10n_pe_biller_state},
            )
            self.env.cr.commit()

    # -------------------------------------------------------- emisión asíncrona
    # Toggle: ir.config_parameter `l10n_pe_ne_biller.async_enabled` = "1".
    # Odoo encola en SQS (rol IAM del EC2, patrón del sibling partner_lookup) y
    # responde al instante; el Lambda facturas-worker procesa contra biller-core
    # con idempotencia (DynamoDB) y deja XML/CDR en S3; el cron de abajo recoge.

    @api.model
    def _l10n_pe_boto_client(self, service, region):
        """Cliente boto3 memoizado por (service, region). Crear un cliente
        cuesta 100-400ms de CPU (carga los modelos JSON del servicio) y se
        pagaba dos veces POR EMISIÓN (dynamodb + sqs). El cache vive por
        worker de Odoo (prefork: se puebla post-fork, sin estado compartido
        entre procesos; los clientes boto3 son thread-safe para invocar).
        Se reconstruye si el módulo boto3 cambió (tests que lo parchean)."""
        key = (service, region)
        cached = _BOTO_CLIENTS.get(key)
        if cached is not None and cached[0] is boto3:
            return cached[1]
        client = boto3.client(service, region_name=region)
        _BOTO_CLIENTS[key] = (boto3, client)
        return client

    def _l10n_pe_enqueue_emission(self, icp):
        self.ensure_one()
        queue_url = icp.get_param("l10n_pe_ne_biller.sqs_queue_url", "")
        region = icp.get_param("l10n_pe_ne_biller.aws_region", "us-east-1")
        if not boto3 or not queue_url:
            self.l10n_pe_biller_state = "error"
            self.l10n_pe_biller_message = _(
                "Modo asíncrono activo pero falta boto3 o el parámetro "
                "l10n_pe_ne_biller.sqs_queue_url."
            )
            return
        endpoint, payload = self._l10n_pe_target()
        serie, correlativo = self._l10n_pe_serie_correlativo()
        msg = {
            "ruc": self.company_id.vat or "",
            "serie_correlativo": "%s-%s" % (serie, correlativo.zfill(8)),
            "db": self.env.cr.dbname,
            "move_id": self.id,
            "path": "/generator/" + endpoint,
            "api_key": self.company_id.sudo().l10n_pe_ne_api_key or "",
            # tipoDoc (01/03/07/08) para que el worker pre-genere el PDF
            "doc_type": self._l10n_pe_document_type(),
            "payload": payload,
        }
        # Reintento tras un rechazo: borra el resultado viejo ANTES de encolar,
        # para que el cron no aplique el resultado obsoleto mientras el worker
        # procesa el intento nuevo (best-effort: si no existe, no pasa nada).
        table = icp.get_param("l10n_pe_ne_biller.results_table", "")
        if table:
            try:
                self._l10n_pe_boto_client("dynamodb", region).delete_item(
                    TableName=table,
                    Key={
                        "ruc_emisor": {"S": msg["ruc"]},
                        "serie_correlativo": {"S": msg["serie_correlativo"]},
                    },
                )
            except Exception as exc:  # noqa: BLE001
                _logger.warning("async biller: no se pudo limpiar resultado previo: %s", exc)
        try:
            self._l10n_pe_boto_client("sqs", region).send_message(
                QueueUrl=queue_url,
                MessageBody=json.dumps(msg, ensure_ascii=False),
            )
        except Exception as exc:  # noqa: BLE001
            self.l10n_pe_biller_state = "error"
            self.l10n_pe_biller_message = _("No se pudo encolar la emisión: %s") % exc
            return
        # Re-emisión tras rechazado/error: el XML firmado y los PDFs del intento
        # anterior quedan obsoletos (el worker firmará uno nuevo). Sin esto, el
        # cache pdfver serviría la representación vieja para siempre y el PDF
        # nuevo del worker jamás se adjuntaría (guard "ya hay PDF" del attach).
        if self.l10n_pe_biller_xml:
            self.l10n_pe_biller_xml.unlink()
        self._l10n_pe_invalidar_pdfs()
        self.l10n_pe_biller_state = "en_proceso"
        self.l10n_pe_biller_message = _(
            "Encolado para envío a SUNAT — el resultado llega en unos minutos "
            "(aparece en el chatter; recargá la vista para ver el estado final)."
        )
        self._l10n_pe_trigger_poll_async(seconds=20)

    @api.model
    def _l10n_pe_trigger_poll_async(self, seconds=20):
        """Adelanta el próximo run del cron de recogida: sin esto el resultado
        espera el beat base de 2 min aunque el worker ya lo haya dejado en
        DynamoDB. Ojo con la expectativa: el scheduler de Odoo duerme beats
        fijos de ~60s y un trigger futuro NO lo despierta a call_at (ni con
        ODOO_NOTIFY_CRON_CHANGES: ese NOTIFY sale al commit, cuando el trigger
        aún no venció) — el pickup real es el primer beat posterior a call_at,
        o sea hasta ~60-70s después. Best-effort: si falla, el beat base sigue."""
        try:
            self.env.ref(
                "l10n_pe_ne_biller.ir_cron_l10n_pe_ne_poll_async"
            ).sudo()._trigger(at=fields.Datetime.now() + timedelta(seconds=seconds))
        except Exception as exc:  # noqa: BLE001
            _logger.warning("async biller: no se pudo adelantar el cron: %s", exc)

    def _l10n_pe_pdf_ver(self):
        """Etiqueta de versión del template de PDF (`description` del adjunto):
        _l10n_pe_get_pdf_attachment solo sirve el cache si coincide con el
        `pdf_ver` vigente; cualquier PDF que se adjunte debe llevarla."""
        return "pdfver:" + self.env["ir.config_parameter"].sudo().get_param(
            "l10n_pe_ne_biller.pdf_ver", "1"
        )

    def _l10n_pe_invalidar_pdfs(self):
        """Descarta los PDFs cacheados (A4 y ticket). Debe llamarse siempre que
        el XML firmado cambie (re-emisión tras rechazo/error): la representación
        impresa de un XML anterior no debe sobrevivir — el cache por `pdfver`
        solo detecta cambios de template, no de contenido."""
        self.ensure_one()
        for campo in ("l10n_pe_biller_pdf", "l10n_pe_biller_pdf_ticket"):
            att = self[campo]
            if att:
                try:
                    att.sudo().unlink()
                except Exception as exc:  # noqa: BLE001 — best-effort
                    _logger.warning(
                        "no se pudo descartar el PDF cacheado de %s: %s",
                        self.name, exc,
                    )

    def _l10n_pe_attach_async_pdf(self, s3c, bucket, item):
        """Adjunta el PDF pre-generado por el worker (pdf_s3_key del item), si
        ya existe y el move no tiene uno. Best-effort: si falta, el botón
        Descargar PDF cae al camino síncrono de siempre."""
        self.ensure_one()
        # El worker pre-genera el A4 SIN logo del emisor ni dirección del cliente (el mensaje
        # de la cola no los lleva, ver _l10n_pe_enqueue_emission). Si el emisor tiene logo o el
        # cliente tiene dirección, ese PDF saldría incompleto: NO lo adjuntamos y dejamos que la
        # descarga lo regenere por la ruta síncrona (_l10n_pe_get_pdf_attachment), que sí los
        # incluye. Si no hay nada que agregar, reusamos el del worker (más rápido, sin diferencia).
        if self.company_id.logo or self.partner_id.street or self.partner_id.street2:
            return
        pdf_s3 = (item.get("pdf_s3_key") or {}).get("S", "")
        if not pdf_s3 or self.l10n_pe_biller_pdf:
            return
        try:
            pdf_bytes = s3c.get_object(Bucket=bucket, Key=pdf_s3)["Body"].read()
            if not pdf_bytes.startswith(b"%PDF"):
                _logger.warning(
                    "async biller: pdf_s3_key de %s no es un PDF; se ignora",
                    self.name,
                )
                return
            serie = self.l10n_pe_ne_serie_emit
            corr = self.l10n_pe_ne_corr_emit
            if not serie or not corr:
                serie, corr = self._l10n_pe_serie_correlativo()
                corr = corr.zfill(8)
            att = self.env["ir.attachment"].create(
                {
                    "name": "%s-%s-%s.pdf"
                    % (self.company_id.vat or "", serie, corr),
                    "res_model": "account.move",
                    "res_id": self.id,
                    "mimetype": "application/pdf",
                    "raw": pdf_bytes,
                    # Sin la etiqueta, la primera descarga vía API lo descartaba
                    # (cache-busting) y re-renderizaba contra el micro (~hasta 60s).
                    "description": self._l10n_pe_pdf_ver(),
                }
            )
            self.l10n_pe_biller_pdf = att.id
        except Exception as exc:  # noqa: BLE001 — PDF es best-effort
            _logger.warning(
                "async biller: PDF no adjuntado en %s: %s", self.name, exc
            )

    def _l10n_pe_async_attach_firmado(self, s3c, bucket, item):
        """Modo async: cuando el worker publica un item intermedio (status no
        terminal, p.ej. "firmado") con `xml_s3_key`, adjunta el XML firmado a
        `l10n_pe_biller_xml` y toma el PDF del worker si ya está (`pdf_s3_key`).
        Con el XML adjunto, la descarga funciona estando en_proceso aunque el PDF
        aún no llegue (el botón cae al camino on-demand de siempre). NO cambia el
        estado (sigue en_proceso) y NO genera el PDF localmente — el worker es el
        único generador; ver nota al final del cuerpo. Best-effort e idempotente:
        sin `xml_s3_key` no hace nada; con el XML ya adjunto solo intenta traer
        el PDF del worker."""
        self.ensure_one()
        if self.l10n_pe_biller_xml:
            # Ya adjuntado en una corrida previa: solo traer el PDF del worker si aún no está.
            self._l10n_pe_attach_async_pdf(s3c, bucket, item)
            return
        xml_key = (item.get("xml_s3_key") or {}).get("S", "")
        if not xml_key:
            return
        try:
            body = (
                s3c.get_object(Bucket=bucket, Key=xml_key)["Body"]
                .read()
                .decode("iso-8859-1")
            )
        except Exception as exc:  # noqa: BLE001 — aún no está en S3: se reintenta al próximo poll
            _logger.warning(
                "async biller: XML firmado aún no disponible en %s: %s", self.name, exc
            )
            return
        if not any(tag in body for tag in ("<Invoice", "<CreditNote", "<DebitNote")):
            return
        serie, correlativo = self._l10n_pe_serie_correlativo()
        att = self.env["ir.attachment"].create(
            {
                "name": "%s-%s-%s.xml"
                % (self.company_id.vat, serie, correlativo.zfill(8)),
                "res_model": "account.move",
                "res_id": self.id,
                "mimetype": "application/xml",
                # Normalizado a utf-8 igual que _l10n_pe_apply_emission_response, para que
                # el render del PDF (que decodifica utf-8) no rompa con tildes/ñ.
                "raw": body.encode("utf-8"),
            }
        )
        self.l10n_pe_biller_xml = att.id
        self.l10n_pe_ne_tipo_doc = self._l10n_pe_document_type()
        self.l10n_pe_ne_serie_emit = serie
        self.l10n_pe_ne_corr_emit = correlativo.zfill(8)
        # PDF: SOLO el pre-generado por el worker (pdf_s3_key). NO generarlo acá:
        # en la ventana "firmado" el worker ya está invocando biller-pdf con este
        # mismo XML — hacerlo también desde el cron duplicaba renders (A4+ticket
        # síncronos de hasta ~60s c/u DENTRO del loop del poll: un import masivo
        # bloqueaba el cron varios minutos) y el PDF del worker terminaba
        # descartado. Si el usuario descarga antes de que llegue, el botón usa el
        # camino on-demand de siempre — posible porque el XML ya quedó adjunto.
        self._l10n_pe_attach_async_pdf(s3c, bucket, item)

    @api.model
    def _l10n_pe_cron_poll_async(self):
        """Recoge resultados de emisiones asíncronas: lee el item del worker en
        DynamoDB (PK ruc_emisor / SK serie_correlativo) y, si terminó, baja el
        XML/CDR de S3 y lo aplica con el mismo código del flujo síncrono."""
        icp = self.env["ir.config_parameter"].sudo()
        if icp.get_param("l10n_pe_ne_biller.async_enabled", "").strip().lower() not in ("1", "true"):
            return
        table = icp.get_param("l10n_pe_ne_biller.results_table", "")
        bucket = icp.get_param("l10n_pe_ne_biller.results_bucket", "")
        region = icp.get_param("l10n_pe_ne_biller.aws_region", "us-east-1")
        if not boto3 or not table or not bucket:
            _logger.warning(
                "async biller: faltan parámetros results_table/results_bucket o boto3"
            )
            return
        ddb = self._l10n_pe_boto_client("dynamodb", region)
        s3c = self._l10n_pe_boto_client("s3", region)
        moves = self.search([("l10n_pe_biller_state", "=", "en_proceso")], limit=25)
        for move in moves:
            try:
                serie, correlativo = move._l10n_pe_serie_correlativo()
                key = {
                    "ruc_emisor": {"S": move.company_id.vat or ""},
                    "serie_correlativo": {"S": "%s-%s" % (serie, correlativo.zfill(8))},
                }
                item = ddb.get_item(TableName=table, Key=key).get("Item")
                if not item:  # aún en cola o procesándose
                    continue
                status = item["status"]["S"]
                if status == "enviado":
                    xml_key = (item.get("xml_s3_key") or {}).get("S", "")
                    body = (
                        s3c.get_object(Bucket=bucket, Key=xml_key)["Body"]
                        .read()
                        .decode("iso-8859-1")
                    )
                    cdr_b64 = ""
                    cdr_key = (item.get("cdr_s3_key") or {}).get("S", "")
                    if cdr_key:
                        cdr_b64 = base64.b64encode(
                            s3c.get_object(Bucket=bucket, Key=cdr_key)["Body"].read()
                        ).decode()
                    move._l10n_pe_apply_emission_response(True, body, cdr_b64)
                    # PDF pre-generado por el worker: el botón "Descargar PDF"
                    # lo sirve cacheado, sin llamada síncrona al facturador.
                    move._l10n_pe_attach_async_pdf(s3c, bucket, item)
                elif status in ("rechazado", "error"):
                    move.l10n_pe_biller_state = status
                    move.l10n_pe_biller_message = (
                        (item.get("message") or {}).get("S") or ""
                    )[:2000]
                else:
                    # Item intermedio (p.ej. "firmado"): el worker ya firmó pero SUNAT aún
                    # no responde. Adjunta el XML firmado + PDF para que ticket/PDF estén
                    # disponibles AL TOQUE en en_proceso, sin esperar el CDR. Sigue en
                    # en_proceso: sin transición de estado no se postea al chatter ni se
                    # notifica (evita spam en cada corrida mientras el item no es final).
                    move._l10n_pe_async_attach_firmado(s3c, bucket, item)
                    continue
                # El form no refresca solo cuando escribe un cron: el chatter sí.
                move.message_post(
                    body=_("Facturador (async): %s — %s")
                    % (
                        dict(
                            move._fields["l10n_pe_biller_state"].selection
                        ).get(move.l10n_pe_biller_state, move.l10n_pe_biller_state),
                        (move.l10n_pe_biller_message or "")[:500],
                    )
                )
                # ...y el statusbar en vivo va por el bus (websocket): el JS
                # biller_live_statusbar recarga el form abierto al recibir esto.
                self.env["bus.bus"]._sendone(
                    "l10n_pe_biller_updates",
                    "l10n_pe_biller_update",
                    {"move_id": move.id, "state": move.l10n_pe_biller_state},
                )
            except Exception as exc:  # noqa: BLE001 — un move malo no frena al resto
                _logger.warning("async biller: error procesando %s: %s", move.name, exc)
        # Segundo pase — PDFs rezagados: el worker publica "enviado" ANTES de
        # generar el PDF, así que el pase de arriba suele aplicar el resultado
        # cuando pdf_s3_key aún no existe; se re-lee el item hasta que aparezca
        # (ventana corta: biller-pdf tarda segundos, ~2 min en cold start).
        sin_pdf = self.search(
            [
                ("l10n_pe_biller_state", "=", "enviado"),
                ("l10n_pe_biller_pdf", "=", False),
                ("write_date", ">=", fields.Datetime.now() - timedelta(minutes=15)),
            ],
            limit=25,
        )
        for move in sin_pdf:
            try:
                serie = move.l10n_pe_ne_serie_emit
                corr = move.l10n_pe_ne_corr_emit
                if not serie or not corr:
                    serie, corr = move._l10n_pe_serie_correlativo()
                    corr = corr.zfill(8)
                item = ddb.get_item(
                    TableName=table,
                    Key={
                        "ruc_emisor": {"S": move.company_id.vat or ""},
                        "serie_correlativo": {"S": "%s-%s" % (serie, corr)},
                    },
                ).get("Item")
                if item:
                    move._l10n_pe_attach_async_pdf(s3c, bucket, item)
            except Exception as exc:  # noqa: BLE001
                _logger.warning(
                    "async biller: reconciliación PDF %s: %s", move.name, exc
                )
        # Re-poll corto mientras quede trabajo FRESCO (emisiones en curso o
        # PDFs por reconciliar). Acotado por edad: si el worker nunca escribió
        # el item (p.ej. mensaje muerto en la DLQ), el move zombi vuelve al
        # beat base de 2 min en vez de re-disparar el cron para siempre.
        limite = fields.Datetime.now() - timedelta(minutes=30)
        pendientes = moves.filtered(
            lambda m: m.l10n_pe_biller_state == "en_proceso"
            and m.write_date
            and m.write_date >= limite
        )
        if pendientes or sin_pdf.filtered(lambda m: not m.l10n_pe_biller_pdf):
            self._l10n_pe_trigger_poll_async(seconds=30)

    # ------------------------------------------------------------------ acción
    def action_l10n_pe_send_to_biller(self):
        _logger.info("Enviando facturas a Biller: %s", self.ids)
        icp = self.env["ir.config_parameter"].sudo()
        base = icp.get_param("l10n_pe_ne_biller.url", "http://localhost:8090").rstrip(
            "/"
        )
        _logger.info("URL: %s", base)
        # >240 es inalcanzable: limit_time_real=240 mata el worker de Odoo
        # antes (SIGKILL con rollback), con el POST quizá ya aceptado en SUNAT.
        timeout = int(icp.get_param("l10n_pe_ne_biller.timeout", "240"))
        _logger.info("Timeout: %s", timeout)
        use_async = icp.get_param(
            "l10n_pe_ne_biller.async_enabled", ""
        ).strip().lower() in ("1", "true")
        use_instant = icp.get_param(
            "l10n_pe_ne_biller.instant_enabled", ""
        ).strip().lower() in ("1", "true")
        for move in self:
            _logger.info(
                "Procesando factura: %s (%s)", move.name, move.l10n_pe_biller_state
            )
            if move.l10n_pe_biller_state in ("enviado", "en_proceso"):
                _logger.info("Factura ya enviada o en proceso: %s", move.name)
                continue
            # Guarda: no aplicar percepción a un cliente exceptuado del régimen (QA-028). El cobro
            # adicional no corresponde; se bloquea con un mensaje claro en vez de emitir mal.
            if move.l10n_pe_ne_percepcion and move.partner_id.l10n_pe_ne_exceptuado_percepcion:
                raise UserError(_(
                    "El cliente %s está exceptuado del régimen de percepciones; no corresponde "
                    "aplicarle percepción. Desactivá la percepción para emitir este comprobante."
                ) % (move.partner_id.display_name or ""))
            # Valida la serie (familia correcta + habilitada, QA-074) ANTES de asignar el
            # correlativo, para no consumir un número si la serie se rechaza.
            move._l10n_pe_check_serie()
            # Fija la serie+correlativo fiscal ANTES de construir el payload/firmar, desde la
            # secuencia POR SERIE (no el folio del diario). A partir de aquí el número es estable
            # e igual en payload, XML firmado, QR, PDF y una eventual baja. Va DESPUÉS del guard
            # para no consumir un correlativo en un comprobante que se bloquea.
            move._l10n_pe_ne_assign_numero()
            if use_async:
                move._l10n_pe_enqueue_emission(icp)
                continue
            if use_instant:
                # Modo instantáneo: FIRMAR (rápido, sin SUNAT) → ticket/PDF ya disponibles y
                # estado 'en_proceso'. El cron _l10n_pe_cron_enviar_pendientes envía a SUNAT.
                endpoint, payload = move._l10n_pe_target()
                headers = {"X-Api-Key": move.company_id.sudo().l10n_pe_ne_api_key or ""}
                try:
                    resp = requests.post(
                        base + "/generator/" + endpoint + "/firmar",
                        json=payload, headers=headers, timeout=(5, 30),
                    )
                    if resp.status_code == 200:
                        move._l10n_pe_apply_signed(resp.json())
                    else:
                        move.l10n_pe_biller_state = "error"
                        move.l10n_pe_biller_message = (
                            "Firma HTTP %s: %s" % (resp.status_code, resp.text)
                        )[:2000]
                except requests.RequestException as exc:
                    move.l10n_pe_biller_state = "error"
                    move.l10n_pe_biller_message = (
                        _("Error de conexión con el facturador (firma): %s") % exc
                    )
                continue
            endpoint, payload = move._l10n_pe_target()
            _logger.info("AAAEnviando %s: %s", endpoint, payload)
            headers = {"X-Api-Key": move.company_id.sudo().l10n_pe_ne_api_key or ""}
            try:
                _logger.info("EEEEnviando %s: %s", endpoint, payload)
                resp = requests.post(
                    base + "/generator/" + endpoint,
                    json=payload,
                    headers=headers,
                    # connect corto aparte: un endpoint inalcanzable (SG, DNS)
                    # falla en 5s en vez de colgar el worker hasta el read.
                    timeout=(5, timeout),
                )
                _logger.info(
                    "RESP %s -> POST %s/generator/%s -> HTTP %s | %s",
                    move.name,
                    base,
                    endpoint,
                    resp.status_code,
                    resp.text[:500],
                )
                _logger.info("Respuesta: %s", resp.text)
            except requests.RequestException as exc:
                move.l10n_pe_biller_state = "error"
                _logger.error("Error: %s", exc)
                move.l10n_pe_biller_message = (
                    _("Error de conexión con el facturador: %s") % exc
                )
                continue
            # El biller devuelve el XML firmado como body y el CDR de SUNAT en
            # el header X-Sunat-Cdr (base64 del zip).
            move._l10n_pe_apply_emission_response(
                resp.status_code == 200, resp.text, resp.headers.get("X-Sunat-Cdr")
            )
        return True

    # ------------------------------------------- API ligera (BFF NE Express, /json/2)
    @api.model
    def l10n_pe_ne_quick_emit(self, payload):
        """Emite un comprobante desde un payload PLANO (sin contexto contable previo): crea/halla el
        cliente, arma el account.move con sus líneas (impuesto por código cat-05), lo postea y lo envía a
        SUNAT vía el facturador. Devuelve el resultado. Lo consume el BFF stateless por /json/2 — así la
        lógica de negocio queda en Odoo (fuente única) y el dato vive en Odoo (upgrade sin migración)."""
        company = self.env.company
        journal = self.env["account.journal"].search(
            [("type", "=", "sale"), ("company_id", "=", company.id)], limit=1
        )
        if not journal:
            raise UserError(_("No hay diario de ventas configurado para la compañía."))
        # Anti-doble-conversión (QA-098): una cotización/orden ya convertida no puede emitir OTRO
        # comprobante (evita duplicar la venta). Se valida antes de armar el move.
        cotid = payload.get("cotizacionId")
        if cotid:
            cot = self.env["l10n_pe_ne.cotizacion"].browse(int(cotid)).exists()
            if cot and cot.estado == "convertida":
                raise UserError(_(
                    "La cotización %s ya fue convertida en el comprobante %s; no se puede emitir otro."
                ) % (cot.name or cot.id, cot._l10n_pe_ne_comprobante_numero() or "—"))
        tipo = payload.get("tipoDoc") or "01"
        # NC motivo 03 = "Corrección por error en la descripción": SOLO corrige el texto,
        # NO cambia importes. La nota va con importe 0.00 (la factura original conserva su
        # valor). Se fuerza aquí para que la correctitud fiscal no dependa del front.
        es_correccion = tipo == "07" and str(payload.get("motivo") or "") == "03"
        # NC (07) / ND (08): resuelven el documento afectado (mismo cliente, serie derivada del original).
        origin = None
        if tipo in ("07", "08"):
            origin = self._l10n_pe_ne_quick_origin(
                payload.get("docAfectado") or payload.get("afectado")
            )
            origin._l10n_pe_check_afectable_con_nota()
        if origin is not None:
            partner = origin.partner_id
        else:
            partner = self._l10n_pe_ne_quick_partner(payload.get("cliente") or {})
        # Descuento global (% sobre toda la operación): se prorratea a cada línea como descuento que
        # afecta la base (cat. 53 código 00), combinándose con el descuento propio de la línea. Produce
        # los mismos totales que un descuento global y reusa la emisión de descuento por ítem ya validada.
        g = float(payload.get("descuentoGlobal") or 0)
        lines = []
        for ln in payload.get("lineas") or []:
            tax = self._l10n_pe_ne_tax_by_code(ln.get("taxCode"))
            # Sin tax resuelta la línea saldría en el XML como 'gravada con IGV 0.00'
            # (rechazo SUNAT 3111): mejor cortar aquí con el dato accionable.
            if not tax:
                raise UserError(
                    _(
                        "No hay un impuesto de venta con código SUNAT %(code)s configurado "
                        "para la compañía (línea «%(linea)s»). Configura el IGV en "
                        "Contabilidad → Impuestos (o ejecuta el setup de la compañía) y "
                        "vuelve a emitir."
                    )
                    % {
                        "code": ln.get("taxCode") or "1000",
                        "linea": (ln.get("descripcion") or "").strip() or "ITEM",
                    }
                )
            taxes = tax
            if ln.get("icbper"):
                # Bolsa plástica: el ICBPER (monto fijo por unidad) se SUMA al IGV de la línea.
                taxes = tax + self._l10n_pe_ne_ensure_icbper_tax()
            isc_rate = float(ln.get("isc") or 0)
            if isc_rate > 0:
                # ISC (ad-valorem): se agrega a la línea; el IGV se recalcula sobre valor + ISC.
                taxes = taxes + self._l10n_pe_ne_ensure_isc_tax(isc_rate)
            # Notas (07/08): solo resolver el producto, nunca crearlo — sus líneas pueden ser
            # espejo o texto sintético (DICE/DEBE DECIR) que no debe entrar al catálogo.
            # precio_con_igv=False: el payload de emisión trae el valor SIN IGV.
            prod = self._l10n_pe_ne_quick_product(
                ln, tax, create=tipo not in ("07", "08"), precio_con_igv=False
            )
            d = float(ln.get("descuento") or 0)
            disc = round(100.0 * (1 - (1 - d / 100.0) * (1 - g / 100.0)), 6) if g else d
            lvals = {
                "name": ln.get("descripcion") or (prod.name if prod else "ITEM"),
                "quantity": float(ln.get("cantidad") or 1),
                # Motivo 03: importe 0 (solo se corrige la descripción, no el monto).
                "price_unit": 0.0 if es_correccion else float(ln.get("precioUnitario") or 0),
                "discount": 0.0 if es_correccion else disc,
                "tax_ids": [(6, 0, taxes.ids if taxes else [])],
            }
            if prod:
                lvals["product_id"] = prod.id
            if ln.get("unidad"):
                lvals["l10n_pe_ne_unit_code"] = ln["unidad"]
            if ln.get("codSunat"):
                lvals["l10n_pe_ne_cod_producto_sunat"] = ln["codSunat"]
            if ln.get("afectacionGratuita"):
                lvals["l10n_pe_ne_afectacion_gratuita"] = ln["afectacionGratuita"]
            lines.append((0, 0, lvals))
        # Otros cargos (que afectan la base imponible): se agregan como una línea gravada adicional, así
        # suben gravada/IGV/total con la maquinaria de líneas ya validada (no se prorratea el desc. global).
        oc = float(payload.get("otrosCargos") or 0)
        if oc > 0:
            lines.append(
                (
                    0,
                    0,
                    {
                        "name": payload.get("otrosCargosDesc") or "OTROS CARGOS",
                        "quantity": 1,
                        "price_unit": oc,
                        "tax_ids": [(6, 0, self._l10n_pe_ne_tax_by_code("1000").ids)],
                    },
                )
            )
        vals = {
            "move_type": "out_refund" if tipo == "07" else "out_invoice",
            "partner_id": partner.id,
            "journal_id": journal.id,
            "invoice_date": payload.get("fechaEmision")
            or self._l10n_pe_ne_today_lima(),
            "l10n_pe_serie": payload.get("serie")
            or self._l10n_pe_ne_default_serie(tipo, origin),
            "invoice_line_ids": lines,
        }
        # Alinear el tipo latam con el tipoDoc pedido: sin esto, una BOLETA a un cliente
        # con RUC se emitiría como Factura (el fallback decide por el documento del cliente).
        es_boleta = tipo == "03" or (
            tipo in ("07", "08")
            and origin is not None
            and (origin.l10n_pe_ne_tipo_doc or origin._l10n_pe_document_type()) == "03"
        )
        doc_xmlid = {
            "01": "l10n_pe.document_type01",
            "03": "l10n_pe.document_type02",
            "07": "l10n_pe.document_type07b" if es_boleta else "l10n_pe.document_type07",
            "08": "l10n_pe.document_type08b" if es_boleta else "l10n_pe.document_type08",
        }.get(tipo)
        doc_type = doc_xmlid and self.env.ref(doc_xmlid, raise_if_not_found=False)
        if doc_type:
            vals["l10n_latam_document_type_id"] = doc_type.id
        if origin is not None and not payload.get("moneda"):
            # NC/ND heredan la moneda del comprobante afectado: SUNAT exige que la
            # nota vaya en la misma moneda que el documento original (sin esto una
            # NC de una factura en USD salía forzada a PEN).
            moneda = origin.currency_id
        else:
            moneda = self._l10n_pe_ne_quick_currency(payload.get("moneda"))
        if moneda:
            vals["currency_id"] = moneda.id
            # Comprobante en dólares: asegura el TC oficial del día en
            # res.currency.rate para que el PLE y la conversión a soles salgan
            # bien. Best-effort: si la red falla, no bloquea la emisión.
            if moneda.name and moneda.name != "PEN":
                try:
                    fecha_tc = vals.get("invoice_date") or fields.Date.context_today(self)
                    self.env.company._l10n_pe_ne_ensure_tc(fecha_tc)
                except Exception as e:  # noqa: BLE001
                    _logger.warning("TC SUNAT: no se pudo asegurar en emisión (%s)", e)
        if origin is not None:
            vals["l10n_pe_motivo_code"] = str(
                payload.get("motivo") or ("01" if tipo == "07" else "02")
            )
            # Motivo/sustento (texto libre): si el front lo envía se usa como desMotivo;
            # si no, _l10n_pe_build_note_request cae a la descripción del catálogo.
            sustento = (payload.get("sustento") or "").strip()
            if sustento:
                vals["l10n_pe_motivo_desc"] = sustento[:250]
            if tipo == "07":
                vals["reversed_entry_id"] = origin.id
            else:
                vals["debit_origin_id"] = origin.id
        if payload.get("correlativo"):
            vals["l10n_pe_correlativo"] = str(payload["correlativo"])
            # Con correlativo MANUAL no aplica la unicidad de la secuencia por diario:
            # dos emisiones forzadas comparten serie+correlativo fiscal pero tienen
            # 'name' internos distintos, así que account_move_unique_name_latam no las
            # detecta. Verificamos el número fiscal (serie_emit+corr_emit) contra los ya
            # emitidos/anulados de la compañía antes de crear y mandar a SUNAT.
            self._l10n_pe_ne_check_numero_libre(
                vals["l10n_pe_serie"], str(payload["correlativo"])
            )
        move = self.env["account.move"].create(vals)
        self._l10n_pe_ne_quick_flags(move, payload)
        move.action_post()
        # Stock: el bien sale (o vuelve, si es NC) cuando la venta existe en Odoo, no cuando
        # SUNAT responde — la mercadería ya cambió de manos. Va después del post y antes de
        # enviar: si SUNAT rechaza, el movimiento se corrige con la NC, igual que el importe.
        move._l10n_pe_ne_mover_stock()
        # Nota de Crédito: no puede acreditar más de lo facturado. Se permiten VARIAS NC
        # sobre el mismo comprobante, pero el ACUMULADO no puede superar su total: el tope
        # de esta nota es el saldo pendiente de acreditar (total − NC previas vigentes).
        # Respaldo del front; la NC de importe 0 (motivo 03) pasa. (La ND suma a la deuda,
        # así que no lleva tope.)
        if tipo == "07" and origin is not None:
            previas = origin._l10n_pe_ne_nc_previas() - move
            acreditado = sum(previas.mapped("amount_total"))
            saldo = (origin.amount_total or 0) - acreditado
            if move.amount_total > saldo + 0.05:
                if previas:
                    raise UserError(
                        _(
                            "El comprobante afectado ya tiene %(n)d nota(s) de crédito por "
                            "%(acred)s (%(lista)s); saldo pendiente de acreditar: %(saldo)s. "
                            "Esta nota (%(nc)s) lo supera."
                        )
                        % {
                            "n": len(previas),
                            "acred": "%.2f" % acreditado,
                            "lista": ", ".join(
                                "%s-%s" % m._l10n_pe_ne_doc_id() for m in previas
                            ),
                            "saldo": "%.2f" % saldo,
                            "nc": "%.2f" % move.amount_total,
                        }
                    )
                raise UserError(
                    _(
                        "La nota de crédito (%(nc)s) no puede superar el total del comprobante "
                        "afectado (%(orig)s)."
                    )
                    % {
                        "nc": "%.2f" % move.amount_total,
                        "orig": "%.2f" % (origin.amount_total or 0),
                    }
                )
        # Si la emisión vino de "Convertir a comprobante", vincula el comprobante
        # recién posteado a la cotización de origen y la marca como 'convertida'.
        cotid = payload.get("cotizacionId")
        if cotid:
            cot = self.env["l10n_pe_ne.cotizacion"].browse(int(cotid)).exists()
            if cot:
                cot.l10n_pe_ne_vincular_comprobante(move.id)
        # Avance de obra (QA-039): la suma de las valorizaciones no puede superar el valor total
        # del contrato. Se valida con el move ya posteado (amount_total disponible); si se pasa,
        # el raise revierte la transacción y no se emite.
        proj = move.l10n_pe_ne_proyecto_id
        if proj:
            otras = move.amount_total or 0.0  # esta valorización
            if round(proj.facturado + otras, 2) > round(proj.valor_total or 0.0, 2) + 0.01:
                raise UserError(_(
                    "Esta valorización (%s) haría que lo facturado del contrato «%s» supere su "
                    "valor total. Facturado: %s · Contrato: %s · Esta: %s."
                ) % (
                    self._l10n_pe_fmt(otras), proj.name,
                    self._l10n_pe_fmt(proj.facturado), self._l10n_pe_fmt(proj.valor_total),
                    self._l10n_pe_fmt(otras),
                ))
        move.action_l10n_pe_send_to_biller()
        return move.l10n_pe_ne_quick_result()

    def _l10n_pe_ne_check_numero_libre(self, serie, correlativo):
        """Impide reutilizar un número fiscal (serie+correlativo) ya emitido/anulado en
        la compañía. Necesario solo con correlativo manual: la unicidad de la secuencia
        del diario no cubre este caso (ver quick_emit)."""
        corr = (correlativo or "").strip().zfill(8)
        dup = self.env["account.move"].sudo().search(
            [
                ("company_id", "=", self.env.company.id),
                ("l10n_pe_ne_serie_emit", "=", serie),
                ("l10n_pe_ne_corr_emit", "=", corr),
                ("l10n_pe_biller_state", "in", ("enviado", "anulado")),
            ],
            limit=1,
        )
        if dup:
            raise UserError(
                _(
                    "Ya existe un comprobante con ese número para ese cliente "
                    "(número duplicado)."
                )
            )

    def _l10n_pe_ne_default_serie(self, tipo, origin=None):
        """Serie por defecto: F001/B001 para factura/boleta; FC01/FD01 (o BC01/BD01 si el afectado es
        boleta) para NC/ND, derivando la familia del documento original."""
        if tipo == "03":
            return "B001"
        if tipo in ("07", "08"):
            base = (
                "B"
                if origin is not None
                and (origin.l10n_pe_serie or "F")[:1].upper() == "B"
                else "F"
            )
            return base + ("C01" if tipo == "07" else "D01")
        return "F001"

    def _l10n_pe_ne_quick_currency(self, moneda):
        """Moneda del comprobante: PEN por defecto; USD si el payload lo pide
        (USD/DOLARES/$). Activa la moneda si está inactiva. El builder ya emite
        tipMoneda desde currency_id."""
        code = (moneda or "PEN").strip().upper()
        code = (
            "USD"
            if code in ("USD", "DOLARES", "DÓLARES", "DOLAR", "US$", "$")
            else "PEN"
        )
        cur = (
            self.env["res.currency"]
            .with_context(active_test=False)
            .search([("name", "=", code)], limit=1)
        )
        if cur and not cur.active:
            cur.sudo().active = True
        return cur

    def _l10n_pe_ne_quick_origin(self, ref):
        """Resuelve el account.move afectado por una NC/ND: por id (lo natural, el emit devuelve 'id') o
        por serie+correlativo. Lanza si no lo encuentra."""
        ref = ref or {}
        Move = self.env["account.move"]
        if ref.get("id"):
            m = Move.browse(int(ref["id"])).exists()
            if m:
                return m
        serie = (ref.get("serie") or "").strip()
        corr = str(ref.get("correlativo") or "").strip().lstrip("0")
        if serie and corr:
            cands = Move.search(
                [
                    ("l10n_pe_serie", "=", serie),
                    ("move_type", "in", ("out_invoice", "out_refund")),
                ],
                order="id desc",
                limit=300,
            )
            for m in cands:
                _s, c = m._l10n_pe_serie_correlativo()
                if (c or "").lstrip("0") == corr:
                    return m
        raise UserError(
            _(
                "No se encontró el documento afectado (envía docAfectado.id o serie+correlativo)."
            )
        )

    def _l10n_pe_check_afectable_con_nota(self):
        """Una NC/ND solo puede emitirse sobre una factura o una boleta: SUNAT rechaza la
        referencia a otra nota. La guarda va en el call site de la emisión y no dentro de
        _l10n_pe_ne_quick_origin porque ese helper lo comparte la anulación, que SÍ acepta
        notas (una NC se anula comunicando su baja)."""
        self.ensure_one()
        tipo = self.l10n_pe_ne_tipo_doc or self._l10n_pe_document_type()
        if tipo not in ("01", "03"):
            docname = {
                "07": _("Nota de Crédito"),
                "08": _("Nota de Débito"),
            }.get(tipo, tipo)
            raise UserError(
                _(
                    "Una nota de crédito o débito solo puede emitirse sobre una factura o una "
                    "boleta; el documento afectado (%(doc)s %(serie)s-%(corr)s) es una nota. "
                    "Para anularla, comunique su baja."
                )
                % {
                    "doc": docname,
                    "serie": self.l10n_pe_ne_serie_emit
                    or self._l10n_pe_serie_correlativo()[0],
                    "corr": self.l10n_pe_ne_corr_emit
                    or self._l10n_pe_serie_correlativo()[1],
                }
            )

    @api.model
    def l10n_pe_ne_quick_anular(self, payload):
        """Anula un comprobante ya emitido a SUNAT: boletas por Resumen Diario (RC, tipEstado 3),
        facturas/NC/ND por Comunicación de Baja (RA). payload: {id | serie+correlativo, motivo}.
        Lo consume el BFF por /json/2."""
        payload = payload or {}
        move = self._l10n_pe_ne_quick_origin(payload.get("comprobante") or payload)
        move.l10n_pe_ne_baja_motivo = (
            payload.get("motivo") or ""
        ).strip() or "Anulacion de la operacion"
        move.action_l10n_pe_send_baja()
        return move._l10n_pe_ne_anular_result()

    def _l10n_pe_ne_anular_result(self):
        self.ensure_one()
        tipo, serie, corr = self._l10n_pe_baja_identidad()
        msg = self.l10n_pe_biller_message or ""
        m = re.search(r"ResponseCode (\d+)", msg)
        anulado = self.l10n_pe_biller_state == "anulado"
        return {
            "id": self.id,
            "tipoAnulacion": "RC" if tipo == "03" else "RA",
            "docAnulacion": self.l10n_pe_ne_baja_doc or "",
            "comprobante": "%s-%s" % (serie, (corr or "").zfill(8)),
            "estado": self.l10n_pe_biller_state,
            "anulado": anulado,
            "responseCode": m.group(1) if m else ("0" if anulado else ""),
            "mensaje": msg,
        }

    def l10n_pe_ne_get_baja_files(self, kind=None):
        """{cdr} base64 de la anulación (RA/RC), para que el BFF lo sirva.

        Acepta e ignora ``kind`` (una baja no tiene ticket): la ruta
        ``/ne/api/anulacion/<id>/cdr`` invoca este método vía
        ``_serve_file`` con ``kind='cdr'`` — simétrico con
        ``l10n_pe_ne_get_files``.
        """
        self.ensure_one()
        out = {}
        att = self.l10n_pe_ne_baja_cdr
        if att:
            v = att.datas
            out["cdr"] = (
                v.decode("ascii") if isinstance(v, (bytes, bytearray)) else (v or "")
            )
        return out

    def _l10n_pe_ne_fetch_direccion_padron(self, num):
        """Domicilio fiscal desde el padrón externo (DynamoDB) o, como respaldo, SUNAT.

        Consulta la fuente directamente (no lee el partner ya guardado, que puede tener
        street vacío). Degrada a "" ante cualquier fallo o si la fuente no está
        configurada — NUNCA bloquea la emisión."""
        num = (num or "").strip()
        if not num:
            return ""
        P = self.env["res.partner"].sudo()
        data = None
        for fetch in (P._l10n_pe_query_external_db, P._l10n_pe_query_sunat):
            try:
                data = fetch(num)
            except Exception:  # noqa: BLE001 — fuente no configurada / red: seguimos
                data = None
            if data:
                break
        return (data or {}).get("address") or ""

    def _l10n_pe_ne_quick_partner(self, c):
        num = (c.get("numDoc") or "").strip()
        nombre = (c.get("razonSocial") or "").strip()
        dire = (c.get("direccion") or "").strip()
        urb = (c.get("urbanizacion") or "").strip()
        Partner = self.env["res.partner"]
        found = Partner.search([("vat", "=", num)], limit=1) if num else Partner.browse()
        if not found and not num and not nombre:
            # Público general SIN documento ni nombre: reusa UN solo 'CONSUMIDOR
            # FINAL' por tenant en vez de crear un partner desechable por venta.
            # (La emisión no reescribe el partner, así que reusarlo es seguro.)
            found = Partner.search([
                ("company_id", "=", self.env.company.id),
                ("vat", "=", False),
                ("name", "=", "CONSUMIDOR FINAL"),
            ], limit=1)
        if not found:
            # company_id del emisor actual: aísla el cliente por RUC (multi-tenant). Sin
            # esto quedaría company_id=False = visible/editable por TODOS los tenants.
            vals = {
                "name": nombre or "CONSUMIDOR FINAL",
                "customer_rank": 1,
                "company_id": self.env.company.id,
            }
            if num:
                vals["vat"] = num
                t = self.env["l10n_latam.identification.type"].search(
                    [("l10n_pe_vat_code", "=", c.get("tipoDoc") or "6")], limit=1
                )
                if t:
                    vals["l10n_latam_identification_type_id"] = t.id
            if dire:
                vals["street"] = dire
            if urb:
                vals["street2"] = urb
            # País del adquirente (exportación / no domiciliado): alimenta codPaisCliente en la
            # cabecera 0200. Solo al crear (la emisión no reescribe un partner ya existente).
            pais = (c.get("pais") or "").strip().upper()
            if pais:
                country = self.env["res.country"].search([("code", "=", pais)], limit=1)
                if country:
                    vals["country_id"] = country.id
            found = Partner.create(vals)
        # Dirección faltante → la completamos (sin pisar una ya guardada). Primero lo que
        # mandó el front; si no vino, el domicilio fiscal del padrón. Así la representación
        # impresa (A4) muestra la dirección de los RUC 20 y de los 10/naturales que la tengan.
        if not found.street:
            addr = dire or self._l10n_pe_ne_fetch_direccion_padron(num)
            if addr:
                found.street = addr
        if urb and not found.street2:
            found.street2 = urb
        return found

    # Afectaciones de tasa 0% (cat-05) que se auto-crean si el plan no las trae. El IGV (1000)
    # y el IVAP (1016) NO están aquí a propósito: su tasa es una decisión contable y crearlos
    # con una tasa adivinada emitiría montos fiscales incorrectos — si faltan, la emisión corta
    # con un error accionable (ver quick_emit).
    _L10N_PE_NE_TAXES_CERO = {
        "9997": "Exonerado",
        "9998": "Inafecto",
        "9995": "Exportación",
        "9996": "Gratuito",
    }

    def _l10n_pe_ne_tax_by_code(self, code):
        """account.tax de venta por código cat-05 (l10n_pe_edi_tax_code); default 1000 (IGV gravado).

        Las taxes 0% (exonerado/inafecto/exportación/gratuito) se crean si faltan, como
        ICBPER/ISC: una BD recién configurada suele traer solo el IGV, y sin esto la línea
        quedaba SIN impuesto → `_l10n_pe_tax_info` la clasificaba con su default 'gravado
        (1000)' a tasa 0 → XML con TaxableAmount>0 y TaxAmount=0.00 → rechazo SUNAT 3111."""
        code = code or "1000"
        tax = self.env["account.tax"].search(
            [
                ("company_id", "=", self.env.company.id),
                ("type_tax_use", "=", "sale"),
                ("l10n_pe_edi_tax_code", "=", code),
            ],
            limit=1,
        )
        if not tax and code in self._L10N_PE_NE_TAXES_CERO:
            label = self._L10N_PE_NE_TAXES_CERO[code]
            tax = self.env["account.tax"].sudo().create(
                {
                    "name": "%s (0%%)" % label,
                    "amount_type": "percent",
                    "amount": 0.0,
                    "type_tax_use": "sale",
                    "l10n_pe_edi_tax_code": code,
                    "company_id": self.env.company.id,
                    "description": label,
                }
            )
        return self._l10n_pe_ne_normalize_tax_excluded(tax)

    @api.model
    def _l10n_pe_ne_normalize_tax_excluded(self, tax):
        """Garantiza que la IGV/IVAP de venta trate el precio como VALOR (sin IGV).

        Contrato del app: `precioUnitario` es el valor unitario SIN IGV — el front
        (Emitir) lo muestra como `Gravada` y suma el IGV 18% por encima. Pero la base
        que emitimos sale de `line.price_subtotal`, que respeta el flag `price_include`
        de la tax: si en la BD la IGV quedó como "precio incluye impuesto"
        (`price_include_override='tax_included'`, o por el default de la compañía),
        Odoo descompone la base dividiendo por 1+tasa (100 -> 84.75) y el comprobante
        emitido NO coincide con el preview (que mostraba 118). Para que preview==emitido
        sin depender de la config ambiente, fijamos tax-excluded en la IGV/IVAP de venta
        de forma idempotente (solo escribe si hace falta; se autocorrige en el 1er emit)."""
        if (
            tax
            and tax.l10n_pe_edi_tax_code in ("1000", "1016")
            and tax.price_include_override != "tax_excluded"
        ):
            tax.sudo().write({"price_include_override": "tax_excluded"})
        return tax

    def _l10n_pe_ne_ensure_icbper_tax(self):
        """Tax ICBPER (cat-05 7152): monto FIJO por unidad (S/ 0.50 vigente desde 2023). Se crea en Odoo
        si no existe — el dato y la lógica viven en Odoo, no en el orquestador."""
        Tax = self.env["account.tax"].sudo()
        company = self.env.company
        tax = Tax.search(
            [
                ("company_id", "=", company.id),
                ("type_tax_use", "=", "sale"),
                ("l10n_pe_edi_tax_code", "=", "7152"),
            ],
            limit=1,
        )
        if tax:
            return tax
        return Tax.create(
            {
                "name": "ICBPER",
                "amount_type": "fixed",
                "amount": 0.50,
                "type_tax_use": "sale",
                "l10n_pe_edi_tax_code": "7152",
                "company_id": company.id,
                "description": "ICBPER",
            }
        )

    def _l10n_pe_ne_ensure_isc_tax(self, rate):
        """Tax ISC (Impuesto Selectivo al Consumo, cat-05 2000) — Sistema al Valor (ad-valorem %).
        Se crea/reusa por tasa. include_base_amount=True y secuencia ANTES del IGV → el IGV se
        computa sobre (valor venta + ISC), como exige SUNAT (mtoBaseIgvItem = base + ISC)."""
        Tax = self.env["account.tax"].sudo()
        company = self.env.company
        rate = round(float(rate or 0), 4)
        tax = Tax.search(
            [
                ("company_id", "=", company.id),
                ("type_tax_use", "=", "sale"),
                ("l10n_pe_edi_tax_code", "=", "2000"),
                ("amount_type", "=", "percent"),
                ("amount", "=", rate),
            ],
            limit=1,
        )
        if tax:
            return tax
        igv = self._l10n_pe_ne_tax_by_code("1000")
        return Tax.create(
            {
                "name": "ISC %g%%" % rate,
                "amount_type": "percent",
                "amount": rate,
                "type_tax_use": "sale",
                "l10n_pe_edi_tax_code": "2000",
                "include_base_amount": True,   # el IGV se calcula sobre valor + ISC
                "sequence": (igv.sequence - 1) if igv else 1,   # ISC se aplica antes que el IGV
                "company_id": company.id,
                "description": "ISC",
            }
        )

    @api.model
    def l10n_pe_ne_config(self):
        """Parámetros que React debe leer DESDE Odoo (no hardcodear): tasa IGV y monto ICBPER por unidad."""
        return {
            "igv": 18.0,
            "icbperRate": self._l10n_pe_ne_ensure_icbper_tax().amount,
            "agentePercepcion": bool(self.env.company.l10n_pe_ne_agente_percepcion),
            # Redondeo de efectivo: el POS lo aplica en vivo con estos parámetros (ver lib/redondeo.ts).
            "redondeoActivo": bool(self.env.company.l10n_pe_ne_redondeo_activo),
            "redondeoModo": self.env.company.l10n_pe_ne_redondeo_modo or "favor",
        }

    @api.model
    def l10n_pe_ne_paises(self):
        """Catálogo de países (ISO 3166 alpha-2) para el selector del cliente extranjero en la
        factura de exportación. Perú primero (default habitual) y el resto por nombre."""
        paises = self.env["res.country"].search([("code", "!=", False)], order="name")
        return [{"code": c.code, "name": c.name} for c in paises]

    @api.model
    def l10n_pe_ne_series(self, limit=None, offset=None):
        """Series realmente en uso, agregadas desde los comprobantes emitidos (la serie la
        fija el emisor al emitir; el correlativo lo autoincrementa Odoo por diario). Por serie:
        tipo, cuántos emitidos, último correlativo y el próximo a emitir. Incluye las series de
        retención/percepción (account.payment). Aislado por RUC vía el contexto de compañía."""
        TIPO = {
            "01": "Factura",
            "03": "Boleta",
            "07": "Nota de crédito",
            "08": "Nota de débito",
            "20": "Retención",
            "40": "Percepción",
        }
        agg = {}

        def add(serie, tipo, corr):
            # Solo cuenta CPE realmente emitidos: con correlativo asignado (n>=1). Un
            # account.payment lleva R001 y P001 por defecto, pero solo se emite uno; el
            # otro queda 'por_enviar' con correlativo vacío y no debe contarse.
            n = int(corr) if (corr or "").strip().isdigit() else 0
            if not serie or n < 1:
                return
            cur = agg.setdefault(
                serie, {"serie": serie, "tipoDoc": tipo, "emitidos": 0, "ultimo": 0}
            )
            cur["emitidos"] += 1
            if n > cur["ultimo"]:
                cur["ultimo"] = n

        for m in self.search([("l10n_pe_ne_serie_emit", "!=", False)]):
            add(
                m.l10n_pe_ne_serie_emit,
                m.l10n_pe_ne_tipo_doc or m._l10n_pe_document_type(),
                m.l10n_pe_ne_corr_emit,
            )
        for p in self.env["account.payment"].search(
            [("company_id", "=", self.env.company.id)]
        ):
            add(p.l10n_pe_ret_serie, "20", p.l10n_pe_ret_correlativo)
            add(p.l10n_pe_per_serie, "40", p.l10n_pe_per_correlativo)

        filas = [
            {
                "serie": s["serie"],
                "tipoDoc": s["tipoDoc"],
                "tipo": TIPO.get(s["tipoDoc"], s["tipoDoc"]),
                "emitidos": s["emitidos"],
                "ultimo": str(s["ultimo"]).zfill(8) if s["ultimo"] else "—",
                "proximo": str(s["ultimo"] + 1).zfill(8),
            }
            for s in sorted(agg.values(), key=lambda x: x["serie"])
        ]
        # Paginación opt-in sobre el agregado ya construido (no hay search directo).
        if offset is None:
            return filas
        return {"items": filas[offset:offset + limit] if limit else filas[offset:],
                "total": len(filas)}

    # ============================================================ datos negocio
    @api.model
    def l10n_pe_ne_negocio(self):
        """Datos del emisor (negocio) que alimentan el bloque `emisor` del XML, leídos desde
        res.company + su partner. El RUC es de solo lectura (identidad del emisor, indexa el
        certificado de firma en el servidor)."""
        company = self.env.company
        p = company.partner_id
        d = p.l10n_pe_district
        return {
            "ruc": p.vat or "",
            "razonSocial": company.name or "",
            "direccion": p.street or "",
            "urbanizacion": p.street2 or "",
            "telefono": p.phone or "",
            "email": p.email or "",
            "distritoId": d.id if d else None,
            "distrito": d.name if d else "",
            "ubigeo": d.code if d else "",
            "provincia": (d.city_id.name if d and d.city_id else (p.city or "")),
            "departamento": p.state_id.name or "",
            "datosPago": company.l10n_pe_ne_datos_pago or "",
            "hasLogo": bool(company.logo),
            "agentePercepcion": bool(company.l10n_pe_ne_agente_percepcion),
            "redondeoActivo": bool(company.l10n_pe_ne_redondeo_activo),
            "redondeoModo": company.l10n_pe_ne_redondeo_modo or "favor",
        }

    def l10n_pe_ne_get_logo(self):
        """(bytes, content_type) del logo del emisor para servirlo por HTTP, o (None, None)."""
        logo = self.env.company.logo
        if not logo:
            return None, None
        raw = base64.b64decode(logo)
        ct = (
            "image/png" if raw[:4] == b"\x89PNG"
            else "image/jpeg" if raw[:2] == b"\xff\xd8"
            else "application/octet-stream"
        )
        return raw, ct

    def _l10n_pe_ne_set_logo(self, company, logo_b64):
        """Valida y guarda el logo del emisor. Vacío/None → lo quita. Acepta data-URI o base64
        pelado. Exige PNG/JPEG y ≤ ~1.4 MB (mismo tope que valida biller-pdf al imprimir)."""
        if not logo_b64:
            company.logo = False
            return
        if isinstance(logo_b64, str) and logo_b64.startswith("data:"):
            logo_b64 = logo_b64.split(",", 1)[-1]
        try:
            raw = base64.b64decode(logo_b64, validate=True)
        except Exception as exc:  # noqa: BLE001
            raise UserError(_("El logo no es una imagen válida.")) from exc
        if len(raw) > 1_400_000:
            raise UserError(_("El logo es demasiado grande (máx. ~1.4 MB)."))
        if not (raw[:4] == b"\x89PNG" or raw[:2] == b"\xff\xd8"):
            raise UserError(_("El logo debe ser PNG o JPEG."))
        company.logo = base64.b64encode(raw)

    @api.model
    def l10n_pe_ne_buscar_distrito(self, q=None, limit=20):
        """Busca distritos (ubigeo) por nombre, código, provincia o departamento — así el
        selector llena el ubigeo automáticamente sin tipear los 6 dígitos (escribes 'Miraflores'
        o 'Arequipa' y sale el distrito con su código)."""
        q = (q or "").strip()
        dom = (["|", "|", "|",
                ("name", "ilike", q), ("code", "ilike", q),
                ("city_id.name", "ilike", q), ("city_id.state_id.name", "ilike", q)]
               if q else [])
        recs = self.env["l10n_pe.res.city.district"].search(dom, limit=limit)
        return [
            {
                "id": r.id,
                "code": r.code or "",
                "name": r.name or "",
                "provincia": r.city_id.name or "",
                "departamento": r.city_id.state_id.name or "",
            }
            for r in recs
        ]

    @api.model
    def l10n_pe_ne_update_negocio(self, vals):
        """Actualiza los datos editables del emisor (razón social, dirección, contacto y
        distrito). El RUC nunca se toca. Al fijar un distrito se sincronizan también provincia
        (city) y departamento (state) para que el bloque `emisor` quede consistente. Los cambios
        fluyen al PRÓXIMO XML emitido vía _l10n_pe_emisor."""
        # env.company lo fija el servidor desde el usuario (with_company), así que estas
        # escrituras SIEMPRE recaen sobre la empresa del propio emisor. res.company solo es
        # escribible por "Access Rights" (que el emisor no tiene); usamos sudo acotado a su
        # propia empresa para no exigirle ese rol global.
        company = self.env.company.sudo()
        p = company.partner_id
        razon = (vals.get("razonSocial") or "").strip()
        if "razonSocial" in vals and razon:
            company.name = razon
        pvals = {}
        for key, field in (
            ("direccion", "street"),
            ("urbanizacion", "street2"),
            ("telefono", "phone"),
            ("email", "email"),
        ):
            if key in vals:
                pvals[field] = (vals.get(key) or "").strip() or False
        did = vals.get("distritoId")
        if did:
            d = self.env["l10n_pe.res.city.district"].sudo().browse(int(did)).exists()
            if d:
                pvals["l10n_pe_district"] = d.id
                if d.city_id:
                    pvals["city"] = d.city_id.name
                    if d.city_id.state_id:
                        pvals["state_id"] = d.city_id.state_id.id
                    if d.city_id.country_id:
                        pvals["country_id"] = d.city_id.country_id.id
        if pvals:
            p.write(pvals)
        if "datosPago" in vals:
            company.l10n_pe_ne_datos_pago = (vals.get("datosPago") or "").strip() or False
        if "agentePercepcion" in vals:
            company.l10n_pe_ne_agente_percepcion = bool(vals.get("agentePercepcion"))
        if "redondeoActivo" in vals:
            company.l10n_pe_ne_redondeo_activo = bool(vals.get("redondeoActivo"))
        if vals.get("redondeoModo") in ("favor", "cercano"):
            company.l10n_pe_ne_redondeo_modo = vals["redondeoModo"]
        if "logo" in vals:
            self._l10n_pe_ne_set_logo(company, vals.get("logo"))
        return self.l10n_pe_ne_negocio()

    # ============================================================ resumen estado
    @api.model
    def l10n_pe_ne_resumen(self):
        """Resumen de estado del emisor, calculado en Odoo (no en React): actividad emitida
        hoy y en el mes en curso —separando PEN/USD para no mezclar monedas— y el desglose por
        estado SUNAT de todos los comprobantes de venta. Aislado por RUC vía la compañía."""
        today = fields.Date.context_today(self)
        mes0 = today.replace(day=1)
        sales = [
            ("move_type", "in", ("out_invoice", "out_refund")),
            ("company_id", "=", self.env.company.id),
        ]
        emitidos = sales + [("l10n_pe_biller_state", "in", ("enviado", "anulado"))]

        def bucket(moves):
            pen = usd = 0.0
            for m in moves:
                if (m.currency_id.name or "PEN") == "USD":
                    usd += m.amount_total or 0.0
                else:
                    pen += m.amount_total or 0.0
            return {"count": len(moves), "pen": round(pen, 2), "usd": round(usd, 2)}

        hoy = self.search(emitidos + [("invoice_date", "=", today)])
        mes = self.search(
            emitidos + [("invoice_date", ">=", mes0), ("invoice_date", "<=", today)]
        )

        # Desglose por estado SUNAT (toda la historia de ventas de la compañía).
        estados = {
            "aceptado": 0,
            "anulado": 0,
            "rechazado": 0,
            "pendiente": 0,
            "error": 0,
        }
        MAP = {
            "enviado": "aceptado",
            "anulado": "anulado",
            "rechazado": "rechazado",
            "por_enviar": "pendiente",
            "error": "error",
        }
        for m in self.search(sales):
            k = MAP.get(m.l10n_pe_biller_state)
            if k:
                estados[k] += 1

        return {
            "hoy": bucket(hoy),
            "mes": dict(bucket(mes), periodo=today.strftime("%Y%m")),
            "estados": estados,
            "porAtender": estados["rechazado"] + estados["error"],
        }

    # ============================================================== PLE 14.1
    # Registro de Ventas e Ingresos Electrónico (PLE, formato 14.1). Estructura
    # oficial SUNAT (Anexo RS 286-2009 y modif.): campos 1-34, separador '|',
    # palote final, líneas CRLF, codificación ISO-8859-1. El archivo que el
    # CONTADOR sube al PLE de SUNAT, generado desde los comprobantes emitidos.

    @staticmethod
    def _l10n_pe_ne_ple_num(v):
        return "%.2f" % (v or 0.0)

    def _l10n_pe_ne_ple_breakdown(self):
        """Desglose por afectación para el PLE (cuadra con el XML emitido)."""
        self.ensure_one()
        gravado = exonerado = inafecto = exportacion = igv = icbper = 0.0
        for ln in self.invoice_line_ids:
            codes = ln.tax_ids.mapped("l10n_pe_edi_tax_code")
            base = ln.price_subtotal or 0.0
            if "9997" in codes:
                exonerado += base
            elif "9998" in codes:
                inafecto += base
            elif "9995" in codes:
                exportacion += base
            elif any(c in codes for c in ("1000", "1016")):
                gravado += base
        for tl in self.line_ids.filtered(lambda l: l.tax_line_id):
            code = tl.tax_line_id.l10n_pe_edi_tax_code or ""
            amt = abs(tl.amount_currency or 0.0)
            if code == "7152":
                icbper += amt
            elif code in ("1000", "1016"):
                igv += amt
        return {
            "gravado": gravado,
            "exonerado": exonerado,
            "inafecto": inafecto,
            "exportacion": exportacion,
            "igv": igv,
            "icbper": icbper,
            "total": self.amount_total or 0.0,
        }

    def _l10n_pe_ne_doc_id(self):
        """(serie, número) del comprobante para el PLE: prefiere el correlativo
        EMITIDO; si no, el folio (parte numérica final) del `name`, zfill 8. NO usa
        l10n_pe_correlativo (datos antiguos lo tienen con basura)."""
        self.ensure_one()
        serie = (
            self.l10n_pe_ne_serie_emit
            or self.l10n_pe_serie
            or self.journal_id.l10n_pe_ne_serie
            or "F001"
        )
        numero = (self.l10n_pe_ne_corr_emit or "").strip()
        if not numero:
            folios = re.findall(r"\d+", (self.name or "").replace(" ", ""))
            numero = (folios[-1] if folios else "1").zfill(8)
        return serie, numero

    def _l10n_pe_ne_ple_origen(self):
        """(fecha, tipo, serie, numero) del comprobante que se modifica (NC/ND)."""
        orig = self.reversed_entry_id or getattr(self, "debit_origin_id", False)
        if not orig:
            return "", "", "", ""
        fecha = orig.invoice_date.strftime("%d/%m/%Y") if orig.invoice_date else ""
        tipo = orig.l10n_pe_ne_tipo_doc or orig._l10n_pe_document_type()
        serie, num = orig._l10n_pe_ne_doc_id()
        return fecha, tipo, serie, num

    def _l10n_pe_ne_ple_linea(self, periodo8, cuo):
        """Una línea del PLE 14.1 (campos 1-34, '|' separador + palote final)."""
        self.ensure_one()
        num = self._l10n_pe_ne_ple_num
        b = self._l10n_pe_ne_ple_breakdown()
        tipo = self.l10n_pe_ne_tipo_doc or self._l10n_pe_document_type()
        serie, corr = self._l10n_pe_ne_doc_id()
        tdoc, ndoc = self._l10n_pe_cliente_doc()
        con_doc = bool(ndoc) and ndoc != "00000000"
        fecha = self.invoice_date.strftime("%d/%m/%Y") if self.invoice_date else ""
        estado = "2" if self.l10n_pe_biller_state == "anulado" else "1"
        of, ot, os_, on = self._l10n_pe_ne_ple_origen()
        moneda = self.currency_id.name or "PEN"
        campos = [
            periodo8,  # 1 Periodo (AAAAMM00)
            str(self.id),  # 2 CUO (único)
            "",  # 3 Nro correlativo (solo estado 8/9)
            fecha,  # 4 Fecha emisión
            "",  # 5 Fecha vencimiento
            tipo,  # 6 Tipo comprobante (tabla 10)
            serie,  # 7 Serie
            corr,  # 8 Número
            "",  # 9 Número final (consolidado)
            tdoc if con_doc else "",  # 10 Tipo doc cliente (tabla 2)
            ndoc if con_doc else "",  # 11 Nro doc cliente
            (self.partner_id.name or "").upper(),  # 12 Razón social
            num(b["exportacion"]),  # 13 Valor exportación
            num(b["gravado"]),  # 14 Base imponible gravada
            "0.00",  # 15 Descuento base
            num(b["igv"]),  # 16 IGV / IPM
            "0.00",  # 17 Descuento IGV
            num(b["exonerado"]),  # 18 Exonerado
            num(b["inafecto"]),  # 19 Inafecto
            "0.00",  # 20 ISC
            "0.00",  # 21 Base IVAP
            "0.00",  # 22 IVAP
            num(b["icbper"]),  # 23 Otros tributos (ICBPER)
            num(b["total"]),  # 24 Importe total
            moneda,  # 25 Moneda (tabla 4)
            "1.000"
            if moneda == "PEN"
            else "%.3f"
            % (
                1.0
                / (self.currency_id.with_context(date=self.invoice_date).rate or 1.0)
            ),  # 26 Tipo cambio
            of,  # 27 Fecha doc modificado
            ot,  # 28 Tipo doc modificado
            os_,  # 29 Serie doc modificado
            on,  # 30 Nro doc modificado
            "",  # 31 Contrato/proyecto
            "",  # 32 Error tipo 1
            "",  # 33 Indicador medio de pago
            estado,  # 34 Estado
        ]
        return "|".join(campos) + "|"

    @api.model
    def _l10n_pe_ne_ventas_periodo(self, periodo):
        """Comprobantes de venta válidos (01/03/07/08) del periodo YYYYMM, ordenados.
        Excluye borradores/rechazados; aislado por compañía. Compartido PLE + SIRE."""
        import calendar

        periodo = (periodo or "").strip()
        if len(periodo) != 6 or not periodo.isdigit():
            raise UserError(_("Periodo inválido. Usa YYYYMM (p.ej. 202606)."))
        year, month = int(periodo[:4]), int(periodo[4:6])
        if not (1 <= month <= 12):
            raise UserError(_("Mes inválido en el periodo."))
        last = calendar.monthrange(year, month)[1]
        d0 = fields.Date.to_date("%04d-%02d-01" % (year, month))
        d1 = fields.Date.to_date("%04d-%02d-%02d" % (year, month, last))
        return self.search(
            [
                ("move_type", "in", ("out_invoice", "out_refund")),
                ("state", "=", "posted"),
                ("invoice_date", ">=", d0),
                ("invoice_date", "<=", d1),
                ("l10n_pe_biller_state", "not in", ("por_enviar", "rechazado", False)),
            ],
            order="invoice_date, id",
        )

    @api.model
    def l10n_pe_ne_ple_ventas(self, periodo):
        """Genera el PLE 14.1 (Registro de Ventas) del periodo YYYYMM desde los
        comprobantes emitidos (01/03/07/08) de la compañía actual. Devuelve
        {filename, contentB64, count, periodo, total}. contentB64 = base64 del
        txt en ISO-8859-1 (lo que sube el contador al PLE de SUNAT)."""
        import base64

        periodo = (periodo or "").strip()
        moves = self._l10n_pe_ne_ventas_periodo(periodo)
        periodo8 = periodo + "00"
        lines = [m._l10n_pe_ne_ple_linea(periodo8, i) for i, m in enumerate(moves, 1)]
        content = ("\r\n".join(lines) + "\r\n") if lines else ""
        ruc = (self.env.company.vat or "").strip()
        ind_cont = "1" if lines else "0"  # contenido: con/sin información
        # LE + RUC + AAAAMM + DD(00) + 140100 + indOper(1) + indCont + moneda(1=PEN) + libro(1)
        filename = "LE%s%s00140100%s11.txt" % (ruc, periodo, "1" + ind_cont)
        return {
            "filename": filename,
            "contentB64": base64.b64encode(content.encode("latin-1", "replace")).decode(
                "ascii"
            ),
            "count": len(lines),
            "periodo": periodo,
            "total": sum(moves.mapped("amount_total")),
        }

    # ------------------------------------------------------------- PLE 8.1 (compras)
    #
    # ⚠ LA ESTRUCTURA DE ESTE FORMATO ESTÁ PENDIENTE DE VALIDACIÓN CONTABLE.
    #
    # Los anexos de SUNAT con el layout del 8.1 (RS 286-2009 anexo 2, y sus modificatorias)
    # se publican como PDF ESCANEADO: no hay de dónde extraerlo de forma fiable. Lo de abajo
    # espeja las convenciones del 14.1 de este mismo addon —que sí está en producción— y el
    # orden de campos que documenta SUNAT para el registro de compras, pero NADIE lo verificó
    # contra la norma vigente.
    #
    # Cada campo va NUMERADO a propósito, para que un contador pueda auditarlo uno por uno y
    # decir "el 14 no es ese" sin leer Python.
    #
    # El modo de fallo es benigno: el validador del PLE de SUNAT revisa la estructura, así que
    # un layout corrido se RECHAZA al subirlo — no entra mal en silencio. Aun así, que nadie
    # lo presente sin que su contador lo confirme primero.

    def _l10n_pe_ne_ple_compra_breakdown(self):
        """Desglose de una compra por afectación. Espeja _l10n_pe_ne_ple_breakdown (ventas),
        pero el registro de compras separa la base gravada por su DESTINO (a operaciones
        gravadas / mixtas / no gravadas), no por el tipo de tributo.

        Hoy todo va al destino "gravadas" (campo 14): es el caso de un negocio que vende
        gravado, que es el de esta app. Prorratear a operaciones no gravadas exige saber a qué
        se destina cada compra, dato que no se pide en ningún lado — sería inventarlo."""
        self.ensure_one()
        gravado = exonerado = inafecto = igv = 0.0
        for ln in self.invoice_line_ids:
            codes = ln.tax_ids.mapped("l10n_pe_edi_tax_code")
            base = ln.price_subtotal or 0.0
            if "1000" in codes:
                gravado += base
            elif "9997" in codes:
                exonerado += base
            else:
                inafecto += base
        igv = (self.amount_total or 0.0) - (self.amount_untaxed or 0.0)
        rnd = self.currency_id.rounding or 0.01
        return {
            "gravado": float_round(gravado, precision_rounding=rnd),
            "exonerado": float_round(exonerado, precision_rounding=rnd),
            "inafecto": float_round(inafecto, precision_rounding=rnd),
            "igv": float_round(igv, precision_rounding=rnd),
            "total": self.amount_total or 0.0,
        }

    def _l10n_pe_ne_ple_compra_linea(self, periodo8, cuo):
        """Una línea del PLE 8.1. Campos numerados: son POSICIONALES y separados por '|', así
        que un campo de más o de menos corre todos los siguientes."""
        self.ensure_one()
        num = self._l10n_pe_ne_ple_num
        b = self._l10n_pe_ne_ple_compra_breakdown()
        doc = self.l10n_latam_document_number or self.ref or ""
        serie, _sep, corr = doc.partition("-")
        tipo = (
            self.l10n_latam_document_type_id.code
            if self.l10n_latam_document_type_id
            else "01"
        )
        fecha = self.invoice_date.strftime("%d/%m/%Y") if self.invoice_date else ""
        moneda = self.currency_id.name or "PEN"
        # Tipo de documento del proveedor (tabla 2): 6 = RUC. Sin RUC no hay crédito fiscal,
        # así que el caso normal de este registro es 6.
        ndoc = (self.partner_id.vat or "").strip()
        tdoc = "6" if len(ndoc) == 11 else ("1" if ndoc else "0")
        campos = [
            periodo8,  # 1  Periodo (AAAAMM00)
            str(self.id),  # 2  CUO (único por operación)
            "",  # 3  Nro correlativo del asiento (solo estados 8/9)
            fecha,  # 4  Fecha de emisión del comprobante
            "",  # 5  Fecha de vencimiento o pago
            tipo,  # 6  Tipo de comprobante (tabla 10)
            serie,  # 7  Serie del comprobante
            "",  # 8  Año de emisión de la DUA/DSI (solo importaciones)
            corr,  # 9  Número del comprobante
            "",  # 10 Número final (rango) / DUA
            tdoc,  # 11 Tipo de documento del proveedor (tabla 2)
            ndoc,  # 12 Número de documento del proveedor
            (self.partner_id.name or "").upper(),  # 13 Razón social del proveedor
            num(b["gravado"]),  # 14 Base imponible destinada a operaciones GRAVADAS
            num(b["igv"]),  # 15 IGV/IPM de 14
            "0.00",  # 16 Base destinada a operaciones gravadas Y no gravadas
            "0.00",  # 17 IGV/IPM de 16
            "0.00",  # 18 Base destinada a operaciones NO gravadas
            "0.00",  # 19 IGV/IPM de 18
            num(b["exonerado"] + b["inafecto"]),  # 20 Valor de adquisiciones no gravadas
            "0.00",  # 21 ISC
            "0.00",  # 22 ICBPER
            "0.00",  # 23 Otros tributos y cargos
            num(b["total"]),  # 24 Importe total
            "",  # 25 Código de la moneda (tabla 4) — ver nota abajo
            "",  # 26 Tipo de cambio
            "",  # 27 Fecha de emisión del comprobante modificado
            "",  # 28 Tipo del comprobante modificado
            "",  # 29 Serie del comprobante modificado
            "",  # 30 Número del comprobante modificado
            "",  # 31 Fecha de la constancia de detracción
            "",  # 32 Número de la constancia de detracción
            "",  # 33 Marca del comprobante sujeto a retención
            "",  # 34 Clasificación de bienes y servicios
            "",  # 35 Identificación del contrato o proyecto
            "",  # 36 Error tipo 1
            "",  # 37 Error tipo 9
            "",  # 38 Errores tipo 4
            "",  # 39 Indicador de comprobante de pago cancelado con medio de pago
            "1",  # 40 Estado (1 = registro que corresponde al periodo)
        ]
        # Moneda y tipo de cambio: se llenan acá y no en la lista para no repetir el cálculo.
        campos[24] = moneda
        campos[25] = (
            "1.000"
            if moneda == "PEN"
            else "%.3f"
            % (1.0 / (self.currency_id.with_context(date=self.invoice_date).rate or 1.0))
        )
        return "|".join(campos) + "|"

    @api.model
    def _l10n_pe_ne_compras_periodo(self, periodo):
        """Compras posteadas del periodo YYYYMM, ordenadas. Aislado por compañía.
        Espeja _l10n_pe_ne_ventas_periodo, con move_type de proveedor."""
        import calendar

        periodo = (periodo or "").strip()
        if len(periodo) != 6 or not periodo.isdigit():
            raise UserError(_("Periodo inválido. Usa YYYYMM (p.ej. 202606)."))
        year, month = int(periodo[:4]), int(periodo[4:6])
        if not (1 <= month <= 12):
            raise UserError(_("Mes inválido en el periodo."))
        last = calendar.monthrange(year, month)[1]
        d0 = fields.Date.to_date("%04d-%02d-01" % (year, month))
        d1 = fields.Date.to_date("%04d-%02d-%02d" % (year, month, last))
        return self.search(
            [
                ("move_type", "in", ("in_invoice", "in_refund")),
                ("state", "=", "posted"),
                ("invoice_date", ">=", d0),
                ("invoice_date", "<=", d1),
            ],
            order="invoice_date, id",
        )

    @api.model
    def l10n_pe_ne_ple_compras(self, periodo):
        """PLE 8.1 (Registro de Compras) del periodo YYYYMM. Devuelve
        {filename, contentB64, count, periodo, total}. Espeja l10n_pe_ne_ple_ventas.

        ⚠ Estructura pendiente de validación contable (ver la nota del bloque)."""
        import base64

        periodo = (periodo or "").strip()
        moves = self._l10n_pe_ne_compras_periodo(periodo)
        periodo8 = periodo + "00"
        lines = [
            m._l10n_pe_ne_ple_compra_linea(periodo8, i) for i, m in enumerate(moves, 1)
        ]
        content = ("\r\n".join(lines) + "\r\n") if lines else ""
        ruc = (self.env.company.vat or "").strip()
        ind_cont = "1" if lines else "0"
        # Mismo patrón que el 14.1, cambiando el código del libro: 140100 → 080100.
        filename = "LE%s%s00080100%s11.txt" % (ruc, periodo, "1" + ind_cont)
        return {
            "filename": filename,
            "contentB64": base64.b64encode(content.encode("latin-1", "replace")).decode(
                "ascii"
            ),
            "count": len(lines),
            "periodo": periodo,
            "total": sum(moves.mapped("amount_total")),
        }

    # ------------------------------------- PLE 12.1 (inventario en unidades físicas)
    #
    # ⚠ ESTRUCTURA PENDIENTE DE VALIDACIÓN CONTABLE, igual que el 8.1: los anexos de SUNAT
    # con el layout se publican como PDF escaneado. Cada campo va numerado para auditarlo.
    #
    # Se hace el de UNIDADES FÍSICAS y NO el valorizado, y no por comodidad: el valorizado
    # exige el costo de cada movimiento, y con la valorización en `periodic` —el default de
    # Odoo, el que esta app deja puesto— `stock.move.value` y `price_unit` salen en CERO.
    # Verificado sobre los movimientos reales de la BD. Inventar ese costo (p. ej. usando el
    # precio de lista) sería declararle a SUNAT un número que no salió de ningún lado.
    # Para el valorizado hay que pasar la compañía a valorización perpetua, que cambia los
    # asientos contables: es una decisión del contador, no un efecto colateral de un reporte.

    @api.model
    def _l10n_pe_ne_kardex_periodo(self, periodo):
        """Movimientos de inventario del periodo YYYYMM, ordenados. Solo los que cruzan la
        frontera del almacén (entradas y salidas reales); los internos no son del kardex."""
        import calendar

        periodo = (periodo or "").strip()
        if len(periodo) != 6 or not periodo.isdigit():
            raise UserError(_("Periodo inválido. Usa YYYYMM (p.ej. 202606)."))
        year, month = int(periodo[:4]), int(periodo[4:6])
        if not (1 <= month <= 12):
            raise UserError(_("Mes inválido en el periodo."))
        last = calendar.monthrange(year, month)[1]
        return self.env["stock.move.line"].search(
            [
                ("state", "=", "done"),
                ("company_id", "=", self.env.company.id),
                ("date", ">=", "%04d-%02d-01 00:00:00" % (year, month)),
                ("date", "<=", "%04d-%02d-%02d 23:59:59" % (year, month, last)),
                ("product_id.is_storable", "=", True),
                # Solo lo que entra o sale del almacén: un traslado interno no es del kardex.
                "|",
                ("location_id.usage", "in", ("supplier", "customer", "inventory")),
                ("location_dest_id.usage", "in", ("supplier", "customer", "inventory")),
            ],
            order="date, id",
        )

    @api.model
    def _l10n_pe_ne_kardex_linea(self, ml, periodo8, cuo, saldo):
        """Una línea del PLE 12.1. Campos POSICIONALES separados por '|'."""
        num = self._l10n_pe_ne_ple_num
        entra = ml.location_dest_id.usage == "internal"
        cant = abs(ml.quantity or 0)
        doc = ml.move_id.l10n_pe_ne_move_id
        if doc and doc.move_type in ("out_invoice", "out_refund"):
            # VENTA: la serie/correlativo salen del mismo helper que usa la emisión y la
            # baja. Partir l10n_latam_document_number por "-" no sirve: en una venta ese
            # campo trae solo el número, y la serie terminaba llevándose el correlativo.
            serie = doc.l10n_pe_ne_serie_emit or doc._l10n_pe_serie_correlativo()[0]
            corr = doc.l10n_pe_ne_corr_emit or doc._l10n_pe_serie_correlativo()[1]
            tipo = doc.l10n_pe_ne_tipo_doc or doc._l10n_pe_document_type()
            fecha = doc.invoice_date.strftime("%d/%m/%Y") if doc.invoice_date else ""
        elif doc:
            # COMPRA: acá el documento es del proveedor y sí viene como "F001-00095001".
            serie, _sep, corr = (doc.l10n_latam_document_number or doc.ref or "").partition("-")
            tipo = (
                doc.l10n_latam_document_type_id.code
                if doc.l10n_latam_document_type_id
                else "01"
            )
            fecha = doc.invoice_date.strftime("%d/%m/%Y") if doc.invoice_date else ""
        else:
            # Ajuste de inventario: no nace de un comprobante. Tipo 00 = "otros" (tabla 10).
            serie, corr, tipo = "", "", "00"
            fecha = ml.date.strftime("%d/%m/%Y") if ml.date else ""
        p = ml.product_id
        campos = [
            periodo8,  # 1  Periodo (AAAAMM00)
            str(ml.id),  # 2  CUO (único por movimiento)
            "",  # 3  Nro correlativo del asiento
            fecha,  # 4  Fecha de emisión del documento
            tipo,  # 5  Tipo de documento (tabla 10)
            serie,  # 6  Serie del documento
            corr,  # 7  Número del documento
            "01" if entra else "02",  # 8  Tipo de operación (tabla 12): 01 entrada, 02 salida
            p.default_code or "",  # 9  Código de la existencia
            "01",  # 10 Tipo de existencia (tabla 5): 01 mercadería
            (p.name or "")[:100],  # 11 Descripción de la existencia
            p.l10n_pe_ne_unit_code or "NIU",  # 12 Unidad de medida (tabla 6)
            "",  # 13 Método de valuación (solo en el valorizado)
            num(cant) if entra else "0.00",  # 14 Entradas — cantidad
            "0.00" if entra else num(cant),  # 15 Salidas — cantidad
            num(saldo),  # 16 Saldo final — cantidad
            "1",  # 17 Estado (1 = del periodo)
        ]
        return "|".join(campos) + "|"

    @api.model
    def l10n_pe_ne_ple_inventario(self, periodo):
        """PLE 12.1 (Registro de Inventario Permanente en Unidades Físicas) del periodo.

        ⚠ Estructura pendiente de validación contable (ver la nota del bloque)."""
        import base64
        from collections import defaultdict

        periodo = (periodo or "").strip()
        lineas_ml = self._l10n_pe_ne_kardex_periodo(periodo)
        periodo8 = periodo + "00"
        # El saldo se arrastra POR PRODUCTO en el orden de los movimientos: es lo que hace
        # legible un kardex — cada renglón muestra con cuánto quedó esa existencia.
        saldos = defaultdict(float)
        lines = []
        for i, ml in enumerate(lineas_ml, 1):
            entra = ml.location_dest_id.usage == "internal"
            cant = abs(ml.quantity or 0)
            saldos[ml.product_id.id] += cant if entra else -cant
            lines.append(
                self._l10n_pe_ne_kardex_linea(ml, periodo8, i, saldos[ml.product_id.id])
            )
        content = ("\r\n".join(lines) + "\r\n") if lines else ""
        ruc = (self.env.company.vat or "").strip()
        ind_cont = "1" if lines else "0"
        # Mismo patrón que el 14.1/8.1, con el código del libro 120100.
        filename = "LE%s%s00120100%s11.txt" % (ruc, periodo, "1" + ind_cont)
        return {
            "filename": filename,
            "contentB64": base64.b64encode(content.encode("latin-1", "replace")).decode(
                "ascii"
            ),
            "count": len(lines),
            "periodo": periodo,
            "total": 0.0,
        }

    @api.model
    def l10n_pe_ne_dashboard(self, periodo=None):
        """Datos del dashboard de ventas del periodo (YYYYMM, default mes actual):
        serie diaria (para el gráfico), desglose por tipo de comprobante y KPIs.
        Reusa el filtro de ventas; aislado por compañía."""
        import calendar

        if not periodo:
            periodo = fields.Date.context_today(self).strftime("%Y%m")
        moves = self._l10n_pe_ne_ventas_periodo(periodo)
        year, month = int(periodo[:4]), int(periodo[4:6])
        por_dia, por_tipo = {}, {}
        total = anulados = 0.0
        for m in moves:
            key = m.invoice_date.strftime("%Y-%m-%d") if m.invoice_date else ""
            por_dia[key] = por_dia.get(key, 0.0) + (m.amount_total or 0.0)
            t = m.l10n_pe_ne_tipo_doc or m._l10n_pe_document_type()
            agg = por_tipo.setdefault(t, {"count": 0, "total": 0.0})
            agg["count"] += 1
            agg["total"] += m.amount_total or 0.0
            total += m.amount_total or 0.0
            if m.l10n_pe_biller_state == "anulado":
                anulados += 1
        ndays = calendar.monthrange(year, month)[1]
        serie = [
            {
                "dia": d,
                "total": round(
                    por_dia.get("%04d-%02d-%02d" % (year, month, d), 0.0), 2
                ),
            }
            for d in range(1, ndays + 1)
        ]
        tipos = [
            {
                "tipoDoc": t,
                "count": v["count"],
                "total": round(v["total"], 2),
            }
            for t, v in sorted(por_tipo.items())
        ]
        gastos = self.env["l10n_pe_ne.gasto"].l10n_pe_ne_total_gastos(periodo)
        return {
            "periodo": periodo,
            "total": round(total, 2),
            "count": len(moves),
            "anulados": int(anulados),
            "gastos": gastos,
            "neto": round(total - gastos, 2),
            "porDia": serie,
            "porTipo": tipos,
        }

    @api.model
    def l10n_pe_ne_reporte_ventas(self, periodo=None):
        """Reportes de ventas del periodo (YYYYMM, default mes actual): resumen,
        ventas de hoy, top por producto y top por cliente. Reusa el filtro de
        ventas; aislado por compañía. (Suma amount_total en su moneda; mezclar
        PEN/USD es aproximado para el MVP.)"""
        if not periodo:
            periodo = fields.Date.context_today(self).strftime("%Y%m")
        moves = self._l10n_pe_ne_ventas_periodo(periodo)
        today = fields.Date.context_today(self)
        prod, cli = {}, {}
        hoy_count, hoy_total = 0, 0.0
        for m in moves:
            for ln in m.invoice_line_ids:
                key = (
                    ln.product_id.display_name if ln.product_id else (ln.name or "ITEM")
                )
                a = prod.setdefault(
                    key, {"cantidad": 0.0, "total": 0.0, "base": 0.0, "costo": 0.0}
                )
                a["cantidad"] += ln.quantity or 0.0
                a["total"] += ln.price_total or 0.0
                # Rentabilidad: valor de venta SIN IGV (price_subtotal) vs costo del
                # producto (standard_price × cantidad). El costo es 0 si el producto no
                # lo tiene registrado → la utilidad de esa línea queda sobrestimada.
                a["base"] += ln.price_subtotal or 0.0
                a["costo"] += (ln.quantity or 0.0) * (
                    ln.product_id.standard_price or 0.0
                )
            kc = (m.partner_id.name or "—", m.partner_id.vat or "")
            c = cli.setdefault(kc, {"count": 0, "total": 0.0})
            c["count"] += 1
            c["total"] += m.amount_total or 0.0
            if m.invoice_date == today:
                hoy_count += 1
                hoy_total += m.amount_total or 0.0
        por_producto = sorted(
            (
                {
                    "producto": k,
                    "cantidad": round(v["cantidad"], 2),
                    "total": round(v["total"], 2),
                    "venta": round(v["base"], 2),
                    "costo": round(v["costo"], 2),
                    "utilidad": round(v["base"] - v["costo"], 2),
                    # Margen % sobre el valor de venta. None si el producto no tiene costo
                    # registrado (no se puede calcular una utilidad real).
                    "margen": round((v["base"] - v["costo"]) / v["base"] * 100, 1)
                    if v["base"] and v["costo"]
                    else None,
                }
                for k, v in prod.items()
            ),
            key=lambda x: -x["total"],
        )[:50]
        # Resumen de rentabilidad del periodo. Se calcula SOLO sobre productos con costo
        # registrado (los de costo 0 inflarían la utilidad como si todo fuera ganancia).
        # `conCosto`/`totalProductos` le dice al front qué tan completa es la estimación.
        rent_venta = sum(v["base"] for v in prod.values() if v["costo"])
        rent_costo = sum(v["costo"] for v in prod.values() if v["costo"])
        rentabilidad = {
            "venta": round(rent_venta, 2),
            "costo": round(rent_costo, 2),
            "utilidad": round(rent_venta - rent_costo, 2),
            "margen": round((rent_venta - rent_costo) / rent_venta * 100, 1)
            if rent_venta and rent_costo
            else None,
            "conCosto": sum(1 for v in prod.values() if v["costo"]),
            "totalProductos": len(prod),
        }
        por_cliente = sorted(
            (
                {
                    "cliente": k[0],
                    "ruc": k[1],
                    "count": v["count"],
                    "total": round(v["total"], 2),
                }
                for k, v in cli.items()
            ),
            key=lambda x: -x["total"],
        )[:50]
        return {
            "periodo": periodo,
            "resumen": {
                "count": len(moves),
                "total": round(sum(moves.mapped("amount_total")), 2),
            },
            "hoy": {"count": hoy_count, "total": round(hoy_total, 2)},
            "rentabilidad": rentabilidad,
            "porProducto": por_producto,
            "porCliente": por_cliente,
        }

    @api.model
    def l10n_pe_ne_export(self, tipo, periodo=None):
        """Centro de descargas: exporta a XLSX. tipo = ventas|productos|clientes
        (ventas usa el periodo). Devuelve {filename, contentB64, count}."""
        import base64
        import io

        import xlsxwriter

        tipo = (tipo or "ventas").strip().lower()
        if tipo == "ventas":
            if not periodo:
                periodo = fields.Date.context_today(self).strftime("%Y%m")
            moves = self._l10n_pe_ne_ventas_periodo(periodo)
            headers = [
                "Serie",
                "Número",
                "Tipo",
                "Fecha",
                "Cliente",
                "Doc. cliente",
                "Gravada",
                "Exonerada",
                "Inafecta",
                "IGV",
                "ICBPER",
                "Total",
                "Moneda",
                "Estado",
            ]
            rows = []
            for m in moves:
                b = m._l10n_pe_ne_ple_breakdown()
                serie, num = m._l10n_pe_ne_doc_id()
                rows.append(
                    [
                        serie,
                        num,
                        m.l10n_pe_ne_tipo_doc or m._l10n_pe_document_type(),
                        m.invoice_date.strftime("%d/%m/%Y") if m.invoice_date else "",
                        m.partner_id.name or "",
                        m.partner_id.vat or "",
                        round(b["gravado"], 2),
                        round(b["exonerado"], 2),
                        round(b["inafecto"], 2),
                        round(b["igv"], 2),
                        round(b["icbper"], 2),
                        round(b["total"], 2),
                        m.currency_id.name or "PEN",
                        m.l10n_pe_biller_state or "",
                    ]
                )
            sheet, base = "Ventas", "ventas-%s" % periodo
        elif tipo == "productos":
            prods = self.l10n_pe_ne_list_productos(limit=10000)
            headers = ["Código", "Descripción", "Precio", "Afectación"]
            rows = [
                [
                    p.get("codigo", ""),
                    p.get("descripcion", ""),
                    p.get("precio", 0),
                    p.get("taxCode", ""),
                ]
                for p in prods
            ]
            sheet, base = "Productos", "productos"
        elif tipo == "clientes":
            clis = self.l10n_pe_ne_list_clientes(limit=10000)
            headers = [
                "Razón social",
                "Tipo doc",
                "Número",
                "Email",
                "Teléfono",
                "Dirección",
            ]
            rows = [
                [
                    c.get("razonSocial", ""),
                    c.get("tipoDocNombre", ""),
                    c.get("numDoc", ""),
                    c.get("email", ""),
                    c.get("telefono", ""),
                    c.get("direccion", ""),
                ]
                for c in clis
            ]
            sheet, base = "Clientes", "clientes"
        else:
            raise UserError(
                _("Tipo de exporte no soportado (ventas|productos|clientes).")
            )
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet(sheet)
        head = wb.add_format(
            {"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1}
        )
        for c, h in enumerate(headers):
            ws.write(0, c, h, head)
            ws.set_column(c, c, max(12, len(h) + 2))
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row):
                ws.write(r, c, val)
        ws.autofilter(0, 0, max(1, len(rows)), len(headers) - 1)
        ws.freeze_panes(1, 0)
        wb.close()
        ruc = (self.env.company.vat or "").strip()
        return {
            "filename": "%s-%s.xlsx" % (base, ruc),
            "count": len(rows),
            "contentB64": base64.b64encode(buf.getvalue()).decode("ascii"),
        }

    @api.model
    def l10n_pe_ne_rvie_reemplazo(self, periodo):
        """SIRE — archivo de REEMPLAZO de la propuesta del RVIE (Registro de Ventas)
        del periodo YYYYMM, empaquetado en ZIP, para que el contador reemplace la
        propuesta de SUNAT. Reusa el motor de líneas del PLE (mismo desglose).

        Nombre SIRE (35 chars): LE + RUC + AAAA + MM + 00 + 140000(libro RVIE) +
        02(reemplazo) + O(operaciones) + I(contenido) + M(moneda) + 2(fijo) + NN(secuencia).
        Devuelve {zipFilename, txtFilename, contentB64 (zip base64), count, total}.
        OJO: el layout EXACTO de campos del Anexo 3 se valida/ajusta contra el PVSIRE."""
        import base64
        import io
        import zipfile

        periodo = (periodo or "").strip()
        moves = self._l10n_pe_ne_ventas_periodo(periodo)
        periodo8 = periodo + "00"
        lines = [m._l10n_pe_ne_ple_linea(periodo8, i) for i, m in enumerate(moves, 1)]
        content = ("\r\n".join(lines) + "\r\n") if lines else ""
        ruc = (self.env.company.vat or "").strip()
        cont = "1" if lines else "0"  # I: con/sin información
        # LE+RUC+AAAAMM+00 + 140000(libro RVIE) + 02(reemplazo) + O=1 + I=cont + M=1(soles) + 2 + NN=00
        txt_name = "LE%s%s00140000021%s1200.txt" % (ruc, periodo, cont)
        zip_name = txt_name[:-4] + ".zip"
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(txt_name, content.encode("latin-1", "replace"))
        return {
            "zipFilename": zip_name,
            "txtFilename": txt_name,
            "contentB64": base64.b64encode(buf.getvalue()).decode("ascii"),
            "count": len(lines),
            "periodo": periodo,
            "total": sum(moves.mapped("amount_total")),
        }

    @api.model
    def _l10n_pe_ne_margen_default(self):
        """Margen por defecto del negocio, en %. Configurable en caliente sin redeploy.
        30% es un punto de partida razonable para el retail peruano, no una verdad: cada
        negocio lo ajusta, y cada producto puede tener el suyo."""
        raw = self.env["ir.config_parameter"].sudo().get_param(
            "l10n_pe_ne.margen_default", "30"
        )
        try:
            return float(raw)
        except (TypeError, ValueError):
            return 30.0

    @api.model
    def _l10n_pe_ne_precio_con_margen(self, costo, margen):
        """Costo → precio de venta, ambos CON IGV.

        El margen se aplica sobre el bruto porque toda la app trabaja con precios de vitrina:
        así el número que sale es el que va en la etiqueta, sin desarmar el impuesto para
        pensar el negocio. Redondea a 2: es un precio, no una base imponible.

        `margen=None` usa el default del negocio; `margen=0` es un margen de CERO (vender al
        costo) y se respeta. Se distingue con `is None` a propósito: en Python 0 == False, y
        un `if not margen` convertiría el 0% en el default — el producto de promoción saldría
        30% más caro sin que nadie lo pidiera.

        El campo del producto es Float y no distingue "sin margen" de "0%": su 0 significa
        "usa el default", y quien llama lo traduce a None."""
        c = float(costo or 0)
        m = self._l10n_pe_ne_margen_default() if margen is None else float(margen)
        return round(c * (1 + m / 100.0), 2)

    @api.model
    def _l10n_pe_ne_rastreo_producto(self, rastreo):
        """Rastreo en Odoo: 'lot' | 'serial' | 'none'. La API habla el vocabulario del
        negocio ("lote"/"serie"), no el de Odoo.

        Solo tiene sentido con existencias: Odoo no rastrea lo que no cuenta. Y el rastreo
        NO se decide solo — es del producto: la misma caja de paracetamol necesita lote y un
        tornillo no."""
        r = (rastreo or "").strip().lower()
        if r in ("lote", "lot"):
            return "lot"
        if r in ("serie", "serial", "imei"):
            return "serial"
        return "none"

    @api.model
    def _l10n_pe_ne_tipo_producto(self, tipo=None, unidad=None):
        """Tipo del producto en Odoo: 'consu' (bien) o 'service' (servicio).

        Manda lo que el usuario eligió (`tipo`: "bien"/"servicio"). Si no eligió —el producto
        se auto-crea al emitir, donde no hay quién responda— se deduce de la UNIDAD, que es la
        única señal real que tiene la línea: ZZ es la unidad de servicio del catálogo 03 de
        SUNAT; cualquier otra (NIU, KGM, …) describe algo tangible.

        Antes esto era "service" fijo, lo que además se contradecía con SUNAT: sin unidad se
        emite NIU (DEFAULT_UNIT_CODE), o sea que al mismo producto se le declaraba BIEN a SUNAT
        y servicio en Odoo. Ahora ambos dicen lo mismo, y el default coincide con el de Odoo.

        No se toca `is_storable` (llevar stock o no): ese campo solo existe con el módulo
        `stock` instalado, y hoy no lo está. Es una decisión aparte, por producto.
        """
        t = (tipo or "").strip().lower()
        if t in ("bien", "bienes", "producto", "consu"):
            return "consu"
        if t in ("servicio", "servicios", "service"):
            return "service"
        return "service" if (unidad or "").strip().upper() == "ZZ" else "consu"
    def _l10n_pe_ne_lineas_con_stock(self):
        """Líneas del comprobante cuyo producto lleva existencias. En Odoo 19 eso es
        type='consu' Y is_storable=True: 'consu' solo dice que es tangible; el booleano decide
        si se rastrea. Un servicio nunca mueve stock."""
        self.ensure_one()
        return self.invoice_line_ids.filtered(
            lambda l: l.product_id
            and l.product_id.type == "consu"
            and l.product_id.is_storable
            and (l.quantity or 0) > 0
        )

    def _l10n_pe_ne_mover_stock(self, reversa=False):
        """Descuenta (o repone) el stock de las líneas de bien del comprobante.

        `reversa=True` invierte la dirección y marca los movimientos como reversa: lo usa
        _l10n_pe_ne_revertir_stock cuando SUNAT rechaza.

        La factura NO mueve stock en Odoo: los movimientos nacen de un stock.picking, que en
        el flujo estándar viene de un sale.order. Esta app no usa sale.order —emite el
        account.move directo— así que el movimiento se crea aquí, igual que hace el POS de
        Odoo (pos_order._create_order_picking() corre aparte de _generate_pos_order_invoice()).

        Dirección según el documento:
          * 01/03 (factura/boleta) → SALIDA: existencias → cliente.
          * 07 (nota de crédito)   → ENTRADA: cliente → existencias. Sin esto, anular una
            venta dejaría el stock descontado para siempre y el kardex se iría en falso.
          * 08 (nota de débito)    → nada: es un cargo (mora, penalidad), no mueve bienes.

        NUNCA bloquea la venta: si no hay existencias el movimiento igual se hace y el stock
        queda negativo — coherente con la caja, que tampoco bloquea. Un negativo es una señal
        visible de que falta un ajuste, y es preferible a impedirle vender a quien tiene el
        producto en la mano. Los fallos se tragan a propósito: el comprobante ya es válido
        ante SUNAT y no puede caerse porque el inventario no cuadre.
        """
        self.ensure_one()
        # Solo documentos de VENTA. _l10n_pe_document_type() no distingue: para un `in_invoice`
        # (una compra) devuelve '03', así que sin esta guarda una compra entraría por acá y
        # SACARÍA el stock en vez de meterlo. La compra va por _l10n_pe_ne_mover_stock_compra.
        if self.move_type not in ("out_invoice", "out_refund"):
            return self.env["stock.move"].browse()
        tipo = self.l10n_pe_ne_tipo_doc or self._l10n_pe_document_type()
        if tipo not in ("01", "03", "07"):
            return self.env["stock.move"].browse()
        lineas = self._l10n_pe_ne_lineas_con_stock()
        if not lineas:
            return self.env["stock.move"].browse()
        wh = self.env["stock.warehouse"].search(
            [("company_id", "=", self.company_id.id)], limit=1
        )
        clientes = self.env.ref(
            "stock.stock_location_customers", raise_if_not_found=False
        )
        if not wh or not clientes:
            _logger.warning(
                "stock: sin almacén o ubicación de clientes para %s; no se mueve stock",
                self.name,
            )
            return self.env["stock.move"].browse()
        # La NC va al revés que la factura; y una reversa va al revés de lo que sea.
        # Los dos XOR: la reversa de una NC vuelve a ser una salida.
        entrada = (tipo == "07") != bool(reversa)
        origen, destino = (
            (clientes, wh.lot_stock_id) if entrada else (wh.lot_stock_id, clientes)
        )
        return self._l10n_pe_ne_stock_aplicar(lineas, origen, destino, reversa=reversa)

    def _l10n_pe_ne_lote_de(self, linea):
        """stock.lot de una línea de compra, creándolo si hace falta. None si el producto no
        se rastrea o la línea no trae lote.

        Solo la ENTRADA define el lote. En la salida no se pide: Odoo lo asigna al reservar,
        por su estrategia de salida — con vencimiento, lo que caduca antes sale primero, que
        es justo lo que una farmacia necesita. Verificado contra Odoo 19."""
        prod = linea.product_id
        if not prod or prod.tracking == "none":
            return None
        nombre = (linea.l10n_pe_ne_lote or "").strip()
        if not nombre:
            return None
        Lot = self.env["stock.lot"]
        lote = Lot.search(
            [("name", "=", nombre), ("product_id", "=", prod.id),
             ("company_id", "=", self.company_id.id)], limit=1)
        if not lote:
            vals = {"name": nombre, "product_id": prod.id, "company_id": self.company_id.id}
            lote = Lot.create(vals)
        # El vencimiento se escribe aparte: el campo lo agrega product_expiry y solo tiene
        # sentido si el producto lo usa. Se pisa solo si la línea trae uno.
        if linea.l10n_pe_ne_vence and prod.use_expiration_date:
            lote.expiration_date = linea.l10n_pe_ne_vence
        return lote

    def _l10n_pe_ne_stock_aplicar(self, lineas, origen, destino, reversa=False, con_lote=False):
        """Motor común: crea y valida los movimientos de `lineas` entre dos ubicaciones.

        Lo comparten la venta (existencias → cliente), la devolución por NC, la reversa de un
        rechazo y la compra (proveedor → existencias). Lo único que cambia entre ellas son las
        dos ubicaciones y el sentido; la mecánica —y el "nunca bloquear"— es la misma.

        `con_lote`: la ENTRADA asigna el lote que trae la línea. La salida no lo necesita —
        Odoo lo asigna al reservar.
        """
        self.ensure_one()
        moves = self.env["stock.move"].browse()
        lotes = {}
        for l in lineas:
            if con_lote:
                lotes[l.id] = self._l10n_pe_ne_lote_de(l)
            # Sin 'name': stock.move no lo tiene en Odoo 19 (su `reference` se computa).
            # `origin` deja el rastro legible; l10n_pe_ne_move_id es el enlace real (por id).
            moves |= self.env["stock.move"].create(
                {
                    "product_id": l.product_id.id,
                    "product_uom_qty": abs(l.quantity),
                    "product_uom": l.product_uom_id.id,
                    "location_id": origen.id,
                    "location_dest_id": destino.id,
                    "company_id": self.company_id.id,
                    "origin": self.name or "",
                    "l10n_pe_ne_move_id": self.id,
                    "l10n_pe_ne_reversa": reversa,
                }
            )
        try:
            moves._action_confirm()
            moves._action_assign()
            for m, l in zip(moves, lineas):
                lote = lotes.get(l.id)
                if lote:
                    # Entrada de un producto rastreado: el lote va en la LÍNEA del movimiento
                    # (stock.move_line), no en el move. Sin esto Odoo lanza "debe proporcionar
                    # un número de serie o lote" y la mercadería no entraría.
                    if not m.move_line_ids:
                        m.move_line_ids = [(0, 0, {
                            "product_id": m.product_id.id,
                            "location_id": m.location_id.id,
                            "location_dest_id": m.location_dest_id.id,
                            "company_id": m.company_id.id,
                        })]
                    m.move_line_ids.write({"lot_id": lote.id})
                # quantity explícito: sin esto _action_done mueve solo lo reservado, y sin
                # existencias no reserva nada → la salida quedaría en 0 y el kardex mentiría.
                m.quantity = m.product_uom_qty
                m.picked = True
            moves._action_done()
            # La fecha del movimiento es la del DOCUMENTO, no la de cuando se registró.
            # Odoo pone `date` = ahora al validar; sin corregirlo, una compra de marzo
            # cargada en julio caería en el kardex de julio y el libro del periodo saldría
            # mal. Se escribe después de _action_done porque antes lo pisa él.
            if self.invoice_date:
                moves.write({"date": self.invoice_date})
                moves.move_line_ids.write({"date": self.invoice_date})
        except Exception as e:  # noqa: BLE001 — el documento ya existe: el stock no lo tumba
            # Se traga a propósito: el comprobante ya es válido ante SUNAT y no puede caerse
            # porque el inventario no cuadre. Pero se deja RASTRO en el documento, no solo en
            # el log: un movimiento que no ocurre y nadie ve es un kardex mintiendo en
            # silencio. El caso típico es un producto rastreado sin existencias en ningún
            # lote — ahí Odoo no puede inventar de dónde sale.
            _logger.exception("stock: no se pudo mover el stock de %s: %s", self.name, e)
            self.l10n_pe_ne_stock_aviso = (
                _("No se pudo mover el inventario de este documento: %s") % e
            )[:500]
            return self.env["stock.move"].browse()
        self.l10n_pe_ne_stock_aviso = False
        return moves

    @api.model
    def _l10n_pe_ne_asegurar_fefo(self, wh):
        """Pone la ubicación de existencias en FEFO: sale primero lo que vence antes.

        El default de Odoo es FIFO —sale lo que entró primero—, y para lo que caduca eso está
        MAL: comprobado con dos lotes (uno vence 2026, otro 2028), la venta se llevó el de
        2028 y dejó el de 2026 pudriéndose en el almacén. En una farmacia eso es plata tirada
        y riesgo sanitario.

        FEFO no perjudica a lo que no vence: sin fecha de caducidad, Odoo cae de vuelta al
        orden de entrada. Por eso se pone en la ubicación y no producto por producto.

        Idempotente: si ya está, no toca nada. Se llama al ingresar mercadería porque es
        cuando la ubicación empieza a importar — no hay un lugar mejor sin un asistente de
        configuración, que esta app no tiene."""
        fefo = self.env.ref("product_expiry.removal_fefo", raise_if_not_found=False)
        loc = wh.lot_stock_id if wh else None
        if fefo and loc and not loc.removal_strategy_id:
            loc.sudo().removal_strategy_id = fefo.id

    def _l10n_pe_ne_mover_stock_compra(self):
        """Entrada de mercadería por una compra: proveedor → existencias.

        Es la otra mitad del kardex. Sin esto solo hay salidas y todo negocio deriva a
        negativo: un inventario permanente es entradas MENOS salidas.

        Va aparte de _l10n_pe_ne_mover_stock y no reusa su dirección a propósito: aquella
        deduce el sentido de _l10n_pe_document_type(), que para un `in_invoice` devuelve '03'
        — o sea que trataría la compra como una boleta y SACARÍA el stock en vez de meterlo.
        """
        self.ensure_one()
        if self.move_type != "in_invoice":
            return self.env["stock.move"].browse()
        lineas = self._l10n_pe_ne_lineas_con_stock()
        if not lineas:
            return self.env["stock.move"].browse()
        wh = self.env["stock.warehouse"].search(
            [("company_id", "=", self.company_id.id)], limit=1
        )
        proveedores = self.env.ref(
            "stock.stock_location_suppliers", raise_if_not_found=False
        )
        if not wh or not proveedores:
            _logger.warning(
                "stock: sin almacén o ubicación de proveedores para %s; no entra mercadería",
                self.name,
            )
            return self.env["stock.move"].browse()
        # La mercadería que entra decide cómo saldrá: FEFO para que lo que vence antes se
        # venda primero (el default de Odoo, FIFO, dejaría caducar el lote más viejo).
        self._l10n_pe_ne_asegurar_fefo(wh)
        # con_lote: la entrada es la única que define el lote (la salida lo asigna Odoo).
        return self._l10n_pe_ne_stock_aplicar(
            lineas, proveedores, wh.lot_stock_id, con_lote=True
        )

    def _l10n_pe_ne_revertir_stock(self):
        """Deshace el movimiento de un comprobante que SUNAT rechazó.

        Un rechazado NO existe para SUNAT: hay que corregir y emitir uno NUEVO. Ese nuevo
        comprobante vuelve a descontar, así que si el rechazado se queda con su movimiento,
        el bien sale DOS VECES del kardex por una sola venta.

        Se REVIERTE, no se borra: el kardex es un libro: se compensa con el movimiento
        contrario y queda el rastro de que hubo un intento. Borrar el original escondería que
        pasó algo, que es justo lo que un inventario permanente no debe hacer.

        Idempotente: si ya se revirtió, no hace nada. Lo llama el write() al detectar la
        transición a 'rechazado' — por ahí pasan los tres caminos que la fijan (el envío
        síncrono, el cron de pendientes y el resumen diario de boletas), y también cualquiera
        que se agregue después.
        """
        self.ensure_one()
        Move = self.env["stock.move"]
        hechos = Move.search(
            [
                ("l10n_pe_ne_move_id", "=", self.id),
                ("l10n_pe_ne_reversa", "=", False),
                ("state", "=", "done"),
            ]
        )
        if not hechos:
            return Move.browse()
        ya = Move.search_count(
            [("l10n_pe_ne_move_id", "=", self.id), ("l10n_pe_ne_reversa", "=", True)]
        )
        if ya:
            return Move.browse()
        return self._l10n_pe_ne_mover_stock(reversa=True)

    def write(self, vals):
        """Revierte el stock al pasar a 'rechazado'.

        Va en el write y no en cada sitio que fija el estado porque son tres (envío síncrono,
        cron de pendientes, resumen diario de boletas) y mañana pueden ser cuatro: la
        invariante no debe depender de que alguien se acuerde de llamar al helper.

        Solo los que ENTRAN a rechazado (los que ya lo estaban no se re-revierten).
        """
        revertir = self.browse()
        if vals.get("l10n_pe_biller_state") == "rechazado":
            revertir = self.filtered(
                lambda m: m.l10n_pe_biller_state != "rechazado"
            )
        res = super().write(vals)
        for m in revertir:
            m._l10n_pe_ne_revertir_stock()
        return res

    def _l10n_pe_ne_quick_product(self, ln, tax=None, create=True, precio_con_igv=True):
        """Resuelve el product.product de una línea para que el documento USE un registro de Odoo:
        busca por id, por código (default_code) o por nombre exacto; si no existe y hay datos, lo
        CREA simplificado y lo enlaza (igual que el cliente por vat). Devuelve recordset vacío si la
        línea no aporta nada por lo que crear (queda como texto libre, compatible hacia atrás).
        Con create=False solo resuelve y NUNCA crea: las líneas de notas (07/08) pueden traer
        texto sintético (p. ej. "DICE: … DEBE DECIR: …" del motivo 03) que no debe convertirse
        en producto del catálogo.
        `precio_con_igv`: la convención del catálogo es list_price CON IGV (Productos, POS e
        import lo tratan como precio de vitrina). El payload de EMISIÓN trae el valor unitario
        SIN IGV (ni ISC): quick_emit pasa False y al crear se repone el impuesto — sin esto el
        producto auto-creado quedaba ~15% más barato al revenderlo desde el catálogo."""
        Product = self.env["product.product"]
        # `conceptoLibre`: el usuario dijo que esto NO es un producto, sino el detalle de un
        # servicio, distinto en cada comprobante ("POR EL SERVICIO DE TRANSPORTE LIMA-JULIACA …
        # DAM NRO. …"). No hay nada que resolver ni que crear, y se respeta al pie de la letra:
        # engancharlo a uno del catálogo que se llame igual movería su stock, que es justo lo
        # que el usuario dijo que no era. Uno por factura, además, volvería basura el catálogo.
        if ln.get("conceptoLibre"):
            return Product.browse()
        pid = ln.get("productId")
        if pid:
            prod = Product.browse(int(pid)).exists()
            if prod:
                return prod
        cod = (ln.get("productCod") or ln.get("codProducto") or "").strip()
        if cod:
            found = Product.search([("default_code", "=", cod)], limit=1)
            if found:
                return found
        desc = (ln.get("descripcion") or "").strip()
        if desc:
            found = Product.search([("name", "=", desc)], limit=1)
            if found:
                return found
        if not (cod or desc) or not create:
            return Product.browse()
        precio = float(ln.get("precioUnitario") or 0)
        if not precio_con_igv and tax and (tax.amount or 0) > 0:
            # Valor SIN IGV (ni ISC) del payload de emisión → precio de vitrina CON IGV.
            isc = float(ln.get("isc") or 0)
            precio = round(precio * (1 + isc / 100.0) * (1 + (tax.amount or 0) / 100.0), 4)
        uni = (ln.get("unidad") or "").strip()
        vals = {
            "name": desc or cod or "PRODUCTO",
            "type": self._l10n_pe_ne_tipo_producto(ln.get("tipo"), uni),
            "sale_ok": True,
            "list_price": precio,
            # is_storable va en False por defecto en Odoo: sin decirlo explícito, el producto
            # NO llevaría existencias y nunca movería stock. El auto-creado al emitir se queda
            # sin stock a propósito (nadie eligió); el catálogo lo manda por llevaStock.
            "is_storable": bool(ln.get("llevaStock")),
            "tracking": self._l10n_pe_ne_rastreo_producto(ln.get("rastreo")),
            # use_expiration_date lo agrega product_expiry; solo tiene sentido con rastreo.
            "use_expiration_date": bool(ln.get("vence")),
            "l10n_pe_ne_margen": float(ln.get("margen") or 0),
            # company_id del emisor: aísla el producto por RUC (igual que el cliente).
            "company_id": self.env.company.id,
        }
        # Costo: solo lo trae quien lo conoce (crear desde una línea de compra sabe cuánto se
        # pagó). Al emitir no viene, y ahí no se toca: el costo de venta no es el de compra.
        costo = float(ln.get("costo") or 0)
        if costo > 0:
            vals["standard_price"] = costo
        if cod:
            vals["default_code"] = cod
        bc = (ln.get("barcode") or "").strip()
        if bc:
            vals["barcode"] = bc
        cs = (ln.get("codSunat") or "").strip()
        if cs:
            vals["l10n_pe_ne_cod_producto_sunat"] = cs
        if ln.get("detraCod"):
            vals["l10n_pe_ne_detraccion_cod"] = str(ln["detraCod"]).strip()
        if ln.get("percepTasa"):
            vals["l10n_pe_ne_percepcion_tasa"] = _percep_float(ln["percepTasa"])
        if uni:
            vals["l10n_pe_ne_unit_code"] = uni
        if tax:
            vals["taxes_id"] = [(6, 0, tax.ids)]
        return Product.create(vals)

    def _l10n_pe_ne_product_dict(self, p):
        sale_taxes = p.taxes_id.filtered(lambda t: t.type_tax_use == "sale")
        # El ICBPER (7152) NO es la afectación: es un tributo aparte (bolsa plástica). Se
        # deriva como flag propio y se excluye al elegir la afectación (IGV) del producto.
        icbper = bool(sale_taxes.filtered(lambda t: t.l10n_pe_edi_tax_code == "7152"))
        tax = sale_taxes.filtered(lambda t: t.l10n_pe_edi_tax_code != "7152")[:1]
        return {
            "id": p.id,
            "descripcion": p.name or "",
            "codigo": p.default_code or "",
            "barcode": p.barcode or "",
            "codSunat": p.l10n_pe_ne_cod_producto_sunat or "",
            "detraCod": p.l10n_pe_ne_detraccion_cod or "",
            "percepTasa": p.l10n_pe_ne_percepcion_tasa or 0.0,
            "precio": p.list_price,
            "taxCode": (tax.l10n_pe_edi_tax_code or "1000") if tax else "1000",
            "unidad": p.l10n_pe_ne_unit_code or "",
            "icbper": icbper,
            # "bien" | "servicio" — el vocabulario del negocio, no el de Odoo (consu/service).
            # 'combo' no lo usa esta app; si apareciera, se trata como bien (es tangible).
            "tipo": "servicio" if p.type == "service" else "bien",
            # ¿Se le llevan existencias? (Odoo: is_storable). Va en False por defecto, así que
            # SIN esto ningún producto movería stock nunca: es lo que activa _l10n_pe_ne_mover_stock.
            "llevaStock": bool(p.is_storable),
            # Existencias actuales. Solo tiene sentido si llevaStock; si no, va en 0 y la UI
            # muestra un guion (no es "cero unidades", es "no aplica").
            "stock": p.qty_available if p.is_storable else 0.0,
            # Costo (con IGV) y margen: lo que hace falta para proponer el precio de venta
            # cuando una compra trae un costo distinto.
            "costo": p.standard_price or 0.0,
            "margen": p.l10n_pe_ne_margen or 0.0,
            # Rastreo por lote o serie (Odoo: tracking). "lote" agrupa unidades (farmacia,
            # alimentos); "serie" es un número por unidad (celulares, equipos).
            "rastreo": {"lot": "lote", "serial": "serie"}.get(p.tracking, "ninguno"),
            # ¿Los lotes llevan vencimiento? Solo aplica con rastreo por lote/serie.
            "vence": bool(p.use_expiration_date),
        }

    def _l10n_pe_ne_partner_dict(self, p):
        return {
            "id": p.id,
            "razonSocial": p.name or "",
            "numDoc": p.vat or "",
            "tipoDoc": p.l10n_latam_identification_type_id.l10n_pe_vat_code or "",
            "tipoDocNombre": p.l10n_latam_identification_type_id.name or "",
            "email": p.email or "",
            "telefono": p.phone or "",
            "direccion": p.street or "",
            "pais": p.country_id.code or "",
            "exceptuadoPercepcion": p.l10n_pe_ne_exceptuado_percepcion,
            "parteVinculada": p.l10n_pe_ne_parte_vinculada,
        }

    def _l10n_pe_ne_ident_type(self, tipoDoc):
        return self.env["l10n_latam.identification.type"].search(
            [("l10n_pe_vat_code", "=", tipoDoc or "6")], limit=1
        )

    def _l10n_pe_ne_partner_apply(self, p, c):
        """Aplica los campos simplificados (los del caso común de facturación) a un res.partner."""
        vals = {}
        if c.get("razonSocial"):
            vals["name"] = c["razonSocial"]
        if c.get("numDoc") is not None:
            vals["vat"] = (c.get("numDoc") or "").strip() or False
        if c.get("tipoDoc"):
            t = self._l10n_pe_ne_ident_type(c["tipoDoc"])
            if t:
                vals["l10n_latam_identification_type_id"] = t.id
        # País del adquirente (exportación / no domiciliado): ISO 3166 alpha-2 = res.country.code.
        # Alimenta codPaisCliente en la cabecera 0200. "" limpia el país.
        if "pais" in c:
            code = (c.get("pais") or "").strip().upper()
            country = self.env["res.country"].search([("code", "=", code)], limit=1) if code else False
            vals["country_id"] = country.id if country else False
        for key, field in (
            ("email", "email"),
            ("telefono", "phone"),
            ("direccion", "street"),
            ("exceptuadoPercepcion", "l10n_pe_ne_exceptuado_percepcion"),
            ("parteVinculada", "l10n_pe_ne_parte_vinculada"),
        ):
            if key in c:
                vals[field] = c.get(key) or False
        if vals:
            p.write(vals)
        return p

    @api.model
    def l10n_pe_ne_list_clientes(self, query=None, limit=50, offset=None):
        """Clientes de Odoo para que React liste/autocomplete (no reinventa el padrón).

        Paginación opt-in: con `offset` (aunque sea 0) devuelve el envelope
        {items, total}; sin `offset` (None) devuelve la lista plana de siempre
        —así el autocomplete del POS/Emitir sigue recibiendo un array."""
        domain = [("customer_rank", ">", 0)]
        if query:
            domain = [
                "&",
                ("customer_rank", ">", 0),
                "|",
                ("name", "ilike", query),
                ("vat", "ilike", query),
            ]
        Partner = self.env["res.partner"]
        parts = Partner.search(domain, order="name", limit=limit, offset=offset or 0)
        items = [self._l10n_pe_ne_partner_dict(p) for p in parts]
        if offset is None:
            return items
        return {"items": items, "total": Partner.search_count(domain)}

    @api.model
    def l10n_pe_ne_create_cliente(self, cliente):
        """Crea (o reusa por vat) un cliente con los campos PE correctos; lo guarda EN Odoo."""
        cliente = cliente or {}
        p = self._l10n_pe_ne_quick_partner(cliente)
        self._l10n_pe_ne_partner_apply(p, cliente)
        if not p.customer_rank:
            p.customer_rank = 1
        return self._l10n_pe_ne_partner_dict(p)

    @api.model
    def l10n_pe_ne_update_cliente(self, cliente):
        """Actualiza un cliente existente (por id) con los campos simplificados."""
        cliente = cliente or {}
        p = self.env["res.partner"].browse(int(cliente.get("id") or 0)).exists()
        if not p:
            raise UserError(_("Cliente no encontrado."))
        self._l10n_pe_ne_partner_apply(p, cliente)
        return self._l10n_pe_ne_partner_dict(p)

    @api.model
    def l10n_pe_ne_delete_cliente(self, rec_id):
        """Elimina el cliente; si está referenciado (comprobantes), lo archiva en su lugar."""
        p = self.env["res.partner"].browse(int(rec_id or 0)).exists()
        if not p:
            return {"ok": True, "modo": "inexistente"}
        try:
            p.unlink()
            return {"ok": True, "modo": "eliminado"}
        except Exception:
            p.active = False
            return {"ok": True, "modo": "archivado"}

    @api.model
    def l10n_pe_ne_list_productos(self, query=None, limit=50, offset=None):
        """Productos de Odoo para que React liste/autocomplete y los documentos los referencien.
        Busca por nombre, código interno (default_code) o código de barras (barcode).

        Paginación opt-in: con `offset` devuelve {items, total}; sin él, lista plana."""
        domain = [("sale_ok", "=", True)]
        if query:
            domain = [
                "&",
                ("sale_ok", "=", True),
                "|",
                "|",
                ("name", "ilike", query),
                ("default_code", "ilike", query),
                ("barcode", "ilike", query),
            ]
        Product = self.env["product.product"]
        prods = Product.search(domain, order="name", limit=limit, offset=offset or 0)
        items = [self._l10n_pe_ne_product_dict(p) for p in prods]
        if offset is None:
            return items
        return {"items": items, "total": Product.search_count(domain)}

    @api.model
    def l10n_pe_ne_producto_por_barcode(self, code):
        """Resuelve UN producto por código de barras exacto (para el escaneo en el POS).
        Devuelve el dict del producto o None si no hay coincidencia. Aislado por compañía."""
        code = (code or "").strip()
        if not code:
            return None
        p = self.env["product.product"].search(
            [("sale_ok", "=", True), ("barcode", "=", code)], limit=1
        )
        return self._l10n_pe_ne_product_dict(p) if p else None

    @api.model
    def l10n_pe_ne_create_producto(self, producto):
        """Crea (o reusa por código/nombre) un producto simplificado; lo guarda EN Odoo."""
        _logger.info("l10n_pe_ne_create_producto: %s", producto)
        producto = producto or {}
        desc = producto.get("descripcion") or producto.get("nombre")
        _logger.info("desc: %s", desc)
        if not desc and not producto.get("codigo"):
            raise UserError(
                _("El producto necesita al menos una descripción o un código.")
            )
        tax = self._l10n_pe_ne_tax_by_code(producto.get("taxCode") or "1000")
        _logger.info("tax: %s", tax)
        p = self._l10n_pe_ne_quick_product(
            {
                "descripcion": desc,
                "productCod": producto.get("codigo"),
                "barcode": producto.get("barcode"),
                "codSunat": producto.get("codSunat"),
                "detraCod": producto.get("detraCod"),
                "percepTasa": producto.get("percepTasa"),
                "precioUnitario": producto.get("precio"),
                "unidad": producto.get("unidad"),
                "tipo": producto.get("tipo"),
                "llevaStock": producto.get("llevaStock"),
                "rastreo": producto.get("rastreo"),
                "vence": producto.get("vence"),
                "margen": producto.get("margen"),
                "costo": producto.get("costo"),
            },
            tax,
        )
        _logger.info("p: %s", p)
        return self._l10n_pe_ne_product_dict(p)

    @api.model
    def l10n_pe_ne_update_producto(self, producto):
        """Actualiza un producto (por id): descripción, código, precio e impuesto (afectación)."""
        producto = producto or {}
        p = self.env["product.product"].browse(int(producto.get("id") or 0)).exists()
        if not p:
            raise UserError(_("Producto no encontrado."))
        vals = {}
        if producto.get("descripcion"):
            vals["name"] = producto["descripcion"]
        if "codigo" in producto:
            vals["default_code"] = (producto.get("codigo") or "").strip() or False
        if "barcode" in producto:
            vals["barcode"] = (producto.get("barcode") or "").strip() or False
        if "codSunat" in producto:
            vals["l10n_pe_ne_cod_producto_sunat"] = (producto.get("codSunat") or "").strip() or False
        if "detraCod" in producto:
            vals["l10n_pe_ne_detraccion_cod"] = (producto.get("detraCod") or "").strip() or False
        if producto.get("percepTasa") is not None:
            vals["l10n_pe_ne_percepcion_tasa"] = _percep_float(producto.get("percepTasa"))
        if "unidad" in producto:
            vals["l10n_pe_ne_unit_code"] = (producto.get("unidad") or "").strip() or False
        if producto.get("tipo"):
            # Solo si viene explícito: aquí NO se deduce de la unidad. Cambiar la unidad de un
            # producto ya clasificado no debe reclasificarlo a su espalda.
            vals["type"] = self._l10n_pe_ne_tipo_producto(producto["tipo"])
        if "llevaStock" in producto:
            vals["is_storable"] = bool(producto.get("llevaStock"))
        if "rastreo" in producto:
            vals["tracking"] = self._l10n_pe_ne_rastreo_producto(producto.get("rastreo"))
        if producto.get("margen") is not None:
            vals["l10n_pe_ne_margen"] = float(producto.get("margen") or 0)
        if producto.get("costo") is not None:
            vals["standard_price"] = float(producto.get("costo") or 0)
        if "vence" in producto:
            vals["use_expiration_date"] = bool(producto.get("vence"))
        if producto.get("precio") is not None:
            vals["list_price"] = float(producto.get("precio") or 0)
        if producto.get("taxCode"):
            tax = self._l10n_pe_ne_tax_by_code(producto["taxCode"])
            vals["taxes_id"] = [(6, 0, tax.ids if tax else [])]
        if vals:
            p.write(vals)
        return self._l10n_pe_ne_product_dict(p)

    @api.model
    def l10n_pe_ne_delete_producto(self, rec_id):
        """Elimina el producto; si está referenciado (en comprobantes), lo archiva en su lugar."""
        p = self.env["product.product"].browse(int(rec_id or 0)).exists()
        if not p:
            return {"ok": True, "modo": "inexistente"}
        try:
            p.unlink()
            return {"ok": True, "modo": "eliminado"}
        except Exception:
            p.active = False
            return {"ok": True, "modo": "archivado"}

    # ------------------------------------------------------- importación productos
    @api.model
    def l10n_pe_ne_plantilla_productos(self):
        """Plantilla xlsx para importar/actualizar el catálogo (hoja 'Productos' con las
        cabeceras + ejemplos + listas de unidad/afectación, y una hoja 'Instrucciones').
        Devuelve {filename, contentB64}. Mismo estilo visual que la plantilla de la masiva."""
        import io
        import base64
        import xlsxwriter

        headers = ["CÓDIGO", "CÓDIGO DE BARRAS", "NOMBRE", "UNIDAD", "PRECIO VENTA", "COSTO", "AFECTACIÓN", "BOLSA", "DETRACCIÓN", "PERCEPCION %"]
        ejemplos = [
            ["PROD0001", "7751234000018", "CEMENTO SOL 42.5 KG", "UNIDAD", 33.90, 28.00, "GRAVADO", "NO", "", ""],
            ["PROD0002", "7751234000025", "FIERRO CORRUGADO 1/2 PULG", "KILOGRAMO", 4.50, 3.20, "GRAVADO", "NO", "", ""],
            ["PROD0004", "", "BOLSA PLÁSTICA", "UNIDAD", 0.50, 0.10, "GRAVADO", "SI", "", ""],
        ]
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet("Productos")
        head = wb.add_format({"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1})
        # El código de barras se escribe como TEXTO para no perder ceros a la izquierda
        # ni que Excel lo pase a notación científica (ej. 7.75E+12).
        txtfmt = wb.add_format({"num_format": "@"})
        for c, h in enumerate(headers):
            ws.write(0, c, h, head)
            # DETRACCIÓN también va como TEXTO: sus códigos (027, 019, 022...) empiezan con
            # cero y Excel se lo comería si la celda quedara en formato numérico.
            ws.set_column(c, c, max(16, len(h) + 4), txtfmt if c in (1, 8) else None)
        for r, row in enumerate(ejemplos, 1):
            ws.write_row(r, 0, row)
        # Comentarios de ayuda al pasar el mouse por la cabecera (el triangulito rojo).
        note = {"x_scale": 2.2, "y_scale": 1.8, "author": "CHASKIFACT"}
        ws.write_comment(0, 1, (
            "Opcional. El código de barras (EAN) que trae el producto, para escanearlo "
            "en el POS. Déjalo vacío si el producto no tiene."), note)
        ws.write_comment(0, 4, "Precio final CON IGV incluido (lo que paga el cliente).", note)
        ws.write_comment(0, 5, "Opcional. Precio de compra referencial. NO afecta la facturación.", note)
        ws.write_comment(0, 6, (
            "Tipo de afectación de IGV. Elígelo del desplegable.\n"
            "• GRAVADO = con IGV 18% (lo normal)\n"
            "• EXONERADO / INAFECTO = sin IGV\n"
            "• EXPORTACION / GRATUITO = casos especiales\n"
            "Si lo dejas vacío se asume GRAVADO."), note)
        ws.write_comment(0, 7, (
            "SI / NO. Márcalo SI solo si el producto es una BOLSA PLÁSTICA: "
            "cobra el ICBPER (monto fijo por unidad) al venderlo. Vacío = NO."), note)
        ws.write_comment(0, 8, (
            "Opcional. Código cat. 54 de SUNAT si el producto está sujeto a detracción "
            "(ej. 027 transporte de carga). Vacío = no sujeto."), note)
        ws.write_comment(0, 9, (
            "Opcional. % de percepción sugerido si el bien está en el Apéndice 1 "
            "(2 general, 1 combustibles). Vacío = no sujeto."), note)
        # Desplegable (select) para UNIDAD, con ayuda al hacer clic en la celda.
        ws.data_validation(1, 3, 1000, 3, {
            "validate": "list", "source": [
                "UNIDAD", "SERVICIO", "KILOGRAMO", "GRAMO", "LITRO", "GALON", "CAJA",
                "METRO", "METRO CUADRADO", "METRO CUBICO", "MILLAR", "DOCENA"],
            "input_title": "Unidad de medida",
            "input_message": "Elige de la lista o escribe el código SUNAT (NIU, KGM…). Vacío = UNIDAD."})
        # Desplegable (select) para AFECTACIÓN, con ayuda + alerta suave si no es de la lista.
        ws.data_validation(1, 6, 1000, 6, {
            "validate": "list", "source": [
                "GRAVADO", "EXONERADO", "INAFECTO", "EXPORTACION", "GRATUITO"],
            "input_title": "Afectación de IGV",
            "input_message": (
                "GRAVADO = con IGV 18% (lo normal).\n"
                "EXONERADO / INAFECTO = sin IGV.\n"
                "Vacío = GRAVADO."),
            "error_type": "information",
            "error_title": "Valor sugerido",
            "error_message": "Usa: GRAVADO, EXONERADO, INAFECTO, EXPORTACION o GRATUITO."})
        # Desplegable (select) SI/NO para BOLSA (ICBPER).
        ws.data_validation(1, 7, 1000, 7, {
            "validate": "list", "source": ["SI", "NO"],
            "input_title": "Bolsa plástica (ICBPER)",
            "input_message": (
                "SI solo si es una bolsa plástica: cobra ICBPER por unidad.\n"
                "Para todo lo demás: NO (o déjalo vacío).")})
        ws.freeze_panes(1, 0)
        wi = wb.add_worksheet("Instrucciones")
        wi.set_column(0, 0, 110)
        for r, line in enumerate([
            "CHASKIFACT — Plantilla de importación de productos",
            "",
            "1. Una fila = un producto. 'CÓDIGO' es la clave: si ya existe, se ACTUALIZA; si no, se CREA.",
            "2. 'CÓDIGO DE BARRAS' es opcional: el EAN del producto para escanearlo en el POS. No puede repetirse entre productos.",
            "3. 'NOMBRE' es obligatorio. 'PRECIO VENTA' es el precio final CON IGV incluido.",
            "4. 'UNIDAD': puedes escribir el nombre (UNIDAD, KILOGRAMO, CAJA…) o el código SUNAT (NIU, KGM, BX…). Vacío = UNIDAD (NIU).",
            "5. 'AFECTACIÓN' (elígela del desplegable de la celda): define el IGV del producto.",
            "     • GRAVADO = lleva IGV 18% (la mayoría de productos).  • EXONERADO / INAFECTO = sin IGV.",
            "     • EXPORTACION / GRATUITO = casos especiales.  Si la dejas vacía se asume GRAVADO.",
            "6. 'COSTO' es opcional (precio de compra, referencial). No afecta la facturación.",
            "7. 'BOLSA' = SI solo para bolsas plásticas (cobran ICBPER por unidad al venderlas). Para el resto: NO o vacío.",
            "8. 'DETRACCIÓN' es opcional: código cat. 54 de SUNAT (3 dígitos, ej. 027 transporte de carga) si el producto está sujeto a detracción. Vacío = no sujeto.",
            "9. 'PERCEPCION %' es opcional: % de percepción sugerido si el bien está en el Apéndice 1 (2 general, 1 combustibles). Vacío = no sujeto.",
            "10. Sube el archivo, revisa el resumen (nuevos / actualizados / errores) y recién ahí confirma.",
        ]):
            wi.write(r, 0, line)
        wb.close()
        return {"filename": "plantilla-productos-chaskifact.xlsx",
                "contentB64": base64.b64encode(buf.getvalue()).decode("ascii")}

    @api.model
    def l10n_pe_ne_revisar_tipos(self, payload=None):
        """Propone reclasificar los productos que quedaron como SERVICIO por el default viejo.

        Hasta hace poco todo producto nacía con type='service' —estuviera bien o no—, así que
        un catálogo existente tiene tornillos declarados como servicios. Y un servicio no
        lleva stock en Odoo: mientras no se corrijan, esos productos no mueven inventario.

        PROPONE, no decide. La deducción usa la misma regla que la creación
        (_l10n_pe_ne_tipo_producto: ZZ → servicio, el resto → bien), pero acá puede
        equivocarse: el formulario trae NIU por defecto, así que una consultora que no lo
        cambió tiene servicios con NIU y saldrían propuestos como bienes. Por eso se devuelve
        la lista para que la revise un humano y se aplica solo lo que confirme —
        `l10n_pe_ne_aplicar_tipos` recibe los ids elegidos, no un "aplicar todo".

        No propone nada sobre `llevaStock`: llevar inventario es una decisión del negocio y
        no hay señal ninguna que la delate. Se activa producto por producto.
        """
        Product = self.env["product.product"]
        # Solo los 'service': un 'consu' ya fue clasificado (por el usuario o por la regla).
        sospechosos = Product.search(
            [("type", "=", "service"), ("company_id", "in", (False, self.env.company.id))],
            order="name",
        )
        propuestas = []
        for p in sospechosos:
            uni = p.l10n_pe_ne_unit_code or ""
            propuesto = self._l10n_pe_ne_tipo_producto(None, uni)
            if propuesto != "service":
                propuestas.append({
                    "id": p.id,
                    "descripcion": p.name or "",
                    "codigo": p.default_code or "",
                    # Sin unidad no significa "servicio": a SUNAT se le declara NIU por
                    # defecto (DEFAULT_UNIT_CODE), o sea un bien. Se muestra para que el
                    # usuario juzgue con el mismo dato que usó la regla.
                    "unidad": uni,
                    "tipoPropuesto": "bien",
                })
        return {
            "propuestas": propuestas,
            "total": len(propuestas),
            "revisados": len(sospechosos),
        }

    @api.model
    def l10n_pe_ne_aplicar_tipos(self, payload):
        """Aplica la reclasificación SOLO a los ids que el usuario confirmó.
        payload = {ids: [...], tipo: "bien"|"servicio"}."""
        payload = payload or {}
        ids = [int(i) for i in (payload.get("ids") or [])]
        if not ids:
            return {"actualizados": 0}
        tipo = self._l10n_pe_ne_tipo_producto(payload.get("tipo") or "bien")
        prods = self.env["product.product"].browse(ids).exists()
        prods.write({"type": tipo})
        return {"actualizados": len(prods)}

    @api.model
    def l10n_pe_ne_importar_productos(self, payload):
        """Importa/actualiza productos desde el xlsx de la plantilla. payload = {contentB64, commit}.
        UPSERT por CÓDIGO. commit=False → solo valida y devuelve el reporte (dry-run, no escribe);
        commit=True → aplica y devuelve creados/actualizados/errores. Aislado por compañía."""
        import io
        import base64
        import unicodedata

        payload = payload or {}
        commit = bool(payload.get("commit"))
        try:
            data = base64.b64decode(payload.get("contentB64") or "")
        except Exception:
            raise UserError(_("Archivo inválido."))

        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception:
            raise UserError(_("No se pudo leer el archivo. Sube un .xlsx válido (no un .xls antiguo)."))
        ws = wb["Productos"] if "Productos" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("El archivo está vacío."))

        def norm(h):
            s = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode("ascii")
            return " ".join(s.lower().split())

        header = [norm(h) for h in rows[0]]
        idx = {h: i for i, h in enumerate(header) if h}
        faltan = [h for h in ("codigo", "nombre") if h not in idx]
        if faltan:
            raise UserError(_("Faltan columnas obligatorias: %s. Usa la plantilla.") % ", ".join(faltan))

        def cell(row, name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        def txt(v):
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        def num(v):
            if v is None or (isinstance(v, str) and not v.strip()):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).strip().replace(" ", "").replace(",", "."))
            except ValueError:
                return "ERROR"

        Product = self.env["product.product"]
        creados = actualizados = 0
        errores = []
        avisos = []
        for n, row in enumerate(rows[1:], start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            cod = txt(cell(row, "codigo"))
            nombre = txt(cell(row, "nombre"))
            if not cod:
                errores.append({"fila": n, "msg": "Falta el CÓDIGO"})
                continue
            if not nombre:
                errores.append({"fila": n, "msg": "Falta el NOMBRE"})
                continue
            precio = num(cell(row, "precio venta"))
            costo = num(cell(row, "costo"))
            if precio == "ERROR" or costo == "ERROR":
                errores.append({"fila": n, "msg": "PRECIO o COSTO no es un número válido"})
                continue
            uni_raw = norm(cell(row, "unidad"))
            if not uni_raw:
                unidad = "NIU"
            elif uni_raw in UNIDAD_IMPORT:
                unidad = UNIDAD_IMPORT[uni_raw]
            elif uni_raw.upper() in _UNIDAD_CODES:
                unidad = uni_raw.upper()
            else:
                unidad = "NIU"
                avisos.append({"fila": n, "msg": "Unidad '%s' no reconocida, se usó UNIDAD (NIU)" % txt(cell(row, "unidad"))})
            afe_raw = norm(cell(row, "afectacion"))
            tax_code = AFECT_IMPORT.get(afe_raw, "1000") if afe_raw else "1000"
            detra_raw = txt(cell(row, "detraccion"))
            if detra_raw and not re.fullmatch(r"[0-9]{3}", detra_raw):
                errores.append({"fila": n, "msg": "DETRACCIÓN debe ser el código de 3 dígitos del catálogo 54 (ej. 027) o vacío"})
                continue
            percep_raw = txt(cell(row, "percepcion %"))
            # 0 = "no sujeto" (help del campo / percepTasa: 0 en la API): limpia el campo igual
            # que la celda vacía, NO es un valor inválido. Solo <0, >10 o no-numérico son error
            # de fila (y ahí sí se descarta la fila completa, incluido precio/nombre).
            percep_val = False
            if percep_raw:
                try:
                    percep_num = float(percep_raw.replace(",", "."))
                except ValueError:
                    percep_num = None
                if percep_num is None or percep_num < 0 or percep_num > 10:
                    errores.append({"fila": n, "msg": "PERCEPCION % debe ser un número mayor a 0 y hasta 10 (ej. 2), 0/vacío para no sujeto."})
                    continue
                percep_val = percep_num or False  # 0 limpia el campo, igual que vacío
            barcode = txt(cell(row, "codigo de barras"))
            bolsa = norm(cell(row, "bolsa")) in ("si", "s")  # ICBPER: SI/NO (vacío = NO)

            existing = Product.search([("default_code", "=", cod)], limit=1)
            # El código de barras no puede pertenecer a OTRO producto (Odoo lo exige único).
            if barcode:
                dup = Product.search([("barcode", "=", barcode)], limit=1)
                if dup and dup.id != existing.id:
                    errores.append({"fila": n, "msg": "El código de barras '%s' ya pertenece a otro producto" % barcode})
                    continue
            if not commit:
                if existing:
                    actualizados += 1
                else:
                    creados += 1
                continue
            vals = {"name": nombre, "l10n_pe_ne_unit_code": unidad}
            # Columna DETRACCIÓN AUSENTE (plantilla vieja) vs. celda VACÍA (el usuario limpió el
            # código) lucen igual por cell() (None en ambos casos): sin este guard contra `idx`
            # (headers reales del archivo) un re-import sin la columna borraba en silencio los
            # códigos de detracción ya guardados vía existing.write(vals).
            if "detraccion" in idx:
                vals["l10n_pe_ne_detraccion_cod"] = detra_raw or False
            if "percepcion %" in idx:
                vals["l10n_pe_ne_percepcion_tasa"] = percep_val
            if precio is not None:
                vals["list_price"] = precio
            if costo is not None:
                vals["standard_price"] = costo
            if barcode:
                vals["barcode"] = barcode
            tax = self._l10n_pe_ne_tax_by_code(tax_code)
            tax_ids = list(tax.ids) if tax else []
            if bolsa:  # bolsa plástica → suma la tax ICBPER (monto fijo por unidad)
                tax_ids += self._l10n_pe_ne_ensure_icbper_tax().ids
            vals["taxes_id"] = [(6, 0, tax_ids)]
            if existing:
                existing.write(vals)
                actualizados += 1
            else:
                # Tipo deducido de la unidad de la fila (ZZ → servicio, resto → bien): el
                # Excel no trae columna de tipo y la unidad es la señal que sí trae.
                vals.update({"default_code": cod,
                             "type": self._l10n_pe_ne_tipo_producto(unidad=unidad),
                             "sale_ok": True, "company_id": self.env.company.id})
                Product.create(vals)
                creados += 1
        return {"commit": commit, "creados": creados, "actualizados": actualizados,
                "errores": errores, "avisos": avisos,
                "totalOk": creados + actualizados, "totalError": len(errores)}

    # ----------------------------------------------------------------- compras
    # Compra = factura de proveedor (account.move in_invoice). TODA la lógica en
    # Odoo; reusa el patrón de in_invoice de retención (campos l10n_latam que PE
    # exige). Aislado por compañía vía reglas multi-compañía nativas de account.move.
    def _l10n_pe_ne_compra_dict(self):
        self.ensure_one()
        return {
            "id": self.id,
            "fecha": self.invoice_date.strftime("%Y-%m-%d")
            if self.invoice_date
            else "",
            "documento": self.l10n_latam_document_number or self.ref or "",
            "tipoComprobante": self.l10n_latam_document_type_id.code
            if self.l10n_latam_document_type_id
            else "",
            "proveedor": self.partner_id.name or "",
            "ruc": self.partner_id.vat or "",
            "total": self.amount_total or 0.0,
            # Base e IGV separados: es lo que pide el Registro de Compras y lo que sostiene el
            # crédito fiscal. Antes la compra iba sin impuesto y solo se guardaba el total.
            # Redondeado a la moneda: la resta en float da 1.1000000000000005 y ese ruido
            # llegaría tal cual a la pantalla.
            "base": self.amount_untaxed or 0.0,
            "igv": float_round(
                (self.amount_total or 0.0) - (self.amount_untaxed or 0.0),
                precision_rounding=self.currency_id.rounding or 0.01,
            ),
            "afectacion": (
                self.invoice_line_ids[:1].tax_ids[:1].l10n_pe_edi_tax_code or "1000"
            ) if self.invoice_line_ids[:1].tax_ids[:1] else "9998",
            "moneda": self.currency_id.name or "PEN",
            "estado": self.state,
            # Descripción = nombre de la línea (para prefill al editar; la línea de
            # detalle simple lleva la descripción original o "COMPRA").
            "descripcion": (self.invoice_line_ids[:1].name or "")
            if self.invoice_line_ids
            else "",
        }

    @api.model
    def l10n_pe_ne_list_compras(self, query=None, periodo=None, limit=200, offset=None):
        """Lista de compras (facturas de proveedor) — opcional por texto o periodo.

        Paginación opt-in: con `offset` devuelve {items, total}; sin él, lista plana."""
        import calendar

        domain = [("move_type", "=", "in_invoice")]
        if query:
            domain += [
                "|",
                ("partner_id.name", "ilike", query),
                ("l10n_latam_document_number", "ilike", query),
            ]
        if periodo and len(str(periodo)) == 6 and str(periodo).isdigit():
            y, m = int(periodo[:4]), int(periodo[4:6])
            last = calendar.monthrange(y, m)[1]
            domain += [
                ("invoice_date", ">=", "%04d-%02d-01" % (y, m)),
                ("invoice_date", "<=", "%04d-%02d-%02d" % (y, m, last)),
            ]
        recs = self.search(
            domain, order="invoice_date desc, id desc", limit=limit, offset=offset or 0
        )
        items = [m._l10n_pe_ne_compra_dict() for m in recs]
        if offset is None:
            return items
        return {"items": items, "total": self.search_count(domain)}

    @api.model
    def l10n_pe_ne_importar_compra_xml(self, payload):
        """Lee el XML de la factura electrónica del PROVEEDOR y devuelve el payload de una
        compra, listo para que el usuario lo revise y guarde. NO registra nada.

        El proveedor está obligado a entregar el XML: es el documento fiscal de verdad (el PDF
        es solo su representación impresa). Leerlo evita teclear —y equivocarse— en el dato
        que va al Registro de Compras.

        Devuelve, no guarda: el mapeo de productos necesita a un humano (ver abajo) y el
        usuario debe poder revisar antes de que entre mercadería al kardex.
        """
        b64 = (payload or {}).get("xml") or ""
        try:
            raw = base64.b64decode(b64)
        except Exception:
            raise UserError(_("El archivo no es un XML válido (base64 ilegible)."))
        return self._l10n_pe_ne_parse_compra_xml(raw)

    @api.model
    def _l10n_pe_ne_parse_compra_xml(self, raw):
        """Parseo puro del UBL 2.1 (Invoice/CreditNote) → payload de compra. Sin ORM salvo el
        match de productos, para poder testearlo con un XML real."""
        from xml.etree import ElementTree as ET

        try:
            root = ET.fromstring(raw)
        except ET.ParseError as e:
            raise UserError(_("No se pudo leer el XML: %s") % e)
        # Los tags vienen con namespace; se ignora el prefijo y se busca por nombre local.
        # Es lo mismo que hace el biller al depurar el XML para el PDF: los namespaces de
        # SUNAT varían por versión y atarse a ellos rompe con el primer proveedor distinto.
        def hijos(el, nombre):
            return [c for c in el if c.tag.rsplit("}", 1)[-1] == nombre] if el is not None else []

        def uno(el, *ruta):
            cur = el
            for nombre in ruta:
                hs = hijos(cur, nombre)
                if not hs:
                    return None
                cur = hs[0]
            return cur

        def txt(el, *ruta):
            n = uno(el, *ruta) if ruta else el
            return (n.text or "").strip() if n is not None and n.text else ""

        if root.tag.rsplit("}", 1)[-1] not in ("Invoice", "CreditNote", "DebitNote"):
            raise UserError(
                _("El XML no es un comprobante electrónico (se esperaba Invoice/CreditNote).")
            )
        sup = uno(root, "AccountingSupplierParty", "Party")
        ruc = txt(sup, "PartyIdentification", "ID")
        razon = txt(sup, "PartyLegalEntity", "RegistrationName") or txt(sup, "PartyName", "Name")
        if not ruc:
            raise UserError(_("El XML no trae el RUC del emisor."))
        doc_id = txt(root, "ID")
        serie, _sep, numero = doc_id.partition("-")
        tipo = txt(root, "InvoiceTypeCode") or "01"
        total = txt(root, "LegalMonetaryTotal", "PayableAmount")
        igv = ""
        for tt in hijos(root, "TaxTotal"):
            igv = txt(tt, "TaxAmount")
            if igv:
                break

        lineas = []
        for ln in hijos(root, "InvoiceLine") + hijos(root, "CreditNoteLine"):
            item = uno(ln, "Item")
            cant = txt(ln, "InvoicedQuantity") or txt(ln, "CreditedQuantity")
            # Precio CON IGV: AlternativeConditionPrice con PriceTypeCode 01 (catálogo 16 de
            # SUNAT) es el precio unitario que incluye el impuesto — la misma convención que
            # usa toda la app. cac:Price (sin IGV) NO sirve acá: el detalle se compara contra
            # el total del documento, que va con IGV.
            precio = ""
            for pr in hijos(uno(ln, "PricingReference") or ln, "AlternativeConditionPrice"):
                if txt(pr, "PriceTypeCode") == "01":
                    precio = txt(pr, "PriceAmount")
                    break
            cod_prov = txt(item, "SellersItemIdentification", "ID")
            barcode = txt(item, "StandardItemIdentification", "ID")
            lineas.append({
                "descripcion": txt(item, "Description"),
                "cantidad": float(cant or 0),
                "precioUnitario": float(precio or 0),
                "codigoProveedor": cod_prov,
                "barcode": barcode,
                # El match con NUESTRO catálogo es una propuesta, no un hecho: el proveedor
                # nombra y codifica los productos a su manera. Sin coincidencia se deja en
                # None y lo elige el usuario — inventar el mapeo ensuciaría el kardex.
                "productId": self._l10n_pe_ne_match_producto(barcode, cod_prov),
            })
        # Afectación: se LEE del XML (TaxScheme/ID, catálogo 05 de la primera línea), no se
        # deduce del IGV. Asumir "gravado" en una factura exonerada le inventaría al usuario
        # un crédito fiscal que no tiene — un error fiscal, no una imprecisión.
        afect = ""
        primera = (hijos(root, "InvoiceLine") + hijos(root, "CreditNoteLine"))[:1]
        if primera:
            st = uno(primera[0], "TaxTotal", "TaxSubtotal")
            afect = txt(st, "TaxCategory", "TaxScheme", "ID")
        return {
            "proveedor": {"tipoDoc": "6", "numDoc": ruc, "razonSocial": razon},
            "tipoComprobante": tipo,
            "serie": serie,
            "numero": numero.lstrip("0") or numero,
            "fecha": txt(root, "IssueDate"),
            "total": float(total or 0),
            "igv": float(igv or 0),
            "afectacion": afect or ("1000" if float(igv or 0) > 0 else "9998"),
            "descripcion": "",
            "lineas": lineas,
        }

    @api.model
    def _l10n_pe_ne_match_producto(self, barcode, codigo):
        """Propone un producto NUESTRO para una línea del XML del proveedor.

        Por código de barras primero (el GTIN es universal: si coincide, es el mismo producto)
        y por código propio después (más débil: 'P001' puede ser cualquier cosa en otro
        catálogo). Sin coincidencia devuelve None y decide el usuario."""
        Product = self.env["product.product"]
        if barcode:
            p = Product.search([("barcode", "=", barcode)], limit=1)
            if p:
                return p.id
        if codigo:
            p = Product.search([("default_code", "=", codigo)], limit=1)
            if p:
                return p.id
        return None

    @api.model
    def _l10n_pe_ne_tax_compra_by_code(self, code):
        """account.tax de COMPRA por código cat-05; default 1000 (IGV gravado).

        Existe aparte del de venta porque el crédito fiscal se imputa con impuestos de
        compra: usar el de venta metería el IGV en la cuenta equivocada. La localización ya
        trae los cuatro (IGV 18%, 0% Exo, 0% Ina, 0% Exp)."""
        return self.env["account.tax"].search(
            [
                ("company_id", "=", self.env.company.id),
                ("type_tax_use", "=", "purchase"),
                ("l10n_pe_edi_tax_code", "=", code or "1000"),
            ],
            limit=1,
        )

    @api.model
    def _l10n_pe_ne_base_sin_igv(self, bruto, tax):
        """Precio CON IGV → base SIN IGV, que es lo que espera `price_unit` con un impuesto
        tax_excluded (la convención de esta app: el usuario ve y teclea precios con IGV).

        No se redondea a 2: con `round_globally` en la compañía, Odoo calcula el impuesto
        sobre la base sin redondear y el total vuelve a dar el bruto redondo. Redondear acá
        rompería justo eso (118 → base 100.00 ✓, pero 7.20 → 6.10 y el total daría 7.198)."""
        rate = (tax.amount or 0) if tax else 0
        return (bruto or 0) / (1 + rate / 100.0) if rate else (bruto or 0)

    @api.model
    def _l10n_pe_ne_compra_lineas(self, compra):
        """invoice_line_ids de una compra, desde `lineas` si vienen, o la línea única del total.

        El detalle es OPCIONAL a propósito: no toda compra es mercadería (luz, alquiler,
        servicios), y el flujo de "solo el total" existe para registrar el crédito fiscal sin
        inventariar nada. Quien necesita kardex, detalla; el resto sigue como siempre.
        Solo las líneas con producto pueden mover stock (ver _l10n_pe_ne_lineas_con_stock)."""
        # Afectación del documento (cat-05): 1000 gravado por defecto — es la compra normal.
        # Va a nivel documento y no por línea: una factura de compra suele ser toda gravada o
        # toda no gravada (un recibo de servicios, un RH). El caso mixto necesita afectación
        # por línea y es otra iteración; hoy no hay dato para adivinarlo.
        tax = self._l10n_pe_ne_tax_compra_by_code(compra.get("afectacion"))
        tax_ids = [(6, 0, tax.ids if tax else [])]
        lineas = compra.get("lineas") or []
        if not lineas:
            total = float(compra.get("total") or 0)
            return [
                (0, 0, {
                    "name": compra.get("descripcion") or "COMPRA",
                    "quantity": 1,
                    # El total va CON IGV: se guarda la base y Odoo repone el impuesto, así
                    # el Registro de Compras tiene base e IGV separados (antes iba sin
                    # impuesto y el crédito fiscal no existía).
                    "price_unit": self._l10n_pe_ne_base_sin_igv(total, tax),
                    "tax_ids": tax_ids,
                })
            ]
        out = []
        suma = 0.0
        for ln in lineas:
            # create=False: una compra NO da de alta productos en el catálogo. El proveedor
            # los llama a su manera y crearlos aquí llenaría el catálogo de duplicados; se
            # elige uno existente desde la UI. Sin producto, la línea es solo un importe.
            prod = self._l10n_pe_ne_quick_product(ln, create=False)
            cant = float(ln.get("cantidad") or 0)
            if cant <= 0:
                raise UserError(_("Cada línea de la compra necesita una cantidad mayor a 0."))
            costo = float(ln.get("precioUnitario") or 0)
            if costo < 0:
                raise UserError(_("El costo de una línea no puede ser negativo."))
            suma += cant * costo
            # El costo de compra se guarda SIEMPRE: es un hecho del documento, no una
            # opinión — es lo que se pagó. El precio de VENTA solo se toca si el usuario lo
            # pidió (actualizarPrecio), porque cambiarlo solo movería la etiqueta de la
            # vitrina sin que nadie se entere.
            if prod and costo > 0:
                prod.sudo().standard_price = costo
                if ln.get("actualizarPrecio"):
                    prod.sudo().list_price = self._l10n_pe_ne_precio_con_margen(
                        costo, prod.l10n_pe_ne_margen or None
                    )
            vals = {
                "name": ln.get("descripcion") or (prod.name if prod else "ITEM"),
                "quantity": cant,
                # `costo` viene CON IGV (la convención de la app y lo que trae el XML del
                # proveedor en AlternativeConditionPrice); se guarda la base y Odoo repone
                # el impuesto. La suma para el cuadre sigue siendo sobre el bruto.
                "price_unit": self._l10n_pe_ne_base_sin_igv(costo, tax),
                "tax_ids": tax_ids,
                # El lote entra con la mercadería: viaja con la línea hasta el movimiento.
                "l10n_pe_ne_lote": (ln.get("lote") or "").strip() or False,
                "l10n_pe_ne_vence": ln.get("vence") or False,
            }
            if prod:
                vals["product_id"] = prod.id
            out.append((0, 0, vals))
        # El detalle MANDA: la compra se registra por la suma de las líneas y el `total` del
        # payload queda ignorado. Si no cuadran, lo que entra al Registro de Compras no es lo
        # que el usuario cree — un error fiscal. Se corta acá y no solo en el front: el
        # backend es la autoridad y a /ne/api/compras puede llamar cualquiera.
        total = float(compra.get("total") or 0)
        if total and abs(suma - total) > 0.01:
            raise UserError(
                _(
                    "El detalle suma %(suma).2f y el total de la compra dice %(total).2f. "
                    "Deben coincidir."
                )
                % {"suma": suma, "total": total}
            )
        return out

    @api.model
    def l10n_pe_ne_create_compra(self, compra):
        """Registra una compra (factura de proveedor). payload: {proveedor:{numDoc,
        razonSocial,tipoDoc}, tipoComprobante(cat.10), serie, numero, fecha, total,
        descripcion, moneda, lineas?}. Sin `lineas` es el registro simple de siempre
        (línea = total); con `lineas` se detalla por producto y la mercadería ENTRA al stock."""
        compra = compra or {}
        prov = self._l10n_pe_ne_quick_partner(compra.get("proveedor") or {})
        if not prov.supplier_rank:
            prov.supplier_rank = 1
        journal = self.env["account.journal"].search(
            [("type", "=", "purchase"), ("company_id", "=", self.env.company.id)],
            limit=1,
        )
        if not journal:
            raise UserError(_("No hay diario de compras configurado para la compañía."))
        serie = (compra.get("serie") or "").strip()
        numero = (compra.get("numero") or "").strip()
        doc_num = ("%s-%s" % (serie, numero)) if serie and numero else (numero or serie)
        if not doc_num:
            raise UserError(_("Indica el número del documento del proveedor."))
        total = float(compra.get("total") or 0)
        if total <= 0:
            raise UserError(_("Indica el monto total de la compra."))
        vals = {
            "move_type": "in_invoice",
            "partner_id": prov.id,
            "journal_id": journal.id,
            "invoice_date": compra.get("fecha") or fields.Date.context_today(self),
            "ref": doc_num,
            "l10n_latam_document_number": doc_num,
            "invoice_line_ids": self._l10n_pe_ne_compra_lineas(compra),
        }
        moneda = self._l10n_pe_ne_quick_currency(compra.get("moneda"))
        if moneda:
            vals["currency_id"] = moneda.id
        dt = self.env["l10n_latam.document.type"].search(
            [
                ("code", "=", compra.get("tipoComprobante") or "01"),
                ("country_id.code", "=", "PE"),
            ],
            limit=1,
        )
        if dt:
            vals["l10n_latam_document_type_id"] = dt.id
        move = self.create(vals)
        move.action_post()
        # La otra mitad del kardex: la mercadería detallada ENTRA al stock. Sin líneas con
        # producto no mueve nada, así que la compra "solo total" de siempre no cambia.
        move._l10n_pe_ne_mover_stock_compra()
        return move._l10n_pe_ne_compra_dict()

    @api.model
    def l10n_pe_ne_update_compra(self, rec_id, compra):
        """Actualiza una compra existente: la pasa a borrador, reescribe cabecera y
        la línea única, y la vuelve a postear. Mismas validaciones que el alta."""
        m = self.browse(int(rec_id or 0)).exists()
        if not m or m.move_type != "in_invoice":
            raise UserError(_("Compra no encontrada."))
        compra = compra or {}
        prov = self._l10n_pe_ne_quick_partner(compra.get("proveedor") or {})
        if not prov.supplier_rank:
            prov.supplier_rank = 1
        serie = (compra.get("serie") or "").strip()
        numero = (compra.get("numero") or "").strip()
        doc_num = ("%s-%s" % (serie, numero)) if serie and numero else (numero or serie)
        if not doc_num:
            raise UserError(_("Indica el número del documento del proveedor."))
        total = float(compra.get("total") or 0)
        if total <= 0:
            raise UserError(_("Indica el monto total de la compra."))
        if m.state == "posted":
            m.button_draft()
        vals = {
            "partner_id": prov.id,
            "invoice_date": compra.get("fecha") or m.invoice_date,
            "ref": doc_num,
            "l10n_latam_document_number": doc_num,
            "invoice_line_ids": [
                (5, 0, 0),
                (
                    0,
                    0,
                    {
                        "name": compra.get("descripcion") or "COMPRA",
                        "quantity": 1,
                        "price_unit": total,
                        "tax_ids": [(6, 0, [])],
                    },
                ),
            ],
        }
        moneda = self._l10n_pe_ne_quick_currency(compra.get("moneda"))
        if moneda:
            vals["currency_id"] = moneda.id
        dt = self.env["l10n_latam.document.type"].search(
            [
                ("code", "=", compra.get("tipoComprobante") or "01"),
                ("country_id.code", "=", "PE"),
            ],
            limit=1,
        )
        if dt:
            vals["l10n_latam_document_type_id"] = dt.id
        m.write(vals)
        m.action_post()
        return m._l10n_pe_ne_compra_dict()

    @api.model
    def l10n_pe_ne_delete_compra(self, rec_id):
        """Elimina la compra; si está posteada, la pasa a borrador y elimina; si no
        se puede, la anula (cancel)."""
        m = self.browse(int(rec_id or 0)).exists()
        if not m or m.move_type != "in_invoice":
            return {"ok": True, "modo": "inexistente"}
        try:
            if m.state == "posted":
                m.button_draft()
            m.unlink()
            return {"ok": True, "modo": "eliminado"}
        except Exception:
            m.button_cancel()
            return {"ok": True, "modo": "anulado"}

    def _l10n_pe_ne_quick_flags(self, move, payload):
        d = payload.get("detraccion")
        if d:
            move.l10n_pe_ne_detraccion = True
            move.l10n_pe_ne_detraccion_code = d.get("codBien") or "037"
            move.l10n_pe_ne_detraccion_rate = float(d.get("tasa") or 12)
            if d.get("medioPago"):
                move.l10n_pe_ne_detraccion_medio_pago = d["medioPago"]
            if d.get("cuentaBN"):
                # La cuenta se guarda EN el comprobante (lo tecleado siempre gana y sale
                # tal cual en su PDF/XML). Además, si la empresa aún no tiene cuenta de
                # detracción por defecto, se fija con la primera para futuras emisiones.
                move.l10n_pe_ne_detraccion_cuenta = d["cuentaBN"]
                if not move.company_id.l10n_pe_ne_cuenta_detraccion:
                    move.company_id.sudo().l10n_pe_ne_cuenta_detraccion = d["cuentaBN"]
        p = payload.get("percepcion")
        if p:
            move.l10n_pe_ne_percepcion = True
            move.l10n_pe_ne_percepcion_rate = float(p.get("tasa") or 2)
        if payload.get("esAnticipo"):
            move.l10n_pe_ne_es_anticipo = True
        # Descuento que NO afecta la base del IGV: el por ítem (descNoAfecta de cada línea) + el global
        # (descuentoGlobalNoAfecta) se agregan en un solo importe. NO reduce gravada/IGV: el emisor lo
        # aplica como AllowanceCharge global que solo baja el total (ver _l10n_pe_desc_no_afecta).
        desc_no_afecta = round(
            sum(float(ln.get("descNoAfecta") or 0) for ln in (payload.get("lineas") or []))
            + float(payload.get("descuentoGlobalNoAfecta") or 0),
            2,
        )
        if desc_no_afecta > 0:
            move.l10n_pe_ne_desc_no_afecta = desc_no_afecta
        # Anticipos regularizados: lista JSON (varios anticipos / pagos escalonados). Retrocompat
        # con el payload viejo de un solo anticipo (objeto): se envuelve en lista de 1.
        anticipos = payload.get("anticipos")
        if anticipos is None and payload.get("anticipo"):
            anticipos = [payload["anticipo"]]
        if anticipos:
            move.l10n_pe_ne_anticipos = [
                {
                    "doc": a.get("doc"),
                    "monto": float(a.get("monto") or a.get("total") or 0),
                    "tipo": a.get("tipo") or "02",
                    # Enlace al anticipo local (doc. A) elegido en el autocompletado, para
                    # llevar su saldo.
                    "origenId": a.get("origenId"),
                }
                for a in anticipos
            ]
        # Forma de pago: Crédito (con cuotas) emite cac:PaymentTerms; medios de pago
        # (efectivo/Yape/…) se guardan como dato interno del POS (no van al XML SUNAT).
        # Establecimiento emisor (sucursal): código de local anexo SUNAT del comprobante.
        if payload.get("codEstablecimiento"):
            move.l10n_pe_ne_cod_establecimiento = payload["codEstablecimiento"]
        # Guía de remisión referenciada (DespatchDocumentReference).
        if payload.get("guiaRef"):
            move.l10n_pe_ne_guia_ref = payload["guiaRef"]
            if payload.get("guiaTipo"):
                move.l10n_pe_ne_guia_tipo = payload["guiaTipo"]
        # Proyecto/contrato (avance de obra).
        if payload.get("proyectoId"):
            move.l10n_pe_ne_proyecto_id = int(payload["proyectoId"])
        fp = payload.get("formaPago") or {}
        if fp.get("tipo") == "Credito" or fp.get("cuotas"):
            move.l10n_pe_ne_forma_pago = "Credito"
            move.l10n_pe_ne_cuotas = fp.get("cuotas") or []
            # Forma de pago mixta: inicial al contado; el saldo a crédito lo llevan las cuotas.
            if fp.get("inicial"):
                move.l10n_pe_ne_inicial_contado = float(fp["inicial"])
            venc = (fp.get("cuotas") or [{}])[-1].get("fecha")
            if venc:
                move.invoice_date_due = venc
        if fp.get("medios"):
            move.l10n_pe_ne_medios_pago = fp.get("medios")
        # Redondeo de efectivo (dato de caja, no del XML): el POS lo calcula en vivo (≤ 0). Se
        # persiste solo si el pago es efectivo y el flag de la compañía está activo; ausente/0 = sin
        # redondeo. El importe entregado en efectivo = amount_total + redondeo.
        red = payload.get("redondeo")
        if red and move.company_id.l10n_pe_ne_redondeo_activo and self._l10n_pe_ne_solo_efectivo(fp.get("medios")):
            move.l10n_pe_ne_redondeo = float(red)

    @staticmethod
    def _l10n_pe_ne_solo_efectivo(medios):
        """¿el pago es 100% efectivo? Sin medios detallados el POS asume efectivo (True). Un solo
        medio no-efectivo o mezcla desactiva el redondeo (espeja lib/redondeo.ts:esSoloEfectivo)."""
        con_monto = [m for m in (medios or []) if float(m.get("monto") or 0) > 0]
        if not con_monto:
            return True
        return all((m.get("medio") or "Efectivo").strip() == "Efectivo" for m in con_monto)

    def l10n_pe_ne_quick_result(self):
        self.ensure_one()
        serie, corr = self._l10n_pe_serie_correlativo()
        m = re.search(r"ResponseCode (\d+)", self.l10n_pe_biller_message or "")
        return {
            "id": self.id,
            "tipoDoc": self.l10n_pe_ne_tipo_doc or self._l10n_pe_document_type(),
            "serie": self.l10n_pe_ne_serie_emit or serie,
            "correlativo": (self.l10n_pe_ne_corr_emit or corr).zfill(8),
            "estado": self.l10n_pe_biller_state,
            "responseCode": m.group(1) if m else "",
            "mensaje": self.l10n_pe_biller_message or "",
            "total": self.amount_total,
            "cliente": self.partner_id.name or "",
            "fechaEmision": self.invoice_date.strftime("%Y-%m-%d")
            if self.invoice_date
            else "",
        }

    @api.model
    def l10n_pe_ne_quick_list(self, query=None, desde=None, hasta=None, estado=None, tipo=None,
                              forma_pago=None, monto_min=None, monto_max=None, serie=None,
                              moneda=None, limit=100, offset=None):
        """Lista de comprobantes emitidos (sin los blobs), para la UI. Filtros
        opcionales: query (cliente/RUC/correlativo), rango de fechas (desde/hasta),
        estado del facturador (por_enviar/en_proceso/enviado/anulado/rechazado/error),
        tipo de comprobante (01/03/07/08), forma de pago (Contado/Credito) y rango de
        monto total (monto_min/monto_max). `estado` y `tipo` aceptan varios valores
        (lista o CSV "a,b") → filtran con `in` (multiselect en la UI).

        Paginación opt-in: con `offset` devuelve {items, total} (total vía
        search_count sobre el mismo dominio); sin él, la lista plana de siempre."""
        def _as_list(v):
            if not v:
                return None
            vals = [x for x in v.split(",") if x] if isinstance(v, str) else list(v)
            return vals or None

        def _num(v):
            try:
                return float(v) if v not in (None, "") else None
            except (TypeError, ValueError):
                return None
        estados = _as_list(estado)
        tipos = _as_list(tipo)
        series = _as_list(serie)
        mmin, mmax = _num(monto_min), _num(monto_max)
        # Se incluyen los 'por_enviar' (pendientes de envío) para que sean visibles y
        # reenviables desde la UI; antes se excluían y quedaban sin dónde verse.
        domain = [("l10n_pe_biller_state", "!=", False)]
        if estados:
            domain.append(("l10n_pe_biller_state", "in", estados))
        if tipos:
            domain.append(("l10n_pe_ne_tipo_doc", "in", tipos))
        if series:
            domain.append(("l10n_pe_ne_serie_emit", "in", series))
        if moneda:
            domain.append(("currency_id.name", "=", moneda))
        if forma_pago:
            domain.append(("l10n_pe_ne_forma_pago", "=", forma_pago))
        if mmin is not None:
            domain.append(("amount_total", ">=", mmin))
        if mmax is not None:
            domain.append(("amount_total", "<=", mmax))
        if desde:
            domain.append(("invoice_date", ">=", desde))
        if hasta:
            domain.append(("invoice_date", "<=", hasta))
        if query:
            q = query.strip()
            domain += [
                "|",
                "|",
                ("partner_id.name", "ilike", q),
                ("partner_id.vat", "ilike", q),
                ("l10n_pe_ne_corr_emit", "ilike", q),
            ]
        moves = self.search(domain, order="id desc", limit=limit, offset=offset or 0)
        # NC vigentes por comprobante en UNA consulta agrupada: la lista marca las
        # facturas/boletas acreditadas ("tiene NC") sin una búsqueda por fila. Mismo
        # criterio de "vigente" que _l10n_pe_ne_nc_previas (las en cola cuentan).
        nc_por_doc = {}
        if moves:
            grupos = self.env["account.move"]._read_group(
                [
                    ("move_type", "=", "out_refund"),
                    ("reversed_entry_id", "in", moves.ids),
                    ("state", "=", "posted"),
                    ("l10n_pe_biller_state", "not in", ("rechazado", "error", "anulado")),
                ],
                groupby=["reversed_entry_id"],
                aggregates=["__count", "amount_total:sum"],
            )
            nc_por_doc = {rev.id: (count, total or 0.0) for rev, count, total in grupos}
        items = [
            {
                "id": m.id,
                "tipoDoc": m.l10n_pe_ne_tipo_doc or m._l10n_pe_document_type(),
                "serie": m.l10n_pe_ne_serie_emit or m.l10n_pe_serie or "",
                "correlativo": m.l10n_pe_ne_corr_emit or "",
                "estado": m.l10n_pe_biller_state,
                "total": m.amount_total,
                "moneda": m.currency_id.name or "PEN",
                "cliente": m.partner_id.name or "",
                "fechaEmision": m.invoice_date.strftime("%Y-%m-%d")
                if m.invoice_date
                else "",
                # Hora de creación del comprobante (≈ emisión), en tz local (Lima).
                "hora": fields.Datetime.context_timestamp(m, m.create_date).strftime("%H:%M")
                if m.create_date
                else "",
                "mensaje": m.l10n_pe_biller_message or "",
                # Notas de crédito vigentes que afectan este comprobante (0 si no tiene).
                "ncCount": nc_por_doc.get(m.id, (0, 0.0))[0],
                "ncTotal": round(nc_por_doc.get(m.id, (0, 0.0))[1], 2),
            }
            for m in moves
        ]
        if offset is None:
            return items
        return {"items": items, "total": self.search_count(domain)}

    def _l10n_pe_ne_nc_previas(self):
        """Notas de crédito VIGENTES que afectan este comprobante: posteadas y no
        rechazadas/anuladas/con error. Las que siguen en cola (por_enviar/en_proceso)
        también cuentan, para que dos NC simultáneas no acrediten más que el total."""
        self.ensure_one()
        return self.env["account.move"].search(
            [
                ("move_type", "=", "out_refund"),
                ("reversed_entry_id", "=", self.id),
                ("state", "=", "posted"),
                ("l10n_pe_biller_state", "not in", ("rechazado", "error", "anulado")),
            ]
        )

    def l10n_pe_ne_comprobante_detalle(self):
        """Detalle completo de un comprobante para la vista de detalle (cabecera +
        líneas + totales por afectación + estado SUNAT). Todo calculado en Odoo."""
        self.ensure_one()
        b = self._l10n_pe_ne_ple_breakdown()
        serie, num = self._l10n_pe_ne_doc_id()
        lineas = []
        for ln in self.invoice_line_ids:
            codes = ln.tax_ids.mapped("l10n_pe_edi_tax_code")
            afect = next(
                (c for c in ("9997", "9998", "9995", "9996") if c in codes), "1000"
            )
            lineas.append(
                {
                    "descripcion": ln.name or "",
                    "cantidad": ln.quantity or 0.0,
                    "precio": ln.price_unit or 0.0,
                    "descuento": ln.discount or 0.0,
                    "afectacion": afect,
                    "unidad": self._l10n_pe_unit_code(ln),
                    "subtotal": ln.price_subtotal or 0.0,
                    # El front lo usa para conservar el producto real al espejar una NC
                    # o al refacturar (post-NC motivo 02).
                    "productId": ln.product_id.id or None,
                }
            )
        of, ot, os_, on = self._l10n_pe_ne_ple_origen()
        # NC previas vigentes (solo aplica a facturas/boletas): el front muestra el saldo
        # pendiente de acreditar y las notas asociadas al elegir el comprobante a afectar.
        ncs = (
            self._l10n_pe_ne_nc_previas()
            if self.move_type == "out_invoice"
            else self.env["account.move"]
        )
        return {
            "id": self.id,
            "tipoDoc": self.l10n_pe_ne_tipo_doc or self._l10n_pe_document_type(),
            "serie": serie,
            "correlativo": num,
            "fecha": self.invoice_date.strftime("%Y-%m-%d")
            if self.invoice_date
            else "",
            "cliente": self.partner_id.name or "",
            "clienteDoc": self.partner_id.vat or "",
            "moneda": self.currency_id.name or "PEN",
            "estado": self.l10n_pe_biller_state or "",
            "mensaje": self.l10n_pe_biller_message or "",
            "formaPago": self.l10n_pe_ne_forma_pago or "Contado",
            "docOrigen": ("%s %s-%s" % (ot, os_, on)) if on else "",
            "anticipos": self._l10n_pe_ne_anticipos_list(),
            "lineas": lineas,
            "notasCredito": [
                {
                    "id": m.id,
                    "numero": "%s-%s" % m._l10n_pe_ne_doc_id(),
                    "total": round(m.amount_total or 0.0, 2),
                    "estado": m.l10n_pe_biller_state or "",
                }
                for m in ncs
            ],
            "saldoAcreditable": round(
                (self.amount_total or 0.0) - sum(ncs.mapped("amount_total")), 2
            ),
            "totales": {
                "gravada": round(b["gravado"], 2),
                "exonerada": round(b["exonerado"], 2),
                "inafecta": round(b["inafecto"], 2),
                "igv": round(b["igv"], 2),
                "icbper": round(b["icbper"], 2),
                "total": round(b["total"], 2),
            },
        }

    def l10n_pe_ne_get_files(self, kind=None):
        """Devuelve {xml, cdr, pdf[, ticket]} en base64 del comprobante, para que el BFF los sirva
        (sin /web/content). `ticket` (80mm) solo se incluye cuando kind == 'ticket' — así una descarga
        normal no dispara el render del ticket."""
        self.ensure_one()

        def b64(att):
            v = att.datas
            return v.decode("ascii") if isinstance(v, (bytes, bytearray)) else (v or "")

        out = {}
        if self.l10n_pe_biller_xml:
            out["xml"] = b64(self.l10n_pe_biller_xml)
        if self.l10n_pe_biller_cdr:
            out["cdr"] = b64(self.l10n_pe_biller_cdr)
        # El PDF/ticket se renderiza contra el micro a partir del XML firmado (no del
        # CDR): está disponible apenas FIRMADO (en_proceso), sin esperar a SUNAT. Se
        # genera SOLO el formato pedido (un pedido de xml/cdr no debe disparar el micro).
        # Antes se generaba el PDF SIEMPRE y se tragaba cualquier fallo → el controller
        # devolvía un 404 opaco ("no tiene pdf") aunque el problema real fuera el micro
        # caído o un timeout. Ahora, si falla justo el formato pedido, se propaga el
        # motivo real (el controller lo traduce a un mensaje legible).
        want_pdf = kind in (None, "pdf", "ticket")
        if want_pdf and self.l10n_pe_biller_xml:
            es_ticket = kind == "ticket"
            try:
                att = self._l10n_pe_get_pdf_attachment(
                    formato="TICKET" if es_ticket else "A4"
                )
                if att:
                    out["ticket" if es_ticket else "pdf"] = b64(att)
            except Exception:
                # Solo se propaga cuando el cliente pedía EXACTAMENTE ese archivo; si el
                # kind era None (uso interno/tests) se degrada en silencio como antes.
                if kind in ("pdf", "ticket"):
                    raise
        return out

    # ------------------------------------------------- descargas / PDF (SFS 2.4)
    @staticmethod
    def _l10n_pe_download_url(attachment):
        """Acción de descarga directa del adjunto vía /web/content."""
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }

    def _l10n_pe_ne_medios_pago_texto(self):
        """Detalle de medios de pago del POS ('Efectivo S/ 50.00, Yape S/ 68.00') para la
        representación impresa. NO va al XML SUNAT (es interno del punto de venta). Devuelve
        "" si no hay medios con importe (medio sin detallar → no se muestra el bloque). Lo usan
        el ticket 80mm (dentro del bloque POS `adicionalTxt`) y el A4 (param `MEDIOS_PAGO`)."""
        self.ensure_one()
        medios = self.l10n_pe_ne_medios_pago or []

        def _txt(m):
            base = "%s S/ %.2f" % (m.get("medio") or "", float(m.get("monto") or 0))
            op = str(m.get("numOp") or "").strip()
            return "%s (Op. %s)" % (base, op) if op else base

        return ", ".join(_txt(m) for m in medios if float(m.get("monto") or 0) > 0)

    def _l10n_pe_ne_ticket_adicional(self):
        """Bloque de pago del ticket 80mm (se manda como `adicionalTxt`): medios de pago del
        POS, vuelto, cajero y nota. Estos datos NO van al XML SUNAT (son internos del punto de
        venta), pero sí a la representación impresa. Devuelve HTML simple (el textField usa
        markup html) o "" si no hay nada que mostrar."""
        self.ensure_one()
        partes = []
        medios = self.l10n_pe_ne_medios_pago or []
        det = self._l10n_pe_ne_medios_pago_texto()
        if det:
            partes.append("Pago: " + det)
            # Redondeo de efectivo (≤ 0): el comprobante mantiene amount_total, pero en efectivo se
            # cobra 'a pagar' = amount_total + redondeo. El vuelto se calcula contra ese importe.
            redondeo = self.l10n_pe_ne_redondeo or 0.0
            a_pagar = round((self.amount_total or 0.0) + redondeo, 2)
            if redondeo:
                partes.append("Redondeo: S/ %.2f" % redondeo)
                partes.append("A pagar efectivo: S/ %.2f" % a_pagar)
            pagado = sum(float(m.get("monto") or 0) for m in medios)
            vuelto = round(pagado - a_pagar, 2)
            if vuelto > 0:
                partes.append("Vuelto: S/ %.2f" % vuelto)
        if self.invoice_user_id:
            partes.append("Atendido por: " + (self.invoice_user_id.name or ""))
        nota = re.sub("<[^>]+>", " ", self.narration or "").strip()
        if nota:
            partes.append("Nota: " + nota)
        # El micro (sanitizarAdicional) escapa el HTML y traduce '\n' -> <br/>; se envía texto plano.
        return "\n".join(partes)

    def _l10n_pe_get_pdf_attachment(self, formato="A4"):
        """Devuelve (o genera y cachea) el PDF de la representación impresa pidiéndolo al micro
        (POST /report/pdf con el XML firmado). El micro lo renderiza con las plantillas del SFS 2.4.
        formato: 'A4' (SFS 2.4) o 'TICKET' (80mm, solo 01/03; otros tipos caen al A4)."""
        self.ensure_one()
        tipo, serie, correlativo = self._l10n_pe_baja_identidad()
        es_ticket = formato == "TICKET" and tipo in ("01", "03")
        if formato == "TICKET" and not es_ticket:
            return self._l10n_pe_get_pdf_attachment()  # fallback A4 (NC/ND/retención…)
        cache_field = "l10n_pe_biller_pdf_ticket" if es_ticket else "l10n_pe_biller_pdf"
        # Cache-busting: el PDF cacheado se etiqueta con la versión del template
        # (config `pdf_ver`). Si esa versión cambió (mejora del template) o el PDF
        # viejo no la trae, se descarta y se regenera → nadie ve representaciones
        # desactualizadas. Para forzar regeneración masiva, subir el parámetro.
        pdf_ver = self._l10n_pe_pdf_ver()
        cached = self[cache_field]
        if cached:
            if cached.description == pdf_ver:
                return cached
            cached.sudo().unlink()  # template cambió → descartar el PDF viejo
        if not self.l10n_pe_biller_xml:
            raise UserError(
                _("El comprobante no tiene XML firmado; envíelo primero a SUNAT.")
            )
        icp = self.env["ir.config_parameter"].sudo()
        base = icp.get_param("l10n_pe_ne_biller.url", "http://localhost:8090").rstrip(
            "/"
        )
        # Clave propia: reusar l10n_pe_ne_biller.timeout hacía que subir el
        # timeout de emisión (240s) arrastrara también la espera de un PDF.
        timeout = int(icp.get_param("l10n_pe_ne_biller.pdf_timeout", "60"))
        payload = {
            "ruc": self.company_id.vat or "",
            "tipoDoc": tipo,
            "xml": (self.l10n_pe_biller_xml.raw or b"").decode("utf-8"),
            # Serie-correlativo AUTORITATIVO desde Odoo (no se depende del xpath /Invoice/ID
            # de la plantilla, que en algún entorno no resolvía y dejaba el número en blanco).
            "numComprobante": "%s-%s" % (serie, (correlativo or "").zfill(8)),
            # Dirección del adquiriente para la representación impresa (no va al XML SUNAT: el
            # biller no la incluye en el bloque del cliente). Toma calle + urbanización si hay.
            "dirCliente": ", ".join(
                p for p in (self.partner_id.street, self.partner_id.street2) if p
            ),
            # Vendedor/cajero que atendió (no va al XML SUNAT): va en ambos formatos como
            # "Atendido por" (el ticket ya lo traía en el bloque POS; ahora también el A4).
            "atendidoPor": self.invoice_user_id.name or "",
            # Medios de pago del POS (Efectivo/Yape/Plin…): NO van al XML SUNAT. El A4 los
            # muestra junto a la forma de pago (param MEDIOS_PAGO); el ticket ya los trae en
            # el bloque POS (adicionalTxt). "" si no hay medios detallados.
            "mediosPago": self._l10n_pe_ne_medios_pago_texto(),
        }
        # Logo del emisor (si lo tiene): va en ambos formatos (A4 y ticket).
        logo = self.company_id.logo
        if logo:
            payload["logo"] = logo.decode() if isinstance(logo, bytes) else logo
        if es_ticket:
            payload["formato"] = "TICKET"
            # Bloque de pago (medios/vuelto/cajero/nota) — solo en el ticket 80mm.
            adic = self._l10n_pe_ne_ticket_adicional()
            if adic:
                payload["adicionalTxt"] = adic
            # Contacto del emisor (no va al XML SUNAT): teléfono y correo de la compañía.
            contacto = "   ·   ".join(
                p for p in (
                    ("Tel: " + self.company_id.phone) if self.company_id.phone else "",
                    self.company_id.email or "",
                ) if p
            )
            if contacto:
                payload["contactoEmisor"] = contacto
        headers = {"X-Api-Key": self.company_id.sudo().l10n_pe_ne_api_key or ""}
        try:
            resp = requests.post(
                base + "/report/pdf",
                json=payload,
                headers=headers,
                timeout=(5, timeout),
            )
        except requests.RequestException as exc:
            raise UserError(_("Error de conexión con el facturador: %s") % exc)
        if resp.status_code != 200 or not resp.content.startswith(b"%PDF"):
            raise UserError(
                _("El facturador no devolvió un PDF (HTTP %s): %s")
                % (resp.status_code, (resp.text or "")[:500])
            )
        att = self.env["ir.attachment"].create(
            {
                "name": "%s-%s-%s%s.pdf"
                % (
                    self.company_id.vat or "",
                    serie,
                    correlativo.zfill(8),
                    "-ticket" if es_ticket else "",
                ),
                "res_model": "account.move",
                "res_id": self.id,
                "mimetype": "application/pdf",
                "raw": resp.content,
                "description": pdf_ver,   # etiqueta de versión para el cache-busting
            }
        )
        self[cache_field] = att.id
        return att

    def _l10n_pe_ne_is_aceptado(self):
        """True solo si el comprobante fue aceptado por SUNAT: estado 'enviado',
        con CDR guardado y ResponseCode 0. Re-parsea el CDR (no confía en el texto
        de l10n_pe_biller_message)."""
        self.ensure_one()
        if self.l10n_pe_biller_state != "enviado" or not self.l10n_pe_biller_cdr:
            return False
        code, _desc = self._l10n_pe_parse_cdr_codes(self.l10n_pe_biller_cdr.raw or b"")
        return code == "0"

    def l10n_pe_ne_email_comprobante(self, to=None, cc=None):
        """Envía el comprobante aceptado al cliente por correo, adjuntando el PDF
        (representación impresa SFS) y el XML firmado, vía la plantilla
        l10n_pe_ne_biller.mail_template_comprobante."""
        self.ensure_one()
        if not self._l10n_pe_ne_is_aceptado():
            raise UserError(
                _("El comprobante no está aceptado por SUNAT; no se puede enviar.")
            )
        to = (to or self.partner_id.email or "").strip()
        if not to:
            raise UserError(
                _("El cliente no tiene correo configurado; indica un destinatario.")
            )
        pdf = self._l10n_pe_get_pdf_attachment()
        xml = self.l10n_pe_biller_xml
        template = self.env.ref("l10n_pe_ne_biller.mail_template_comprobante")
        template.send_mail(
            self.id,
            force_send=True,
            raise_exception=True,
            email_values={
                "email_to": to,
                "email_cc": cc or "",
                "attachment_ids": [(6, 0, [pdf.id, xml.id])],
            },
        )
        self.message_post(body=_("Comprobante enviado por correo a %s") % to)
        return {"ok": True, "to": to}

    def action_l10n_pe_download_pdf(self):
        self.ensure_one()
        return self._l10n_pe_download_url(self._l10n_pe_get_pdf_attachment())

    def action_l10n_pe_download_ticket(self):
        self.ensure_one()
        return self._l10n_pe_download_url(self._l10n_pe_get_pdf_attachment(formato="TICKET"))

    def action_l10n_pe_download_xml(self):
        self.ensure_one()
        if not self.l10n_pe_biller_xml:
            raise UserError(_("El comprobante no tiene XML firmado."))
        return self._l10n_pe_download_url(self.l10n_pe_biller_xml)

    def action_l10n_pe_download_cdr(self):
        self.ensure_one()
        if not self.l10n_pe_biller_cdr:
            raise UserError(_("El comprobante no tiene CDR de SUNAT."))
        return self._l10n_pe_download_url(self.l10n_pe_biller_cdr)

    def action_l10n_pe_download_zip(self):
        """Empaqueta en un ZIP el XML firmado, el CDR y el PDF de los comprobantes seleccionados.
        El PDF se genera al vuelo (best-effort) si aún no existe."""
        buf = io.BytesIO()
        incluidos = 0
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for move in self:
                if move.l10n_pe_biller_xml:
                    zf.writestr(
                        move.l10n_pe_biller_xml.name, move.l10n_pe_biller_xml.raw or b""
                    )
                    incluidos += 1
                if move.l10n_pe_biller_cdr:
                    zf.writestr(
                        move.l10n_pe_biller_cdr.name, move.l10n_pe_biller_cdr.raw or b""
                    )
                    incluidos += 1
                try:
                    pdf = (
                        move._l10n_pe_get_pdf_attachment()
                        if move.l10n_pe_biller_xml
                        else False
                    )
                    if pdf:
                        zf.writestr(pdf.name, pdf.raw or b"")
                        incluidos += 1
                except UserError as exc:
                    # PDF best-effort: si el micro falla, el ZIP igual lleva XML + CDR; se registra el motivo.
                    _logger.warning(
                        "No se pudo generar el PDF de %s para el ZIP: %s",
                        move.name,
                        exc,
                    )
        if not incluidos:
            raise UserError(
                _("Los comprobantes seleccionados no tienen archivos para descargar.")
            )
        att = self.env["ir.attachment"].create(
            {
                "name": "comprobantes_sunat.zip",
                "mimetype": "application/zip",
                "raw": buf.getvalue(),
            }
        )
        return self._l10n_pe_download_url(att)

    # ---------------------------------------------- comunicación de baja (RA)
    @api.depends(
        "l10n_pe_ne_baja_fecha", "l10n_pe_ne_baja_correlativo", "l10n_pe_ne_tipo_doc"
    )
    def _compute_l10n_pe_ne_baja_doc(self):
        for m in self:
            if m.l10n_pe_ne_baja_fecha and m.l10n_pe_ne_baja_correlativo:
                # Boletas (03) se anulan por Resumen Diario (RC); el resto por Comunicación de Baja (RA).
                prefijo = "RC" if m.l10n_pe_ne_tipo_doc == "03" else "RA"
                m.l10n_pe_ne_baja_doc = "%s-%s-%s" % (
                    prefijo,
                    m.l10n_pe_ne_baja_fecha.strftime("%Y%m%d"),
                    m.l10n_pe_ne_baja_correlativo,
                )
            else:
                m.l10n_pe_ne_baja_doc = False

    def _l10n_pe_baja_identidad(self):
        """(tipo, serie, correlativo) realmente emitidos. Usa lo congelado al enviar; si falta (comprobante
        enviado por una versión previa), recae en el cálculo del momento."""
        self.ensure_one()
        tipo = self.l10n_pe_ne_tipo_doc or self._l10n_pe_document_type()
        serie = self.l10n_pe_ne_serie_emit or self._l10n_pe_serie_correlativo()[0]
        correlativo = self.l10n_pe_ne_corr_emit or self._l10n_pe_serie_correlativo()[1]
        return tipo, serie, correlativo

    def _l10n_pe_check_baja(self):
        """Guardas de la comunicación de baja: solo factura/NC/ND ya enviadas, con serie válida, motivo,
        fecha y (para facturas) dentro del plazo de 7 días."""
        self.ensure_one()
        if self.l10n_pe_biller_state not in ("enviado", "anulado"):
            raise UserError(
                _(
                    "Solo puede comunicarse la baja de un comprobante ya enviado a SUNAT."
                )
            )
        tipo, serie, _corr = self._l10n_pe_baja_identidad()
        if tipo not in ("01", "03", "07", "08"):
            raise UserError(
                _(
                    "La anulación aplica a factura, boleta, nota de crédito y nota de débito."
                )
            )
        # Una factura/boleta con NC VIGENTES no se da de baja: la baja anula el documento
        # COMPLETO y las notas ya acreditaron parte (crédito duplicado), además de dejar
        # esas NC referenciando un comprobante dado de baja. Primero se anulan las NC,
        # o se acredita el saldo con otra NC en lugar de la baja.
        if tipo in ("01", "03"):
            ncs = self._l10n_pe_ne_nc_previas()
            if ncs:
                raise UserError(
                    _(
                        "No se puede anular %(doc)s: tiene %(n)d nota(s) de crédito "
                        "vigente(s) por %(monto)s (%(lista)s). Anularla duplicaría el "
                        "crédito — anule primero esas notas, o acredite el saldo con "
                        "una nota de crédito en lugar de la baja."
                    )
                    % {
                        "doc": "%s-%s" % (serie or "", (_corr or "").zfill(8)),
                        "n": len(ncs),
                        "monto": "%.2f" % sum(ncs.mapped("amount_total")),
                        "lista": ", ".join(
                            "%s-%s" % m._l10n_pe_ne_doc_id() for m in ncs
                        ),
                    }
                )
        # Serie con prefijo B (boleta) / F / S, o numérica: refleja el formato del comprobante emitido.
        if not re.match(r"^([BFS][A-Z0-9]{3}|\d{1,4})$", serie or ""):
            raise UserError(
                _(
                    "La serie del comprobante (%s) no tiene un formato válido para la anulación."
                )
                % serie
            )
        if not (self.l10n_pe_ne_baja_motivo or "").strip():
            raise UserError(_("Indique el motivo de la baja."))
        if not self.invoice_date:
            raise UserError(_("El comprobante no tiene fecha de emisión."))
        # Plazo de 7 días calendario para anular una FACTURA por baja (las NC/ND no tienen este límite).
        if tipo == "01":
            limite = int(
                self.env["ir.config_parameter"]
                .sudo()
                .get_param("l10n_pe_ne_biller.baja_plazo_dias", "7")
            )
            dias = (fields.Date.context_today(self) - self.invoice_date).days
            if dias > limite:
                raise UserError(
                    _(
                        "Fuera del plazo de baja: han pasado %s días desde la emisión de la factura "
                        "(máximo %s días calendario). Para anularla, emita una nota de crédito."
                    )
                    % (dias, limite)
                )
        # Boleta mayor a S/ 700 exige el documento de identidad del adquirente (igual que en su emisión).
        if (
            tipo == "03"
            and (self.amount_total or 0.0) > 700
            and not (self.partner_id.vat or "").strip()
        ):
            raise UserError(
                _(
                    "Una boleta mayor a S/ 700 requiere el documento de identidad del cliente para anularse "
                    "por resumen diario."
                )
            )

    def _l10n_pe_build_baja_request(self):
        """Comunicación de Baja (RA) — endpoint /generator/resumenBaja. Da de baja este comprobante."""
        self.ensure_one()
        tipo, serie, correlativo = self._l10n_pe_baja_identidad()
        return {
            "id": {
                "ruc": self.company_id.vat or "",
                # El DTO del facturador deserializa estas fechas con patrón yyyyMMdd (sin guiones).
                "fechaGeneracion": self.l10n_pe_ne_baja_fecha.strftime("%Y%m%d"),
                "correlativo": self.l10n_pe_ne_baja_correlativo or "1",
            },
            "emisor": self._l10n_pe_emisor(),
            # ReferenceDate = fecha de emisión del comprobante que se anula; IssueDate = fecha de la baja.
            "fecGeneracion": self.invoice_date.strftime("%Y%m%d"),
            "fecComunicacion": self.l10n_pe_ne_baja_fecha.strftime("%Y%m%d"),
            "resumenBajas": [
                {
                    "tipDocBaja": tipo,
                    "numDocBaja": "%s-%s" % (serie, correlativo.zfill(8)),
                    "desMotivoBaja": self.l10n_pe_ne_baja_motivo or "",
                }
            ],
        }

    # --- Resumen Diario de Boletas (RC): anula boletas (tipEstado 3) ---
    _RC_CATEGORIA = {
        "1000": "gravado",
        "1016": "gravado",
        "9997": "exonerado",
        "9998": "inafecto",
        "9995": "exportado",
        "9996": "gratuito",
    }

    def _l10n_pe_rc_totales(self):
        """Totales de valor por categoría (gravado/exonerado/inafecto/exportado/gratuito) para el RC."""
        self.ensure_one()
        cats = dict.fromkeys(
            ("gravado", "exonerado", "inafecto", "exportado", "gratuito"), 0.0
        )
        for line in self._l10n_pe_product_lines():
            (_tip, cod_tri, *_), _por = self._l10n_pe_tax_info(line)
            base, _igv, _isc, _icb = self._l10n_pe_line_amounts(line)
            cats[self._RC_CATEGORIA.get(cod_tri, "gravado")] += base
        return {k: round(v, 2) for k, v in cats.items()}

    def _l10n_pe_build_rc_request(self):
        """Resumen Diario de Boletas (RC) — endpoint /generator/resumenBoleta. Anula esta boleta (tipEstado 3)."""
        self.ensure_one()
        fmt = self._l10n_pe_fmt
        _tipo, serie, correlativo = self._l10n_pe_baja_identidad()
        partner = self.partner_id
        cats = self._l10n_pe_rc_totales()
        tributos = [
            {
                "idLineaRd": "1",
                "ideTributoRd": t["ideTributo"],
                "nomTributoRd": t["nomTributo"],
                "codTipTributoRd": t["codTipTributo"],
                "mtoBaseImponibleRd": t["mtoBaseImponible"],
                "mtoTributoRd": t["mtoTributo"],
            }
            for t in self._l10n_pe_tributos()
        ]
        # ICBPER (7152): NO se agrega aquí. _l10n_pe_tributos() ya lo incluye (regla 3279) y entra al
        # RC por el comprehension de arriba. Duplicarlo generaba un segundo cac:TaxTotal con el mismo
        # código de tributo → SUNAT rechazaba el RC (obs 2355: un solo TaxTotal por tributo/ítem). La
        # suma de componentes con totImpCpe (obs 4027) sigue cuadrando: el ICBPER está presente una vez.
        # El validador exige que CADA línea del RC tenga el tributo IGV '1000'. Si la boleta no es
        # gravada (exo/inafecto/exportación/gratuita/IVAP), se agrega uno en cero (regla 2278).
        if not any(t["ideTributoRd"] == "1000" for t in tributos):
            tributos.append(
                {
                    "idLineaRd": "1",
                    "ideTributoRd": "1000",
                    "nomTributoRd": "IGV",
                    "codTipTributoRd": "VAT",
                    "mtoBaseImponibleRd": "0.00",
                    "mtoTributoRd": "0.00",
                }
            )
        # Adquirente: consumidor final sin documento → tipo "0" y número "00000000" (catálogo SUNAT).
        vat = (partner.vat or "").strip()
        cod_doc = partner.l10n_latam_identification_type_id.l10n_pe_vat_code or ""
        if not vat:
            cod_doc, vat = "0", "00000000"
        elif not cod_doc:
            cod_doc = "6" if (len(vat) == 11 and vat.isdigit()) else "1"
        return {
            "id": {
                "ruc": self.company_id.vat or "",
                "fechaGeneracion": self.l10n_pe_ne_baja_fecha.strftime("%Y%m%d"),
                "correlativo": self.l10n_pe_ne_baja_correlativo or "1",
            },
            "emisor": self._l10n_pe_emisor(),
            "resumenDiario": [
                {
                    # ReferenceDate = emisión de la boleta; IssueDate = fecha del resumen (ISO en el XML).
                    "fecEmision": self.invoice_date.strftime("%Y-%m-%d"),
                    "fecResumen": self.l10n_pe_ne_baja_fecha.strftime("%Y-%m-%d"),
                    "tipDocResumen": "03",
                    "idDocResumen": "%s-%s" % (serie, correlativo.zfill(8)),
                    "tipDocUsuario": cod_doc,
                    "numDocUsuario": vat,
                    "tipMoneda": self.currency_id.name or "PEN",
                    "totValGrabado": fmt(cats["gravado"]),
                    "totValExoneado": fmt(cats["exonerado"]),
                    "totValInafecto": fmt(cats["inafecto"]),
                    "totValExportado": fmt(cats["exportado"]),
                    "totValGratuito": fmt(cats["gratuito"]),
                    "totOtroCargo": "0.00",
                    "totImpCpe": fmt(self.amount_total),
                    "tipDocModifico": "",
                    "serDocModifico": "",
                    "numDocModifico": "",
                    "tipRegPercepcion": "",
                    "porPercepcion": "",
                    "monBasePercepcion": "",
                    "monPercepcion": "",
                    "monTotIncPercepcion": "",
                    "tipEstado": "3",  # 3 = anulación/baja de la boleta
                    "tributosDocResumen": tributos,
                }
            ],
        }

    def _l10n_pe_rc_emision_item(self, id_linea):
        """Un ítem del Resumen Diario para EMISIÓN (tipEstado 1 = registrar la boleta ante SUNAT).
        Misma estructura que el de anulación pero con la identidad EMITIDA y estado 1."""
        self.ensure_one()
        fmt = self._l10n_pe_fmt
        serie = self.l10n_pe_ne_serie_emit or self._l10n_pe_serie_correlativo()[0]
        correlativo = self.l10n_pe_ne_corr_emit or self._l10n_pe_serie_correlativo()[1]
        partner = self.partner_id
        cats = self._l10n_pe_rc_totales()
        idl = str(id_linea)
        tributos = [
            {
                "idLineaRd": idl, "ideTributoRd": t["ideTributo"], "nomTributoRd": t["nomTributo"],
                "codTipTributoRd": t["codTipTributo"], "mtoBaseImponibleRd": t["mtoBaseImponible"],
                "mtoTributoRd": t["mtoTributo"],
            }
            for t in self._l10n_pe_tributos()
        ]
        # ICBPER (7152): ya viene de _l10n_pe_tributos() por el comprehension; no re-agregarlo o SUNAT
        # rechaza el RC con un TaxTotal duplicado (obs 2355).
        if not any(t["ideTributoRd"] == "1000" for t in tributos):
            tributos.append({"idLineaRd": idl, "ideTributoRd": "1000", "nomTributoRd": "IGV",
                             "codTipTributoRd": "VAT", "mtoBaseImponibleRd": "0.00", "mtoTributoRd": "0.00"})
        vat = (partner.vat or "").strip()
        cod_doc = partner.l10n_latam_identification_type_id.l10n_pe_vat_code or ""
        if not vat:
            cod_doc, vat = "0", "00000000"
        elif not cod_doc:
            cod_doc = "6" if (len(vat) == 11 and vat.isdigit()) else "1"
        return {
            "fecEmision": self.invoice_date.strftime("%Y-%m-%d"),
            "fecResumen": fields.Date.context_today(self).strftime("%Y-%m-%d"),
            "tipDocResumen": "03",
            "idDocResumen": "%s-%s" % (serie, (correlativo or "").zfill(8)),
            "tipDocUsuario": cod_doc, "numDocUsuario": vat,
            "tipMoneda": self.currency_id.name or "PEN",
            "totValGrabado": fmt(cats["gravado"]), "totValExoneado": fmt(cats["exonerado"]),
            "totValInafecto": fmt(cats["inafecto"]), "totValExportado": fmt(cats["exportado"]),
            "totValGratuito": fmt(cats["gratuito"]), "totOtroCargo": "0.00",
            "totImpCpe": fmt(self.amount_total),
            "tipDocModifico": "", "serDocModifico": "", "numDocModifico": "",
            "tipRegPercepcion": "", "porPercepcion": "", "monBasePercepcion": "",
            "monPercepcion": "", "monTotIncPercepcion": "",
            "tipEstado": "1",  # 1 = adicionar/registrar la boleta
            "tributosDocResumen": tributos,
        }

    def _l10n_pe_build_rc_emision(self, fecha_gen, correlativo):
        """RC de EMISIÓN para un CONJUNTO de boletas (self = recordset; misma compañía y fecha)."""
        first = self[0]
        return {
            "id": {
                "ruc": first.company_id.vat or "",
                "fechaGeneracion": fecha_gen.strftime("%Y%m%d"),
                "correlativo": str(correlativo),
            },
            "emisor": first._l10n_pe_emisor(),
            "resumenDiario": [b._l10n_pe_rc_emision_item(i + 1) for i, b in enumerate(self)],
        }

    @api.model
    def _l10n_pe_cron_resumen_boletas(self):
        """Boletas por Resumen Diario (RC, tipEstado 1) IDEMPOTENTE, en dos fases:
        A) ENVIAR: agrupa las boletas firmadas SIN ticket (por compañía+fecha), manda el RC vía
           /resumenBoleta/enviar (firma + sendSummary) y GUARDA el ticket en cada boleta. No las
           re-envía en la próxima corrida (ya tienen ticket) → no duplica el resumen en SUNAT.
        B) CONSULTAR: pollea los grupos que ya tienen ticket vía /ticket/estado; al llegar el CDR
           marca las boletas aceptado/rechazado y libera el ticket. Requiere instant + boletas_resumen."""
        icp = self.env["ir.config_parameter"].sudo()
        if icp.get_param("l10n_pe_ne_biller.boletas_resumen", "").strip().lower() not in ("1", "true"):
            return
        base = icp.get_param("l10n_pe_ne_biller.url", "http://localhost:8090").rstrip("/")
        timeout = int(icp.get_param("l10n_pe_ne_biller.resumen_timeout", "80"))
        STATUS_EN_PROCESO = 98

        def _bus(b):
            self.env["bus.bus"]._sendone(
                "l10n_pe_biller_updates", "l10n_pe_biller_update",
                {"move_id": b.id, "state": b.l10n_pe_biller_state})

        # ── FASE B — consultar los grupos que YA tienen ticket (idempotente: no re-envía) ──
        con_ticket = self.search(
            [("l10n_pe_biller_state", "=", "en_proceso"), ("l10n_pe_ne_rc_ticket", "!=", False)],
            limit=200,
        )
        por_ticket = {}
        for m in con_ticket:
            por_ticket.setdefault(m.l10n_pe_ne_rc_ticket, self.browse())
            por_ticket[m.l10n_pe_ne_rc_ticket] |= m
        for ticket, boletas in por_ticket.items():
            company = boletas[0].company_id
            headers = {"X-Api-Key": company.sudo().l10n_pe_ne_api_key or ""}
            body = {"ruc": company.vat or "", "ticket": ticket, "canal": "GEM"}
            try:
                resp = requests.post(base + "/generator/ticket/estado", json=body, headers=headers, timeout=(5, timeout))
            except Exception as e:  # noqa: BLE001 — red: reintenta con el MISMO ticket
                _logger.warning("ticket %s: %s (reintenta)", ticket, e)
                continue
            if resp.status_code != 200:
                continue  # transitorio: reintenta con el mismo ticket
            data = resp.json() or {}
            status = int(data.get("statusCode") or -1)
            cdr = data.get("cdr") or ""
            if status == STATUS_EN_PROCESO:
                continue  # SUNAT aún procesa el resumen: reintenta luego (mismo ticket)
            if cdr:
                code, desc = boletas[0]._l10n_pe_parse_cdr_codes(base64.b64decode(cdr))
                estado = "enviado" if code == "0" else "rechazado"
                for b in boletas:
                    b.l10n_pe_biller_state = estado
                    b.l10n_pe_biller_message = (
                        _("Aceptado por SUNAT vía Resumen Diario (RC corr %s). %s") % (b.l10n_pe_ne_rc_correlativo or "", desc or "")
                        if code == "0" else
                        _("Rechazado en el Resumen Diario (RC corr %s): ResponseCode %s. %s") % (b.l10n_pe_ne_rc_correlativo or "", code, desc or ""))[:2000]
                    b.l10n_pe_ne_rc_ticket = False
                    b.l10n_pe_ne_envi_zip = False
                    b._l10n_pe_store_cdr(cdr)
                    _bus(b)
            else:
                for b in boletas:
                    b.l10n_pe_biller_state = "rechazado"
                    b.l10n_pe_biller_message = _("Resumen Diario RC: SUNAT terminó con statusCode %s sin CDR.") % status
                    b.l10n_pe_ne_rc_ticket = False
                    _bus(b)
            self.env.cr.commit()

        # ── FASE A — enviar las boletas firmadas SIN ticket todavía (una llamada = un ticket) ──
        sin_ticket = self.search(
            [("l10n_pe_biller_state", "=", "en_proceso"), ("l10n_pe_ne_tipo_doc", "=", "03"),
             ("l10n_pe_ne_serie_emit", "!=", False), ("l10n_pe_ne_rc_ticket", "=", False)],
            limit=200,
        )
        grupos = {}
        for m in sin_ticket:
            grupos.setdefault((m.company_id.id, m.invoice_date), self.browse())
            grupos[(m.company_id.id, m.invoice_date)] |= m
        for (cid, fecha), boletas in grupos.items():
            company = boletas[0].company_id
            correlativo = self.env["ir.sequence"].next_by_code("l10n_pe.ne.rc") or "1"
            fecha_gen = fields.Date.context_today(boletas[0])
            payload = boletas._l10n_pe_build_rc_emision(fecha_gen, correlativo)
            headers = {"X-Api-Key": company.sudo().l10n_pe_ne_api_key or ""}
            try:
                resp = requests.post(base + "/generator/resumenBoleta/enviar", json=payload, headers=headers, timeout=(5, timeout))
            except Exception as e:  # noqa: BLE001 — no se envió: reintenta con un correlativo fresco
                _logger.warning("resumen boletas %s/%s: %s (reintenta)", cid, fecha, e)
                continue
            if resp.status_code == 200 and (resp.json() or {}).get("ticket"):
                ticket = resp.json()["ticket"]
                for b in boletas:
                    b.l10n_pe_ne_rc_ticket = ticket
                    b.l10n_pe_ne_rc_correlativo = str(correlativo)
                    b.l10n_pe_ne_rc_fecha = fecha_gen
                    b.l10n_pe_biller_message = _("Resumen Diario enviado (RC corr %s), ticket %s — esperando SUNAT.") % (correlativo, ticket)
            else:
                msg = ("Resumen RC HTTP %s: %s" % (resp.status_code, resp.text))[:1500]
                for b in boletas:
                    b.l10n_pe_biller_message = msg
            self.env.cr.commit()

    def _l10n_pe_store_baja_cdr(self, cdr_b64):
        """Guarda el CDR de la baja en un adjunto propio (no pisa el CDR original) y devuelve (code, desc)."""
        self.ensure_one()
        try:
            cdr_bytes = base64.b64decode(cdr_b64)
        except Exception:
            return "", ""
        att = self.env["ir.attachment"].create(
            {
                "name": "R%s-%s.zip"
                % (self.company_id.vat or "", self.l10n_pe_ne_baja_doc or "RA"),
                "res_model": "account.move",
                "res_id": self.id,
                "mimetype": "application/zip",
                "raw": cdr_bytes,
            }
        )
        self.l10n_pe_ne_baja_cdr = att.id
        return self._l10n_pe_parse_cdr_codes(cdr_bytes)

    def action_l10n_pe_send_baja(self):
        """Anula en SUNAT cada comprobante seleccionado: boletas por Resumen Diario (RC, tipEstado 3),
        facturas/NC/ND por Comunicación de Baja (RA)."""
        icp = self.env["ir.config_parameter"].sudo()
        base = icp.get_param("l10n_pe_ne_biller.url", "http://localhost:8090").rstrip(
            "/"
        )
        # En producción el biller está tras API Gateway (tope duro 30s): esperar
        # 120s solo alargaba el error. 40 = margen sobre el 504 del gateway; el
        # cron/reintento resuelve las bajas que SUNAT termina aceptando después.
        timeout = int(icp.get_param("l10n_pe_ne_biller.baja_timeout", "40"))
        for move in self:
            if move.l10n_pe_biller_state == "anulado":
                continue  # ya dado de baja: no reenviar el mismo RA (SUNAT lo rechaza por duplicado)
            move._l10n_pe_check_baja()
            es_boleta = move._l10n_pe_baja_identidad()[0] == "03"
            # Boletas → Resumen Diario (RC, tipEstado 3); el resto → Comunicación de Baja (RA).
            seq, endpoint, root = (
                ("l10n_pe.ne.rc", "/generator/resumenBoleta", "<SummaryDocuments")
                if es_boleta
                else ("l10n_pe.ne.ra", "/generator/resumenBaja", "<VoidedDocuments")
            )
            # Correlativo del resumen asignado una sola vez; un reintento limpio reusa el mismo.
            if not move.l10n_pe_ne_baja_correlativo:
                move.l10n_pe_ne_baja_correlativo = (
                    self.env["ir.sequence"].next_by_code(seq) or "1"
                )
                move.l10n_pe_ne_baja_fecha = fields.Date.context_today(move)
            ra = move.l10n_pe_ne_baja_doc
            payload = (
                move._l10n_pe_build_rc_request()
                if es_boleta
                else move._l10n_pe_build_baja_request()
            )
            headers = {"X-Api-Key": move.company_id.sudo().l10n_pe_ne_api_key or ""}
            try:
                resp = requests.post(
                    base + endpoint,
                    json=payload,
                    headers=headers,
                    timeout=(5, timeout),
                )
            except requests.RequestException as exc:
                # Nunca llegó a SUNAT: libera el resumen para reintentar con uno fresco.
                move._l10n_pe_baja_liberar()
                move.l10n_pe_biller_message = _(
                    "Error de conexión con el facturador (anulación %s): %s"
                ) % (ra, exc)
                continue
            if resp.status_code == 200 and root in resp.text:
                cdr_b64 = resp.headers.get("X-Sunat-Cdr")
                code, desc = (
                    move._l10n_pe_store_baja_cdr(cdr_b64) if cdr_b64 else ("", "")
                )
                if code == "0":
                    move.l10n_pe_biller_state = "anulado"
                    move.l10n_pe_biller_message = _(
                        "Anulación %s aceptada por SUNAT (ResponseCode 0). %s"
                    ) % (ra, desc or "")
                elif code:
                    # SUNAT lo recibió y rechazó: libera el resumen; el usuario corrige y reintenta con uno nuevo.
                    move._l10n_pe_baja_liberar()
                    move.l10n_pe_biller_message = _(
                        "Anulación %s rechazada por SUNAT (ResponseCode %s). %s"
                    ) % (ra, code, desc or "")
                else:
                    # 200 + XML firmado pero sin CDR legible: resultado indeterminado; no marcar anulado ni
                    # liberar el resumen (SUNAT pudo haberlo aceptado). El usuario verifica antes de reintentar.
                    move.l10n_pe_biller_message = (
                        _(
                            "Anulación %s enviada; respuesta de SUNAT indeterminada (sin CDR legible). Verifique en SUNAT."
                        )
                        % ra
                    )
            else:
                move._l10n_pe_baja_liberar()
                move.l10n_pe_biller_message = _(
                    "Anulación %s rechazada por el facturador: %s"
                ) % (ra, (resp.text or "")[:1200])
        return True

    def _l10n_pe_baja_liberar(self):
        """Libera el correlativo/fecha del resumen tras un fallo, para un reintento limpio."""
        self.l10n_pe_ne_baja_correlativo = False
        self.l10n_pe_ne_baja_fecha = False
