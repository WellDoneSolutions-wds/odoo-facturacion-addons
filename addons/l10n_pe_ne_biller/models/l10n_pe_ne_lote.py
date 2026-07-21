# -*- coding: utf-8 -*-
"""Emisión masiva (NE Express) — carga de ventas desde Excel → N comprobantes.

Modelo propio API-only (patrón l10n_pe_ne.gasto): TODA la lógica (parseo openpyxl,
validación fiscal, emisión vía quick_emit, plantilla/resultados xlsxwriter) vive en el
addon; React solo llama. El estado del lote (reanudable) vive en el servidor. Aislado por
compañía (regla multi-compañía global en security). CERO lógica de emisión nueva: reusa
account.move.l10n_pe_ne_quick_emit / action_l10n_pe_send_to_biller."""
import base64
import hashlib
import io
import json
import os
import re
import unicodedata
from datetime import date, datetime

from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError, ValidationError

# Cabeceras de la plantilla/hoja 'Ventas' (orden exacto; el parseo lee por NOMBRE normalizado,
# no por posición, pero la plantilla las escribe en este orden).
_HEADERS = ["venta", "tipo", "serie", "fecha", "tipo doc cliente", "num doc cliente",
            "cliente", "codigo producto", "producto", "cantidad", "precio unitario",
            "descuento %", "afectacion", "bolsa", "moneda"]
# Defaults de los límites (ir.config_parameter con override en caliente, sin data XML).
_MASIVO_DEFAULTS = {
    "masivo_max_filas": 500,
    "masivo_max_comprobantes": 200,
    "masivo_max_chunk": 5,
    "masivo_max_bytes": 2097152,
}
# afectación humana (columna Excel) → taxCode catálogo-05 que consume quick_emit.
_AFECTACION_TAXCODE = {
    "GRAVADO": "1000", "EXONERADO": "9997", "INAFECTO": "9998",
    "EXPORTACION": "9995", "GRATUITO": "9996",
    "1000": "1000", "9997": "9997", "9998": "9998", "9995": "9995", "9996": "9996",
}
# tipo doc cliente humano → código catálogo-06 (l10n_pe_vat_code) que consume quick_partner.
_TIPODOC_CLIENTE = {
    "RUC": "6", "DNI": "1", "CE": "4", "PASAPORTE": "7", "SD": "0",
    "6": "6", "1": "1", "4": "4", "7": "7", "0": "0",
}
# tipo comprobante humano → código catálogo-01 (solo 01/03 en la masiva v1).
_TIPO_COMPROBANTE = {"FACTURA": "01", "BOLETA": "03", "01": "01", "03": "03"}
# l10n_pe_biller_state (que devuelve quick_result como 'estado') → estado de fila del lote.
# 'en_proceso' es el estado del modo ASÍNCRONO (SQS): al emitir, el biller encola y responde
# 'en_proceso'; el resultado real de SUNAT llega minutos después vía cron. Antes no estaba en el
# mapa y caía al default 'error', mostrando como fallidos comprobantes que en realidad se aceptaban.
_ESTADO_MAP = {"enviado": "emitido", "rechazado": "rechazado", "error": "error",
               "en_proceso": "en_proceso"}

# ------------------------------------------------------------------ Lote de GUÍAS
# Cabeceras de la plantilla/hoja 'Guias' (MVP: GRE remitente 09, motivo 01 Venta,
# modalidad 02 transporte privado, un destinatario, un vehículo + un conductor, uno o más
# bienes agrupados por 'guia'). El parseo lee por NOMBRE normalizado, no por posición.
# 'destinatario' (razón social) va más allá del mínimo pedido porque una guía necesita el
# nombre del destinatario para SUNAT cuando el partner es nuevo (rznSocialDestinatario).
_HEADERS_GUIA = ["guia", "destinatario doc", "destinatario", "motivo",
                 "ubigeo partida", "dir partida", "ubigeo llegada", "dir llegada",
                 "peso bruto", "placa", "conductor tipo doc", "conductor num doc",
                 "conductor nombres", "conductor apellidos", "conductor licencia",
                 "descripcion", "cantidad", "unidad"]
# tipo doc del conductor humano → código cat.06 que consume el modelo de guía.
_CONDUCTOR_TIPODOC = {"DNI": "1", "CE": "4", "PASAPORTE": "7",
                      "1": "1", "4": "4", "7": "7"}
# estado de la guía (l10n_pe_ne.guia_remision.estado, que devuelven quick_guia/emitir_guia) →
# estado de fila del lote. 'enviado' = Aceptada por SUNAT; 'en_proceso' = ticket pendiente (lo
# resuelve el cron de guías _cron_consultar_en_proceso, no un cron nuevo del lote).
_ESTADO_MAP_GUIA = {"enviado": "emitido", "en_proceso": "en_proceso",
                    "rechazado": "rechazado", "error": "error"}


class L10nPeNeLote(models.Model):
    _name = "l10n_pe_ne.lote"
    _description = "Lote de emisión masiva (NE Express)"
    _order = "id desc"

    name = fields.Char(string="Archivo", required=True)          # filename original
    sha256 = fields.Char(index=True)                             # detección de re-subida
    attachment_id = fields.Many2one("ir.attachment")            # xlsx original (soporte/debug)
    # Discriminador del lote: emite comprobantes (facturas/boletas) o guías de remisión. Los
    # lotes existentes quedan 'comprobante' (default) y se comportan EXACTAMENTE igual que hoy;
    # solo el tipo 'guia' ramifica plantilla/parseo/validación/procesado hacia el modelo de guía.
    tipo = fields.Selection([
        ("comprobante", "Comprobantes"),
        ("guia", "Guías de remisión"),
    ], default="comprobante", required=True)
    estado = fields.Selection([
        ("con_errores", "Con errores"),   # terminal: solo consulta del reporte
        ("validado", "Validado"),         # listo para procesar
        ("en_proceso", "En proceso"),
        ("terminado", "Terminado"),
        ("cancelado", "Cancelado"),
    ], default="validado", required=True)
    company_id = fields.Many2one("res.company", required=True, index=True,
                                 default=lambda s: s.env.company)
    fila_ids = fields.One2many("l10n_pe_ne.lote.fila", "lote_id")
    # Denormalizados del reporte (para listar sin re-parsear el reporte_json).
    total_filas = fields.Integer(default=0)          # filas de datos del xlsx
    total_comprobantes = fields.Integer(default=0)   # comprobantes agrupados
    reporte_json = fields.Text()                     # {errores:[], advertencias:[], duplicadoDe:int|None}

    # ------------------------------------------------------------- helpers de config
    def _masivo_param(self, key, default):
        """Lee l10n_pe_ne.<key> (int) con default; valor inválido o <=0 → default.
        Patrón de _ttl_hours() en controllers/main.py (aquí int en vez de float)."""
        raw = self.env["ir.config_parameter"].sudo().get_param("l10n_pe_ne.%s" % key, default)
        try:
            v = int(raw)
        except (TypeError, ValueError):
            v = int(default)
        return v if v > 0 else int(default)

    def _masivo_can_commit(self):
        """Commit por fila SOLO fuera de tests (test_enable) y del harness E2E (E2E_NO_COMMIT).
        Así un doc aceptado por SUNAT no se pierde por rollback de una fila posterior, pero los
        unit tests y el shell E2E mantienen el rollback transaccional."""
        return not tools.config["test_enable"] and not os.environ.get("E2E_NO_COMMIT")

    # ------------------------------------------------------------- helpers de parseo
    @staticmethod
    def _l10n_pe_ne_norm(h):
        """Cabecera normalizada: minúsculas, sin tildes, espacios colapsados."""
        s = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode("ascii")
        return " ".join(s.lower().split())

    @staticmethod
    def _l10n_pe_ne_txt(v):
        """Celda → texto; enteros float ('45678912.0') → '45678912' (openpyxl lee números)."""
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    @staticmethod
    def _l10n_pe_ne_num(v):
        """Número tolerante: acepta coma decimal ('12,50') y espacios; None si vacío/ilegible."""
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def _l10n_pe_ne_fecha(v):
        """Fecha tolerante: datetime/date de openpyxl, DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY.
        Devuelve date, None (vacío) o 'ERROR' (ilegible)."""
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return "ERROR"

    @staticmethod
    def _l10n_pe_ne_ruc_valido(ruc):
        """Dígito verificador módulo-11 del RUC (espejo de validateRucCheckDigit del SPA)."""
        ruc = (ruc or "").strip()
        if not (len(ruc) == 11 and ruc.isdigit()):
            return False
        pesos = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
        s = sum(int(ruc[i]) * pesos[i] for i in range(10))
        r = 11 - (s % 11)
        dv = 0 if r == 10 else (1 if r == 11 else r)
        return dv == int(ruc[10])

    def _l10n_pe_ne_parse_xlsx(self, data):
        """Lee la hoja 'Ventas' (o la primera) por NOMBRE de cabecera normalizado (no por
        posición). Ignora columnas desconocidas y filas totalmente vacías. data_only=True lee
        el valor calculado de fórmulas. Falta una cabecera obligatoria → UserError global."""
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception:
            raise UserError(_("No se pudo leer el archivo. Sube un .xlsx válido (no un .xls antiguo)."))
        ws = wb["Ventas"] if "Ventas" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("El archivo está vacío."))
        header = [self._l10n_pe_ne_norm(h) for h in rows[0]]
        idx = {h: i for i, h in enumerate(header) if h}
        faltan = [h for h in ("tipo", "producto", "cantidad", "precio unitario") if h not in idx]
        if faltan:
            raise UserError(_("Faltan columnas obligatorias en la hoja: %s") % ", ".join(faltan))

        def raw(row, name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        filas = []
        for n, row in enumerate(rows[1:], start=2):   # n = fila Excel real (1 = cabecera)
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            filas.append({
                "filaExcel": n,
                "venta": self._l10n_pe_ne_txt(raw(row, "venta")),
                "tipo": self._l10n_pe_ne_txt(raw(row, "tipo")).upper(),
                "serie": self._l10n_pe_ne_txt(raw(row, "serie")).upper(),
                "fecha_raw": raw(row, "fecha"),
                "cli_tipo": self._l10n_pe_ne_txt(raw(row, "tipo doc cliente")).upper(),
                "cli_num": self._l10n_pe_ne_txt(raw(row, "num doc cliente")),
                "cli_nombre": self._l10n_pe_ne_txt(raw(row, "cliente")),
                "cod": self._l10n_pe_ne_txt(raw(row, "codigo producto")),
                "producto": self._l10n_pe_ne_txt(raw(row, "producto")),
                "cantidad_raw": raw(row, "cantidad"),
                "precio_raw": raw(row, "precio unitario"),
                "descuento_raw": raw(row, "descuento %"),
                "afectacion": self._l10n_pe_ne_txt(raw(row, "afectacion")).upper(),
                "bolsa": self._l10n_pe_ne_txt(raw(row, "bolsa")).upper(),
                "moneda": self._l10n_pe_ne_txt(raw(row, "moneda")).upper(),
            })
        return filas

    def _l10n_pe_ne_validar(self, filas):
        """Agrupa por 'venta' (contiguas; vacío = comprobante propio), valida cada grupo y
        construye el payload subset de quick_emit. Devuelve (comprobantes, errores, advertencias).
        v1 todo-o-nada: si hay ≥1 error, el llamador no crea filas procesables."""
        errores, advertencias, comprobantes = [], [], []
        grupos, vistos, i = [], set(), 0
        while i < len(filas):
            v = filas[i]["venta"]
            if not v:
                grupos.append((v, [filas[i]])); i += 1; continue
            j, bloque = i, []
            while j < len(filas) and filas[j]["venta"] == v:
                bloque.append(filas[j]); j += 1
            if v in vistos:
                for fx in bloque:
                    errores.append({"filaExcel": fx["filaExcel"], "venta": v,
                                    "mensaje": _("La venta '%s' aparece en filas no contiguas") % v})
            vistos.add(v)
            grupos.append((v, bloque)); i = j
        for v, bloque in grupos:
            comp = self._l10n_pe_ne_validar_grupo(v, bloque, errores, advertencias)
            if comp:
                comprobantes.append(comp)
        return comprobantes, errores, advertencias

    def _l10n_pe_ne_validar_grupo(self, venta, bloque, errores, advertencias):
        local, first, fn = [], bloque[0], bloque[0]["filaExcel"]
        tipo = _TIPO_COMPROBANTE.get(first["tipo"])
        if not tipo:
            for f in bloque:
                local.append((f["filaExcel"], _("Tipo no soportado en emisión masiva; usa el formulario individual")))
        for f in bloque[1:]:
            mismo = ((f["tipo"], f["serie"], f["cli_num"], f["moneda"]) ==
                     (first["tipo"], first["serie"], first["cli_num"], first["moneda"])
                     and self._l10n_pe_ne_txt(f["fecha_raw"]) == self._l10n_pe_ne_txt(first["fecha_raw"]))
            if not mismo:
                local.append((f["filaExcel"], _("Las filas de la venta '%s' deben compartir tipo, serie, fecha, cliente y moneda") % venta))
        serie = first["serie"]
        if serie:
            pref = "F" if tipo == "01" else "B"
            if not re.match(r"^[BF][A-Z0-9]{3}$", serie) or serie[:1] != pref:
                local.append((fn, _("La serie '%s' no es válida para ese tipo de comprobante") % serie))
        fecha = None
        if first["fecha_raw"] not in (None, ""):
            fecha = self._l10n_pe_ne_fecha(first["fecha_raw"])
            hoy = self.env["account.move"]._l10n_pe_ne_today_lima()
            if fecha == "ERROR":
                local.append((fn, _("La fecha no es válida (usa DD/MM/AAAA)"))); fecha = None
            elif fecha > hoy:
                local.append((fn, _("La fecha no puede ser futura")))
            elif (hoy - fecha).days > 3:
                advertencias.append({"filaExcel": fn, "venta": venta,
                                     "mensaje": _("Fecha con más de 3 días de antigüedad (plazo de envío SUNAT)")})
        cli_num, cli_nombre = first["cli_num"], first["cli_nombre"]
        cli_code = _TIPODOC_CLIENTE.get(first["cli_tipo"], "") if first["cli_tipo"] else ""
        if tipo == "01":
            if not self._l10n_pe_ne_ruc_valido(cli_num):
                local.append((fn, _("RUC inválido: el dígito verificador no coincide")))
            cli_code = "6"
        elif cli_num:
            if first["cli_tipo"] and cli_code == "":
                local.append((fn, _("Tipo de documento de cliente no soportado: %s") % first["cli_tipo"]))
            if not cli_code:
                cli_code = "1" if len(cli_num) == 8 else "6"
            if cli_code == "1" and not (len(cli_num) == 8 and cli_num.isdigit()):
                local.append((fn, _("El DNI debe tener 8 dígitos")))
        else:
            cli_code = "0"   # boleta a público general (quick_emit → CONSUMIDOR FINAL)
        if cli_num and not cli_nombre and not self.env["res.partner"].search([("vat", "=", cli_num)], limit=1):
            local.append((fn, _("Falta el nombre/razón social del cliente")))
        lineas, total_est = [], 0.0
        for f in bloque:
            fx = f["filaExcel"]
            if not f["producto"]:
                local.append((fx, _("El producto (descripción) es requerido")))
            cant = self._l10n_pe_ne_num(f["cantidad_raw"])
            if cant is None or cant <= 0:
                local.append((fx, _("La cantidad debe ser mayor a 0"))); cant = cant or 0.0
            pu = self._l10n_pe_ne_num(f["precio_raw"])
            if pu is None or pu < 0:
                local.append((fx, _("El precio unitario no puede ser negativo"))); pu = pu or 0.0
            desc = self._l10n_pe_ne_num(f["descuento_raw"]) or 0.0
            if desc < 0 or desc > 100:
                local.append((fx, _("El descuento debe estar entre 0 y 100")))
            afect = f["afectacion"] or "GRAVADO"
            taxcode = _AFECTACION_TAXCODE.get(afect)
            if not taxcode:
                local.append((fx, _("Afectación no válida: %s") % afect)); taxcode = "1000"
            if f["bolsa"] not in ("SI", "NO", ""):
                local.append((fx, _("El valor de 'bolsa' debe ser SI o NO")))
            # El 'precio' del CSV es CON IGV (precio final de venta). A SUNAT va el valor
            # unitario SIN IGV: gravado (1000) se divide entre 1.18; el resto ya es base.
            pu_base = pu / 1.18 if taxcode == "1000" else pu
            linea = {"descripcion": f["producto"], "cantidad": cant, "precioUnitario": pu_base,
                     "descuento": desc, "taxCode": taxcode, "icbper": f["bolsa"] == "SI"}
            if f["cod"]:
                linea["productCod"] = f["cod"]
            lineas.append(linea)
            # 'pu' ya incluye IGV, así que el estimado del total no vuelve a sumar el 18%.
            total_est += cant * pu * (1 - desc / 100.0)
        if tipo == "03" and not cli_num and total_est >= 700:
            advertencias.append({"filaExcel": fn, "venta": venta,
                                 "mensaje": _("Boleta ≥ S/ 700 sin documento de identidad (SUNAT puede rechazarla)")})
        if local:
            for fx, msg in local:
                errores.append({"filaExcel": fx, "venta": venta, "mensaje": msg})
            return None
        cliente = {"tipoDoc": cli_code}
        if cli_num:
            cliente["numDoc"] = cli_num
        if cli_nombre:
            cliente["razonSocial"] = cli_nombre
        payload = {"tipoDoc": tipo, "cliente": cliente, "lineas": lineas, "moneda": first["moneda"] or "PEN"}
        if serie:
            payload["serie"] = serie
        if fecha:
            payload["fechaEmision"] = fecha.strftime("%Y-%m-%d")
        filas_excel = "%s" % fn if len(bloque) == 1 else "%s-%s" % (fn, bloque[-1]["filaExcel"])
        return {"payload": payload, "total": round(total_est, 2), "filas_excel": filas_excel}

    # ------------------------------------------------------- parseo/validación GUÍAS
    def _l10n_pe_ne_parse_xlsx_guia(self, data):
        """Lee la hoja 'Guias' (o la primera) por NOMBRE de cabecera normalizado. Una fila = un
        bien; la columna 'guia' agrupa varios bienes en una misma guía (filas contiguas). Falta
        una cabecera obligatoria → UserError global. Espeja _l10n_pe_ne_parse_xlsx."""
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception:
            raise UserError(_("No se pudo leer el archivo. Sube un .xlsx válido (no un .xls antiguo)."))
        ws = wb["Guias"] if "Guias" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("El archivo está vacío."))
        header = [self._l10n_pe_ne_norm(h) for h in rows[0]]
        idx = {h: i for i, h in enumerate(header) if h}
        faltan = [h for h in ("destinatario doc", "ubigeo partida", "ubigeo llegada",
                              "descripcion", "cantidad") if h not in idx]
        if faltan:
            raise UserError(_("Faltan columnas obligatorias en la hoja: %s") % ", ".join(faltan))

        def raw(row, name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        filas = []
        for n, row in enumerate(rows[1:], start=2):   # n = fila Excel real (1 = cabecera)
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            filas.append({
                "filaExcel": n,
                "guia": self._l10n_pe_ne_txt(raw(row, "guia")),
                "dest_doc": self._l10n_pe_ne_txt(raw(row, "destinatario doc")),
                "dest_nombre": self._l10n_pe_ne_txt(raw(row, "destinatario")),
                "motivo": self._l10n_pe_ne_txt(raw(row, "motivo")),
                "ubi_partida": self._l10n_pe_ne_txt(raw(row, "ubigeo partida")),
                "dir_partida": self._l10n_pe_ne_txt(raw(row, "dir partida")),
                "ubi_llegada": self._l10n_pe_ne_txt(raw(row, "ubigeo llegada")),
                "dir_llegada": self._l10n_pe_ne_txt(raw(row, "dir llegada")),
                "peso_raw": raw(row, "peso bruto"),
                "placa": self._l10n_pe_ne_txt(raw(row, "placa")).upper(),
                "cond_tipo": self._l10n_pe_ne_txt(raw(row, "conductor tipo doc")).upper(),
                "cond_num": self._l10n_pe_ne_txt(raw(row, "conductor num doc")),
                "cond_nombres": self._l10n_pe_ne_txt(raw(row, "conductor nombres")),
                "cond_apellidos": self._l10n_pe_ne_txt(raw(row, "conductor apellidos")),
                "cond_licencia": self._l10n_pe_ne_txt(raw(row, "conductor licencia")),
                "producto": self._l10n_pe_ne_txt(raw(row, "descripcion")),
                "cantidad_raw": raw(row, "cantidad"),
                "unidad": self._l10n_pe_ne_txt(raw(row, "unidad")).upper(),
            })
        return filas

    def _l10n_pe_ne_validar_guia(self, filas):
        """Agrupa por 'guia' (contiguas; vacío = guía propia), valida cada grupo y construye el
        payload de l10n_pe_ne_quick_guia. Devuelve (guias, errores, advertencias). Mismo contrato
        y mecanismo de errores que _l10n_pe_ne_validar (todo-o-nada: ≥1 error → sin filas)."""
        errores, advertencias, guias = [], [], []
        grupos, vistos, i = [], set(), 0
        while i < len(filas):
            v = filas[i]["guia"]
            if not v:
                grupos.append((v, [filas[i]])); i += 1; continue
            j, bloque = i, []
            while j < len(filas) and filas[j]["guia"] == v:
                bloque.append(filas[j]); j += 1
            if v in vistos:
                for fx in bloque:
                    errores.append({"filaExcel": fx["filaExcel"], "venta": v,
                                    "mensaje": _("La guía '%s' aparece en filas no contiguas") % v})
            vistos.add(v)
            grupos.append((v, bloque)); i = j
        for v, bloque in grupos:
            g = self._l10n_pe_ne_validar_grupo_guia(v, bloque, errores, advertencias)
            if g:
                guias.append(g)
        return guias, errores, advertencias

    def _l10n_pe_ne_validar_grupo_guia(self, guia, bloque, errores, advertencias):
        """Valida un grupo de bienes de una misma guía y arma el payload de quick_guia (MVP:
        GRE 09, motivo 01, modalidad 02 privada, un destinatario, un vehículo + un conductor por
        campos legados, uno o más bienes). Los errores usan la clave 'venta' (agrupador genérico)
        para que la tabla de errores de la SPA los muestre sin cambios."""
        local, first, fn = [], bloque[0], bloque[0]["filaExcel"]
        doc = first["dest_doc"]
        dest_tipo = "6"
        if not doc:
            local.append((fn, _("El documento del destinatario es requerido")))
        elif len(doc) == 11:
            if not self._l10n_pe_ne_ruc_valido(doc):
                local.append((fn, _("RUC del destinatario inválido: el dígito verificador no coincide")))
        elif len(doc) == 8 and doc.isdigit():
            dest_tipo = "1"
        else:
            local.append((fn, _("El documento del destinatario debe ser RUC (11 dígitos) o DNI (8)")))
        nombre = first["dest_nombre"]
        if not nombre and not (doc and self.env["res.partner"].search([("vat", "=", doc)], limit=1)):
            local.append((fn, _("El nombre/razón social del destinatario es requerido")))
        # SUNAT 2555 (motivo 01 Venta): el destinatario no puede ser la propia empresa.
        ruc_emisor = (self.env.company.vat or "").strip()
        if len(doc) == 11 and doc == ruc_emisor:
            local.append((fn, _("Para el motivo Venta el destinatario no puede ser tu propia empresa")))
        # Normaliza el motivo: Excel suele traerlo numérico y sin el cero a la izquierda
        # (1 / 1.0 → "01"), lo que antes disparaba un rechazo confuso "solo 01 soportado".
        motivo_raw = first["motivo"]
        if isinstance(motivo_raw, float) and motivo_raw.is_integer():
            motivo_raw = int(motivo_raw)
        motivo = (str(motivo_raw).strip() or "01") if motivo_raw not in (None, "") else "01"
        if motivo.isdigit():
            motivo = motivo.zfill(2)
        if motivo != "01":
            local.append((fn, _("Solo el motivo 01 (Venta) está soportado en emisión masiva de "
                                "guías; usa el formulario individual")))
        for campo, etiqueta in (("ubi_partida", "partida"), ("ubi_llegada", "llegada")):
            if not re.fullmatch(r"\d{6}", first[campo] or ""):
                local.append((fn, _("El ubigeo de %s debe tener 6 dígitos") % etiqueta))
        if not first["dir_partida"]:
            local.append((fn, _("La dirección de partida es requerida")))
        if not first["dir_llegada"]:
            local.append((fn, _("La dirección de llegada es requerida")))
        peso = self._l10n_pe_ne_num(first["peso_raw"])
        if peso is None or peso <= 0:
            local.append((fn, _("El peso bruto debe ser mayor a 0"))); peso = peso or 0.0
        placa = (first["placa"] or "").strip().upper()
        if not re.match(r"^(?!0+$)[0-9A-Z]{6,8}$", placa):
            local.append((fn, _("La placa debe tener de 6 a 8 caracteres alfanuméricos (SUNAT 2567)")))
        cond_tipo = _CONDUCTOR_TIPODOC.get(first["cond_tipo"], "") if first["cond_tipo"] else "1"
        if first["cond_tipo"] and not cond_tipo:
            local.append((fn, _("Tipo de documento del conductor no soportado: %s") % first["cond_tipo"]))
            cond_tipo = "1"
        for campo, etiqueta in (("cond_num", _("el documento")), ("cond_nombres", _("los nombres")),
                                ("cond_apellidos", _("los apellidos")), ("cond_licencia", _("la licencia"))):
            if not (first[campo] or "").strip():
                local.append((fn, _("Falta %s del conductor") % etiqueta))
        items = []
        for f in bloque:
            fx = f["filaExcel"]
            if not f["producto"]:
                local.append((fx, _("La descripción del bien es requerida")))
            cant = self._l10n_pe_ne_num(f["cantidad_raw"])
            if cant is None or cant <= 0:
                local.append((fx, _("La cantidad debe ser mayor a 0"))); cant = cant or 0.0
            # Unidad SUNAT (cat.06): normaliza a mayúscula ("niu" → "NIU") con default NIU;
            # una unidad inválida la rechaza el biller/SUNAT por fila (no corrompe el lote).
            unidad = (str(f["unidad"]).strip().upper() if f["unidad"] else "NIU") or "NIU"
            items.append({"descripcion": f["producto"], "cantidad": cant, "unidad": unidad})
        if local:
            for fx, msg in local:
                errores.append({"filaExcel": fx, "venta": guia, "mensaje": msg})
            return None
        destinatario = {"numDoc": doc, "tipoDoc": dest_tipo}
        if nombre:
            destinatario["razonSocial"] = nombre
        payload = {
            "tipoGre": "09", "modalidadTraslado": "02", "motivoTraslado": motivo,
            "destinatario": destinatario,
            "ubigeoPartida": first["ubi_partida"], "dirPartida": first["dir_partida"],
            "ubigeoLlegada": first["ubi_llegada"], "dirLlegada": first["dir_llegada"],
            "pesoBruto": peso, "uniMedidaPeso": "KGM",
            "numPlaca": placa, "conductorTipoDoc": cond_tipo,
            "conductorNumDoc": first["cond_num"], "conductorNombres": first["cond_nombres"],
            "conductorApellidos": first["cond_apellidos"], "conductorLicencia": first["cond_licencia"],
            "items": items,
        }
        filas_excel = "%s" % fn if len(bloque) == 1 else "%s-%s" % (fn, bloque[-1]["filaExcel"])
        return {"payload": payload, "total": round(peso, 2),
                "cliente": nombre or doc, "filas_excel": filas_excel}

    @api.model
    def l10n_pe_ne_crear_lote(self, payload):
        """Sube+valida un xlsx (NO emite). Valida extensión/tamaño, sha256 (advertencia de
        re-subida), parsea (openpyxl), agrupa+valida, y crea el lote + filas pendientes con su
        payload_json. Devuelve el reporte de validación. `payload['tipo']` ∈ (comprobante, guia)
        decide qué parser/validador se usa; el esqueleto (dedup/estados/filas) es idéntico."""
        tipo = payload.get("tipo") or "comprobante"
        if tipo not in ("comprobante", "guia"):
            raise UserError(_("Tipo de lote no soportado: %s") % tipo)
        es_guia = tipo == "guia"
        filename = (payload.get("filename") or "").strip()
        if not filename.lower().endswith(".xlsx"):
            raise UserError(_("Sube un archivo Excel (.xlsx) — descarga la plantilla si tienes dudas."))
        try:
            data = base64.b64decode(payload.get("contentB64") or "")
        except Exception:
            raise UserError(_("El archivo no se pudo leer (base64 inválido)."))
        max_bytes = self._masivo_param("masivo_max_bytes", _MASIVO_DEFAULTS["masivo_max_bytes"])
        if len(data) > max_bytes:
            raise UserError(_("El archivo no puede superar %s MB.") % round(max_bytes / 1048576.0, 1))
        sha = hashlib.sha256(data).hexdigest()
        dup = self.search([("sha256", "=", sha), ("company_id", "=", self.env.company.id)], limit=1)
        filas = (self._l10n_pe_ne_parse_xlsx_guia(data) if es_guia
                 else self._l10n_pe_ne_parse_xlsx(data))
        max_filas = self._masivo_param("masivo_max_filas", _MASIVO_DEFAULTS["masivo_max_filas"])
        if len(filas) > max_filas:
            raise UserError(_("El archivo supera el máximo de %s filas") % max_filas)
        comprobantes, errores, advertencias = (self._l10n_pe_ne_validar_guia(filas) if es_guia
                                               else self._l10n_pe_ne_validar(filas))
        max_comp = self._masivo_param("masivo_max_comprobantes", _MASIVO_DEFAULTS["masivo_max_comprobantes"])
        if len(comprobantes) > max_comp:
            raise UserError(_("El archivo supera el máximo de %s %s")
                            % (max_comp, _("guías") if es_guia else _("comprobantes")))
        estado = "con_errores" if errores else "validado"
        att = self.env["ir.attachment"].create({
            "name": filename, "res_model": "l10n_pe_ne.lote", "mimetype":
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "raw": data})
        reporte = {"errores": errores, "advertencias": advertencias, "duplicadoDe": dup.id or None}
        lote = self.create({
            "name": filename, "sha256": sha, "attachment_id": att.id, "estado": estado, "tipo": tipo,
            "total_filas": len(filas), "total_comprobantes": len(comprobantes),
            "reporte_json": json.dumps(reporte)})
        att.res_id = lote.id
        if estado == "validado":
            for i, comp in enumerate(comprobantes, 1):
                p = comp["payload"]
                vals = {
                    "lote_id": lote.id, "secuencia": i, "filas_excel": comp["filas_excel"],
                    "payload_json": json.dumps(p), "estado": "pendiente", "total": comp["total"]}
                if es_guia:
                    # La serie (T###) se conoce recién al crear la guía (secuencia por serie);
                    # se rellena al procesar. Aquí quedan tipo_doc/destinatario para el listado.
                    vals.update({"tipo_doc": "09", "serie": "", "moneda": "",
                                 "cliente": comp.get("cliente") or ""})
                else:
                    vals.update({"tipo_doc": p["tipoDoc"], "serie": p.get("serie") or "",
                                 "cliente": p["cliente"].get("razonSocial") or "",
                                 "moneda": p.get("moneda") or "PEN"})
                self.env["l10n_pe_ne.lote.fila"].create(vals)
        return dict(reporte, id=lote.id, estado=estado, filename=filename, tipo=tipo,
                    totalFilas=len(filas), totalComprobantes=len(comprobantes))

    @api.model
    def l10n_pe_ne_plantilla(self, tipo="comprobante"):
        """Plantilla xlsx (mismo estilo visual que l10n_pe_ne_export: cabecera azul #2563eb):
        hoja 'Ventas' con 15 cabeceras + 3-4 ejemplos + listas desplegables (data_validation),
        y hoja 'Instrucciones'. Devuelve {filename, contentB64}. `tipo='guia'` devuelve la
        plantilla de guías de remisión."""
        if tipo == "guia":
            return self._l10n_pe_ne_plantilla_guia()
        import xlsxwriter
        ejemplos = [
            ["V-001", "FACTURA", "F001", "01/07/2026", "RUC", "20100070970", "FERRETERIA LA UNION SAC", "CEM-01", "CEMENTO SOL 42.5KG", 2, 33.90, 0, "GRAVADO", "NO", "PEN"],
            ["V-001", "FACTURA", "F001", "01/07/2026", "RUC", "20100070970", "FERRETERIA LA UNION SAC", "CLV-02", "CLAVO 2 PULGADAS X KG", 5, 4.50, 10, "GRAVADO", "NO", "PEN"],
            ["", "BOLETA", "B001", "01/07/2026", "DNI", "45678912", "ROSA QUISPE", "", "PINTURA LATEX BLANCO 1GAL", 1, 45.00, 0, "GRAVADO", "SI", "PEN"],
            ["", "BOLETA", "", "", "", "", "", "", "GASEOSA 500ML", 3, 2.50, 0, "GRAVADO", "NO", "PEN"],
        ]
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet("Ventas")
        head = wb.add_format({"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1})
        for c, h in enumerate(_HEADERS):
            ws.write(0, c, h, head)
            ws.set_column(c, c, max(12, len(h) + 2))
        for r, row in enumerate(ejemplos, 1):
            ws.write_row(r, 0, row)

        def dv(col, opts):
            ws.data_validation(1, col, 500, col, {"validate": "list", "source": opts})
        dv(1, ["FACTURA", "BOLETA"])
        dv(4, ["RUC", "DNI", "CE", "PASAPORTE", "SD"])
        dv(12, ["GRAVADO", "EXONERADO", "INAFECTO", "EXPORTACION", "GRATUITO"])
        dv(13, ["SI", "NO"])
        dv(14, ["PEN", "USD"])
        ws.freeze_panes(1, 0)
        wi = wb.add_worksheet("Instrucciones")
        wi.set_column(0, 0, 105)
        for r, line in enumerate([
            "CHASKIFACT — Plantilla de emisión masiva (boletas y facturas)",
            "",
            "1. Una fila = una línea de venta. Usa la columna 'venta' para agrupar varias líneas en un mismo comprobante (mismo código en filas contiguas).",
            "2. 'tipo': FACTURA (01) o BOLETA (03). La factura exige RUC de cliente válido; la boleta acepta DNI/CE/PASAPORTE o queda a público general si dejas el cliente en blanco.",
            "3. 'serie' vacía = automática (F001 factura, B001 boleta). Puedes abrir una serie nueva (p.ej. B002) y aparecerá luego en Series.",
            "4. 'fecha' vacía = hoy. No puede ser futura; si tiene más de 3 días te avisamos (plazo de envío SUNAT).",
            "5. 'precio unitario' es CON IGV incluido (precio final de venta). 'afectacion' por línea (GRAVADO ya trae el 18%). 'bolsa' = SI cobra ICBPER por unidad.",
            "6. Límite: 500 filas / 200 comprobantes por archivo, hasta 2 MB.",
            "7. Sube el archivo, revisa el reporte de validación y recién ahí emite. Si hay errores, corrige el Excel y vuelve a subir.",
        ]):
            wi.write(r, 0, line)
        wb.close()
        return {"filename": "plantilla-ventas-chaskifact.xlsx",
                "contentB64": base64.b64encode(buf.getvalue()).decode("ascii")}

    def _l10n_pe_ne_plantilla_guia(self):
        """Plantilla xlsx de GUÍAS de remisión (MVP): hoja 'Guias' con las cabeceras de guía +
        ejemplos + dropdowns, y hoja 'Instrucciones'. Mismo estilo visual que la de ventas."""
        import xlsxwriter
        ejemplos = [
            ["G-001", "20601030013", "COMERCIAL LOS ANDES SAC", "01", "150101", "AV. AREQUIPA 100, LIMA", "150203", "JR. LORETO 450, BARRANCO", 25.5, "ABC123", "DNI", "45678912", "JUAN CARLOS", "PEREZ QUISPE", "Q12345678", "CAJA DE TORNILLOS 1/2", 10, "NIU"],
            ["G-001", "20601030013", "COMERCIAL LOS ANDES SAC", "01", "150101", "AV. AREQUIPA 100, LIMA", "150203", "JR. LORETO 450, BARRANCO", 25.5, "ABC123", "DNI", "45678912", "JUAN CARLOS", "PEREZ QUISPE", "Q12345678", "CLAVOS 3 PULGADAS X KG", 5, "KGM"],
            ["G-002", "10456789123", "MARIA ROSA FLORES", "01", "150101", "AV. BRASIL 900, LIMA", "070101", "AV. SAENZ PENA 200, CALLAO", 12.0, "XY4821", "DNI", "70123456", "PEDRO", "GOMEZ RIOS", "P70123456", "MOCHILA ESCOLAR", 3, "NIU"],
        ]
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet("Guias")
        head = wb.add_format({"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1})
        for c, h in enumerate(_HEADERS_GUIA):
            ws.write(0, c, h, head)
            ws.set_column(c, c, max(12, len(h) + 2))
        for r, row in enumerate(ejemplos, 1):
            ws.write_row(r, 0, row)
        # dropdowns: motivo (solo 01 en el MVP) y tipo de doc del conductor.
        ws.data_validation(1, 3, 500, 3, {"validate": "list", "source": ["01"]})
        ws.data_validation(1, 10, 500, 10, {"validate": "list", "source": ["DNI", "CE", "PASAPORTE"]})
        ws.freeze_panes(1, 0)
        wi = wb.add_worksheet("Instrucciones")
        wi.set_column(0, 0, 108)
        for r, line in enumerate([
            "CHASKIFACT — Plantilla de emisión masiva de GUÍAS de remisión (remitente)",
            "",
            "1. Una fila = un bien. Usa la columna 'guia' para agrupar varios bienes en una misma guía (mismo código en filas contiguas). Los datos de cabecera (destinatario, puntos, peso, vehículo, conductor) se toman de la PRIMERA fila del grupo.",
            "2. MVP: guía de REMITENTE (tipo 09), motivo 01 (Venta), transporte PRIVADO (un vehículo + un conductor). Para transportista, transporte público, comercio exterior o comprobante relacionado, usa el formulario individual de guías.",
            "3. 'destinatario doc': RUC (11 dígitos) o DNI (8); no puede ser tu propia empresa. 'destinatario' = razón social o nombre (obligatorio si el destinatario es nuevo).",
            "4. 'ubigeo partida'/'ubigeo llegada': código de 6 dígitos (catálogo 13 de SUNAT). 'dir partida'/'dir llegada': dirección de cada punto.",
            "5. 'peso bruto' en kilogramos (mayor a 0). 'placa': 6 a 8 caracteres. Conductor: tipo doc (DNI/CE/PASAPORTE), número, nombres, apellidos y licencia.",
            "6. 'descripcion'/'cantidad'/'unidad' por cada bien. La serie se asigna automáticamente (T001) y la fecha de emisión es la del envío a SUNAT.",
            "7. Límite: 500 filas / 200 guías por archivo, hasta 2 MB. Sube el archivo, revisa el reporte de validación y recién ahí emite.",
        ]):
            wi.write(r, 0, line)
        wb.close()
        return {"filename": "plantilla-guias-chaskifact.xlsx",
                "contentB64": base64.b64encode(buf.getvalue()).decode("ascii")}

    # ------------------------------------------------------------- serializadores
    def _l10n_pe_ne_sync_async(self):
        """Reconcilia las filas 'en_proceso' (encoladas en modo async) con el estado FINAL del
        move, que el cron de biller (_l10n_pe_cron_poll_async) actualiza minutos después. Sin
        esto, una fila encolada se quedaba en 'en_proceso' aunque su comprobante ya estuviera
        aceptado/rechazado. Idempotente: solo escribe cuando el move ya resolvió.

        Lote de guías: la guía ya es asíncrona y su cron propio (_cron_consultar_en_proceso)
        actualiza el estado; aquí solo refrescamos la fila leyendo guia_id.estado (no hay cron
        del lote ni consulta nueva a SUNAT)."""
        self.ensure_one()
        if self.tipo == "guia":
            for f in self.fila_ids.filtered(lambda x: x.estado == "en_proceso" and x.guia_id):
                nuevo = _ESTADO_MAP_GUIA.get(f.guia_id.estado)
                if nuevo and nuevo != "en_proceso":
                    f.write({"estado": nuevo,
                             "mensaje": f.guia_id.l10n_pe_biller_message or f.mensaje or ""})
            return
        for f in self.fila_ids.filtered(lambda x: x.estado == "en_proceso" and x.move_id):
            nuevo = _ESTADO_MAP.get(f.move_id.l10n_pe_biller_state)
            if nuevo and nuevo != "en_proceso":
                f.write({"estado": nuevo,
                         "mensaje": f.move_id.l10n_pe_biller_message or f.mensaje or ""})

    def _l10n_pe_ne_contadores(self):
        self.ensure_one()
        d = {"pendientes": 0, "enProceso": 0, "emitidos": 0, "rechazados": 0, "errores": 0, "cancelados": 0}
        key = {"pendiente": "pendientes", "en_proceso": "enProceso", "emitido": "emitidos",
               "rechazado": "rechazados", "error": "errores", "cancelado": "cancelados"}
        for f in self.fila_ids:
            k = key.get(f.estado)
            if k:
                d[k] += 1
        return d

    def l10n_pe_ne_lote_detalle(self):
        self.ensure_one()
        self._l10n_pe_ne_sync_async()   # trae el resultado async ya resuelto antes de serializar
        c = self._l10n_pe_ne_contadores()
        rep = json.loads(self.reporte_json or "{}")
        return {
            "id": self.id, "filename": self.name, "tipo": self.tipo,
            "fecha": self.create_date.strftime("%Y-%m-%d") if self.create_date else "",
            "estado": self.estado, "totalFilas": self.total_filas,
            "totalComprobantes": self.total_comprobantes,
            "pendientes": c["pendientes"], "enProceso": c["enProceso"], "emitidos": c["emitidos"],
            "rechazados": c["rechazados"], "errores": c["errores"], "cancelados": c["cancelados"],
            "filas": [f._l10n_pe_ne_fila_dict() for f in self.fila_ids.sorted("secuencia")],
            "erroresValidacion": rep.get("errores") or [],
            "advertencias": rep.get("advertencias") or [],
            "duplicadoDe": rep.get("duplicadoDe"),
        }

    @api.model
    def l10n_pe_ne_list_lotes(self):
        out = []
        for l in self.search([], limit=100):
            l._l10n_pe_ne_sync_async()   # reconcilia async antes de contar
            c = l._l10n_pe_ne_contadores()
            out.append({
                "id": l.id, "filename": l.name, "tipo": l.tipo,
                "fecha": l.create_date.strftime("%Y-%m-%d") if l.create_date else "",
                "estado": l.estado, "totalComprobantes": l.total_comprobantes,
                "emitidos": c["emitidos"], "enProceso": c["enProceso"],
                "rechazados": c["rechazados"], "errores": c["errores"]})
        return out

    # ------------------------------------------------------------- procesamiento
    def l10n_pe_ne_procesar(self, max_filas=1):
        """Procesa hasta min(max_filas, masivo_max_chunk) filas 'pendiente' por secuencia. Cada
        fila en su propio savepoint; commit por fila (si _masivo_can_commit) para no perder un doc
        aceptado por SUNAT ante el rollback de una fila posterior. Devuelve el progreso con
        `filas` = SOLO las procesadas en esta llamada."""
        self.ensure_one()
        if self.estado == "con_errores":
            raise UserError(_("El lote tiene errores de validación; corrige el Excel y súbelo de nuevo"))
        if self.estado in ("validado", "en_proceso"):
            self.estado = "en_proceso"
        # terminado/cancelado: sin pendientes → no-op idempotente (no re-emite).
        chunk = max(1, min(int(max_filas or 1), self._masivo_param("masivo_max_chunk", _MASIVO_DEFAULTS["masivo_max_chunk"])))
        pendientes = self.fila_ids.filtered(lambda f: f.estado == "pendiente").sorted("secuencia")[:chunk]
        procesadas = self.env["l10n_pe_ne.lote.fila"]
        for fila in pendientes:
            try:
                with self.env.cr.savepoint():
                    # Lock por fila (NO por chunk): cierra la carrera de doble-emisión cuando dos
                    # llamadas a procesar() concurrentes (dos pestañas "Reanudar", o un retry que
                    # pisa una llamada en vuelo) leen la misma fila 'pendiente' antes de que
                    # cualquiera de las dos comitee. SKIP LOCKED salta la fila si otra transacción
                    # la está emitiendo AHORA MISMO; estado='pendiente' la salta si una llamada
                    # anterior YA la emitió y comiteó. El lock se toma aquí adentro -y no al
                    # principio de procesar()- porque con commit-por-fila el primer
                    # self.env.cr.commit() libera cualquier lock tomado antes; debe vivir
                    # exactamente entre esta adquisición y el commit de ESTA fila más abajo.
                    self.env.cr.execute(
                        "SELECT id FROM l10n_pe_ne_lote_fila WHERE id = %s AND estado = 'pendiente' "
                        "FOR UPDATE SKIP LOCKED",
                        (fila.id,),
                    )
                    if not self.env.cr.fetchone():
                        continue   # otra transacción concurrente ya la tomó, o ya no está pendiente
                    fila.invalidate_recordset(["estado", "move_id"])   # relee el estado real bajo el lock
                    if fila.estado != "pendiente":
                        continue
                    fila._l10n_pe_ne_procesar_fila()
            except (UserError, ValidationError) as exc:
                fila.write({"estado": "error", "mensaje": exc.args[0] if exc.args else str(exc)})
            procesadas |= fila
            if self._masivo_can_commit():
                self.env.cr.commit()
        # Guarda: solo pasa a 'terminado' si el lote seguía en curso. Un lote 'cancelado' (o ya
        # 'terminado') tampoco tiene filas 'pendiente' -> sin este guard, un procesar() de más
        # (doble submit del SPA, poll fuera de tiempo) pisaba el estado 'cancelado' con
        # 'terminado' aunque no se reemitiera nada (revisión QW10 Task 3).
        if self.estado not in ("cancelado", "terminado") and not self.fila_ids.filtered(
            lambda f: f.estado == "pendiente"
        ):
            self.estado = "terminado"
            if self._masivo_can_commit():
                self.env.cr.commit()
        res = self.l10n_pe_ne_lote_detalle()
        res["filas"] = [f._l10n_pe_ne_fila_dict() for f in procesadas.sorted("secuencia")]
        return res

    def l10n_pe_ne_reintentar(self):
        """Filas 'error'/'rechazado' → 'pendiente' CONSERVANDO move_id (reenvía el mismo move con
        la misma serie-correlativo). Nunca limpia move_id."""
        self.ensure_one()
        reenc = self.fila_ids.filtered(lambda f: f.estado in ("error", "rechazado"))
        reenc.write({"estado": "pendiente"})
        if reenc:
            self.estado = "en_proceso"
        res = self.l10n_pe_ne_lote_detalle()
        res["reencoladas"] = len(reenc)
        return res

    def l10n_pe_ne_cancelar(self):
        """Filas 'pendiente' → 'cancelado'; el lote → 'cancelado'. Lo ya emitido queda emitido."""
        self.ensure_one()
        self.fila_ids.filtered(lambda f: f.estado == "pendiente").write({"estado": "cancelado"})
        self.estado = "cancelado"
        return self.l10n_pe_ne_lote_detalle()

    def l10n_pe_ne_resultados(self):
        """xlsx de resultados (mismo estilo que l10n_pe_ne_export). {filename, count, contentB64}."""
        self.ensure_one()
        import xlsxwriter
        if self.tipo == "guia":
            headers = ["Guía", "Filas Excel", "Tipo", "Serie", "Número", "Destinatario",
                       "Peso (kg)", "Unidad", "Estado", "Mensaje SUNAT"]
        else:
            headers = ["Venta", "Filas Excel", "Tipo", "Serie", "Número", "Cliente", "Total",
                       "Moneda", "Estado", "Mensaje SUNAT"]
        estados = dict(self.fila_ids._fields["estado"].selection)
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet("Resultados")
        head = wb.add_format({"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1})
        for c, h in enumerate(headers):
            ws.write(0, c, h, head)
            ws.set_column(c, c, max(12, len(h) + 2))
        filas = self.fila_ids.sorted("secuencia")
        es_guia = self.tipo == "guia"
        for r, f in enumerate(filas, 1):
            # En guías la col. 7 es la unidad de peso (KGM) y no hay moneda; en comprobantes es
            # la moneda (default PEN). El resto de columnas coincide (total = peso bruto en guías).
            unidad_o_moneda = ("KGM" if es_guia else (f.moneda or "PEN"))
            ws.write_row(r, 0, ["#%s" % f.secuencia, f.filas_excel or "", f.tipo_doc or "",
                                f.serie or "", f.correlativo or "", f.cliente or "", f.total or 0.0,
                                unidad_o_moneda, estados.get(f.estado, f.estado), f.mensaje or ""])
        ws.autofilter(0, 0, max(1, len(filas)), len(headers) - 1)
        ws.freeze_panes(1, 0)
        wb.close()
        ruc = (self.env.company.vat or "").strip()
        return {"filename": "resultados-lote-%s-%s.xlsx" % (self.id, ruc),
                "count": len(filas),
                "contentB64": base64.b64encode(buf.getvalue()).decode("ascii")}


class L10nPeNeLoteFila(models.Model):
    _name = "l10n_pe_ne.lote.fila"
    _description = "Comprobante de un lote masivo (NE Express)"
    _order = "lote_id, secuencia"

    lote_id = fields.Many2one("l10n_pe_ne.lote", required=True, index=True, ondelete="cascade")
    company_id = fields.Many2one(related="lote_id.company_id", store=True, index=True)
    secuencia = fields.Integer(required=True)        # nº de comprobante dentro del lote (1..N)
    filas_excel = fields.Char()                      # "5-7": filas de origen en el xlsx
    payload_json = fields.Text(required=True)        # payload EXACTO para l10n_pe_ne_quick_emit
    estado = fields.Selection([
        ("error_validacion", "Error de validación"),
        ("pendiente", "Pendiente"),
        ("en_proceso", "En proceso"),  # modo async: encolado a SUNAT, resultado por cron
        ("emitido", "Emitido"),
        ("rechazado", "Rechazado"),   # SUNAT/biller rechazó (move existe)
        ("error", "Error"),           # error de conexión/infra (move puede existir)
        ("cancelado", "Cancelado"),
    ], default="pendiente", required=True)
    mensaje = fields.Text()
    move_id = fields.Many2one("account.move")        # ancla de idempotencia (lote comprobante)
    guia_id = fields.Many2one("l10n_pe_ne.guia_remision")  # ancla de idempotencia (lote guía)
    tipo_doc = fields.Char()
    serie = fields.Char()
    correlativo = fields.Char()
    cliente = fields.Char()
    total = fields.Float()
    moneda = fields.Char(default="PEN")

    # ------------------------------------------------------------- procesamiento de fila
    def _l10n_pe_ne_procesar_fila(self):
        """Emite (o reenvía) el comprobante de esta fila. Si ya hay move_id → reenvía el MISMO
        move (idempotente, misma serie-correlativo); si no → quick_emit crea+postea+envía. Mapea
        el biller_state a estado de fila. Sin lógica de emisión nueva. Lote de guías → delega en
        _l10n_pe_ne_procesar_fila_guia."""
        self.ensure_one()
        if self.lote_id.tipo == "guia":
            return self._l10n_pe_ne_procesar_fila_guia()
        if self.move_id:
            self.move_id.action_l10n_pe_send_to_biller()
            result = self.move_id.l10n_pe_ne_quick_result()
        else:
            result = self.env["account.move"].l10n_pe_ne_quick_emit(json.loads(self.payload_json))
            self.move_id = result["id"]
        self.write({
            "estado": _ESTADO_MAP.get(result.get("estado"), "error"),
            "mensaje": result.get("mensaje") or "",
            "tipo_doc": result.get("tipoDoc") or self.tipo_doc,
            "serie": result.get("serie") or self.serie,
            "correlativo": result.get("correlativo") or self.correlativo,
            "total": result.get("total") or self.total,
            "cliente": result.get("cliente") or self.cliente,
        })

    def _l10n_pe_ne_procesar_fila_guia(self):
        """Emite (o reenvía) la guía de esta fila reusando el modelo de guía SIN cambios: si ya
        hay guia_id → re-emite la MISMA guía (idempotente; ESTADOS_EMITIBLES deja reintentar
        error/rechazado y una ya 'enviado' se corta arriba); si no → quick_guia crea el borrador
        y guarda guia_id, luego emitir_guia. Mapea el estado de la guía a estado de fila."""
        self.ensure_one()
        Guia = self.env["l10n_pe_ne.guia_remision"].with_company(self.company_id)
        if self.guia_id:
            guia = self.guia_id
        else:
            d = Guia.l10n_pe_ne_quick_guia(json.loads(self.payload_json))
            guia = Guia.browse(d["id"])
            self.guia_id = guia.id
        result = guia.l10n_pe_ne_emitir_guia()
        self.write({
            "estado": _ESTADO_MAP_GUIA.get(result.get("estado"), "error"),
            "mensaje": result.get("mensaje") or "",
            "tipo_doc": "09",
            "serie": guia.serie or self.serie,
            "correlativo": guia.correlativo or self.correlativo,
            "cliente": result.get("destinatario") or self.cliente,
            "total": guia.peso_bruto or self.total,
        })

    def _l10n_pe_ne_fila_dict(self):
        self.ensure_one()
        return {"secuencia": self.secuencia, "filasExcel": self.filas_excel or "",
                "estado": self.estado, "tipoDoc": self.tipo_doc or "", "serie": self.serie or "",
                "correlativo": self.correlativo or "", "cliente": self.cliente or "",
                "total": self.total or 0.0, "moneda": self.moneda or "PEN",
                "mensaje": self.mensaje or "", "moveId": self.move_id.id or False,
                "guiaId": self.guia_id.id or False}
