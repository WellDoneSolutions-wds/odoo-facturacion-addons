# -*- coding: utf-8 -*-
"""Emisión masiva de GUÍAS de remisión (lote tipo 'guia').

Cubre el slice MVP: GRE remitente (09), motivo 01 (Venta), modalidad 02 (transporte
privado), un destinatario, un vehículo + un conductor, uno o más bienes agrupados por
'guia'. NO se golpea SUNAT: se moquea l10n_pe_ne_emitir_guia (la emisión real de la guía
está probada en test_guia); aquí se prueba el pegamento del lote (plantilla, parseo,
validación, procesar-fila, reconciliación async) y la regresión del camino de comprobantes.
"""
import base64
import io
import json
from unittest.mock import patch

import xlsxwriter

from odoo.tests import TransactionCase, tagged

from odoo.addons.l10n_pe_ne_biller.models.l10n_pe_ne_lote import _HEADERS_GUIA

# Cabeceras del lote de COMPROBANTES (para la regresión del camino existente).
_HEADERS_COMP = ["venta", "tipo", "serie", "fecha", "tipo doc cliente", "num doc cliente",
                 "cliente", "codigo producto", "producto", "cantidad", "precio unitario",
                 "descuento %", "afectacion", "bolsa", "moneda"]
# Ruta de la clase de la guía para moquear la emisión real (no golpear SUNAT).
_EMITIR = ("odoo.addons.l10n_pe_ne_biller.models.l10n_pe_ne_guia_remision."
           "L10nPeNeGuiaRemision.l10n_pe_ne_emitir_guia")

# Fila de guía válida (un bien). Orden = _HEADERS_GUIA.
_G1 = ["G-001", "20601030013", "COMERCIAL LOS ANDES SAC", "01", "150101", "AV. UNO 100",
       "150203", "AV. DOS 200", 25.5, "ABC123", "DNI", "45678912", "JUAN CARLOS",
       "PEREZ QUISPE", "Q12345678", "TORNILLOS 1/2", 10, "NIU"]
# Segundo bien de la MISMA guía (cabecera vacía; el parser la toma de la primera fila).
_G1_BIEN2 = ["G-001", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "CLAVOS 3PULG", 5, "KGM"]


def _xlsx_b64(rows, headers):
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet("Guias" if headers is _HEADERS_GUIA else "Ventas")
    for c, h in enumerate(headers):
        ws.write(0, c, h)
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row):
            ws.write(r, c, val)
    wb.close()
    return base64.b64encode(buf.getvalue()).decode("ascii")


def _fake_emitir(estado="enviado", mensaje="Aceptada por SUNAT — CDR ResponseCode 0"):
    """Devuelve un reemplazo de l10n_pe_ne_emitir_guia que fija el estado sin red."""
    def _emitir(self):
        self.estado = estado
        self.l10n_pe_biller_message = mensaje
        return self._l10n_pe_ne_guia_dict()
    return _emitir


@tagged("post_install", "-at_install")
class TestLoteGuia(TransactionCase):
    def setUp(self):
        super().setUp()
        # Compañía propia: da series T### vírgenes y un RUC de emisor distinto del
        # destinatario (la validación SUNAT 2555 exige destinatario != emisor).
        self.company = self.env["res.company"].with_context(
            l10n_pe_ne_allow_company_create=True).create(
            {"name": "TRANSPORTES MVP SAC", "vat": "20111111117"})
        self.Lote = self.env["l10n_pe_ne.lote"].with_company(self.company)
        self.Guia = self.env["l10n_pe_ne.guia_remision"].with_company(self.company)

    def _crear(self, rows, tipo="guia", headers=_HEADERS_GUIA):
        return self.Lote.l10n_pe_ne_crear_lote(
            {"filename": "g.xlsx", "contentB64": _xlsx_b64(rows, headers), "tipo": tipo})

    # ------------------------------------------------------------------ plantilla
    def test_plantilla_guia_descargable(self):
        import openpyxl
        out = self.Lote.l10n_pe_ne_plantilla("guia")
        self.assertTrue(out["filename"].endswith(".xlsx"))
        self.assertIn("guias", out["filename"])
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(out["contentB64"])))
        self.assertIn("Guias", wb.sheetnames)
        self.assertIn("Instrucciones", wb.sheetnames)
        ws = wb["Guias"]
        self.assertEqual([ws.cell(1, c + 1).value for c in range(len(_HEADERS_GUIA))], _HEADERS_GUIA)
        # La plantilla de comprobantes sigue intacta (default y explícita).
        self.assertIn("ventas", self.Lote.l10n_pe_ne_plantilla()["filename"])
        self.assertIn("ventas", self.Lote.l10n_pe_ne_plantilla("comprobante")["filename"])

    # ------------------------------------------------------- parseo + validación
    def test_crear_lote_guia_valida(self):
        rep = self._crear([list(_G1), list(_G1_BIEN2)])
        self.assertEqual(rep["estado"], "validado")
        self.assertEqual(rep["tipo"], "guia")
        self.assertEqual(rep["errores"], [])
        self.assertEqual(rep["totalFilas"], 2)
        self.assertEqual(rep["totalComprobantes"], 1, "los 2 bienes se agrupan en 1 guía")
        lote = self.Lote.browse(rep["id"])
        self.assertEqual(lote.tipo, "guia")
        fila = lote.fila_ids
        self.assertEqual(len(fila), 1)
        self.assertEqual(fila.tipo_doc, "09")
        self.assertEqual(fila.cliente, "COMERCIAL LOS ANDES SAC")
        p = json.loads(fila.payload_json)
        self.assertEqual(p["tipoGre"], "09")
        self.assertEqual(p["modalidadTraslado"], "02")
        self.assertEqual(p["motivoTraslado"], "01")
        self.assertEqual(p["destinatario"], {"numDoc": "20601030013", "tipoDoc": "6",
                                             "razonSocial": "COMERCIAL LOS ANDES SAC"})
        self.assertEqual(p["ubigeoPartida"], "150101")
        self.assertEqual(p["ubigeoLlegada"], "150203")
        self.assertEqual(p["numPlaca"], "ABC123")
        self.assertEqual(p["conductorNumDoc"], "45678912")
        self.assertEqual(p["conductorTipoDoc"], "1")
        self.assertEqual(p["pesoBruto"], 25.5)
        self.assertEqual([b["descripcion"] for b in p["items"]], ["TORNILLOS 1/2", "CLAVOS 3PULG"])
        self.assertEqual([b["unidad"] for b in p["items"]], ["NIU", "KGM"])
        # El detalle expone el tipo para que la SPA rotule columnas/errores.
        self.assertEqual(lote.l10n_pe_ne_lote_detalle()["tipo"], "guia")

    def test_guia_motivo_numerico_se_normaliza(self):
        # Excel suele traer el motivo como número sin cero a la izquierda (1): debe
        # normalizarse a "01" y validar, no rechazarse con "solo 01 soportado".
        row = list(_G1); row[3] = 1
        rep = self._crear([row])
        self.assertEqual(rep["estado"], "validado", rep.get("errores"))
        p = json.loads(self.Lote.browse(rep["id"]).fila_ids.payload_json)
        self.assertEqual(p["motivoTraslado"], "01")

    def test_guia_unidad_se_normaliza_a_mayuscula(self):
        row = list(_G1); row[17] = "niu"
        rep = self._crear([row])
        self.assertEqual(rep["estado"], "validado", rep.get("errores"))
        p = json.loads(self.Lote.browse(rep["id"]).fila_ids.payload_json)
        self.assertEqual(p["items"][0]["unidad"], "NIU")

    def test_guia_fila_error_validacion(self):
        # Sin destinatario doc → error de validación en esa guía; lote con_errores, sin filas.
        mal = list(_G1); mal[1] = ""
        rep = self._crear([mal])
        self.assertEqual(rep["estado"], "con_errores")
        self.assertTrue(any("destinatario" in e["mensaje"].lower() for e in rep["errores"]))
        self.assertFalse(self.Lote.browse(rep["id"]).fila_ids)

    def test_guia_validaciones_por_campo(self):
        casos = [
            (4, "150101X", "ubigeo"),      # ubigeo partida no 6 dígitos
            (9, "AB", "placa"),            # placa muy corta
            (11, "", "conductor"),         # falta num doc conductor
            (8, 0, "peso"),                # peso 0
        ]
        for col, val, aguja in casos:
            fila = list(_G1); fila[col] = val
            rep = self._crear([fila])
            self.assertEqual(rep["estado"], "con_errores", "col %s" % col)
            self.assertTrue(any(aguja in e["mensaje"].lower() for e in rep["errores"]),
                            "col %s → %s" % (col, [e["mensaje"] for e in rep["errores"]]))

    def test_guia_destinatario_no_puede_ser_emisor(self):
        mismo = list(_G1); mismo[1] = self.company.vat
        rep = self._crear([mismo])
        self.assertEqual(rep["estado"], "con_errores")
        self.assertTrue(any("propia empresa" in e["mensaje"].lower() for e in rep["errores"]))

    def test_guia_motivo_no_soportado_difiere(self):
        otro = list(_G1); otro[3] = "04"
        rep = self._crear([otro])
        self.assertEqual(rep["estado"], "con_errores")
        self.assertTrue(any("motivo" in e["mensaje"].lower() for e in rep["errores"]))

    # ------------------------------------------------------- procesar (emisión)
    def test_procesar_guia_crea_borrador_y_emite(self):
        lote = self.Lote.browse(self._crear([list(_G1), list(_G1_BIEN2)])["id"])
        with patch(_EMITIR, _fake_emitir()) as _mk:
            lote.l10n_pe_ne_procesar(max_filas=1)
        fila = lote.fila_ids
        self.assertEqual(fila.estado, "emitido")
        self.assertEqual(fila.tipo_doc, "09")
        self.assertTrue(fila.serie.startswith("T"))
        self.assertTrue(fila.guia_id, "la fila guarda la guía como ancla de idempotencia")
        g = fila.guia_id
        self.assertEqual(g.tipo_gre, "09")
        self.assertEqual(g.motivo_traslado, "01")
        self.assertEqual(g.modalidad_traslado, "02")
        self.assertEqual(g.partner_id.vat, "20601030013")
        self.assertEqual(g.ubigeo_partida, "150101")
        self.assertEqual(g.dir_partida, "AV. UNO 100")
        self.assertEqual(g.ubigeo_llegada, "150203")
        self.assertEqual(g.num_placa, "ABC123")
        self.assertEqual(g.conductor_num_doc, "45678912")
        self.assertEqual(g.conductor_nombres, "JUAN CARLOS")
        self.assertEqual(g.conductor_apellidos, "PEREZ QUISPE")
        self.assertEqual(g.conductor_licencia, "Q12345678")
        self.assertEqual(g.peso_bruto, 25.5)
        self.assertEqual(len(g.line_ids), 2)
        self.assertEqual(sorted(g.line_ids.mapped("descripcion")), ["CLAVOS 3PULG", "TORNILLOS 1/2"])
        self.assertEqual(lote.estado, "terminado")

    def test_procesar_guia_rechazo(self):
        lote = self.Lote.browse(self._crear([list(_G1)])["id"])
        with patch(_EMITIR, _fake_emitir(estado="rechazado", mensaje="XSLT 3364")):
            lote.l10n_pe_ne_procesar(max_filas=1)
        self.assertEqual(lote.fila_ids.estado, "rechazado")
        self.assertIn("3364", lote.fila_ids.mensaje)

    def test_procesar_guia_async_reconcilia(self):
        # emitir deja la guía 'en_proceso' (ticket pendiente) → fila 'en_proceso'; cuando el cron
        # de guías la acepta, releer el detalle reconcilia la fila a 'emitido' (sin cron del lote).
        lote = self.Lote.browse(self._crear([list(_G1)])["id"])
        with patch(_EMITIR, _fake_emitir(estado="en_proceso", mensaje="ticket pendiente")):
            lote.l10n_pe_ne_procesar(max_filas=1)
        fila = lote.fila_ids
        self.assertEqual(fila.estado, "en_proceso")
        self.assertTrue(fila.guia_id)
        det = lote.l10n_pe_ne_lote_detalle()
        self.assertEqual(det["enProceso"], 1)
        fila.guia_id.write({"estado": "enviado", "l10n_pe_biller_message": "Aceptada"})
        det = lote.l10n_pe_ne_lote_detalle()
        self.assertEqual(fila.estado, "emitido")
        self.assertEqual(det["emitidos"], 1)
        self.assertIn("Aceptada", fila.mensaje)

    def test_procesar_guia_reintento_reusa_guia(self):
        lote = self.Lote.browse(self._crear([list(_G1)])["id"])
        with patch(_EMITIR, _fake_emitir(estado="error", mensaje="conexión caída")):
            lote.l10n_pe_ne_procesar(max_filas=1)
        fila = lote.fila_ids
        self.assertEqual(fila.estado, "error")
        guia_id = fila.guia_id.id
        self.assertTrue(guia_id)
        n0 = self.Guia.search_count([])
        lote.l10n_pe_ne_reintentar()
        self.assertEqual(fila.estado, "pendiente")
        self.assertEqual(fila.guia_id.id, guia_id, "reintentar conserva la guía")
        with patch(_EMITIR, _fake_emitir()):
            lote.l10n_pe_ne_procesar(max_filas=1)
        self.assertEqual(fila.estado, "emitido")
        self.assertEqual(fila.guia_id.id, guia_id, "reusa la MISMA guía (no crea otra)")
        self.assertEqual(self.Guia.search_count([]), n0, "reintentar+procesar no crea guía nueva")

    # ------------------------------------------------------------- regresión
    def test_comprobante_sigue_funcionando(self):
        # El camino de comprobantes debe seguir idéntico: tipo default 'comprobante'.
        boleta = ["", "BOLETA", "", "", "", "", "", "", "AGUA 625ML", 1, 1.50, 0, "GRAVADO", "NO", "PEN"]
        rep = self.Lote.l10n_pe_ne_crear_lote(
            {"filename": "v.xlsx", "contentB64": _xlsx_b64([boleta], _HEADERS_COMP)})
        self.assertEqual(rep["estado"], "validado")
        self.assertEqual(rep["tipo"], "comprobante")
        self.assertEqual(rep["totalComprobantes"], 1)
        lote = self.Lote.browse(rep["id"])
        self.assertEqual(lote.tipo, "comprobante")
        self.assertEqual(lote.fila_ids.tipo_doc, "03")
        self.assertFalse(lote.fila_ids.guia_id)
