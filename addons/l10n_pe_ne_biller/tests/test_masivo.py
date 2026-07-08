# -*- coding: utf-8 -*-
import base64
import io
import json
from datetime import timedelta

import xlsxwriter
from unittest.mock import patch, MagicMock

from odoo import fields
from odoo.exceptions import AccessError, UserError
from odoo.tests import HttpCase, TransactionCase, tagged

# Patch donde se USA requests.post (dentro del módulo del modelo base), igual que test_send.py.
_TARGET = "odoo.addons.l10n_pe_ne_biller.models.account_move_biller.requests.post"
_HEADERS = ["venta", "tipo", "serie", "fecha", "tipo doc cliente", "num doc cliente",
            "cliente", "codigo producto", "producto", "cantidad", "precio unitario",
            "descuento %", "afectacion", "bolsa", "moneda"]
# XML "firmado" mínimo que action_l10n_pe_send_to_biller reconoce como aceptado (contiene <Invoice).
_SIGNED = '<?xml version="1.0"?><Invoice xmlns="urn:x"><ext:UBLExtensions/></Invoice>'


def _xlsx_b64(rows, headers=_HEADERS):
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet("Ventas")
    for c, h in enumerate(headers):
        ws.write(0, c, h)
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row):
            ws.write(r, c, val)
    wb.close()
    return base64.b64encode(buf.getvalue()).decode("ascii")


def _resp(code, text, headers=None):
    return type("R", (), {"status_code": code, "text": text, "headers": headers or {}})()


@tagged("post_install", "-at_install")
class TestMasivo(TransactionCase):
    def setUp(self):
        super().setUp()
        self.Lote = self.env["l10n_pe_ne.lote"]
        self.Fila = self.env["l10n_pe_ne.lote.fila"]
        self.company_b = self.env["res.company"].create(
            {"name": "OTRO RUC SAC", "vat": "20999999991"})
        grp = self.env.ref("l10n_pe_ne_biller.group_l10n_pe_ne_emisor")
        self.user_b = self.env["res.users"].create({
            "name": "Emisor B", "login": "emisor_b_masivo",
            "company_id": self.company_b.id, "company_ids": [(6, 0, [self.company_b.id])],
            "group_ids": [(4, grp.id)]})

    def _lote_simple(self, estado="validado"):
        lote = self.Lote.create(
            {"name": "x.xlsx", "estado": estado, "total_comprobantes": 1, "total_filas": 1})
        self.Fila.create(
            {"lote_id": lote.id, "secuencia": 1, "payload_json": "{}", "estado": "pendiente"})
        return lote

    def test_defaults_y_aislamiento(self):
        lote = self._lote_simple()
        self.assertEqual(lote.estado, "validado")
        self.assertEqual(lote.company_id, self.env.company)
        self.assertEqual(lote.fila_ids.company_id, self.env.company,
                         "la fila hereda company_id del lote (related store)")
        # Compañía B no ve el lote ni la fila (regla ir.rule global).
        self.assertFalse(self.Lote.with_user(self.user_b).search([("id", "=", lote.id)]))
        self.assertFalse(self.Fila.with_user(self.user_b).search([("lote_id", "=", lote.id)]))
        with self.assertRaises(AccessError):
            lote.with_user(self.user_b).read(["estado"])
        with self.assertRaises(AccessError):
            lote.fila_ids.with_user(self.user_b).read(["estado"])

    def test_masivo_param_default_y_override(self):
        lote = self._lote_simple()
        self.assertEqual(lote._masivo_param("masivo_max_filas", 500), 500)
        self.env["ir.config_parameter"].sudo().set_param("l10n_pe_ne.masivo_max_filas", "10")
        self.assertEqual(lote._masivo_param("masivo_max_filas", 500), 10)
        self.env["ir.config_parameter"].sudo().set_param("l10n_pe_ne.masivo_max_filas", "0")
        self.assertEqual(lote._masivo_param("masivo_max_filas", 500), 500, "valor <=0 cae al default")

    # -------------------------------------------------- QW10 Task 2: parse + validación
    def test_plantilla_descargable(self):
        out = self.Lote.l10n_pe_ne_plantilla()
        self.assertTrue(out["filename"].endswith(".xlsx"))
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(out["contentB64"])))
        self.assertIn("Ventas", wb.sheetnames)
        self.assertIn("Instrucciones", wb.sheetnames)
        ws = wb["Ventas"]
        self.assertEqual([ws.cell(1, c + 1).value for c in range(15)], _HEADERS)

    def test_parse_y_agrupado(self):
        b64 = _xlsx_b64([
            ["V-1", "FACTURA", "F001", "01/07/2026", "RUC", "20100070970", "FERRETERIA LA UNION SAC", "CEM-01", "CEMENTO", 2, 33.90, 0, "GRAVADO", "NO", "PEN"],
            ["V-1", "FACTURA", "F001", "01/07/2026", "RUC", "20100070970", "FERRETERIA LA UNION SAC", "CLV-02", "CLAVOS", 5, 4.50, 10, "GRAVADO", "NO", "PEN"],
            ["", "BOLETA", "B001", "01/07/2026", "DNI", "45678912", "ROSA QUISPE", "", "PINTURA", 1, 45.00, 0, "GRAVADO", "SI", "PEN"],
        ])
        rep = self.Lote.l10n_pe_ne_crear_lote({"filename": "v.xlsx", "contentB64": b64})
        self.assertEqual(rep["estado"], "validado")
        self.assertEqual(rep["errores"], [])
        self.assertEqual(rep["totalFilas"], 3)
        self.assertEqual(rep["totalComprobantes"], 2)
        lote = self.Lote.browse(rep["id"])
        f1 = lote.fila_ids.filtered(lambda f: f.secuencia == 1)
        p1 = json.loads(f1.payload_json)
        self.assertEqual(p1["tipoDoc"], "01")
        self.assertEqual(p1["serie"], "F001")
        self.assertEqual(p1["cliente"], {"tipoDoc": "6", "numDoc": "20100070970", "razonSocial": "FERRETERIA LA UNION SAC"})
        self.assertEqual(len(p1["lineas"]), 2)
        self.assertEqual(p1["lineas"][0]["taxCode"], "1000")
        self.assertEqual(p1["lineas"][0]["productCod"], "CEM-01")
        self.assertEqual(f1.filas_excel, "2-3")
        p2 = json.loads(lote.fila_ids.filtered(lambda f: f.secuencia == 2).payload_json)
        self.assertEqual(p2["tipoDoc"], "03")
        self.assertEqual(p2["cliente"]["tipoDoc"], "1")
        self.assertTrue(p2["lineas"][0]["icbper"])

    def test_validacion_reporte_por_fila(self):
        b64 = _xlsx_b64([
            ["", "FACTURA", "F001", "", "RUC", "20100070971", "MAL RUC SAC", "", "ITEM", 1, 10.0, 0, "GRAVADO", "NO", "PEN"],   # fila 2: RUC dígito malo
            ["", "BOLETA", "", "", "", "", "", "", "ITEM", 0, 5.0, 0, "GRAVADO", "NO", "PEN"],                                    # fila 3: cantidad 0
            ["", "FACTURA", "B001", "", "RUC", "20100070970", "OK SAC", "", "ITEM", 1, 10.0, 0, "GRAVADO", "NO", "PEN"],          # fila 4: serie B en factura
        ])
        rep = self.Lote.l10n_pe_ne_crear_lote({"filename": "v.xlsx", "contentB64": b64})
        self.assertEqual(rep["estado"], "con_errores")
        filas_err = {e["filaExcel"] for e in rep["errores"]}
        self.assertTrue({2, 3, 4} <= filas_err)
        msg2 = next(e["mensaje"] for e in rep["errores"] if e["filaExcel"] == 2)
        self.assertIn("dígito verificador", msg2)
        lote = self.Lote.browse(rep["id"])
        self.assertFalse(lote.fila_ids, "un lote con errores no crea filas procesables")

    def test_limites_y_duplicado(self):
        self.env["ir.config_parameter"].sudo().set_param("l10n_pe_ne.masivo_max_filas", "2")
        big = _xlsx_b64([["", "BOLETA", "", "", "", "", "", "", "IT", 1, 1.0, 0, "GRAVADO", "NO", "PEN"]] * 3)
        with self.assertRaises(UserError):
            self.Lote.l10n_pe_ne_crear_lote({"filename": "big.xlsx", "contentB64": big})
        self.env["ir.config_parameter"].sudo().set_param("l10n_pe_ne.masivo_max_filas", "500")
        b64 = _xlsx_b64([["", "BOLETA", "", "", "", "", "", "", "AGUA", 1, 1.5, 0, "GRAVADO", "NO", "PEN"]])
        r1 = self.Lote.l10n_pe_ne_crear_lote({"filename": "a.xlsx", "contentB64": b64})
        r2 = self.Lote.l10n_pe_ne_crear_lote({"filename": "a2.xlsx", "contentB64": b64})
        self.assertIsNone(r1["duplicadoDe"])
        self.assertEqual(r2["duplicadoDe"], r1["id"])

    def test_validacion_reporte_por_fila_procesar_raises(self):
        b64 = _xlsx_b64([
            ["", "FACTURA", "F001", "", "RUC", "20100070971", "MAL RUC SAC", "", "ITEM", 1, 10.0, 0, "GRAVADO", "NO", "PEN"],
        ])
        rep = self.Lote.l10n_pe_ne_crear_lote({"filename": "v.xlsx", "contentB64": b64})
        self.assertEqual(rep["estado"], "con_errores")
        lote = self.Lote.browse(rep["id"])
        with self.assertRaises(UserError):
            lote.l10n_pe_ne_procesar()

    # -------------------------------------------------- QW10 Task 3: proceso del lote
    def _crear(self, rows):
        rep = self.Lote.l10n_pe_ne_crear_lote({"filename": "v.xlsx", "contentB64": _xlsx_b64(rows)})
        return self.Lote.browse(rep["id"])

    _BOLETA = ["", "BOLETA", "", "", "", "", "", "", "AGUA 625ML", 1, 1.50, 0, "GRAVADO", "NO", "PEN"]

    def test_procesar_emite_secuencial(self):
        lote = self._crear([list(self._BOLETA), list(self._BOLETA), list(self._BOLETA)])
        self.assertEqual(lote.estado, "validado")
        with patch(_TARGET, return_value=_resp(200, _SIGNED)) as mp:
            for _ in range(5):
                if lote.estado == "terminado":
                    break
                lote.l10n_pe_ne_procesar(max_filas=1)
        self.assertEqual(lote.estado, "terminado")
        self.assertEqual([f.estado for f in lote.fila_ids.sorted("secuencia")], ["emitido"] * 3)
        self.assertTrue(all(f.move_id for f in lote.fila_ids))
        self.assertEqual(mp.call_count, 3)
        corr = [f.move_id.l10n_pe_ne_corr_emit for f in lote.fila_ids.sorted("secuencia")]
        self.assertEqual(len(set(corr)), 3, "correlativos consecutivos y únicos")

    def test_rechazo_no_detiene(self):
        lote = self._crear([list(self._BOLETA), list(self._BOLETA), list(self._BOLETA)])
        resps = [_resp(200, _SIGNED), _resp(400, "XSLT error 2017"), _resp(200, _SIGNED)]
        with patch(_TARGET, side_effect=resps):
            for _ in range(3):
                lote.l10n_pe_ne_procesar(max_filas=1)
        filas = lote.fila_ids.sorted("secuencia")
        self.assertEqual([f.estado for f in filas], ["emitido", "rechazado", "emitido"])
        self.assertIn("2017", filas[1].mensaje)
        self.assertEqual(lote.estado, "terminado")

    def test_procesar_idempotente(self):
        lote = self._crear([list(self._BOLETA)])
        with patch(_TARGET, return_value=_resp(200, _SIGNED)) as mp:
            lote.l10n_pe_ne_procesar(max_filas=1)
            self.assertEqual(lote.estado, "terminado")
            lote.l10n_pe_ne_procesar(max_filas=1)   # tras terminar: no-op idempotente
        self.assertEqual(mp.call_count, 1)
        self.assertEqual(lote.fila_ids.estado, "emitido")

    def test_async_en_proceso_y_reconciliacion(self):
        """Modo async (SQS): al emitir, el biller encola y el move queda 'en_proceso'. La fila NO
        debe caer a 'error' (bug previo: _ESTADO_MAP no conocía 'en_proceso') sino quedar
        'en_proceso'; y cuando el cron resuelve el move, releer el detalle reconcilia la fila con
        el estado final del comprobante."""
        icp = self.env["ir.config_parameter"].sudo()
        icp.set_param("l10n_pe_ne_biller.async_enabled", "1")
        icp.set_param("l10n_pe_ne_biller.sqs_queue_url", "https://sqs.fake/q")
        icp.set_param("l10n_pe_ne_biller.results_table", "")   # sin DynamoDB (skip delete_item)
        lote = self._crear([list(self._BOLETA)])
        with patch("odoo.addons.l10n_pe_ne_biller.models.account_move_biller.boto3") as mb:
            mb.client.return_value = MagicMock()   # sqs.send_message = no-op
            lote.l10n_pe_ne_procesar(max_filas=1)
        fila = lote.fila_ids
        # 1) encolado → fila 'en_proceso' (NO 'error'); el move quedó 'en_proceso'
        self.assertEqual(fila.estado, "en_proceso")
        self.assertTrue(fila.move_id)
        self.assertEqual(fila.move_id.l10n_pe_biller_state, "en_proceso")
        det = lote.l10n_pe_ne_lote_detalle()
        self.assertEqual(det["enProceso"], 1)
        self.assertEqual(det["errores"], 0)
        self.assertEqual(det["emitidos"], 0)
        # 2) el cron resuelve el move → releer el detalle reconcilia la fila a 'emitido'
        fila.move_id.write({
            "l10n_pe_biller_state": "enviado",
            "l10n_pe_biller_message": "Aceptado por SUNAT — CDR ResponseCode 0",
        })
        det = lote.l10n_pe_ne_lote_detalle()
        self.assertEqual(fila.estado, "emitido")
        self.assertEqual(det["enProceso"], 0)
        self.assertEqual(det["emitidos"], 1)
        self.assertIn("Aceptado", fila.mensaje)

    def test_procesar_chunk_multiple(self):
        """QW10 Task 3 review (cobertura): con max_filas=3 (<= masivo_max_chunk=5) el chunk
        procesa las 3 filas EN UNA sola llamada. El mock devuelve [200, 400, 200]: la fila 2
        queda 'rechazado' pero el for-loop sigue con la fila 3 dentro de la MISMA llamada (no
        solo entre llamadas top-level, como ya cubre test_rechazo_no_detiene)."""
        lote = self._crear([list(self._BOLETA), list(self._BOLETA), list(self._BOLETA)])
        resps = [_resp(200, _SIGNED), _resp(400, "XSLT error 2017"), _resp(200, _SIGNED)]
        with patch(_TARGET, side_effect=resps) as mp:
            lote.l10n_pe_ne_procesar(max_filas=3)
        self.assertEqual(mp.call_count, 3, "las 3 filas se procesaron en UNA sola llamada")
        filas = lote.fila_ids.sorted("secuencia")
        self.assertEqual([f.estado for f in filas], ["emitido", "rechazado", "emitido"])
        self.assertIn("2017", filas[1].mensaje)
        self.assertEqual(lote.estado, "terminado")

    def test_reintento_reusa_move(self):
        import requests as _rq
        lote = self._crear([list(self._BOLETA)])
        with patch(_TARGET, side_effect=_rq.RequestException("conexión caída")):
            lote.l10n_pe_ne_procesar(max_filas=1)
        fila = lote.fila_ids
        self.assertEqual(fila.estado, "error")
        self.assertTrue(fila.move_id, "el move persiste (send_to_biller no relanza RequestException)")
        move_id = fila.move_id.id
        n0 = self.env["account.move"].search_count([])
        lote.l10n_pe_ne_reintentar()
        self.assertEqual(fila.estado, "pendiente")
        self.assertEqual(fila.move_id.id, move_id, "reintentar NO limpia move_id")
        with patch(_TARGET, return_value=_resp(200, _SIGNED)) as mp:
            lote.l10n_pe_ne_procesar(max_filas=1)
        self.assertEqual(fila.estado, "emitido")
        self.assertEqual(fila.move_id.id, move_id, "reusa el MISMO move (misma serie-correlativo)")
        mp.assert_called_once()
        self.assertEqual(self.env["account.move"].search_count([("id", "=", move_id)]), 1)
        self.assertEqual(self.env["account.move"].search_count([]), n0,
                         "reintentar+procesar no crea un move nuevo (defensa en profundidad)")

    def test_cancelar(self):
        lote = self._crear([list(self._BOLETA), list(self._BOLETA)])
        with patch(_TARGET, return_value=_resp(200, _SIGNED)):
            lote.l10n_pe_ne_procesar(max_filas=1)
        lote.l10n_pe_ne_cancelar()
        self.assertEqual([f.estado for f in lote.fila_ids.sorted("secuencia")], ["emitido", "cancelado"])
        self.assertEqual(lote.estado, "cancelado")
        # QW10 Task 3 review (Important): la fila 2 quedó 'cancelado' (no 'pendiente'), así que
        # sin el guard de estado terminal el bloque final "sin pendientes -> terminado" pisaba
        # el 'cancelado' con 'terminado' en un procesar() de más (doble submit/poll tardío).
        n0 = self.env["account.move"].search_count([])
        with patch(_TARGET, return_value=_resp(200, _SIGNED)) as mp:
            lote.l10n_pe_ne_procesar()
        self.assertEqual(lote.estado, "cancelado", "procesar() no debe voltear cancelado -> terminado")
        mp.assert_not_called()
        self.assertEqual(self.env["account.move"].search_count([]), n0, "no se emite nada nuevo")

    def test_resultados_xlsx(self):
        import openpyxl
        lote = self._crear([list(self._BOLETA)])
        with patch(_TARGET, return_value=_resp(200, _SIGNED)):
            lote.l10n_pe_ne_procesar(max_filas=1)
        out = lote.l10n_pe_ne_resultados()
        self.assertEqual(out["count"], 1)
        self.assertTrue(out["filename"].endswith(".xlsx"))
        ws = openpyxl.load_workbook(io.BytesIO(base64.b64decode(out["contentB64"]))).active
        self.assertEqual(ws.cell(1, 1).value, "Venta")
        self.assertEqual(ws.max_row, 2)   # cabecera + 1 fila


# -------------------------------------------------- QW10 Task 4: rutas HTTP /ne/api/lotes*
@tagged("post_install", "-at_install")
class TestMasivoHttp(HttpCase):
    def setUp(self):
        super().setUp()
        from odoo.addons.l10n_pe_ne_biller.controllers.main import _SCOPE
        self.company = self.env.company
        if not self.company.vat:
            self.company.vat = "20321856145"
        grp = self.env.ref("l10n_pe_ne_biller.group_l10n_pe_ne_emisor")
        self.user = self.env["res.users"].create({
            "name": "Emisor HTTP", "login": "emisor_http_masivo",
            "company_id": self.company.id, "company_ids": [(6, 0, [self.company.id])],
            "group_ids": [(4, grp.id)]})
        self.key = self.env["res.users.apikeys"].with_user(self.user)._generate(
            _SCOPE, "masivo-http", fields.Datetime.now() + timedelta(days=1))

    def _get(self, path):
        return self.url_open(path, headers={"Authorization": "Bearer %s" % self.key})

    def _post(self, path, payload):
        return self.url_open(path, data=json.dumps(payload), headers={
            "Authorization": "Bearer %s" % self.key, "Content-Type": "application/json"})

    def test_sin_token_401(self):
        self.assertEqual(self.url_open("/ne/api/lotes").status_code, 401)

    def test_list_lotes_vacio(self):
        r = self._get("/ne/api/lotes")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.json(), [])

    def test_plantilla_descarga(self):
        r = self._get("/ne/api/lotes/plantilla")
        self.assertEqual(r.status_code, 200)
        d = r.json()
        self.assertTrue(d["filename"].endswith(".xlsx"))
        self.assertTrue(d["contentB64"])

    def test_lote_inexistente_404(self):
        self.assertEqual(self._get("/ne/api/lotes/999999").status_code, 404)

    def test_crear_lote_valida(self):
        b64 = _xlsx_b64([["", "BOLETA", "", "", "", "", "", "", "AGUA 625ML", 2, 1.50, 0, "GRAVADO", "NO", "PEN"]])
        r = self._post("/ne/api/lotes", {"filename": "ventas.xlsx", "contentB64": b64})
        self.assertEqual(r.status_code, 200)
        d = r.json()
        self.assertEqual(d["estado"], "validado")
        self.assertEqual(d["totalComprobantes"], 1)
        self.assertEqual(d["errores"], [])

    def test_aislamiento_403(self):
        company_b = self.env["res.company"].create({"name": "RUC B", "vat": "20999999991"})
        lote_b = self.env["l10n_pe_ne.lote"].with_company(company_b).create(
            {"name": "b.xlsx", "estado": "validado", "company_id": company_b.id})
        self.assertEqual(self._get("/ne/api/lotes/%s" % lote_b.id).status_code, 403)
