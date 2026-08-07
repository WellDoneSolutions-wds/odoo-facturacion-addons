# -*- coding: utf-8 -*-
"""account.move — Importación masiva de productos.
Extraído de account_move_biller.py (refactor sin cambio de comportamiento)."""
import base64
import io
import re

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from .account_move_biller import DEFAULT_UNIT_CODE, UNIDAD_IMPORT, AFECT_IMPORT, TIPO_IMPORT, _UNIDAD_CODES, DETRACCION_DESC


class AccountMove(models.Model):
    _inherit = "account.move"

    # ------------------------------------------------------- importación productos
    @api.model
    def l10n_pe_ne_plantilla_productos(self):
        """Plantilla xlsx para importar/actualizar el catálogo (hoja 'Productos' con las
        cabeceras + ejemplos + listas de unidad/afectación, y una hoja 'Instrucciones').
        Devuelve {filename, contentB64}. Mismo estilo visual que la plantilla de la masiva."""
        import io
        import base64
        import unicodedata
        import xlsxwriter
        from xlsxwriter.utility import xl_col_to_name

        # Plantilla COMPLETA: además del caso común (código/nombre/precio) trae categoría,
        # subcategoría, marca, control de stock, rastreo, existencia inicial/mínima y activo — para
        # migrar un catálogo entero de una. Solo CÓDIGO y NOMBRE son obligatorios; el resto es
        # opcional ("vacío = default al crear / mantener al actualizar"). El parser mapea por NOMBRE
        # de cabecera, así que el orden es libre y las plantillas viejas (10 columnas) siguen valiendo.
        headers = [
            "CÓDIGO", "CÓDIGO DE BARRAS", "NOMBRE",
            "CATEGORÍA", "SUBCATEGORÍA", "MARCA",
            "TIPO", "UNIDAD", "CONTROLA STOCK", "RASTREO",
            "STOCK INICIAL", "STOCK MÍNIMO",
            "COSTO", "PRECIO VENTA", "INCLUYE IGV",
            "AFECTACIÓN", "BOLSA", "DETRACCIÓN", "ACTIVO",
        ]
        col = {h: i for i, h in enumerate(headers)}
        ejemplos = [
            ["PROD0001", "7751234000018", "CEMENTO SOL 42.5 KG", "Bazar y hogar", "Ferretería básica", "Sol",
             "BIEN", "UNIDAD", "Sí", "Sin rastreo", 500, 50, 28.00, 33.90, "Sí", "GRAVADO", "NO", "", "Sí"],
            ["PROD0002", "7751234000025", "LECHE GLORIA 400 G", "Lácteos y huevos", "Leche", "Gloria",
             "BIEN", "UNIDAD", "Sí", "Por lote", 300, 24, 3.20, 4.50, "Sí", "GRAVADO", "NO", "", "Sí"],
            ["SERV0001", "", "SERVICIO DE CONSULTORÍA", "", "", "",
             "SERVICIO", "", "No", "Sin rastreo", "", "", "", 150.00, "Sí", "GRAVADO", "NO", "", "Sí"],
            ["PROD0004", "", "BOLSA PLÁSTICA", "Bazar y hogar", "Bolsas", "",
             "BIEN", "UNIDAD", "No", "Sin rastreo", "", "", 0.10, 0.50, "Sí", "GRAVADO", "SI", "", "Sí"],
        ]
        # Listas de REUSO del propio negocio para los desplegables de categoría/subcategoría/marca:
        # que el usuario elija lo que ya tiene (y no duplique 'Coca Cola' vs 'Coca-Cola'). El árbol se
        # siembra la primera vez si está vacío. Son sugerencias: también puede escribir una nueva.
        Cat = self.env["product.category"]
        root = Cat._l10n_pe_ne_root()
        deptos = [d.name for d in root.child_id]
        subcats = sorted({s.name for d in root.child_id for s in d.child_id})
        # Mapa categoría → sus subcategorías, para el desplegable DEPENDIENTE (la subcategoría
        # muestra solo las de la categoría elegida) vía rangos con nombre + INDIRECT.
        cat_subs = [(d.name, [s.name for s in d.child_id]) for d in root.child_id]
        marcas = [m.name for m in self.env["l10n_pe_ne.marca"].search([])]
        detras = ["%s · %s" % (c, d) for c, d in DETRACCION_DESC.items()]

        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet("Productos")
        head = wb.add_format({"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1})
        # Código de barras y detracción como TEXTO: sus valores empiezan con cero / son largos y
        # Excel los rompería (ceros perdidos o notación científica 7.75E+12).
        txtfmt = wb.add_format({"num_format": "@"})
        for c, h in enumerate(headers):
            ws.write(0, c, h, head)
            as_text = h in ("CÓDIGO DE BARRAS", "DETRACCIÓN")
            ws.set_column(c, c, max(13, len(h) + 3), txtfmt if as_text else None)
        for r, row in enumerate(ejemplos, 1):
            ws.write_row(r, 0, row)

        # Hoja OCULTA con las listas de reuso, referenciada por los desplegables dinámicos.
        wl = wb.add_worksheet("Listas")
        for cidx, (titulo, valores) in enumerate([
            ("CATEGORIAS", deptos), ("SUBCATEGORIAS", subcats), ("MARCAS", marcas),
            ("DETRACCIONES", detras)]):
            wl.write(0, cidx, titulo)
            for i, v in enumerate(valores, 1):
                wl.write(i, cidx, v)
        # Una columna por categoría (a partir de la F) con SUS subcategorías, y un rango con nombre
        # por categoría, para que la SUBCATEGORÍA se filtre por la CATEGORÍA elegida (INDIRECT).
        # El nombre de rango de Excel no admite espacios (sí acentos): espacio / '/' / '-' → '_';
        # el mismo reemplazo se hace en la fórmula, así ambos lados coinciden.
        def nombre_rango(cat):
            # Nombre de rango de Excel: sin acentos (á→a) y con espacio / '/' / '-' → '_', 1:1 con la
            # fórmula INDIRECT. Cualquier otro char no válido se descarta (esa categoría no filtra,
            # pero no rompe la generación). No colapsa separadores, para calzar exacto con la fórmula.
            s = unicodedata.normalize("NFKD", (cat or "").strip()).encode("ascii", "ignore").decode("ascii")
            for ch in (" ", "/", "-"):
                s = s.replace(ch, "_")
            s = "".join(c for c in s if c.isalnum() or c == "_")
            return "cat_" + s if s else ""
        usados = set()
        for j, (dname, subs) in enumerate(cat_subs):
            if not subs:
                continue
            ci = 5 + j  # columnas F, G, H, … (deja E libre como separador visual)
            L = xl_col_to_name(ci)
            wl.write(0, ci, dname)
            for i, s in enumerate(subs, 1):
                wl.write(i, ci, s)
            nm = nombre_rango(dname)
            if nm and nm not in usados:
                wb.define_name(nm, "=Listas!$%s$2:$%s$%d" % (L, L, len(subs) + 1))
                usados.add(nm)
        wl.hide()

        def rango(cidx, n):
            L = xl_col_to_name(cidx)
            return "=Listas!$%s$2:$%s$%d" % (L, L, n + 1)

        last = len(ejemplos) + 1000  # el desplegable aplica a un buen rango de filas por llenar

        def dv(nombre, opts):
            c = col[nombre]
            ws.data_validation(1, c, last, c, opts)

        # Comentarios de ayuda al pasar el mouse por la cabecera (el triangulito rojo).
        note = {"x_scale": 2.4, "y_scale": 2.0, "author": "Ekipu"}
        ws.write_comment(0, col["CÓDIGO DE BARRAS"], (
            "Opcional. El código de barras (EAN) del producto, para escanearlo en el POS. "
            "Vacío si no tiene. No puede repetirse entre productos."), note)
        ws.write_comment(0, col["CATEGORÍA"], (
            "Opcional. Departamento del catálogo (Abarrotes, Bebidas…). Elige uno de la lista "
            "o escribe uno nuevo: si no existe, se crea."), note)
        ws.write_comment(0, col["SUBCATEGORÍA"], (
            "Opcional. El desplegable muestra solo las subcategorías de la CATEGORÍA que elegiste "
            "en esta fila. Existente o nueva (se crea bajo esa categoría). Elige primero la CATEGORÍA."), note)
        ws.write_comment(0, col["MARCA"], (
            "Opcional. Marca comercial (Gloria, Sol…). Elige una existente o escribe una nueva; "
            "se reutiliza para no duplicar."), note)
        ws.write_comment(0, col["TIPO"], (
            "BIEN o SERVICIO. Un SERVICIO no lleva stock y va con unidad ZZ a SUNAT.\n"
            "Vacío = se deduce de la UNIDAD (SERVICIO/ZZ → servicio; el resto → bien)."), note)
        ws.write_comment(0, col["CONTROLA STOCK"], (
            "Sí/No. Sí = se le llevan existencias (inventario/kardex). Servicios: No.\n"
            "Vacío = No al crear."), note)
        ws.write_comment(0, col["RASTREO"], (
            "Solo si controla stock. 'Por lote' (alimentos/farmacia) o 'Por serie' (IMEI, uno por "
            "unidad). Vacío = Sin rastreo."), note)
        ws.write_comment(0, col["STOCK INICIAL"], (
            "Opcional y solo al CREAR: existencia con la que arranca el producto (si controla "
            "stock). Al actualizar se ignora — usa 'Ajustar stock'."), note)
        ws.write_comment(0, col["STOCK MÍNIMO"], (
            "Opcional. Umbral de reposición: la app avisa 'bajo mínimo' cuando el stock cae a este "
            "valor o menos."), note)
        ws.write_comment(0, col["COSTO"], "Opcional. Precio de compra referencial. NO afecta la facturación.", note)
        ws.write_comment(0, col["PRECIO VENTA"], "Precio de venta. Por defecto se entiende CON IGV incluido (ver INCLUYE IGV).", note)
        ws.write_comment(0, col["INCLUYE IGV"], (
            "Sí = el PRECIO VENTA ya incluye IGV (lo normal, lo que paga el cliente).\n"
            "No = es sin IGV; se le suma el 18% si la afectación es GRAVADO. Vacío = Sí."), note)
        ws.write_comment(0, col["AFECTACIÓN"], (
            "Afectación de IGV. GRAVADO = con IGV 18% (lo normal). EXONERADO/INAFECTO = sin IGV. "
            "EXPORTACION/GRATUITO = casos especiales. Vacío = GRAVADO."), note)
        ws.write_comment(0, col["BOLSA"], (
            "SI solo si es una BOLSA PLÁSTICA (cobra ICBPER por unidad al venderla). "
            "Para todo lo demás: NO o vacío."), note)
        ws.write_comment(0, col["DETRACCIÓN"], (
            "Opcional. Elige del desplegable el bien/servicio sujeto a detracción (catálogo 54 de "
            "SUNAT). Al importar se guarda solo el código (ej. 027). Vacío = no sujeto."), note)
        ws.write_comment(0, col["ACTIVO"], "Sí/No. No = archivado (no aparece en el catálogo ni al emitir). Vacío = Sí.", note)

        # ── Desplegables de valores fijos ──
        dv("TIPO", {"validate": "list", "source": ["BIEN", "SERVICIO"],
            "input_title": "Bien o servicio",
            "input_message": "SERVICIO no lleva stock (unidad ZZ a SUNAT). Vacío = se deduce de la UNIDAD."})
        dv("UNIDAD", {"validate": "list", "source": [
                "UNIDAD", "KILOGRAMO", "GRAMO", "LITRO", "GALON", "CAJA",
                "METRO", "METRO CUADRADO", "METRO CUBICO", "MILLAR", "DOCENA", "HORA", "DIA"],
            "input_title": "Unidad de medida",
            "input_message": "Elige de la lista o escribe el código SUNAT (NIU, KGM…). Vacío = UNIDAD."})
        dv("CONTROLA STOCK", {"validate": "list", "source": ["Sí", "No"], "error_type": "information",
            "input_title": "¿Controla stock?",
            "input_message": "Sí = se le llevan existencias. Servicios: No. Vacío = No."})
        dv("RASTREO", {"validate": "list", "source": ["Sin rastreo", "Por lote", "Por serie"],
            "error_type": "information", "input_title": "Rastreo",
            "input_message": "Solo si controla stock. Por lote / Por serie. Vacío = Sin rastreo."})
        dv("INCLUYE IGV", {"validate": "list", "source": ["Sí", "No"], "error_type": "information",
            "input_title": "¿El precio incluye IGV?",
            "input_message": "Sí = el precio ya trae IGV (lo normal). No = sin IGV (se suma 18% si es GRAVADO). Vacío = Sí."})
        dv("AFECTACIÓN", {"validate": "list", "source": [
                "GRAVADO", "EXONERADO", "INAFECTO", "EXPORTACION", "GRATUITO"],
            "input_title": "Afectación de IGV",
            "input_message": "GRAVADO = con IGV 18%. EXONERADO/INAFECTO = sin IGV. Vacío = GRAVADO.",
            "error_type": "information", "error_title": "Valor sugerido",
            "error_message": "Usa: GRAVADO, EXONERADO, INAFECTO, EXPORTACION o GRATUITO."})
        dv("BOLSA", {"validate": "list", "source": ["SI", "NO"],
            "input_title": "Bolsa plástica (ICBPER)",
            "input_message": "SI solo si es una bolsa plástica. Para el resto: NO o vacío."})
        dv("ACTIVO", {"validate": "list", "source": ["Sí", "No"], "error_type": "information",
            "input_title": "¿Activo?",
            "input_message": "No = archivado (no aparece). Vacío = Sí."})

        # ── Desplegables DINÁMICOS (reuso del catálogo del negocio; permiten escribir uno nuevo) ──
        pick = {"error_type": "information", "error_title": "Se creará",
                "error_message": "Si no está en la lista, se creará al importar."}
        if deptos:
            dv("CATEGORÍA", {"validate": "list", "source": rango(0, len(deptos)),
                "input_title": "Categoría", "input_message": "Elige una existente o escribe una nueva (se crea).", **pick})
        if subcats:
            # DEPENDIENTE: muestra solo las subcategorías de la CATEGORÍA de la misma fila. INDIRECT
            # arma el nombre de rango "cat_<Categoría con _>". Si la categoría está vacía o es nueva
            # (sin rango), no filtra pero deja escribir (no bloqueante) — la subcategoría igual se crea.
            catL = xl_col_to_name(col["CATEGORÍA"])
            # INDIRECT arma "cat_<Categoría>" con el MISMO saneo que nombre_rango: quita acentos
            # (á→a, ñ→n, ü→u) y pasa espacio / '/' / '-' → '_'. Así el nombre calza aunque la
            # categoría tenga acentos o espacios.
            cref = "$%s2" % catL
            subcat_src = (
                '=INDIRECT("cat_"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
                'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(%s,"á","a"),"é","e"),"í","i"),"ó","o"),'
                '"ú","u"),"ñ","n"),"ü","u")," ","_"),"/","_"),"-","_"))' % cref)
            dv("SUBCATEGORÍA", {"validate": "list", "source": subcat_src,
                "input_title": "Subcategoría", "error_type": "information", "error_title": "Se creará",
                "input_message": "Se filtra por la CATEGORÍA elegida. Elige una o escribe una nueva (se crea)."})
        if marcas:
            dv("MARCA", {"validate": "list", "source": rango(2, len(marcas)),
                "input_title": "Marca", "input_message": "Existente o nueva (se reutiliza/crea).", **pick})
        # DETRACCIÓN: catálogo 54 (código · descripción). Al importar se toma solo el código.
        dv("DETRACCIÓN", {"validate": "list", "source": rango(3, len(detras)),
            "error_type": "information", "input_title": "Detracción (cat. 54)",
            "input_message": "Elige el bien/servicio sujeto a detracción. Vacío = no sujeto. "
                             "Al importar se usa solo el código (los 3 dígitos)."})

        ws.freeze_panes(1, 0)
        wi = wb.add_worksheet("Instrucciones")
        wi.set_column(0, 0, 115)
        for r, line in enumerate([
            "Ekipu — Plantilla de importación de productos (completa)",
            "",
            "Solo CÓDIGO y NOMBRE son obligatorios. Todo lo demás es opcional.",
            "  • Al CREAR: una celda vacía toma el valor por defecto.",
            "  • Al ACTUALIZAR (el CÓDIGO ya existe): una celda vacía MANTIENE el valor actual (no lo borra).",
            "",
            "1. CÓDIGO es la clave: si ya existe se ACTUALIZA; si no, se CREA.",
            "2. CÓDIGO DE BARRAS: el EAN para escanear en el POS. No puede repetirse entre productos.",
            "3. CATEGORÍA / SUBCATEGORÍA: elige del desplegable o escribe una nueva (si no existe, se crea). La SUBCATEGORÍA se filtra por la CATEGORÍA elegida — elige primero la categoría.",
            "4. MARCA: igual; se reutiliza la existente (sin acento/mayúsculas no duplican) o se crea la nueva.",
            "5. TIPO: BIEN o SERVICIO. Un servicio no lleva stock y va con unidad ZZ a SUNAT. Vacío = se deduce de la UNIDAD.",
            "6. UNIDAD: el nombre (UNIDAD, KILOGRAMO, HORA…) o el código SUNAT (NIU, KGM…). Vacío = UNIDAD (NIU), o ZZ si es SERVICIO.",
            "7. CONTROLA STOCK: Sí = se le llevan existencias. RASTREO (Por lote / Por serie) solo aplica si controla stock.",
            "8. STOCK INICIAL: existencia con la que arranca — SOLO al crear y si controla stock. STOCK MÍNIMO: umbral de aviso de reposición.",
            "9. COSTO: precio de compra referencial (no factura). PRECIO VENTA: el precio de venta.",
            "10. INCLUYE IGV: Sí = el precio ya trae IGV (lo normal). No = sin IGV; se le suma 18% si la afectación es GRAVADO.",
            "11. AFECTACIÓN: GRAVADO (con IGV 18%, lo normal), EXONERADO/INAFECTO (sin IGV), EXPORTACION/GRATUITO. Vacío = GRAVADO.",
            "12. BOLSA = SI solo para bolsas plásticas (cobran ICBPER). DETRACCIÓN: elige del desplegable (catálogo 54); se guarda solo el código.",
            "13. ACTIVO = No archiva el producto (no aparece en el catálogo ni al emitir).",
            "14. Sube el archivo, revisa el resumen (nuevos / actualizados / errores) y recién ahí confirma.",
        ]):
            wi.write(r, 0, line)
        wb.close()
        return {"filename": "plantilla-productos-ekipu.xlsx",
                "contentB64": base64.b64encode(buf.getvalue()).decode("ascii")}

    @api.model
    def l10n_pe_ne_revisar_tipos(self, payload=None):
        """Propone reclasificar los productos que quedaron como SERVICIO por el default viejo.

        Hasta hace poco todo producto nacía con type='service' —estuviera bien o no—, así que
        un catálogo existente tiene tornillos declarados como servicios. Y un servicio no
        lleva stock en Odoo: mientras no se corrijan, esos productos no mueven inventario.

        PROPONE, no decide. La deducción usa la misma regla que la creación
        (_l10n_pe_ne_tipo_producto: ZZ → servicio, el resto → bien), pero acá puede
        equivocarse: el formulario trae NIU por defecto, así que una consultora que no lo
        cambió tiene servicios con NIU y saldrían propuestos como bienes. Por eso se devuelve
        la lista para que la revise un humano y se aplica solo lo que confirme —
        `l10n_pe_ne_aplicar_tipos` recibe los ids elegidos, no un "aplicar todo".

        No propone nada sobre `llevaStock`: llevar inventario es una decisión del negocio y
        no hay señal ninguna que la delate. Se activa producto por producto.
        """
        Product = self.env["product.product"]
        # Solo los 'service': un 'consu' ya fue clasificado (por el usuario o por la regla).
        sospechosos = Product.search(
            [("type", "=", "service"), ("company_id", "in", (False, self.env.company.id))],
            order="name",
        )
        propuestas = []
        for p in sospechosos:
            uni = p.l10n_pe_ne_unit_code or ""
            propuesto = self._l10n_pe_ne_tipo_producto(None, uni)
            if propuesto != "service":
                propuestas.append({
                    "id": p.id,
                    "descripcion": p.name or "",
                    "codigo": p.default_code or "",
                    # Sin unidad no significa "servicio": a SUNAT se le declara NIU por
                    # defecto (DEFAULT_UNIT_CODE), o sea un bien. Se muestra para que el
                    # usuario juzgue con el mismo dato que usó la regla.
                    "unidad": uni,
                    "tipoPropuesto": "bien",
                })
        return {
            "propuestas": propuestas,
            "total": len(propuestas),
            "revisados": len(sospechosos),
        }

    @api.model
    def l10n_pe_ne_aplicar_tipos(self, payload):
        """Aplica la reclasificación SOLO a los ids que el usuario confirmó.
        payload = {ids: [...], tipo: "bien"|"servicio"}."""
        payload = payload or {}
        ids = [int(i) for i in (payload.get("ids") or [])]
        if not ids:
            return {"actualizados": 0}
        tipo = self._l10n_pe_ne_tipo_producto(payload.get("tipo") or "bien")
        prods = self.env["product.product"].browse(ids).exists()
        prods.write({"type": tipo})
        return {"actualizados": len(prods)}

    @api.model
    def _l10n_pe_ne_leer_xlsx(self, contentB64, hoja):
        """Decodifica el base64 y devuelve las filas (lista de tuplas; rows[0] = cabecera)."""
        import io
        import base64
        import openpyxl
        try:
            data = base64.b64decode(contentB64 or "")
        except Exception:
            raise UserError(_("Archivo inválido."))
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception:
            raise UserError(_("No se pudo leer el archivo. Sube un .xlsx válido (no un .xls antiguo)."))
        ws = wb[hoja] if hoja in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise UserError(_("El archivo está vacío."))
        return rows

    @api.model
    def l10n_pe_ne_importar_productos(self, payload):
        """Compat single-shot: decodifica el archivo y procesa un tramo (offset/limit) o todo.
        El camino rápido por lotes es preparar()+lote() (parsea el archivo UNA sola vez). UPSERT
        por CÓDIGO; commit=False = dry-run; aislado por compañía."""
        payload = payload or {}
        rows = self._l10n_pe_ne_leer_xlsx(payload.get("contentB64"), "Productos")
        return self._l10n_pe_ne_importar_productos_core(
            rows, bool(payload.get("commit")), offset=int(payload.get("offset") or 0),
            window=int(payload.get("limit") or 0), budget=0.0)

    def _l10n_pe_ne_importar_productos_core(self, rows, commit, offset=0, window=0, budget=0.0):
        """Núcleo del import de productos sobre filas YA parseadas. `window` acota el prefetch
        (0 = todo el resto); `budget` en segundos corta el lote por tiempo (0 = sin corte). Devuelve
        el reporte + nextOffset/done para que el front continúe la importación por lotes."""
        import time
        import unicodedata
        t_ini = time.time()

        def norm(h):
            s = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode("ascii")
            return " ".join(s.lower().split())

        header = [norm(h) for h in rows[0]]
        idx = {h: i for i, h in enumerate(header) if h}
        faltan = [h for h in ("codigo", "nombre") if h not in idx]
        if faltan:
            raise UserError(_("Faltan columnas obligatorias: %s. Usa la plantilla.") % ", ".join(faltan))

        def cell(row, name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        def txt(v):
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        def num(v):
            if v is None or (isinstance(v, str) and not v.strip()):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).strip().replace(" ", "").replace(",", "."))
            except ValueError:
                return "ERROR"

        Product = self.env["product.product"]
        icbper_tax = self._l10n_pe_ne_ensure_icbper_tax()

        def afe_bolsa_actuales(product):
            """(código de afectación, ¿tiene ICBPER?) de un producto ya existente. Afectación y
            bolsa viven en el MISMO campo (taxes_id); esto permite pisar solo la que el usuario trajo
            en el Excel sin borrar la otra (ver 'vacío = mantener')."""
            sale = product.taxes_id.filtered(lambda t: t.type_tax_use == "sale")
            tiene_icbper = bool(sale.filtered(lambda t: t.l10n_pe_edi_tax_code == "7152"))
            afe = sale.filtered(lambda t: t.l10n_pe_edi_tax_code and t.l10n_pe_edi_tax_code != "7152")[:1]
            return (afe.l10n_pe_edi_tax_code or "1000"), tiene_icbper

        def tax_ids_de(afe_code, bolsa):
            tax = self._l10n_pe_ne_tax_by_code(afe_code)
            ids = list(tax.ids) if tax else []
            if bolsa:  # bolsa plástica → suma la tax ICBPER (monto fijo por unidad)
                ids += icbper_tax.ids
            return ids

        # get-or-create de categoría/subcategoría y marca por NOMBRE (case-insensitive), para que el
        # Excel reutilice lo que el negocio ya tiene y no duplique por tipeo. Solo se invoca al
        # COMMIT (el dry-run no crea nada). Cache por archivo para no repetir búsquedas/altas.
        Category = self.env["product.category"]
        Marca = self.env["l10n_pe_ne.marca"]
        _cat_cache = {}
        _marca_cache = {}

        def categ_id_de(categoria, subcategoria):
            """categ_id más específico (subcategoría si vino; si no, la categoría) bajo la raíz
            'Supermercado', creando lo que falte. None si no vino ninguna."""
            categoria = (categoria or "").strip()
            subcategoria = (subcategoria or "").strip()
            if not categoria and not subcategoria:
                return None
            key = (categoria.lower(), subcategoria.lower())
            if key in _cat_cache:
                return _cat_cache[key]
            root = Category._l10n_pe_ne_root()
            dep = None
            if categoria:
                dep = Category.search(
                    [("name", "=ilike", categoria), ("parent_id", "=", root.id)], limit=1)
                if not dep:
                    dep = Category.browse(Category._l10n_pe_ne_crear_bajo_super(categoria)["id"])
            parent = dep or root
            target = dep
            if subcategoria:
                sub = Category.search(
                    [("name", "=ilike", subcategoria), ("parent_id", "=", parent.id)], limit=1)
                if not sub:
                    sub = Category.browse(Category._l10n_pe_ne_crear_bajo_super(
                        subcategoria, parent_id=dep.id if dep else None)["id"])
                target = sub
            res = target.id if target else None
            _cat_cache[key] = res
            return res

        def marca_id_de(marca):
            marca = (marca or "").strip()
            if not marca:
                return None
            k = marca.lower()
            if k not in _marca_cache:
                _marca_cache[k] = Marca.l10n_pe_ne_crear_marca(marca)["id"]
            return _marca_cache[k]

        def precio_con_igv(precio, incluye_igv, afe_code):
            """El catálogo guarda list_price CON IGV. Si el usuario declaró el precio SIN IGV
            (INCLUYE IGV = No), se le suma el 18% solo cuando la afectación es GRAVADO."""
            if precio is None:
                return None
            if incluye_igv or afe_code != "1000":
                return precio
            return round(precio * 1.18, 6)

        # Procesamiento POR LOTES (evita el 504 del gateway con archivos grandes): el front sube en
        # tandas con offset/limit y agrega el reporte; `totalFilas` le dice cuántas quedan. Cada lote
        # es una transacción atómica y, como el UPSERT es por CÓDIGO, reintentar un lote no duplica.
        data_rows = rows[1:]
        total_filas = len(data_rows)
        offset = max(0, offset)
        fin = (offset + window) if window > 0 else total_filas
        sub = data_rows[offset:fin]

        # Prefetch del lote en 2 consultas (no una por fila): existentes por código y por barcode.
        cods_lote = {txt(cell(r, "codigo")) for r in sub if txt(cell(r, "codigo"))}
        bars_lote = {txt(cell(r, "codigo de barras")) for r in sub if txt(cell(r, "codigo de barras"))}
        pre_cod = {p.default_code: p for p in Product.with_context(active_test=False).search(
            [("default_code", "in", list(cods_lote))])} if cods_lote else {}
        pre_bar = {p.barcode: p for p in Product.search(
            [("barcode", "in", list(bars_lote))])} if bars_lote else {}

        creados = actualizados = 0
        errores = []
        avisos = []
        vistos = {}  # CÓDIGO → fila donde apareció por primera vez (duplicados DENTRO del lote)
        # Aviso (no bloqueante) de NOMBRE repetido: se calcula sobre TODO el archivo y se reporta UNA
        # vez (en el primer lote). Nombres iguales con CÓDIGO distinto = productos DISTINTOS por diseño;
        # solo se avisa por si fue sin querer (higiene de catálogo).
        if offset == 0:
            from collections import Counter
            noms = Counter(txt(cell(r, "nombre")) for r in data_rows if txt(cell(r, "nombre")))
            reps = sorted(nom for nom, c in noms.items() if c > 1)
            if reps:
                avisos.append({"fila": 0, "msg": "%d nombre(s) se repiten en el archivo (p. ej. %s) — se importan como productos DISTINTOS por tener código distinto; revisa si fue intencional" % (len(reps), ", ".join(reps[:3]))})
        procesadas = len(sub)  # cuántas filas del tramo se consumieron (para nextOffset)
        for k, row in enumerate(sub):
            n = k + 2 + offset  # número de fila real en la hoja (1 = cabecera)
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            cod = txt(cell(row, "codigo"))
            nombre = txt(cell(row, "nombre"))
            if not cod:
                errores.append({"fila": n, "msg": "Falta el CÓDIGO"})
                continue
            if not nombre:
                errores.append({"fila": n, "msg": "Falta el NOMBRE"})
                continue
            precio = num(cell(row, "precio venta"))
            costo = num(cell(row, "costo"))
            if precio == "ERROR" or costo == "ERROR":
                errores.append({"fila": n, "msg": "PRECIO o COSTO no es un número válido"})
                continue
            # UNIDAD (código SUNAT cat. 03). Vacía = NO se toca al actualizar; al CREAR, el default
            # se alinea al tipo: servicio → ZZ, si no → NIU (así Odoo y SUNAT no se contradicen).
            tipo = TIPO_IMPORT.get(norm(cell(row, "tipo")), "")  # "" | "bien" | "servicio"
            uni_raw = norm(cell(row, "unidad"))
            uni_provisto = bool(uni_raw)
            if not uni_raw:
                unidad = "ZZ" if tipo == "servicio" else "NIU"
            elif uni_raw in UNIDAD_IMPORT:
                unidad = UNIDAD_IMPORT[uni_raw]
            elif uni_raw.upper() in _UNIDAD_CODES:
                unidad = uni_raw.upper()
            else:
                unidad = "NIU"
                avisos.append({"fila": n, "msg": "Unidad '%s' no reconocida, se usó UNIDAD (NIU)" % txt(cell(row, "unidad"))})
            afe_raw = norm(cell(row, "afectacion"))
            afe_provisto = bool(afe_raw)
            if afe_provisto and afe_raw not in AFECT_IMPORT:
                avisos.append({"fila": n, "msg": "Afectación '%s' no reconocida, se usó GRAVADO" % txt(cell(row, "afectacion"))})
            afe_code = AFECT_IMPORT.get(afe_raw, "1000")
            bolsa_raw = norm(cell(row, "bolsa"))
            bolsa_provisto = bool(bolsa_raw)
            bolsa = bolsa_raw in ("si", "s")  # ICBPER: SI/NO
            # DETRACCIÓN: el desplegable trae "027 · Servicio…"; se toma solo el código (los 3
            # primeros dígitos). También acepta el código pelado "027" (plantilla vieja / a mano).
            detra_cell = txt(cell(row, "detraccion"))
            detra_provisto = bool(detra_cell)
            m_detra = re.match(r"^\s*(\d{3})\b", detra_cell)
            detra_raw = m_detra.group(1) if m_detra else detra_cell
            if detra_provisto and not re.fullmatch(r"[0-9]{3}", detra_raw):
                errores.append({"fila": n, "msg": "DETRACCIÓN inválida: elige del desplegable (catálogo 54) o escribe el código de 3 dígitos (ej. 027)"})
                continue
            barcode = txt(cell(row, "codigo de barras"))
            # ── Columnas de la plantilla completa (todas opcionales; "provisto" = trae valor) ──
            categoria = txt(cell(row, "categoria"))
            subcategoria = txt(cell(row, "subcategoria"))
            marca = txt(cell(row, "marca"))
            ctrl_raw = norm(cell(row, "controla stock"))
            ctrl_provisto = bool(ctrl_raw)
            controla_stock = ctrl_raw in ("si", "s", "sí", "true", "1", "x")
            ras_raw = norm(cell(row, "rastreo"))
            ras_provisto = bool(ras_raw)
            if "lote" in ras_raw or "lot" in ras_raw:
                rastreo_val = "lote"
            elif "serie" in ras_raw or "serial" in ras_raw or "imei" in ras_raw:
                rastreo_val = "serie"
            else:
                rastreo_val = ""  # sin rastreo / ninguno
            s_inicial = num(cell(row, "stock inicial"))
            s_minimo = num(cell(row, "stock minimo"))
            if s_inicial == "ERROR" or s_minimo == "ERROR":
                errores.append({"fila": n, "msg": "STOCK INICIAL o STOCK MÍNIMO no es un número válido"})
                continue
            act_raw = norm(cell(row, "activo"))
            act_provisto = bool(act_raw)
            activo = act_raw not in ("no", "n", "false", "0")  # vacío/ Sí → activo
            igv_raw = norm(cell(row, "incluye igv"))
            incluye_igv = igv_raw not in ("no", "n", "false", "0")  # vacío = Sí (con IGV)
            precio = precio_con_igv(precio, incluye_igv, afe_code)

            existing = pre_cod.get(cod) or Product.browse()
            # El código de barras no puede pertenecer a OTRO producto (Odoo lo exige único).
            if barcode:
                dup = pre_bar.get(barcode)
                if dup and dup.id != existing.id:
                    errores.append({"fila": n, "msg": "El código de barras '%s' ya pertenece a otro producto" % barcode})
                    continue
            # CÓDIGO repetido dentro del MISMO archivo: la última fila manda (se aplica igual), pero
            # se avisa y NO se cuenta dos veces (el preview marcaría 2 'nuevos' para un solo producto).
            repetido = cod in vistos
            if repetido:
                avisos.append({"fila": n, "msg": "CÓDIGO '%s' repetido (ya venía en la fila %d); vale la última fila" % (cod, vistos[cod])})
            else:
                vistos[cod] = n
            if not commit:
                if not repetido:
                    actualizados += 1 if existing else 0
                    creados += 0 if existing else 1
                continue
            if existing:
                # Actualización: "vacío = mantener". Solo se pisa lo que el usuario TRAJO con valor;
                # afectación y bolsa comparten campo, así que la que no venga se lee del producto.
                vals = {"name": nombre}
                if uni_provisto:
                    vals["l10n_pe_ne_unit_code"] = unidad
                if tipo:
                    vals["type"] = self._l10n_pe_ne_tipo_producto(tipo)
                if detra_provisto:
                    vals["l10n_pe_ne_detraccion_cod"] = detra_raw
                if precio is not None:
                    vals["list_price"] = precio
                if costo is not None:
                    vals["standard_price"] = costo
                if barcode:
                    vals["barcode"] = barcode
                if afe_provisto or bolsa_provisto:
                    cur_afe, cur_bolsa = afe_bolsa_actuales(existing)
                    vals["taxes_id"] = [(6, 0, tax_ids_de(
                        afe_code if afe_provisto else cur_afe,
                        bolsa if bolsa_provisto else cur_bolsa))]
                if ctrl_provisto:
                    vals["is_storable"] = controla_stock
                if ras_provisto:
                    vals["tracking"] = self._l10n_pe_ne_rastreo_producto(rastreo_val)
                if s_minimo is not None:
                    vals["l10n_pe_ne_stock_minimo"] = s_minimo
                if act_provisto:
                    vals["active"] = activo
                cat_id = categ_id_de(categoria, subcategoria)
                if cat_id:
                    vals["categ_id"] = cat_id
                mid = marca_id_de(marca)
                if mid:
                    vals["l10n_pe_ne_marca_id"] = mid
                existing.write(vals)
                # STOCK INICIAL solo tiene sentido al CREAR (fijaría el stock, pisando el real).
                if s_inicial is not None and s_inicial > 0:
                    avisos.append({"fila": n, "msg": "STOCK INICIAL ignorado: '%s' ya existe (usa Ajustar stock)" % cod})
                if not repetido:
                    actualizados += 1
            else:
                # Alta: se aplican los defaults. El tipo lo manda la columna TIPO si vino; si no, se
                # deduce de la unidad (ZZ → servicio, resto → bien), la señal que sí trae la fila.
                vals = {"name": nombre, "default_code": cod, "l10n_pe_ne_unit_code": unidad,
                        "type": self._l10n_pe_ne_tipo_producto(tipo or None, unidad),
                        "taxes_id": [(6, 0, tax_ids_de(afe_code, bolsa))],
                        "sale_ok": True, "company_id": self.env.company.id}
                if precio is not None:
                    vals["list_price"] = precio
                if costo is not None:
                    vals["standard_price"] = costo
                if detra_provisto:
                    vals["l10n_pe_ne_detraccion_cod"] = detra_raw
                if barcode:
                    vals["barcode"] = barcode
                if ctrl_provisto:
                    vals["is_storable"] = controla_stock
                if ras_provisto:
                    vals["tracking"] = self._l10n_pe_ne_rastreo_producto(rastreo_val)
                if s_minimo is not None:
                    vals["l10n_pe_ne_stock_minimo"] = s_minimo
                if act_provisto:
                    vals["active"] = activo
                cat_id = categ_id_de(categoria, subcategoria)
                if cat_id:
                    vals["categ_id"] = cat_id
                mid = marca_id_de(marca)
                if mid:
                    vals["l10n_pe_ne_marca_id"] = mid
                p = Product.create(vals)
                # Registra el recién creado en la cache del lote: si el mismo CÓDIGO/barcode vuelve
                # a aparecer en este lote, se ACTUALIZA (no se crea un duplicado).
                pre_cod[cod] = p
                if barcode:
                    pre_bar[barcode] = p
                # Existencia inicial: solo si el producto lleva stock y vino > 0 (reusa el motor de
                # ajuste 'fijar', que deja el stock EN esa cantidad).
                if s_inicial and s_inicial > 0 and vals.get("is_storable"):
                    self._l10n_pe_ne_ajustar_stock(p.id, "fijar", s_inicial, _("Existencia inicial (importación)"))
                if not repetido:
                    creados += 1
            # Presupuesto de tiempo: corta el lote tras una fila completa si se pasó, para que NINGUNA
            # request se acerque al timeout del gateway aunque las filas sean caras (stock, categorías).
            if budget and (time.time() - t_ini) > budget:
                procesadas = k + 1
                break
        next_offset = offset + procesadas
        return {"commit": commit, "creados": creados, "actualizados": actualizados,
                "errores": errores, "avisos": avisos,
                "totalOk": creados + actualizados, "totalError": len(errores),
                "totalFilas": total_filas, "nextOffset": next_offset,
                "done": next_offset >= total_filas}

    # ------------------------------------------------------- importación clientes
    @api.model
    def l10n_pe_ne_plantilla_clientes(self):
        """Plantilla xlsx para importar/actualizar clientes (los campos del caso común de
        facturación). Hoja 'Clientes' con cabeceras + ejemplos + desplegable de tipo de documento,
        y una hoja 'Instrucciones'. Devuelve {filename, contentB64}."""
        import io
        import base64
        import xlsxwriter

        headers = ["TIPO DE DOCUMENTO", "NÚMERO DE DOCUMENTO", "RAZÓN SOCIAL O NOMBRE",
                   "EMAIL", "TELÉFONO", "DIRECCIÓN"]
        col = {h: i for i, h in enumerate(headers)}
        ejemplos = [
            ["RUC", "20123456789", "COMERCIAL LOS ANDES S.A.C.", "ventas@losandes.pe", "01 555 1234", "Av. Industrial 123, Lima"],
            ["DNI", "12345678", "JUAN PÉREZ QUISPE", "", "987654321", ""],
            ["Carné de extranjería", "001234567", "MARÍA GARCÍA", "maria@correo.com", "", "Calle Las Flores 456"],
        ]
        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {"in_memory": True})
        ws = wb.add_worksheet("Clientes")
        head = wb.add_format({"bold": True, "bg_color": "#2563eb", "font_color": "white", "border": 1})
        # El número de documento va como TEXTO: los DNI/CE pueden empezar con cero y Excel se los
        # comería, y un RUC de 11 dígitos se iría a notación científica en formato numérico.
        txtfmt = wb.add_format({"num_format": "@"})
        for c, h in enumerate(headers):
            ws.write(0, c, h, head)
            ws.set_column(c, c, max(16, len(h) + 4), txtfmt if h == "NÚMERO DE DOCUMENTO" else None)
        for r, row in enumerate(ejemplos, 1):
            ws.write_row(r, 0, row)
        note = {"x_scale": 2.4, "y_scale": 1.8, "author": "Ekipu"}
        ws.write_comment(0, col["TIPO DE DOCUMENTO"], (
            "Elige del desplegable: RUC, DNI, Carné de extranjería o Pasaporte. Si lo dejas vacío "
            "se deduce del NÚMERO (11 dígitos → RUC, 8 → DNI)."), note)
        ws.write_comment(0, col["NÚMERO DE DOCUMENTO"], (
            "Obligatorio. Es la CLAVE: si ya existe un cliente con ese documento, se ACTUALIZA; si "
            "no, se CREA. RUC = 11 dígitos, DNI = 8."), note)
        ws.write_comment(0, col["RAZÓN SOCIAL O NOMBRE"], "Obligatorio. La razón social (empresa) o el nombre de la persona.", note)
        ws.write_comment(0, col["DIRECCIÓN"], "Opcional. Domicilio del cliente (se usa en el comprobante).", note)
        # Desplegable de tipo de documento.
        ws.data_validation(1, col["TIPO DE DOCUMENTO"], 1005, col["TIPO DE DOCUMENTO"], {
            "validate": "list", "source": ["RUC", "DNI", "Carné de extranjería", "Pasaporte"],
            "error_type": "information", "input_title": "Tipo de documento",
            "input_message": "RUC (empresa, 11 díg.), DNI (persona, 8 díg.), Carné de extranjería o "
                             "Pasaporte. Vacío = se deduce del número."})
        ws.freeze_panes(1, 0)
        wi = wb.add_worksheet("Instrucciones")
        wi.set_column(0, 0, 115)
        for r, line in enumerate([
            "Ekipu — Plantilla de importación de clientes",
            "",
            "Obligatorios: TIPO DE DOCUMENTO (o se deduce del número), NÚMERO DE DOCUMENTO y RAZÓN SOCIAL O NOMBRE.",
            "",
            "1. NÚMERO DE DOCUMENTO es la clave: si ya existe un cliente con ese documento se ACTUALIZA; si no, se CREA.",
            "   Al ACTUALIZAR, una celda vacía MANTIENE el valor actual (no lo borra).",
            "2. TIPO DE DOCUMENTO: elige del desplegable. Vacío = se deduce (11 dígitos → RUC, 8 → DNI).",
            "3. RUC = 11 dígitos; DNI = 8 dígitos. Carné de extranjería / Pasaporte: tal como figura.",
            "4. EMAIL, TELÉFONO y DIRECCIÓN son opcionales. La dirección se usa en el comprobante.",
            "5. Sube el archivo, revisa el resumen (nuevos / actualizados / errores) y recién ahí confirma.",
        ]):
            wi.write(r, 0, line)
        wb.close()
        return {"filename": "plantilla-clientes-ekipu.xlsx",
                "contentB64": base64.b64encode(buf.getvalue()).decode("ascii")}

    @api.model
    def l10n_pe_ne_importar_clientes(self, payload):
        """Compat single-shot. UPSERT por NÚMERO DE DOCUMENTO (vat); commit=False = dry-run. El
        camino rápido por lotes es preparar()+lote()."""
        payload = payload or {}
        rows = self._l10n_pe_ne_leer_xlsx(payload.get("contentB64"), "Clientes")
        return self._l10n_pe_ne_importar_clientes_core(
            rows, bool(payload.get("commit")), offset=int(payload.get("offset") or 0),
            window=int(payload.get("limit") or 0), budget=0.0)

    def _l10n_pe_ne_importar_clientes_core(self, rows, commit, offset=0, window=0, budget=0.0):
        """Núcleo del import de clientes sobre filas YA parseadas (ver _importar_productos_core)."""
        import time
        import unicodedata
        t_ini = time.time()

        def norm(h):
            s = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode("ascii")
            return " ".join(s.lower().split())

        header = [norm(h) for h in rows[0]]
        idx = {h: i for i, h in enumerate(header) if h}
        faltan = [etq for etq, k in (("NÚMERO DE DOCUMENTO", "numero de documento"),
                                     ("RAZÓN SOCIAL O NOMBRE", "razon social o nombre")) if k not in idx]
        if faltan:
            raise UserError(_("Faltan columnas obligatorias: %s. Usa la plantilla.") % ", ".join(faltan))

        def cell(row, name):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        def txt(v):
            if v is None:
                return ""
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        # Etiqueta del desplegable / código del catálogo → l10n_pe_vat_code (6 RUC, 1 DNI, 4 CE, 7 pas).
        TDOC = {
            "ruc": "6", "dni": "1", "pasaporte": "7", "carne de extranjeria": "4",
            "carnet de extranjeria": "4", "ce": "4", "6": "6", "1": "1", "4": "4", "7": "7",
        }
        Partner = self.env["res.partner"]
        company = self.env.company
        # Procesamiento POR LOTES (evita el 504): el front sube en tandas con offset/limit; `totalFilas`
        # le dice cuántas quedan. UPSERT por documento → reintentar un lote no duplica.
        data_rows = rows[1:]
        total_filas = len(data_rows)
        offset = max(0, offset)
        fin = (offset + window) if window > 0 else total_filas
        sub = data_rows[offset:fin]
        # Prefetch de existentes por documento (vat) en una sola consulta.
        docs_lote = {txt(cell(r, "numero de documento")) for r in sub if txt(cell(r, "numero de documento"))}
        pre_doc = {p.vat: p for p in Partner.with_context(active_test=False).search(
            [("vat", "in", list(docs_lote)), ("company_id", "in", (False, company.id))])} if docs_lote else {}
        creados = actualizados = 0
        errores = []
        avisos = []
        vistos = {}
        procesadas = len(sub)
        for k, row in enumerate(sub):
            n = k + 2 + offset  # número de fila real en la hoja
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            num = txt(cell(row, "numero de documento"))
            nombre = txt(cell(row, "razon social o nombre"))
            if not num:
                errores.append({"fila": n, "msg": "Falta el NÚMERO DE DOCUMENTO"})
                continue
            if not nombre:
                errores.append({"fila": n, "msg": "Falta la RAZÓN SOCIAL O NOMBRE"})
                continue
            code = TDOC.get(norm(cell(row, "tipo de documento")), "")
            if not code:  # deducir del número si no vino el tipo
                if re.fullmatch(r"\d{11}", num):
                    code = "6"
                elif re.fullmatch(r"\d{8}", num):
                    code = "1"
                else:
                    errores.append({"fila": n, "msg": "Indica el TIPO DE DOCUMENTO (RUC/DNI/Carné de extranjería/Pasaporte)"})
                    continue
            if code == "6" and not re.fullmatch(r"\d{11}", num):
                errores.append({"fila": n, "msg": "El RUC debe tener 11 dígitos"})
                continue
            if code == "1" and not re.fullmatch(r"\d{8}", num):
                errores.append({"fila": n, "msg": "El DNI debe tener 8 dígitos"})
                continue
            repetido = num in vistos
            if repetido:
                avisos.append({"fila": n, "msg": "Documento '%s' repetido (ya venía en la fila %d); vale la última fila" % (num, vistos[num])})
            else:
                vistos[num] = n
            existing = pre_doc.get(num) or Partner.browse()
            if not commit:
                if not repetido:
                    actualizados += 1 if existing else 0
                    creados += 0 if existing else 1
                continue
            # "vacío = mantener": solo se envían a _partner_apply los campos que trajeron valor.
            c = {"razonSocial": nombre, "numDoc": num, "tipoDoc": code}
            for key, name in (("email", "email"), ("telefono", "telefono"), ("direccion", "direccion")):
                v = txt(cell(row, name))
                if v:
                    c[key] = v
            if existing:
                self._l10n_pe_ne_partner_apply(existing, c)
                if not existing.customer_rank:
                    existing.customer_rank = 1
                if not repetido:
                    actualizados += 1
            else:
                t = self._l10n_pe_ne_ident_type(code)
                p = Partner.create({
                    "name": nombre, "vat": num, "customer_rank": 1, "company_id": company.id,
                    "l10n_latam_identification_type_id": t.id if t else False,
                })
                self._l10n_pe_ne_partner_apply(p, {k: v for k, v in c.items()
                                                   if k in ("email", "telefono", "direccion")})
                pre_doc[num] = p  # el mismo documento repetido en el lote ACTUALIZA, no duplica
                if not repetido:
                    creados += 1
            if budget and (time.time() - t_ini) > budget:
                procesadas = k + 1
                break
        next_offset = offset + procesadas
        return {"commit": commit, "creados": creados, "actualizados": actualizados,
                "errores": errores, "avisos": avisos,
                "totalOk": creados + actualizados, "totalError": len(errores),
                "totalFilas": total_filas, "nextOffset": next_offset,
                "done": next_offset >= total_filas}

    # ------------------------------------------------------- lotes: parse-once + tiempo
    # Ventana de prefetch y presupuesto de tiempo por lote. La ventana acota cuántas filas se
    # prefetchean/consideran de una; el presupuesto corta el lote por tiempo (robustez ante filas
    # caras: ninguna request se acerca al timeout del gateway). El front avanza por nextOffset.
    _IMPORT_WINDOW = 800
    _IMPORT_BUDGET_S = 18.0

    @api.model
    def l10n_pe_ne_import_preparar(self, payload):
        """Sube el archivo UNA vez: lo parsea, valida cabeceras y guarda las filas bajo un token.
        payload = {kind: 'productos'|'clientes', contentB64}. Devuelve {token, totalFilas}."""
        payload = payload or {}
        kind = payload.get("kind")
        if kind not in ("productos", "clientes"):
            raise UserError(_("Tipo de importación no soportado."))
        hoja, oblig = ("Productos", ("codigo", "nombre")) if kind == "productos" else \
            ("Clientes", ("numero de documento", "razon social o nombre"))
        rows = self._l10n_pe_ne_leer_xlsx(payload.get("contentB64"), hoja)
        import unicodedata

        def norm(h):
            s = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode("ascii")
            return " ".join(s.lower().split())
        idx = {norm(h) for h in rows[0] if norm(h)}
        faltan = [k for k in oblig if k not in idx]
        if faltan:
            raise UserError(_("Faltan columnas obligatorias en la plantilla. Usa la plantilla de %s.") % kind)
        token = self.env["l10n_pe_ne.import.session"]._crear(kind, rows)
        return {"token": token, "totalFilas": max(0, len(rows) - 1)}

    @api.model
    def l10n_pe_ne_import_lote(self, payload):
        """Procesa UN lote del archivo ya subido. payload = {token, commit, offset}. Corta por
        ventana/tiempo y devuelve el reporte + nextOffset/done; el front repite hasta done."""
        payload = payload or {}
        kind, rows, _total = self.env["l10n_pe_ne.import.session"]._rows(payload.get("token"))
        commit = bool(payload.get("commit"))
        offset = int(payload.get("offset") or 0)
        core = (self._l10n_pe_ne_importar_productos_core if kind == "productos"
                else self._l10n_pe_ne_importar_clientes_core)
        return core(rows, commit, offset=offset, window=self._IMPORT_WINDOW, budget=self._IMPORT_BUDGET_S)

